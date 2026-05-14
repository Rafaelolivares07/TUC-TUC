"""
server.py — Servidor Flask de reportes multi-origen
Puerto: 5003 (no pisa el administrator_web.py actual en :5002)

Uso:
    python server.py

Rutas:
    GET  /                          → menú de reportes
    GET  /reporte/<id>              → UI del reporte
    POST /api/reporte/<id>          → ejecutar reporte, retorna JSON
    GET  /api/empresas              → lista de empresas (fuente local)
"""

import os, sys, datetime, time, io, configparser
sys.path.insert(0, os.path.dirname(__file__))

from flask import Flask, render_template, request, jsonify, send_file, session
from data_layer import DataLayer, leer_ruta_bd
from reportes import CATALOGO

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

app = Flask(__name__, template_folder='templates')
app.secret_key = 'sar-reportes-2026'

# ── Configuración AWS ─────────────────────────────────────────────────────────
_cfg = configparser.ConfigParser()
_cfg.read(os.path.join(os.path.dirname(__file__), 'server.ini'), encoding='utf-8')
BASE_URL_AWS = _cfg.get('aws', 'base_url',   fallback='https://admin.tuc-tuc.co')
AWS_USUARIO  = _cfg.get('aws', 'usuario',   fallback='')
AWS_PASSWORD = _cfg.get('aws', 'password',  fallback='')
CLIENTE_ID   = _cfg.get('aws', 'cliente_id', fallback='').strip()

_aws_session     = None   # requests.Session con cookie de AWS reutilizable
_aws_session_ok  = False  # True después de login exitoso
_permisos_cache  = None   # {fuente: set(reporte_id)}
_permisos_ts     = 0.0    # timestamp de la última carga
_PERMISOS_TTL    = 30     # segundos antes de refrescar


def _cargar_permisos(cliente_id=None):
    """Carga permisos desde AWS para cliente_id dado (o CLIENTE_ID del ini).
    Sin cliente_id → sin restricciones. Se cachea TTL segundos."""
    global _permisos_cache, _permisos_ts
    cid = (cliente_id or CLIENTE_ID).strip()
    if not cid:
        _permisos_cache = None
        _permisos_ts    = time.time()
        return
    try:
        r = _aws_get(f'/api/admin-agent/reportes-permisos/{cid}', timeout=8)
        data = r.json()
        if data.get('ok'):
            cache = {'local': set(), 'remoto': set()}
            for p in data.get('permisos', []):
                fuente = p.get('fuente', 'local')
                if fuente in cache:
                    cache[fuente].add(p['reporte_id'])
            _permisos_cache = cache
            _permisos_ts    = time.time()
    except Exception:
        _permisos_cache = None  # si falla AWS, no bloquear — mostrar todo


def _permisos_vigentes(cliente_id=None):
    """Recarga si el cache expiró o el cliente_id cambia."""
    if time.time() - _permisos_ts > _PERMISOS_TTL:
        _cargar_permisos(cliente_id)
    return _permisos_cache

def _get_aws_session():
    """Retorna una requests.Session autenticada en AWS. Re-login si la cookie expiró."""
    import requests as req
    global _aws_session, _aws_session_ok
    if _aws_session and _aws_session_ok:
        return _aws_session
    s = req.Session()
    s.post(f'{BASE_URL_AWS}/admin/login',
           data={'usuario': AWS_USUARIO, 'password': AWS_PASSWORD},
           timeout=10, allow_redirects=True)
    _aws_session_ok = 'session' in s.cookies
    _aws_session = s
    return s


def _aws_get(path, timeout=10):
    """GET autenticado a AWS. Re-login automático si devuelve 401."""
    global _aws_session_ok
    for _ in range(2):
        s = _get_aws_session()
        r = s.get(f'{BASE_URL_AWS}{path}', timeout=timeout)
        if r.status_code == 401:
            _aws_session_ok = False  # forzar re-login en siguiente intento
            continue
        return r
    raise RuntimeError('No se pudo autenticar en AWS')


# ── Helpers ───────────────────────────────────────────────────────────────────

def _get_data_layer():
    """Construye DataLayer según parámetros de la request."""
    fuente     = request.args.get('fuente', 'local')
    cliente_id = request.args.get('cliente_id', '').strip()
    if fuente == 'remoto':
        return DataLayer(fuente='remoto', base_url=BASE_URL_AWS,
                         cliente_id=cliente_id,
                         session_token=request.cookies.get('session'))
    ruta_bd = leer_ruta_bd()
    return DataLayer(fuente='local', ruta_bd=ruta_bd)


def _get_filtros():
    """Extrae filtros comunes del request (GET o POST JSON)."""
    if request.method == 'POST':
        data = request.get_json() or {}
    else:
        data = request.args.to_dict()
    return data


# ── Fuente de datos (sesión) ──────────────────────────────────────────────────

@app.route('/api/fuente', methods=['POST'])
def set_fuente():
    data = request.get_json() or {}
    session['fuente']     = data.get('fuente', 'local')
    session['cliente_id'] = data.get('cliente_id', '')
    return jsonify({'ok': True})

@app.route('/api/agentes')
def api_agentes():
    """Proxy hacia AWS — retorna agentes conectados."""
    try:
        r = _aws_get('/api/admin-agent/agentes', timeout=8)
        return jsonify(r.json())
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e), 'agentes': []})


# ── Rutas principales ─────────────────────────────────────────────────────────

def _catalogo_filtrado(fuente, cliente_id=None):
    """Devuelve el catálogo filtrado según permisos. Sin cliente_id efectivo → todo visible."""
    cache = _permisos_vigentes(cliente_id)
    if cache is None:
        return CATALOGO
    permitidos = cache.get(fuente, set())
    return {k: v for k, v in CATALOGO.items() if k in permitidos}


@app.route('/')
def menu():
    from collections import defaultdict
    fuente     = session.get('fuente', 'local')
    cliente_id = session.get('cliente_id', '') or CLIENTE_ID
    catalogo   = _catalogo_filtrado(fuente, cliente_id)
    categorias = defaultdict(list)
    for r in catalogo.values():
        categorias[getattr(r, 'CATEGORIA', 'General')].append(
            {'id': r.ID, 'nombre': r.NOMBRE}
        )
    return render_template('menu.html', categorias=dict(categorias),
                           fuente=fuente,
                           cliente_id=session.get('cliente_id', ''))


@app.route('/reporte/<reporte_id>')
def ver_reporte(reporte_id):
    if reporte_id not in CATALOGO:
        return 'Reporte no encontrado', 404
    try:
        empresas = _get_empresas_local()
    except Exception:
        empresas = []
    hoy   = datetime.date.today()
    desde = hoy.replace(day=1).isoformat()
    hasta = hoy.isoformat()
    template = f'reporte_{reporte_id}.html'
    return render_template(template, empresas=empresas, desde=desde, hasta=hasta)


# ── API datos ─────────────────────────────────────────────────────────────────

@app.route('/api/reporte/<reporte_id>', methods=['POST'])
def api_reporte(reporte_id):
    if reporte_id not in CATALOGO:
        return jsonify({'ok': False, 'error': 'Reporte no encontrado'}), 404
    modulo  = CATALOGO[reporte_id]
    filtros = request.get_json() or {}
    t0      = time.time()
    try:
        dl     = _get_data_layer()
        tablas = modulo.tablas_requeridas(filtros)
        datos  = dl.leer(tablas)
        rows   = modulo.calcular(datos, filtros)
        elapsed = round(time.time() - t0, 1)
        return jsonify({'ok': True, 'rows': rows, 'elapsed': elapsed,
                        'total': len(rows)})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})


@app.route('/api/reporte/<reporte_id>/excel', methods=['POST'])
def api_reporte_excel(reporte_id):
    if not HAS_OPENPYXL:
        return 'openpyxl no instalado', 500
    if reporte_id not in CATALOGO:
        return 'Reporte no encontrado', 404
    modulo  = CATALOGO[reporte_id]
    filtros = request.get_json() or {}
    try:
        dl     = _get_data_layer()
        tablas = modulo.tablas_requeridas(filtros)
        datos  = dl.leer(tablas)
        rows   = modulo.calcular(datos, filtros)
        buf    = _generar_excel(reporte_id, rows, filtros)
        fname  = f'{reporte_id}_{datetime.date.today()}.xlsx'
        return send_file(buf, as_attachment=True, download_name=fname,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        return str(e), 500


@app.route('/api/empresas')
def api_empresas():
    try:
        empresas = _get_empresas_local()
        return jsonify({'ok': True, 'empresas': empresas})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e), 'empresas': []})


# ── Helpers internos ──────────────────────────────────────────────────────────

def _get_empresas_local():
    """Muestrea PROD_FACT1 para obtener códigos de empresa únicos."""
    import struct
    from agente import leer_header
    ENC     = 'cp1252'
    ruta_bd = leer_ruta_bd()
    path    = os.path.join(ruta_bd, 'PROD_FACT1.DBF')
    num, hsz, rec_size, campos = leer_header(path)
    campos_map = {c['nombre']: c for c in campos}
    emp_c = campos_map.get('EMPRESA')
    if not emp_c:
        return []
    with open(path, 'rb') as f:
        f.seek(hsz)
        raw = f.read(num * rec_size)
    empresas = set()
    paso = max(1, num // 500)
    for i in range(0, num, paso):
        b = raw[i*rec_size:(i+1)*rec_size]
        if b[0] == 0x2A:
            continue
        o, n = emp_c['offset'], emp_c['longitud']
        emp = b[o:o+n].rstrip(b' ').decode(ENC, 'replace').strip()
        if emp:
            empresas.add(emp)
    return sorted(empresas)


def _generar_excel(reporte_id, rows, filtros):
    """Genera Excel básico para cualquier reporte."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = reporte_id
    hdr_font = Font(bold=True, color='FFFFFF')
    hdr_fill = PatternFill('solid', fgColor='003F7F')

    if not rows:
        ws.append(['Sin resultados'])
    else:
        cols = list(rows[0].keys())
        ws.append(cols)
        for cell in ws[1]:
            cell.font = hdr_font
            cell.fill = hdr_fill
        for r in rows:
            ws.append([r.get(c, '') for c in cols])

    ws.append([])
    ws.append([f'Generado: {datetime.datetime.now().strftime("%Y-%m-%d %H:%M")}'])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


if __name__ == '__main__':
    import socket as _socket
    if os.environ.get('WERKZEUG_RUN_MAIN') != 'true':
        _lock = _socket.socket(_socket.AF_INET, _socket.SOCK_STREAM)
        try:
            _lock.bind(('127.0.0.1', 47837))
            _lock.listen(1)
        except OSError:
            sys.exit(0)
        print('=' * 50)
        print('  Reportes SAR — localhost:5003')
        if CLIENTE_ID:
            print(f'  Cliente: {CLIENTE_ID} — permisos se cargan al primer acceso')
        else:
            print('  Modo Rafael — sin restricciones (cliente_id vacío en server.ini)')
        print('='  * 50)
    app.run(host='127.0.0.1', port=5003, debug=True, use_reloader=True)
