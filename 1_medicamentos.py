from flask import Flask, request, redirect, url_for,send_from_directory, jsonify, render_template, session, flash, send_file
import sqlalchemy
import pandas
import sqlite3
import psycopg2
import uuid
import json
import os
import re
from werkzeug.utils import secure_filename
import hashlib
import requests
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO
import traceback
#  IMPORTAR EL INICIALIZADOR DE DATOS EXTERNO
from data_initializer import initialize_full_db
from bs4 import BeautifulSoup
import time
import math
import unicodedata
import re
from dotenv import load_dotenv
import cloudinary
import cloudinary.uploader
from apscheduler.schedulers.background import BackgroundScheduler
from apscheduler.triggers.interval import IntervalTrigger
import atexit

# Cargar variables de entorno
load_dotenv()

# Configurar Cloudinary
cloudinary.config(
    cloud_name=os.getenv('CLOUDINARY_CLOUD_NAME'),
    api_key=os.getenv('CLOUDINARY_API_KEY'),
    api_secret=os.getenv('CLOUDINARY_API_SECRET')
)

# -------------------------------------------------------------------
# --- ZONA 1: CONFIGURACIN INICIAL Y CONEXIN A LA BASE DE DATOS ---
# -------------------------------------------------------------------
app = Flask(__name__)

#  Clave secreta requerida para firmar las cookies de sesin
app.config['SECRET_KEY'] = os.getenv('SECRET_KEY', 'mi_clave_secreta_para_probar_el_acceso_temporal')

#  Configurar carpeta de subidas
UPLOAD_FOLDER = os.path.join(app.root_path, 'static', 'uploads')
ALLOWED_EXT = {'png', 'jpg', 'jpeg', 'gif'}
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = app.config.get('MAX_CONTENT_LENGTH', 16 * 1024 * 1024)

# La migracin automtica se ejecutar manualmente va endpoint /api/run-migration-now
# porque necesita que get_db_connection() est definida primero



def calcular_precio_segun_politica(medicamento_id, fabricante_id, conn):
    """
    Calcula el precio de venta segn polticas configuradas basndose en cotizaciones.

    Retorna: precio_nuevo (float) o None si no se puede calcular
    """
    # Cargar configuracin
    config_row = conn.execute("SELECT * FROM CONFIGURACION_PRECIOS LIMIT 1").fetchone()
    if not config_row:
        return None

    CONFIG = dict(config_row)

    # Obtener cotizaciones ordenadas por precio
    cotizaciones = conn.execute("""
        SELECT precio FROM precios_competencia
        WHERE medicamento_id = ? AND fabricante_id = ?
        ORDER BY precio ASC
    """, (medicamento_id, fabricante_id)).fetchall()

    precios_cot = [c['precio'] for c in cotizaciones]
    num_cot = len(precios_cot)

    if num_cot == 0:
        return None  # Sin cotizaciones, no calcular precio

    precio_nuevo = 0

    if num_cot == 1:
        # 1 cotizacin: aplicar recargo_1_cotizacion
        precio_base = precios_cot[0]
        recargo = precio_base * (CONFIG['recargo_1_cotizacion'] / 100)
        precio_nuevo = precio_base + max(recargo, CONFIG['ganancia_min_escaso'])
        precio_nuevo = min(precio_nuevo, precio_base + CONFIG['ganancia_max_escaso'])

    elif num_cot == 2:
        # 2 cotizaciones: siempre usar la mayor (cot#2 es proveedor fijo potencial)
        precio_base = precios_cot[1]
        recargo = precio_base * (CONFIG['recargo_escaso'] / 100)
        precio_nuevo = precio_base + max(recargo, CONFIG['ganancia_min_escaso'])
        precio_nuevo = min(precio_nuevo, precio_base + CONFIG['ganancia_max_escaso'])

    else:  # 3 o ms cotizaciones - FRMULA UNIFICADA
        # Ordenadas: [cot1, cot2, cot3, ...]
        # ndices:    [0,    1,    2,   ...]

        cotizacion_2 = precios_cot[1]  # Segunda cotizacin (ndice 1)
        cotizacion_3 = precios_cot[2]  # Tercera cotizacin (ndice 2)

        descuento_competencia = CONFIG.get('descuento_competencia', 200)
        precio_domicilio = CONFIG.get('precio_domicilio', 5000)
        costo_operario_domicilio = CONFIG.get('costo_operario_domicilio', 3333)
        ganancia_min_escaso = CONFIG.get('ganancia_min_escaso', 1500)
        ganancia_max_escaso = CONFIG.get('ganancia_max_escaso', 50000)
        pedido_min_domicilio_gratis = CONFIG.get('pedido_min_domicilio_gratis', 50000)

        brecha2_3 = cotizacion_3 - cotizacion_2

        if brecha2_3 < descuento_competencia:
            precio_base = cotizacion_2
            precio_nuevo = cotizacion_2 + (brecha2_3 / 2)
        else:
            PRECIO_PARA_CALCULO = cotizacion_3 - descuento_competencia

            if PRECIO_PARA_CALCULO < pedido_min_domicilio_gratis:
                ingreso_domicilio = precio_domicilio
            else:
                ingreso_domicilio = 0

            DIFERENCIA_PARA_CALCULO = (PRECIO_PARA_CALCULO - cotizacion_2 + ingreso_domicilio - costo_operario_domicilio)

            if DIFERENCIA_PARA_CALCULO > ganancia_min_escaso:
                precio_nuevo = PRECIO_PARA_CALCULO
            else:
                DIFERENCIA_FINAL = ganancia_min_escaso - DIFERENCIA_PARA_CALCULO
                precio_nuevo = PRECIO_PARA_CALCULO + DIFERENCIA_FINAL

            precio_base = cotizacion_3

        precio_nuevo = min(precio_nuevo, precio_base + ganancia_max_escaso)

    # Aplicar redondeo superior si est configurado
    redondeo_superior = CONFIG.get('redondeo_superior', 0)
    if redondeo_superior > 0:
        # Redondear hacia arriba al mltiplo de redondeo_superior
        import math
        precio_nuevo = math.ceil(precio_nuevo / redondeo_superior) * redondeo_superior

    return round(precio_nuevo)


def normalizar_texto(texto):
    """
    Normaliza texto: quita tildes, convierte a minsculas, quita caracteres especiales
    Ejemplos:
    - "Ibuprofn" -> "ibuprofeno"
    - "Acetaminofn" -> "acetaminofen"
    - "Diclofenaco-Sdico" -> "diclofenaco sodico"
    """
    if not texto:
        return ""

    # Convertir a minsculas
    texto = texto.lower()
    
    # Quitar tildes usando unicodedata
    texto = ''.join(
        c for c in unicodedata.normalize('NFD', texto)
        if unicodedata.category(c) != 'Mn'
    )
    
    # Reemplazar guiones y caracteres especiales por espacios
    texto = texto.replace('-', ' ').replace('_', ' ')
    
    # Quitar caracteres especiales excepto letras, nmeros y espacios
    texto = ''.join(c if c.isalnum() or c.isspace() else ' ' for c in texto)
    
    # Normalizar espacios mltiples
    texto = ' '.join(texto.split())
    
    return texto



@app.route('/admin/test_simple')
def test_simple():
    return "Funciona!"


# Asegurar que la carpeta exista
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
app.logger.debug(f"Upload folder: {app.config['UPLOAD_FOLDER']} (exist or created)")


#  Configuracin de sesiones para produccin en Render
# Usar sesiones del lado del cliente (cookies firmadas) para compatibilidad con mltiples workers
app.config['SESSION_PERMANENT'] = True
app.config['PERMANENT_SESSION_LIFETIME'] = 30 * 24 * 60 * 60  # 30 das en segundos
# SESSION_COOKIE_SECURE solo en produccin (HTTPS). En desarrollo local usar HTTP.
app.config['SESSION_COOKIE_SECURE'] = os.getenv('RENDER', None) is not None
app.config['SESSION_COOKIE_HTTPONLY'] = True  # Prevenir XSS
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'  # Proteccin CSRF

#  Nombre de la base de datos
DB_NAME = 'medicamentos.db'

class PostgreSQLRow:
    """Row object que simula sqlite3.Row - soporta acceso por ndice y por nombre"""
    def __init__(self, cursor, values):
        self._cursor = cursor
        self._values = values

    def __getitem__(self, key):
        if isinstance(key, int):
            return self._values[key]
        # Acceso por nombre de columna
        if self._cursor.description:
            for i, col in enumerate(self._cursor.description):
                if col[0] == key:
                    return self._values[i]
        raise KeyError(f"No such column: {key}")

    def keys(self):
        if self._cursor.description:
            return [col[0] for col in self._cursor.description]
        return []

    def __iter__(self):
        # Retornar iterator sobre las keys para que dict(row) funcione
        return iter(self.keys())

    def __len__(self):
        return len(self._values)


class PostgreSQLCursorWrapper:
    """Wrapper para cursor de PostgreSQL que simula lastrowid de SQLite"""
    def __init__(self, pg_cursor, last_insert_id=None):
        self._cursor = pg_cursor
        self.lastrowid = last_insert_id

    def fetchone(self):
        row = self._cursor.fetchone()
        if row is None:
            return None
        return PostgreSQLRow(self._cursor, row)

    def fetchall(self):
        rows = self._cursor.fetchall()
        return [PostgreSQLRow(self._cursor, row) for row in rows]

    def __getattr__(self, name):
        return getattr(self._cursor, name)


class PostgreSQLConnectionWrapper:
    """Wrapper para que PostgreSQL funcione como SQLite con conn.execute()"""
    def __init__(self, pg_conn):
        self._conn = pg_conn

    def execute(self, query, params=()):
        # Convertir ? a %s para PostgreSQL
        query = query.replace('?', '%s')

        # Convertir datetime('now') de SQLite a CURRENT_TIMESTAMP de PostgreSQL
        import re
        query = re.sub(r"datetime\(['\"]now['\"]\)", "CURRENT_TIMESTAMP", query, flags=re.IGNORECASE)

        # Tablas que estn en MAYSCULAS en PostgreSQL
        tablas_mayusculas = [
            'usuarios', 'medicamentos', 'sintomas', 'fabricantes', 'precios',
            'diagnosticos', 'recetas', 'configuracion_precios', 'configuracion_sistema',
            'medicamento_sintoma', 'diagnostico_sintoma', 'diagnostico_medicamento',
            'navegacion_menu', 'requerimientos', 'notificaciones',
            'requerimiento_referencias', 'usuario_dispositivo',
            'pedidos_productos', 'usuarios_direcciones', 'usuarios_favoritos'
        ]

        # Tablas que estn en minsculas en PostgreSQL (excepciones)
        tablas_minusculas = [
            'precios_competencia', 'precios_competencia_new',
            'existencias', 'terceros', 'terceros_competidores', 'terceros_direcciones', 'alertas_admin', 'archivos',
            'componentes_activos_sugerencias', 'indicaciones_rechazadas',
            'medicamentos_top', 'navegacion_anonima', 'pastillero_usuarios',
            'sugerir_sintomas', 'pedidos', 'promos_carousel',
            'categorias', 'medicamento_categoria'
        ]

        # Convertir tablas que deben ir en MAYSCULAS
        for tabla in tablas_mayusculas:
            pattern = r'\b' + tabla + r'\b'
            query = re.sub(pattern, f'"{tabla.upper()}"', query, flags=re.IGNORECASE)

        # Convertir tablas que deben permanecer en minsculas
        for tabla in tablas_minusculas:
            pattern = r'\b' + tabla + r'\b'
            query = re.sub(pattern, f'"{tabla}"', query, flags=re.IGNORECASE)

        cursor = self._conn.cursor()

        # Detectar INSERT para agregar RETURNING id
        is_insert = re.match(r'^\s*INSERT\s+INTO', query, re.IGNORECASE)
        last_insert_id = None

        if is_insert and 'RETURNING' not in query.upper():
            # Agregar RETURNING id al final del INSERT
            query = query.rstrip(';').rstrip() + ' RETURNING id'
            cursor.execute(query, params)
            result = cursor.fetchone()
            if result:
                last_insert_id = result[0]
        else:
            cursor.execute(query, params)

        return PostgreSQLCursorWrapper(cursor, last_insert_id)

    def commit(self):
        self._conn.commit()

    def rollback(self):
        self._conn.rollback()

    def close(self):
        self._conn.close()

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        if exc_type:
            self.rollback()
        self.close()

def get_db_connection():
    """
    Conexin a base de datos PostgreSQL.
    - En produccin (Render): usa DATABASE_URL del entorno
    - En local: usa conexion PostgreSQL local
    """
    database_url = os.getenv('DATABASE_URL')

    if not database_url:
        # LOCAL: PostgreSQL local
        database_url = os.getenv('LOCAL_DATABASE_URL', 'postgresql://postgres:postgres@localhost:5432/medicamentos')

    # Render usa postgres://, pero psycopg2 necesita postgresql://
    if database_url.startswith('postgres://'):
        database_url = database_url.replace('postgres://', 'postgresql://', 1)

    import psycopg2
    # Conectar sin RealDictCursor para compatibilidad con el wrapper
    pg_conn = psycopg2.connect(database_url)
    pg_conn.cursor().execute("SET TIME ZONE 'America/Bogota'")
    # Envolver la conexin para que funcione como SQLite
    return PostgreSQLConnectionWrapper(pg_conn)

ALLOWED_EXT = {"png", "jpg", "jpeg", "gif"}


# -------------------------------------------------------------------
# --- FUNCIN HELPER: NOTIFICACIONES TELEGRAM ---
# -------------------------------------------------------------------

def enviar_notificacion_telegram(mensaje):
    """
    Enva una notificacin a Telegram usando el bot configurado.
    Retorna True si se envi correctamente, False si fall.
    """
    try:
        # Obtener configuracin desde la base de datos
        conn = get_db_connection()
        config = conn.execute('SELECT telegram_token, telegram_chat_id, notificaciones_activas FROM CONFIGURACION_SISTEMA WHERE id = 1').fetchone()
        conn.close()

        if not config or not config[2]:  # notificaciones_activas es el ndice 2
            print(" Notificaciones Telegram desactivadas")
            return False

        token = config[0]  # telegram_token
        chat_id = config[1]  # telegram_chat_id

        if not token or not chat_id:
            print(" Token o Chat ID de Telegram no configurado")
            return False

        # Enviar mensaje a Telegram
        url = f"https://api.telegram.org/bot{token}/sendMessage"
        data = {
            'chat_id': chat_id,
            'text': mensaje,
            'parse_mode': 'HTML'
        }

        response = requests.post(url, json=data, timeout=10)

        if response.status_code == 200:
            print(f" Notificacin Telegram enviada correctamente")
            return True
        else:
            print(f" Error enviando Telegram: {response.status_code} - {response.text}")
            return False

    except Exception as e:
        print(f" Excepcin enviando Telegram: {e}")
        import traceback
        traceback.print_exc()
        return False


# -------------------------------------------------------------------
# --- FUNCIN DE RECORDATORIOS AUTOMTICOS (APScheduler) ---
# -------------------------------------------------------------------

def verificar_y_enviar_recordatorios():
    """
    Función que se ejecuta periódicamente para verificar si hay medicamentos
    que necesitan recordatorio y envía mensajes de Telegram con botones interactivos.
    """
    try:
        print("\n[SCHEDULER] Verificando recordatorios pendientes...")

        conn = get_db_connection()

        # Obtener token de Telegram
        config = conn.execute('SELECT telegram_token FROM CONFIGURACION_SISTEMA WHERE id = 1').fetchone()

        if not config:
            token = "8486881295:AAFjs-SU74er_shs4KnQYImMtyU5OTXycng"  # Fallback
        else:
            token = config[0]

        # Buscar medicamentos con recordatorio activo cuya próxima toma ya pasó
        ahora = datetime.now()

        medicamentos_pendientes = conn.execute('''
            SELECT
                p.id,
                p.nombre,
                p.cantidad,
                p.horas_entre_tomas,
                p.proxima_toma,
                t.telegram_chat_id,
                t.nombre as usuario_nombre
            FROM pastillero_usuarios p
            INNER JOIN terceros t ON p.usuario_id = t.id
            WHERE p.recordatorio_activo = TRUE
              AND p.proxima_toma IS NOT NULL
              AND p.proxima_toma <= %s
              AND t.telegram_chat_id IS NOT NULL
              AND t.telegram_chat_id != ''
        ''', (ahora,)).fetchall()

        if not medicamentos_pendientes:
            print("[SCHEDULER] No hay recordatorios pendientes")
            conn.close()
            return

        print(f"[SCHEDULER] Encontrados {len(medicamentos_pendientes)} recordatorios pendientes")

        # Enviar recordatorio para cada medicamento
        for med in medicamentos_pendientes:
            try:
                medicamento_id = med['id']
                nombre = med['nombre']
                cantidad = med['cantidad']
                horas_entre_tomas = med['horas_entre_tomas']
                chat_id_usuario = med['telegram_chat_id']
                usuario_nombre = med['usuario_nombre']

                # Obtener usuario_id para buscar contactos adicionales
                usuario_id = conn.execute('''
                    SELECT usuario_id FROM pastillero_usuarios WHERE id = %s
                ''', (medicamento_id,)).fetchone()['usuario_id']

                # Crear botones interactivos (InlineKeyboard)
                keyboard = {
                    'inline_keyboard': [
                        [
                            {'text': '✓ Ya tomé', 'callback_data': f'tomar_{medicamento_id}'},
                            {'text': '❌ Cancelar hoy', 'callback_data': f'cancelar_{medicamento_id}'}
                        ]
                    ]
                }

                # 1. Enviar mensaje al usuario principal (segunda persona: "debes tomar")
                mensaje_usuario = f"⏰ <b>Recordatorio de Medicamento</b>\n\n"
                mensaje_usuario += f"{usuario_nombre} debes tomar:\n"
                mensaje_usuario += f"<b>{nombre}</b> - 1 pastilla\n\n"
                mensaje_usuario += f"Quedan: {cantidad} pastillas\n"
                mensaje_usuario += f"Frecuencia: cada {horas_entre_tomas} horas"

                url = f"https://api.telegram.org/bot{token}/sendMessage"
                data_usuario = {
                    'chat_id': chat_id_usuario,
                    'text': mensaje_usuario,
                    'parse_mode': 'HTML',
                    'reply_markup': keyboard
                }

                response = requests.post(url, json=data_usuario, timeout=10)

                if response.status_code == 200:
                    print(f"[OK] Recordatorio enviado al usuario: {nombre} -> chat_id={chat_id_usuario}")
                else:
                    print(f"[ERROR] No se pudo enviar al usuario: {response.status_code}")

                # 2. Enviar mensajes a contactos adicionales (tercera persona: "debe tomar")
                contactos_adicionales = conn.execute('''
                    SELECT
                        pca.contacto_id,
                        t.nombre as contacto_nombre,
                        t.telegram_chat_id
                    FROM pastillero_contactos_adicionales pca
                    INNER JOIN terceros t ON pca.contacto_id = t.id
                    WHERE pca.usuario_id = %s
                      AND t.telegram_chat_id IS NOT NULL
                      AND t.telegram_chat_id != ''
                ''', (usuario_id,)).fetchall()

                for contacto in contactos_adicionales:
                    mensaje_contacto = f"⏰ <b>Recordatorio de Medicamento</b>\n\n"
                    mensaje_contacto += f"{usuario_nombre} debe tomar:\n"
                    mensaje_contacto += f"<b>{nombre}</b> - 1 pastilla\n\n"
                    mensaje_contacto += f"Quedan: {cantidad} pastillas\n"
                    mensaje_contacto += f"Frecuencia: cada {horas_entre_tomas} horas"

                    data_contacto = {
                        'chat_id': contacto['telegram_chat_id'],
                        'text': mensaje_contacto,
                        'parse_mode': 'HTML',
                        'reply_markup': keyboard
                    }

                    response_contacto = requests.post(url, json=data_contacto, timeout=10)

                    if response_contacto.status_code == 200:
                        print(f"[OK] Recordatorio enviado a contacto '{contacto['contacto_nombre']}': {nombre}")
                    else:
                        print(f"[ERROR] No se pudo enviar a contacto: {response_contacto.status_code}")

                # 3. Actualizar próxima toma (posponer por las horas configuradas)
                nueva_proxima_toma = ahora + timedelta(hours=horas_entre_tomas)

                conn.execute('''
                    UPDATE pastillero_usuarios
                    SET proxima_toma = %s
                    WHERE id = %s
                ''', (nueva_proxima_toma, medicamento_id))

                conn.commit()

            except Exception as e:
                print(f"[ERROR] Error procesando recordatorio para {med['nombre']}: {e}")
                import traceback
                traceback.print_exc()

        conn.close()
        print("[SCHEDULER] Verificación de recordatorios completada\n")

    except Exception as e:
        print(f"[ERROR] Error en verificar_y_enviar_recordatorios: {e}")
        import traceback
        traceback.print_exc()


# Inicializar el scheduler en background
scheduler = BackgroundScheduler()

# Ejecutar cada 5 minutos
scheduler.add_job(
    func=verificar_y_enviar_recordatorios,
    trigger=IntervalTrigger(minutes=5),
    id='verificar_recordatorios',
    name='Verificar y enviar recordatorios de medicamentos',
    replace_existing=True
)

# Iniciar el scheduler
scheduler.start()

# Asegurar que el scheduler se detenga cuando la aplicación se cierre
atexit.register(lambda: scheduler.shutdown())

print("[SCHEDULER] APScheduler inicializado - Verificando recordatorios cada 5 minutos")


# -------------------------------------------------------------------
# --- ZONA 2: MIDDLEWARE Y LGICA DE AUTENTICACIN (before_request) ---
# -------------------------------------------------------------------

def admin_required(f):
    """
    Decorador para proteger rutas que requieren rol de Administrador.
    Compatible con peticiones AJAX (devuelve JSON) y navegacin normal (redirige).
    """
    def wrapper(*args, **kwargs):

        print(f"DECORADOR EJECUTADO para funcion: {f.__name__}")
        print(f"   Sesion actual: {dict(session)}")
        print(f"   Rol en sesion: {session.get('rol', 'NO TIENE ROL')}")

        if 'rol' not in session or session['rol'] != 'Administrador':
            print(f"   ACCESO DENEGADO - Redirigiendo...")
            # Detectar si es una peticin AJAX
            is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest' or \
                      request.headers.get('Accept', '').find('application/json') != -1 or \
                      request.path.startswith('/admin/') and (
                          request.path.endswith('/sintomas') or 
                          '/sintomas' in request.path or
                          request.is_json
                      )
            print(f"   Es peticion AJAX: {is_ajax}")

            if is_ajax:
                # Si es AJAX, devolver JSON con error
                return jsonify({
                    'ok': False, 
                    'error': 'Acceso denegado. Se requiere ser Administrador.'
                }), 403
            else:
                # Si es navegacin normal, redirigir
                flash("Acceso denegado. Se requiere ser Administrador.", 'danger')
                return redirect(url_for('index'))
          
        print(f"   ACCESO PERMITIDO - Ejecutando funcion {f.__name__}")
        return f(*args, **kwargs)
    
    wrapper.__name__ = f.__name__
    return wrapper



@app.before_request
def check_device_access():
    if request.path.startswith('/static'):
        return  # Ignorar peticiones a recursos estticos

    # RUTAS PBLICAS: permitir acceso sin registro a la tienda
    rutas_publicas = ['/tienda', '/favicon.ico', '/']
    for ruta in rutas_publicas:
        if request.path.startswith(ruta) or request.path == ruta:
            # Para rutas pblicas, solo crear dispositivo_id si no existe
            if 'dispositivo_id' not in session:
                session['dispositivo_id'] = str(uuid.uuid4())
            return  # Permitir acceso sin verificar usuario

    device_id = session.get('dispositivo_id')
    if not device_id:
        device_id = str(uuid.uuid4())
        session['dispositivo_id'] = device_id
        app.logger.debug(f"Nuevo dispositivo_id creado: {device_id}")

    conn = get_db_connection()
    usuario = conn.execute('SELECT id, nombre, rol FROM usuarios WHERE dispositivo_id = ?', (device_id,)).fetchone()
    conn.close()

    if usuario is None:
        # Usuario nuevo
        if 'rol_temporal' not in session:
            if request.path.startswith('/admin'):
                rol_temporal = "Administrador"
            else:
                rol_temporal = "Paciente"
            session['rol_temporal'] = rol_temporal
            app.logger.debug(f"Nuevo usuario detectado, rol temporal asignado: {rol_temporal}")

        # Evitar loop: permitir POST a etapa1_nuevo_registro
        if request.endpoint not in ['etapa1_nuevo_registro', 'etapa1_registro_completo']:
            app.logger.debug(f"Redirigiendo usuario nuevo a etapa1_nuevo_registro desde {request.endpoint}")
            return redirect(url_for('etapa1_nuevo_registro'))
    else:
        # Usuario registrado
        session['rol'] = usuario['rol']
        session['nombre'] = usuario['nombre']
        session['usuario_id'] = usuario['id']
        app.logger.debug(f"Usuario existente: {usuario['nombre']} con rol: {usuario['rol']}")

        # Proteccin de rutas admin
        if request.path.startswith('/admin') and usuario['rol'] != 'Administrador':
            flash("Acceso denegado: No eres administrador.", "danger")
            app.logger.debug(f"Intento de acceso a /admin denegado para {usuario['nombre']}")
            return redirect(url_for('index'))

    # Log final del before_request
    app.logger.debug(f" Sesin actual despus de check_device_access: {dict(session)}")


# -------------------------------------------------------------------
# --- ZONA 3: RUTAS PBLICAS Y REDIRECCIN INICIAL ---
# -------------------------------------------------------------------

@app.route('/')
def index():
    """Pgina principal: muestra directamente la tienda de medicamentos."""
    # Redirigir directamente a la tienda
    return redirect(url_for('tienda_home'))

@app.route('/admin/acceso/<codigo>')
def admin_acceso_directo(codigo):
    """Acceso directo al panel admin con cdigo secreto."""
    # Cdigo secreto para acceso rpido
    CODIGO_ADMIN = 'tuctuc2025'

    print(f" Intento de acceso admin con cdigo: {codigo}")

    if codigo == CODIGO_ADMIN:
        # Buscar el usuario administrador
        conn = get_db_connection()
        admin = conn.execute('SELECT * FROM usuarios WHERE rol = ? LIMIT 1', ['Administrador']).fetchone()
        conn.close()

        if admin:
            # Establecer sesin como administrador
            session.clear()  # Limpiar sesin anterior
            session['dispositivo_id'] = admin['dispositivo_id']
            session['usuario_id'] = admin['id']
            session['nombre'] = admin['nombre']
            session['rol'] = 'Administrador'
            session.modified = True  # Forzar guardado de sesin

            print(f" Sesin admin establecida: {dict(session)}")

            flash('Acceso administrativo concedido', 'success')
            return redirect(url_for('admin_area'))
        else:
            print(f" No se encontr usuario administrador en la BD")
            flash('No se encontr usuario administrador', 'danger')
            return redirect(url_for('index'))
    else:
        print(f" Cdigo invlido: {codigo}")
        flash('Cdigo de acceso invlido', 'danger')
        return redirect(url_for('index'))

@app.route('/admin')
def admin_redirect():
    """Redirige a la pgina de login o al rea admin si ya est logueado."""
    if session.get('rol') == 'Administrador':
        return redirect(url_for('admin_area'))
    return redirect(url_for('admin_login'))

@app.route('/admin/login', methods=['GET'])
def admin_login():
    """Muestra el formulario de login para administradores."""
    # Si ya est logueado, redirigir al rea admin
    if session.get('rol') == 'Administrador':
        return redirect(url_for('admin_area'))

    return render_template('admin_login.html')

@app.route('/admin/login', methods=['POST'])
def admin_login_post():
    """Procesa el login de administradores."""
    usuario = request.form.get('usuario', '').strip()
    password = request.form.get('password', '').strip()
    recordar = request.form.get('recordar') == '1'

    print(f" Intento de login: usuario={usuario}, recordar={recordar}")

    if not usuario or not password:
        flash('Por favor completa todos los campos', 'danger')
        return redirect(url_for('admin_login'))

    # Buscar usuario admin en la base de datos
    conn = get_db_connection()
    try:
        admin = conn.execute("""
            SELECT * FROM usuarios
            WHERE usuario = ? AND password = ? AND rol = 'Administrador'
        """, (usuario, password)).fetchone()

        if admin:
            # Login exitoso
            session.clear()
            session['dispositivo_id'] = admin['dispositivo_id']
            session['usuario_id'] = admin['id']
            session['nombre'] = admin['nombre']
            session['rol'] = 'Administrador'
            session.modified = True

            # Configurar duracin de sesin segn checkbox
            if recordar:
                session.permanent = True
                app.permanent_session_lifetime = timedelta(days=30)
                print(f" Login exitoso - Sesin de 30 das")
            else:
                session.permanent = True
                app.permanent_session_lifetime = timedelta(hours=24)
                print(f" Login exitoso - Sesin de 24 horas")

            flash('Bienvenido! Has iniciado sesin correctamente', 'success')
            return redirect(url_for('admin_area'))
        else:
            # Login fallido
            print(f" Login fallido - Usuario o contrasea incorrectos")
            flash('Usuario o contrasea incorrectos', 'danger')
            return redirect(url_for('admin_login'))

    finally:
        conn.close()

@app.route('/admin/promos')
@admin_required
def admin_promos():
    """Pgina de administracin de promos del carousel"""
    return render_template('admin_promos.html')

@app.route('/paciente_saludo_continuar')
def paciente_saludo_continuar():
    """Muestra un saludo personalizado y redirecciona a la siguiente etapa de registro."""
    if session.get('rol') != 'Paciente':
        flash("Acceso no autorizado.", 'danger')
        return redirect(url_for('index'))
    
    siguiente_ruta = session.get('next_etapa', 'paciente_area') 
    
    return render_template('3_saludo_continuar.html', 
                           nombre=session.get('nombre'),
                           siguiente_ruta=url_for(siguiente_ruta))

@app.route('/inicio')
def inicio():
    """Ruta para 'Volver a la Pgina de Inicio'."""
    return redirect(url_for('index'))


# ========================================
# ECOMMERCE - TIENDA
# ========================================

@app.route('/tienda')
def tienda_home():
    """Catlogo de productos ecommerce"""
    try:
        # Obtener todos los sntomas para filtros
        conn = get_db_connection()
        sintomas = conn.execute("SELECT id, nombre FROM sintomas ORDER BY nombre").fetchall()
        conn.close()

        # Pasar rol si est en sesin (para funcionalidades especiales de admin)
        es_admin = session.get('rol') == 'Administrador'
        return render_template('tienda_home.html', sintomas=[dict(s) for s in sintomas], es_admin=es_admin)
    except Exception as e:
        print(f"ERROR en tienda_home: {e}")
        import traceback
        traceback.print_exc()
        # Devolver pgina sin sntomas si hay error
        return render_template('tienda_home.html', sintomas=[], es_admin=False)


@app.route('/tienda/carrito')
def carrito():
    """Pgina del carrito de compras"""
    return render_template('carrito.html')

@app.route('/admin/setup_promos_table')
@admin_required
def setup_promos_table():
    """Crea la tabla promos_carousel si no existe (solo para setup inicial)"""
    try:
        conn = get_db_connection()

        # Crear tabla (sin foreign key por ahora, la agregamos despus)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS promos_carousel (
                id SERIAL PRIMARY KEY,
                imagen_url VARCHAR(255) NOT NULL,
                titulo VARCHAR(255) NOT NULL,
                activa BOOLEAN DEFAULT TRUE,
                medicamento_id INTEGER,
                orden INTEGER DEFAULT 0,
                fecha_inicio TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                fecha_fin TIMESTAMP,
                fecha_creacion TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                fecha_actualizacion TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
        """)

        # Crear ndices
        conn.execute("CREATE INDEX IF NOT EXISTS idx_promos_activa ON promos_carousel(activa);")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_promos_orden ON promos_carousel(orden);")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_promos_medicamento ON promos_carousel(medicamento_id);")

        # Insertar promos iniciales si la tabla est vaca
        count = conn.execute("SELECT COUNT(*) as cnt FROM promos_carousel").fetchone()['cnt']

        if count == 0:
            conn.execute("""
                INSERT INTO promos_carousel (imagen_url, titulo, activa, orden)
                VALUES
                    ('logo1.png', 'Logo TUC-TUC', TRUE, 1),
                    ('logo_navidad.png', 'Promo Navidad', TRUE, 2),
                    ('logo_promo.png', 'Promocin Especial', TRUE, 3)
            """)

        conn.commit()

        return f"""
        <h1> Tabla promos_carousel configurada</h1>
        <p>Tabla creada correctamente</p>
        <p>ndices creados</p>
        <p>Promos iniciales: {count} existentes antes del setup</p>
        <br>
        <a href="/admin/promos">Ir a Gestin de Promos</a>
        """
    except Exception as e:
        import traceback
        traceback.print_exc()
        return f"<h1> Error:</h1><pre>{str(e)}</pre>", 500

# ==========================================
# ENDPOINTS PARA SISTEMA DE DIRECCIONES
# ==========================================

@app.route('/api/verificar-telefono', methods=['GET'])
def verificar_telefono():
    """
    ltima revisin: 2025-12-15 20:00
    Usado en: templates/checkout.html
    Verifica si un telfono ya existe en terceros y devuelve sus datos
    """
    telefono = request.args.get('telefono', '').strip()

    if not telefono:
        return jsonify({'ok': False, 'error': 'Telfono requerido'}), 400

    try:
        conn = get_db_connection()

        # Buscar tercero por telfono
        tercero = conn.execute("""
            SELECT id, nombre, telefono, id_usuario
            FROM terceros
            WHERE telefono = ?
            LIMIT 1
        """, (telefono,)).fetchone()

        if tercero:
            tercero_id = tercero['id']

            # Obtener direcciones del tercero
            direcciones = conn.execute("""
                SELECT id, alias, nombre_completo, telefono, direccion,
                       latitud, longitud, es_principal
                FROM terceros_direcciones
                WHERE tercero_id = ?
                ORDER BY es_principal DESC, fecha_actualizacion DESC
            """, (tercero_id,)).fetchall()

            conn.close()

            return jsonify({
                'ok': True,
                'existe': True,
                'tercero_id': tercero_id,
                'nombre': tercero['nombre'],
                'direcciones': [dict(d) for d in direcciones]
            })
        else:
            conn.close()
            return jsonify({
                'ok': True,
                'existe': False
            })

    except Exception as e:
        print(f"Error verificar-telefono: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'ok': False, 'error': str(e)}), 500

@app.route('/api/enviar-codigo-whatsapp', methods=['POST'])
def enviar_codigo_whatsapp():
    """
    ltima revisin: 2025-12-15 20:00
    Usado en: templates/checkout.html
    Genera y enva cdigo de verificacin por WhatsApp
    """
    try:
        data = request.get_json()
        telefono = data.get('telefono', '').strip()

        if not telefono:
            return jsonify({'ok': False, 'error': 'Telfono requerido'}), 400

        # Generar cdigo de 6 dgitos
        import random
        codigo = ''.join([str(random.randint(0, 9)) for _ in range(6)])

        # Guardar cdigo en session con timestamp
        import time
        session['codigo_verificacion'] = codigo
        session['codigo_telefono'] = telefono
        session['codigo_timestamp'] = time.time()

        # Enviar por WhatsApp
        conn = get_db_connection()
        config = conn.execute('SELECT whatsapp_numero FROM "CONFIGURACION_SISTEMA" LIMIT 1').fetchone()
        conn.close()

        numero_destino = telefono if not telefono.startswith('+') else telefono[1:]
        mensaje = f" Tu cdigo de verificacin TUC-TUC es: *{codigo}*\n\nVlido por 10 minutos."

        # Enviar WhatsApp
        enviar_whatsapp(numero_destino, mensaje)

        return jsonify({'ok': True, 'mensaje': 'Cdigo enviado por WhatsApp'})

    except Exception as e:
        print(f"Error enviar-codigo-whatsapp: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'ok': False, 'error': str(e)}), 500

@app.route('/api/verificar-codigo', methods=['POST'])
def verificar_codigo():
    """
    ltima revisin: 2025-12-15 20:00
    Usado en: templates/checkout.html
    Verifica el cdigo ingresado por el usuario
    """
    try:
        data = request.get_json()
        codigo = data.get('codigo', '').strip()
        telefono = data.get('telefono', '').strip()

        if not codigo or not telefono:
            return jsonify({'ok': False, 'error': 'Cdigo y telfono requeridos'}), 400

        # Verificar cdigo de session
        codigo_guardado = session.get('codigo_verificacion')
        telefono_guardado = session.get('codigo_telefono')
        timestamp = session.get('codigo_timestamp', 0)

        import time
        tiempo_transcurrido = time.time() - timestamp

        # Validar
        if not codigo_guardado:
            return jsonify({'ok': False, 'error': 'No hay cdigo pendiente'}), 400

        if tiempo_transcurrido > 600:  # 10 minutos
            return jsonify({'ok': False, 'error': 'Cdigo expirado'}), 400

        if codigo != codigo_guardado or telefono != telefono_guardado:
            return jsonify({'ok': False, 'error': 'Cdigo incorrecto'}), 400

        # Cdigo vlido - limpiar session y marcar como verificado
        session.pop('codigo_verificacion', None)
        session.pop('codigo_telefono', None)
        session.pop('codigo_timestamp', None)
        session['telefono_verificado'] = telefono
        session['telefono_verificado_timestamp'] = time.time()

        return jsonify({'ok': True, 'mensaje': 'Cdigo verificado'})

    except Exception as e:
        print(f"Error verificar-codigo: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'ok': False, 'error': str(e)}), 500

@app.route('/api/tercero/direcciones', methods=['GET'])
def obtener_direcciones_tercero():
    """
    ltima revisin: 2025-12-15 20:00
    Usado en: templates/checkout.html
    Obtiene las direcciones de un tercero por telfono
    """
    telefono = request.args.get('telefono', '').strip()

    if not telefono:
        return jsonify({'ok': False, 'error': 'Telfono requerido'}), 400

    try:
        conn = get_db_connection()

        # Buscar tercero
        tercero = conn.execute("""
            SELECT id FROM terceros WHERE telefono = ? LIMIT 1
        """, (telefono,)).fetchone()

        if not tercero:
            conn.close()
            return jsonify({'ok': True, 'direcciones': []})

        # Obtener direcciones
        direcciones = conn.execute("""
            SELECT id, alias, nombre_completo, telefono, direccion,
                   latitud, longitud, es_principal
            FROM terceros_direcciones
            WHERE tercero_id = ?
            ORDER BY es_principal DESC, fecha_actualizacion DESC
        """, (tercero['id'],)).fetchall()

        conn.close()

        return jsonify({
            'ok': True,
            'direcciones': [dict(d) for d in direcciones]
        })

    except Exception as e:
        print(f"Error obtener direcciones: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'ok': False, 'error': str(e)}), 500

@app.route('/api/tercero/direcciones', methods=['POST'])
def agregar_direccion_tercero():
    """
    ltima revisin: 2025-12-15 20:00
    Usado en: templates/checkout.html
    Agrega una nueva direccin a un tercero
    """
    try:
        data = request.get_json()

        telefono = data.get('telefono', '').strip()
        alias = data.get('alias', '').strip()
        nombre_completo = data.get('nombre_completo', '').strip()
        direccion = data.get('direccion', '').strip()
        latitud = data.get('latitud')
        longitud = data.get('longitud')
        es_principal = data.get('es_principal', False)

        if not all([telefono, alias, nombre_completo, direccion]):
            return jsonify({'ok': False, 'error': 'Datos incompletos'}), 400

        conn = get_db_connection()

        # Buscar o crear tercero
        tercero = conn.execute("""
            SELECT id FROM terceros WHERE telefono = ? LIMIT 1
        """, (telefono,)).fetchone()

        if tercero:
            tercero_id = tercero['id']
        else:
            # Crear nuevo tercero
            cursor = conn.execute("""
                INSERT INTO terceros (nombre, telefono, fecha_creacion)
                VALUES (?, ?, CURRENT_TIMESTAMP)
            """, (nombre_completo, telefono))
            tercero_id = cursor.lastrowid

        # Si es principal, quitar principal de las dems
        if es_principal:
            conn.execute("""
                UPDATE terceros_direcciones
                SET es_principal = false
                WHERE tercero_id = ?
            """, (tercero_id,))

        # Insertar direccin
        cursor = conn.execute("""
            INSERT INTO terceros_direcciones
            (tercero_id, alias, nombre_completo, telefono, direccion, latitud, longitud, es_principal)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """, (tercero_id, alias, nombre_completo, telefono, direccion, latitud, longitud, es_principal))

        direccion_id = cursor.lastrowid
        conn.close()

        return jsonify({'ok': True, 'direccion_id': direccion_id})

    except Exception as e:
        print(f"Error agregar direccin: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'ok': False, 'error': str(e)}), 500

# ==========================================
# FIN ENDPOINTS DIRECCIONES
# ==========================================

# ==========================================
# ENDPOINTS PROMOS CAROUSEL
# ==========================================

@app.route('/api/promos', methods=['GET'])
def get_promos():
    """Obtiene todas las promos activas ordenadas"""
    try:
        conn = get_db_connection()
        promos = conn.execute("""
            SELECT p.*, m.nombre as medicamento_nombre
            FROM promos_carousel p
            LEFT JOIN MEDICAMENTOS m ON p.medicamento_id = m.id
            WHERE p.activa IS TRUE
            AND (p.fecha_fin IS NULL OR p.fecha_fin > CURRENT_TIMESTAMP)
            ORDER BY p.orden ASC
        """).fetchall()

        return jsonify({
            'ok': True,
            'promos': [dict(p) for p in promos]
        })
    except Exception as e:
        traceback.print_exc()
        return jsonify({'ok': False, 'error': str(e)}), 500

@app.route('/api/promos/admin', methods=['GET'])
@admin_required
def get_promos_admin():
    """Obtiene todas las promos (admin)"""
    try:
        conn = get_db_connection()
        promos = conn.execute("""
            SELECT p.*, m.nombre as medicamento_nombre
            FROM promos_carousel p
            LEFT JOIN MEDICAMENTOS m ON p.medicamento_id = m.id
            ORDER BY p.orden ASC, p.id DESC
        """).fetchall()

        return jsonify({
            'ok': True,
            'promos': [dict(p) for p in promos]
        })
    except Exception as e:
        traceback.print_exc()
        return jsonify({'ok': False, 'error': str(e)}), 500

@app.route('/api/promos', methods=['POST'])
@admin_required
def crear_promo():
    """Crea una nueva promo"""
    try:
        data = request.get_json()

        imagen_url = data.get('imagen_url', '').strip()
        titulo = data.get('titulo', '').strip()
        medicamento_id = data.get('medicamento_id')
        orden = data.get('orden', 0)
        fecha_fin = data.get('fecha_fin')
        intervalo_carousel = data.get('intervalo_carousel', 5)

        if not imagen_url or not titulo:
            return jsonify({'ok': False, 'error': 'Imagen y ttulo son requeridos'}), 400

        conn = get_db_connection()

        # Push: incrementar orden de todas las promos >= al nuevo orden
        conn.execute("""
            UPDATE promos_carousel
            SET orden = orden + 1
            WHERE orden >= ?
        """, (orden,))

        cursor = conn.execute("""
            INSERT INTO promos_carousel
            (imagen_url, titulo, medicamento_id, orden, fecha_fin, activa, intervalo_carousel)
            VALUES (?, ?, ?, ?, ?, true, ?)
        """, (imagen_url, titulo, medicamento_id, orden, fecha_fin, intervalo_carousel))

        conn.commit()

        return jsonify({
            'ok': True,
            'promo_id': cursor.lastrowid
        })
    except Exception as e:
        traceback.print_exc()
        return jsonify({'ok': False, 'error': str(e)}), 500

@app.route('/api/promos/<int:promo_id>', methods=['PUT'])
@admin_required
def actualizar_promo(promo_id):
    """Actualiza una promo existente"""
    try:
        data = request.get_json()

        conn = get_db_connection()

        # Si se est cambiando el orden, hacer push de otros rdenes
        if 'orden' in data:
            nuevo_orden = data['orden']

            # Obtener orden actual de esta promo
            promo_actual = conn.execute("SELECT orden FROM promos_carousel WHERE id = ?", (promo_id,)).fetchone()
            orden_actual = promo_actual['orden'] if promo_actual else None

            if orden_actual is not None and nuevo_orden != orden_actual:
                if nuevo_orden < orden_actual:
                    # Moviendo hacia arriba: incrementar rdenes entre nuevo y actual (inclusive)
                    conn.execute("""
                        UPDATE promos_carousel
                        SET orden = orden + 1
                        WHERE orden >= ? AND orden <= ? AND id != ?
                    """, (nuevo_orden, orden_actual, promo_id))
                else:
                    # Moviendo hacia abajo: decrementar rdenes entre actual y nuevo
                    conn.execute("""
                        UPDATE promos_carousel
                        SET orden = orden - 1
                        WHERE orden > ? AND orden <= ? AND id != ?
                    """, (orden_actual, nuevo_orden, promo_id))

        # Construir query dinmicamente segn campos presentes
        updates = []
        params = []

        if 'imagen_url' in data:
            updates.append('imagen_url = ?')
            params.append(data['imagen_url'])
        if 'titulo' in data:
            updates.append('titulo = ?')
            params.append(data['titulo'])
        if 'medicamento_id' in data:
            updates.append('medicamento_id = ?')
            params.append(data['medicamento_id'])
        if 'orden' in data:
            updates.append('orden = ?')
            params.append(data['orden'])
        if 'activa' in data:
            updates.append('activa = ?')
            params.append(data['activa'])
        if 'fecha_fin' in data:
            updates.append('fecha_fin = ?')
            params.append(data['fecha_fin'])
        if 'intervalo_carousel' in data:
            updates.append('intervalo_carousel = ?')
            params.append(data['intervalo_carousel'])

        updates.append('fecha_actualizacion = CURRENT_TIMESTAMP')
        params.append(promo_id)

        query = f"UPDATE promos_carousel SET {', '.join(updates)} WHERE id = ?"

        conn.execute(query, params)
        conn.commit()

        return jsonify({'ok': True})
    except Exception as e:
        traceback.print_exc()
        return jsonify({'ok': False, 'error': str(e)}), 500

@app.route('/api/promos/<int:promo_id>', methods=['DELETE'])
@admin_required
def eliminar_promo(promo_id):
    """Elimina una promo"""
    try:
        conn = get_db_connection()
        conn.execute("DELETE FROM promos_carousel WHERE id = ?", (promo_id,))
        conn.commit()

        return jsonify({'ok': True})
    except Exception as e:
        traceback.print_exc()
        return jsonify({'ok': False, 'error': str(e)}), 500

@app.route('/api/promos/upload', methods=['POST'])
@admin_required
def upload_promo_image():
    """Sube una imagen para promo"""
    try:
        if 'imagen' not in request.files:
            return jsonify({'ok': False, 'error': 'No se envi archivo'}), 400

        file = request.files['imagen']

        if file.filename == '':
            return jsonify({'ok': False, 'error': 'Archivo vaco'}), 400

        # Validar extensin
        allowed_extensions = {'png', 'jpg', 'jpeg', 'gif', 'webp'}
        file_ext = file.filename.rsplit('.', 1)[1].lower() if '.' in file.filename else ''

        if file_ext not in allowed_extensions:
            return jsonify({'ok': False, 'error': 'Formato no permitido. Use: png, jpg, jpeg, gif, webp'}), 400

        # Subir a Cloudinary
        filename = secure_filename(file.filename)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        public_id = f"tuctuc/promos/{timestamp}_{filename.rsplit('.', 1)[0]}"

        upload_result = cloudinary.uploader.upload(
            file,
            public_id=public_id,
            folder="tuctuc/promos",
            resource_type="image"
        )

        # Devolver URL completa de Cloudinary
        return jsonify({
            'ok': True,
            'imagen_url': upload_result['secure_url']
        })
    except Exception as e:
        traceback.print_exc()
        return jsonify({'ok': False, 'error': str(e)}), 500

# ==========================================
# FIN ENDPOINTS PROMOS CAROUSEL
# ==========================================

@app.route('/tienda/checkout')
def checkout():
    """Pgina de checkout"""
    return render_template('checkout.html')

@app.route('/tienda/procesar_pedido', methods=['POST'])
def procesar_pedido():
    """Procesa el pedido y crea el registro en BD + enva WhatsApp"""
    try:
        data = request.get_json()
        
        # Validar datos
        nombre = data.get('nombre', '').strip()
        telefono = data.get('telefono', '').strip()
        direccion = data.get('direccion', '').strip()
        metodo_pago = data.get('metodo_pago')
        items = data.get('items', [])
        
        if not all([nombre, telefono, direccion, metodo_pago, items]):
            return jsonify({'ok': False, 'error': 'Datos incompletos'}), 400
        
        # Geocoding - obtener coordenadas
        GOOGLE_API_KEY = 'AIzaSyCiAtNFl95bJJFuqiNsiYynBS3LuDisq9g'
        # Mejorar formato de direccin
        direccion_completa = f"{direccion}, Cali, Valle del Cauca, Colombia"
        geocode_url = f"https://maps.googleapis.com/maps/api/geocode/json?address={direccion_completa}&key={GOOGLE_API_KEY}"

        print(f" Intentando geocoding para: {direccion_completa}")  #  AGREGAR

        try:
            geo_resp = requests.get(geocode_url, timeout=5)
            geo_data = geo_resp.json()
            
            print(f" Respuesta Google: {geo_data.get('status')}")  #  AGREGAR


            if geo_data.get('status') == 'OK' and len(geo_data.get('results', [])) > 0:
                location = geo_data['results'][0]['geometry']['location']
                latitud = location['lat']
                longitud = location['lng']
                direccion_completa = geo_data['results'][0]['formatted_address']
                print(f" Coordenadas: {latitud}, {longitud}")  #  AGREGAR

            else:
                print(f" Google no encontr la direccin. Status: {geo_data.get('status')}")  #  AGREGAR
                latitud = None
                longitud = None
                direccion_completa = direccion
                
        except Exception as e:
            print(f" Error en geocoding: {e}")
            latitud = None
            longitud = None
            direccion_completa = direccion
        
        conn = get_db_connection()

        # 1. Buscar o crear TERCERO (cliente)
        # Primero buscar si ya existe por telfono
        tercero_existente = conn.execute("""
            SELECT id, nombre FROM terceros WHERE telefono = ? LIMIT 1
        """, (telefono,)).fetchone()

        if tercero_existente:
            tercero_id = tercero_existente['id']
            print(f" Tercero existente encontrado: ID {tercero_id} ({tercero_existente['nombre']})")

            # Actualizar nombre si cambi
            if tercero_existente['nombre'] != nombre:
                conn.execute("""
                    UPDATE terceros
                    SET nombre = ?, fecha_actualizacion = CURRENT_TIMESTAMP
                    WHERE id = ?
                """, (nombre, tercero_id))
                print(f"  Nombre actualizado a: {nombre}")

        else:
            # Crear nuevo tercero
            cursor_seq = conn.execute("SELECT COALESCE(MAX(id), 0) + 1 FROM terceros")
            next_tercero_id = cursor_seq.fetchone()[0]

            print(f" Insertando nuevo tercero con ID {next_tercero_id}...")
            cursor = conn.execute("""
                INSERT INTO terceros (id, nombre, telefono, fecha_creacion)
                VALUES (?, ?, ?, CURRENT_TIMESTAMP)
            """, (next_tercero_id, nombre, telefono))
            tercero_id = next_tercero_id
            print(f" Tercero creado con ID: {tercero_id}")

        # 2. Buscar o crear DIRECCIN en terceros_direcciones
        alias_direccion = data.get('alias_direccion', 'Principal')

        # Buscar si existe esta direccin exacta para este tercero
        direccion_existente = conn.execute("""
            SELECT id FROM terceros_direcciones
            WHERE tercero_id = ? AND direccion = ?
            LIMIT 1
        """, (tercero_id, direccion)).fetchone()

        if direccion_existente:
            direccion_id = direccion_existente['id']
            print(f" Direccin existente encontrada: ID {direccion_id}")

            # Actualizar coordenadas si cambiaron
            conn.execute("""
                UPDATE terceros_direcciones
                SET latitud = ?, longitud = ?, fecha_actualizacion = CURRENT_TIMESTAMP
                WHERE id = ?
            """, (latitud, longitud, direccion_id))

        else:
            # Crear nueva direccin
            print(f" Creando nueva direccin '{alias_direccion}'...")
            cursor = conn.execute("""
                INSERT INTO terceros_direcciones
                (tercero_id, alias, nombre_completo, telefono, direccion, latitud, longitud, es_principal)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """, (tercero_id, alias_direccion, nombre, telefono, direccion, latitud, longitud, True))
            direccion_id = cursor.lastrowid
            print(f" Direccin creada con ID: {direccion_id}")
        
        # 2. Calcular totales
        subtotal = sum(item['precio'] * item['cantidad'] for item in items)
        costo_domicilio = 0 if subtotal >= 50000 else 5000
        total = subtotal + costo_domicilio
        
        # 3. Crear PEDIDO
        # Obtener el siguiente ID manualmente (la tabla no tiene secuencia)
        cursor_seq = conn.execute("SELECT COALESCE(MAX(id), 0) + 1 FROM pedidos")
        next_pedido_id = cursor_seq.fetchone()[0]

        print(f" Insertando pedido con ID {next_pedido_id}...")
        cursor = conn.execute("""
            INSERT INTO pedidos (
                id, id_tercero, fecha, total, metodo_pago, costo_domicilio,
                direccion_entrega, latitud_entrega, longitud_entrega,
                estado, tiempo_estimado_entrega
            ) VALUES (?, ?, CURRENT_TIMESTAMP, ?, ?, ?, ?, ?, ?, 'pendiente', '30 minutos')
        """, (next_pedido_id, tercero_id, total, metodo_pago, costo_domicilio, direccion, latitud, longitud))
        pedido_id = next_pedido_id
        print(f" Pedido creado con ID: {pedido_id}")
        
        # 4. Crear EXISTENCIAS (salidas) para cada item
        for item in items:
            # Obtener siguiente ID para existencias
            cursor_seq = conn.execute("SELECT COALESCE(MAX(id), 0) + 1 FROM existencias")
            next_existencia_id = cursor_seq.fetchone()[0]

            conn.execute("""
                INSERT INTO existencias (
                    id, medicamento_id, fabricante_id, tipo_movimiento,
                    cantidad, fecha, id_tercero, pedido_id
                ) VALUES (?, ?, ?, 'salida', ?, CURRENT_TIMESTAMP, ?, ?)
            """, (next_existencia_id, item['medicamento_id'], item['fabricante_id'], item['cantidad'], tercero_id, pedido_id))
        
        conn.commit()
        conn.close()

        # 5. ENVIAR NOTIFICACIN TELEGRAM AL ADMIN
        try:
            # Construir lista de productos
            items_texto = "\n".join([
                f" {item['nombre']} ({item['fabricante']}) x{item['cantidad']} = ${item['precio'] * item['cantidad']:,}"
                for item in items
            ])

            # Link a Google Maps
            maps_link = f"https://www.google.com/maps?q={latitud},{longitud}" if latitud and longitud else "Sin coordenadas"

            # Mensaje Telegram (con formato HTML)
            mensaje = f""" <b>NUEVO PEDIDO #{pedido_id}</b>

 <b>Cliente:</b> {nombre}
 <b>Telfono:</b> {telefono}
 <b>Direccin:</b> {direccion}

 <b>Productos:</b>
{items_texto}

 <b>Subtotal:</b> ${subtotal:,}
 <b>Domicilio:</b> ${costo_domicilio:,}
 <b>TOTAL:</b> ${total:,}

 <b>Mtodo de pago:</b> {metodo_pago.upper()}

 <b>Ver ubicacin:</b>
{maps_link}

 <b>Tiempo estimado:</b> 30 minutos"""

            # Enviar notificacin a Telegram
            enviar_notificacion_telegram(mensaje)

        except Exception as e:
            print(f" Error enviando notificacin Telegram: {e}")
            # No fallar el pedido si la notificacin falla
        
        return jsonify({
            'ok': True,
            'pedido_id': pedido_id,
            'mensaje': 'Pedido creado exitosamente'
        })
        
    except Exception as e:
        print(f" Error procesando pedido: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'ok': False, 'error': str(e)}), 500



@app.route('/tienda/solicitud_especial', methods=['POST'])
def solicitud_especial():
    """Captura solicitudes de productos no catalogados o necesidades especiales"""
    try:
        data = request.get_json()
        
        # Validar datos recibidos
        telefono = data.get('telefono', '').strip()
        busqueda = data.get('busqueda', '').strip()
        tipo_solicitud = data.get('tipo', 'producto')  # 'producto' o 'sintoma'
        
        # Validar telfono
        if not telefono or len(telefono) != 10 or not telefono.startswith('3'):
            return jsonify({'ok': False, 'error': 'Telfono invlido'}), 400
        
        if not busqueda:
            return jsonify({'ok': False, 'error': 'Bsqueda vaca'}), 400
        
        # Obtener dispositivo_id de la sesin
        dispositivo_id = session.get('dispositivo_id')
        if not dispositivo_id:
            return jsonify({'ok': False, 'error': 'Sesin no vlida'}), 400
        
        conn = get_db_connection()
        
        # 1. CREAR/ACTUALIZAR USUARIO
        usuario = conn.execute(
            "SELECT id FROM usuarios WHERE dispositivo_id = ?", 
            (dispositivo_id,)
        ).fetchone()
        
        if not usuario:
            # Crear usuario nuevo
            cursor = conn.execute("""
                INSERT INTO usuarios (dispositivo_id, nombre, fecha_registro, rol)
                VALUES (?, 'Cliente sin nombre', CURRENT_TIMESTAMP, 'Paciente')
            """, (dispositivo_id,))
            usuario_id = cursor.lastrowid
            print(f" Usuario creado: ID={usuario_id}, dispositivo={dispositivo_id}")
        else:
            usuario_id = usuario['id']
            print(f" Usuario existente: ID={usuario_id}")
        
        # 2. CREAR TERCERO (con telfono, nombre temporal)
        cursor = conn.execute("""
            INSERT INTO terceros (nombre, telefono, id_usuario, fecha_creacion)
            VALUES ('Cliente sin nombre', ?, ?, CURRENT_TIMESTAMP)
        """, (telefono, usuario_id))
        tercero_id = cursor.lastrowid
        print(f" Tercero creado: ID={tercero_id}, telfono={telefono}")
        
        # 3. CREAR PEDIDO con estado especial
        etiqueta = "Producto" if tipo_solicitud == 'producto' else "Sntomas"
        notas = f"{etiqueta}: {busqueda}"
        
        cursor = conn.execute("""
            INSERT INTO pedidos (
                id_tercero, 
                fecha, 
                total, 
                metodo_pago,
                costo_domicilio,
                direccion_entrega,
                estado,
                notas,
                tiempo_estimado_entrega
            ) VALUES (?, CURRENT_TIMESTAMP, 0, 'pendiente', 0, '',
                      'VERIFICANDO_DISPONIBILIDAD', ?, 'Por confirmar')
        """, (tercero_id, notas))
        pedido_id = cursor.lastrowid
        
        conn.commit()
        conn.close()
        
        print(f" Pedido especial creado: ID={pedido_id}")
        print(f"   Tipo: {tipo_solicitud}")
        print(f"   Bsqueda: {busqueda}")
        print(f"   Telfono: {telefono}")
        
        # 4. ENVIAR NOTIFICACIN TELEGRAM AL ADMIN
        try:
            emoji = "" if tipo_solicitud == 'producto' else ""
            mensaje = f""" <b>SOLICITUD ESPECIAL #{pedido_id}</b>

{emoji} <b>Tipo:</b> {etiqueta}
 <b>Telfono:</b> {telefono}
 <b>Bsqueda:</b> {busqueda}

 <b>Requiere contacto inmediato</b>

 Ver pedido:
https://tuc-tuc.onrender.com/admin/pedidos"""

            # Enviar notificacin a Telegram
            enviar_notificacion_telegram(mensaje)

        except Exception as e:
            print(f" Error enviando notificacin Telegram: {e}")
            # No fallar la solicitud si la notificacin falla
        
        return jsonify({
            'ok': True,
            'pedido_id': pedido_id,
            'mensaje': 'Solicitud recibida. Te contactaremos en menos de 5 minutos'
        })
        
    except Exception as e:
        print(f" Error en solicitud_especial: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'ok': False, 'error': str(e)}), 500



@app.route('/tienda/confirmacion')
def confirmacion():
    """Pgina de confirmacin del pedido"""
    pedido_id = request.args.get('pedido')
    
    if not pedido_id:
        return redirect('/tienda')
    
    # Obtener datos del pedido
    conn = get_db_connection()
    pedido = conn.execute("""
        SELECT p.*, t.nombre as cliente_nombre, t.telefono, t.direccion
        FROM pedidos p
        INNER JOIN terceros t ON p.id_tercero = t.id
        WHERE p.id = ?
    """, (pedido_id,)).fetchone()
    conn.close()
    
    if not pedido:
        return redirect('/tienda')
    
    return render_template('confirmacion.html', pedido=pedido)


@app.route('/admin/pedidos')
@admin_required
def admin_pedidos():
    """Panel de administracin de pedidos"""
    return render_template('admin_pedidos.html')

@app.route('/admin/pedidos/lista')
@admin_required
def admin_pedidos_lista():
    """API para listar pedidos"""
    try:
        estado = request.args.get('estado', 'todos')
        
        conn = get_db_connection()
        
        query = """
            SELECT 
                p.id,
                p.fecha,
                p.total,
                p.estado,
                p.metodo_pago,
                p.tiempo_estimado_entrega,
                t.nombre as cliente_nombre,
                t.telefono,
                p.direccion_entrega,
                p.latitud_entrega,
                p.longitud_entrega,
                p.notas
            FROM pedidos p
            INNER JOIN terceros t ON p.id_tercero = t.id
        """
        
        if estado != 'todos':
            query += " WHERE p.estado = ?"
            pedidos = conn.execute(query + " ORDER BY p.fecha DESC", (estado,)).fetchall()
        else:
            pedidos = conn.execute(query + " ORDER BY p.fecha DESC").fetchall()
        
        conn.close()
        
        return jsonify({
            'ok': True,
            'pedidos': [dict(p) for p in pedidos]
        })
        
    except Exception as e:
        print(f"Error listando pedidos: {e}")
        return jsonify({'ok': False, 'error': str(e)}), 500

@app.route('/admin/pedidos/<int:pedido_id>/detalles')
@admin_required
def admin_pedido_detalles(pedido_id):
    """Obtiene detalles completos de un pedido"""
    try:
        conn = get_db_connection()
        
        # Datos del pedido
        pedido = conn.execute("""
            SELECT 
                p.*,
                t.nombre as cliente_nombre,
                t.telefono,
                t.direccion
            FROM pedidos p
            INNER JOIN terceros t ON p.id_tercero = t.id
            WHERE p.id = ?
        """, (pedido_id,)).fetchone()
        
        if not pedido:
            conn.close()
            return jsonify({'ok': False, 'error': 'Pedido no encontrado'}), 404
        
        # Productos del pedido
        productos = conn.execute("""
            SELECT 
                e.cantidad,
                m.nombre as medicamento,
                f.nombre as fabricante,
                m.concentracion,
                m.presentacion,
                pr.precio,
                m.imagen as imagen_generica,
                pr.imagen as imagen_especifica
            FROM existencias e
            INNER JOIN medicamentos m ON e.medicamento_id = m.id
            INNER JOIN fabricantes f ON e.fabricante_id = f.id
            LEFT JOIN precios pr ON pr.medicamento_id = m.id AND pr.fabricante_id = f.id
            WHERE e.pedido_id = ? AND e.tipo_movimiento = 'salida'
        """, (pedido_id,)).fetchall()
        
        conn.close()
        
        return jsonify({
            'ok': True,
            'pedido': dict(pedido),
            'productos': [dict(p) for p in productos]
        })
        
    except Exception as e:
        print(f"Error obteniendo detalles: {e}")
        return jsonify({'ok': False, 'error': str(e)}), 500

@app.route('/admin/pedidos/<int:pedido_id>/cambiar_estado', methods=['POST'])
@admin_required
def admin_cambiar_estado_pedido(pedido_id):
    """Cambia el estado de un pedido y agrega medicamentos al pastillero cuando se entrega"""
    conn = None
    try:
        data = request.get_json()
        nuevo_estado = data.get('estado')

        if nuevo_estado not in ['pendiente', 'en_camino', 'entregado', 'cancelado']:
            return jsonify({'ok': False, 'error': 'Estado invlido'}), 400

        conn = get_db_connection()

        # Actualizar estado del pedido
        conn.execute("""
            UPDATE pedidos
            SET estado = %s
            WHERE id = %s
        """, (nuevo_estado, pedido_id))

        # Si el estado es "entregado", agregar medicamentos al pastillero
        if nuevo_estado == 'entregado':
            # 1. Obtener información del pedido y el cliente
            pedido = conn.execute("""
                SELECT p.id_tercero, t.telefono, t.nombre
                FROM pedidos p
                INNER JOIN terceros t ON p.id_tercero = t.id
                WHERE p.id = %s
            """, (pedido_id,)).fetchone()

            if pedido and pedido['telefono']:
                # 2. Buscar si el cliente tiene cuenta de pastillero (por teléfono)
                usuario_pastillero = conn.execute("""
                    SELECT id FROM terceros
                    WHERE telefono = %s AND id IN (
                        SELECT DISTINCT usuario_id FROM pastillero_usuarios WHERE usuario_id IS NOT NULL
                    )
                    LIMIT 1
                """, (pedido['telefono'],)).fetchone()

                # Si no tiene cuenta de pastillero, usar el tercero_id como usuario_id
                usuario_id = usuario_pastillero['id'] if usuario_pastillero else pedido['id_tercero']

                # 3. Obtener los medicamentos del pedido (desde existencias)
                medicamentos_pedido = conn.execute("""
                    SELECT e.medicamento_id, e.cantidad, m.nombre
                    FROM existencias e
                    INNER JOIN medicamentos m ON e.medicamento_id = m.id
                    WHERE e.pedido_id = %s AND e.tipo_movimiento = 'salida'
                """, (pedido_id,)).fetchall()

                # 4. Agregar cada medicamento al pastillero
                medicamentos_agregados = []
                for med in medicamentos_pedido:
                    medicamento_id = med['medicamento_id']
                    cantidad = med['cantidad']
                    nombre = med['nombre']

                    # Extraer nombre base normalizado
                    nombre_normalizado = extraer_nombre_base(nombre)

                    # Verificar si ya existe en el pastillero
                    existe = conn.execute("""
                        SELECT id, cantidad FROM pastillero_usuarios
                        WHERE usuario_id = %s AND medicamento_id = %s
                    """, (usuario_id, medicamento_id)).fetchone()

                    if existe:
                        # Sumar a la cantidad existente
                        conn.execute("""
                            UPDATE pastillero_usuarios
                            SET cantidad = cantidad + %s,
                                fecha_actualizado = CURRENT_TIMESTAMP
                            WHERE id = %s
                        """, (cantidad, existe['id']))
                    else:
                        # Insertar nuevo medicamento
                        conn.execute("""
                            INSERT INTO pastillero_usuarios
                            (usuario_id, medicamento_id, nombre, cantidad, unidad, fecha_agregado)
                            VALUES (%s, %s, %s, %s, 'pastillas', CURRENT_TIMESTAMP)
                        """, (usuario_id, medicamento_id, nombre_normalizado, cantidad))

                    medicamentos_agregados.append(nombre_normalizado)

                print(f"[PASTILLERO] Agregados {len(medicamentos_agregados)} medicamentos al usuario {usuario_id}")

        conn.commit()
        conn.close()

        mensaje = f'Estado cambiado a: {nuevo_estado}'
        if nuevo_estado == 'entregado':
            mensaje += ' - Medicamentos agregados al pastillero del cliente'

        return jsonify({
            'ok': True,
            'mensaje': mensaje
        })

    except Exception as e:
        if conn:
            conn.rollback()
            conn.close()
        print(f"Error cambiando estado: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'ok': False, 'error': str(e)}), 500
    

# ============================================
# FUNCIN HELPER 1: NORMALIZAR PALABRAS (SINGULAR/PLURAL)
#  VERSIN: 2025-11-08 18:30 - Deteccin flexible de plurales
# ============================================

def normalizar_palabra_busqueda(palabra):
    """
    Genera variantes de bsqueda (singular y plural).
    
    Ejemplos:
    - "hormonas" -> ["hormonas", "hormona"]
    - "hormona" -> ["hormona", "hormonas"]
    - "dolor" -> ["dolor", "dolores"]
    - "dolores" -> ["dolores", "dolor"]
    - "infecciones" -> ["infecciones", "infeccion"]
    
    Retorna: lista de variantes nicas
    
     HUELLA: 2025-11-08 18:30
    """
    from datetime import datetime
    
    variantes = [palabra]  # Siempre incluir la original
    
    # DEBUG: Print de entrada
    # print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] normalizar_palabra_busqueda('{palabra}')")
    
    # Caso 1: Termina en "es" (plural de palabras con -n, -or, consonante)
    # Ejemplos: "infecciones" -> "infeccion", "dolores" -> "dolor"
    if len(palabra) > 3 and palabra.endswith('es'):
        # Quitar "es"
        singular = palabra[:-2]
        variantes.append(singular)
        
        # Si la palabra sin "es" termina en vocal, tambin agregamos sin la vocal final
        # Ejemplo: "infecciones" -> "infeccion" y tambin podra ser "infeccin" en BD
        if singular and singular[-1] in 'aeiou':
            variantes.append(singular)
    
    # Caso 2: Termina en "s" simple (plural regular)
    # Ejemplos: "hormonas" -> "hormona", "sntomas" -> "sntoma"
    elif len(palabra) > 2 and palabra.endswith('s') and not palabra.endswith('es'):
        singular = palabra[:-1]
        variantes.append(singular)
    
    # Caso 3: No termina en "s" (es singular)
    # Agregar variante plural
    else:
        # Si termina en vocal, agregar solo "s"
        if palabra[-1] in 'aeiou':
            variantes.append(palabra + 's')
        # Si termina en consonante (excepto 's'), agregar "es"
        else:
            variantes.append(palabra + 'es')
    
    # Retornar lista sin duplicados
    resultado = list(set(variantes))
    print(f"   -> Variantes generadas: {resultado}")
    return resultado


# ============================================
# FUNCIN HELPER 2: DISTANCIA DE LEVENSHTEIN
#  VERSIN: 2025-11-09 07:30
# ============================================

def distancia_levenshtein(s1, s2):
    """
    Calcula la distancia de edicin entre dos strings.
    Retorna: nmero de operaciones (insercin, eliminacin, sustitucin) necesarias.
    
    Ejemplos:
    - distancia_levenshtein("hormona", "hormonal") = 1
    - distancia_levenshtein("hormonas", "hormonal") = 2
    """
    if len(s1) > len(s2):
        s1, s2 = s2, s1
    
    distances = range(len(s1) + 1)
    for i2, c2 in enumerate(s2):
        distances_ = [i2 + 1]
        for i1, c1 in enumerate(s1):
            if c1 == c2:
                distances_.append(distances[i1])
            else:
                distances_.append(1 + min((distances[i1], distances[i1 + 1], distances_[-1])))
        distances = distances_
    return distances[-1]


# ============================================
# FUNCIN HELPER 3: DETECTAR DIAGNSTICO POR PALABRAS CONTINUAS
#  VERSIN: 2025-11-09 07:30 - CON SUBSTRING + LEVENSHTEIN
# ============================================

def detectar_diagnostico_por_palabras(nombre_diagnostico, palabras_usuario, umbral=0.8):
    """
    Detecta si el diagnstico est presente en las palabras del usuario.
    
    Reglas:
    - Para diagnsticos de 1 palabra: requiere 100% (la palabra completa)
    - Para diagnsticos multi-palabra: requiere 80% de las palabras (umbral configurable)
    - Las palabras pueden estar en desorden
    - Usa SUBSTRING + DISTANCIA DE LEVENSHTEIN MS ESTRICTO
    
    Retorna: True si cumple las reglas, False si no
    
     HUELLA: 2025-11-16 - Refinado para evitar falsos positivos
    """
    palabras_diag = nombre_diagnostico.lower().split()
    num_palabras_diag = len(palabras_diag)
    
    # Palabras a ignorar (conectores comunes)
    palabras_ignorar = {'de', 'del', 'la', 'el', 'los', 'las', 'en', 'por', 'para', 'con', 
                       'y', 'a', 'un', 'una', 'al'}
    palabras_diag_filtradas = [p for p in palabras_diag if p not in palabras_ignorar]
    
    if not palabras_diag_filtradas:
        return False
    
    # Calcular cuntas palabras necesitamos (mnimo segn umbral)
    
    palabras_requeridas = max(1, math.ceil(len(palabras_diag_filtradas) * umbral))
    
    # Buscar secuencias continuas de palabras del diagnstico
    palabras_encontradas = []
    
    for palabra_diag in palabras_diag_filtradas:
        encontrada = False
        
        for palabra_usuario in palabras_usuario:
            if encontrada:
                break
            
            # Mtodo 1: Normalizacin de plurales (exacto)
            variantes_diag = normalizar_palabra_busqueda(palabra_diag)
            variantes_usuario = normalizar_palabra_busqueda(palabra_usuario)
            
            # Comparacin exacta con variantes
            for var_diag in variantes_diag:
                for var_user in variantes_usuario:
                    if var_diag == var_user:
                        encontrada = True
                        print(f"       Match exacto: '{palabra_usuario}' ~ '{palabra_diag}'")
                        break
                if encontrada:
                    break
            
            if encontrada:
                continue
            
            # Mtodo 2: SUBSTRING (una palabra contiene a la otra)
            for var_diag in variantes_diag:
                for var_user in variantes_usuario:
                    # Si una est contenida en la otra
                    if len(var_diag) >= 6 and len(var_user) >= 6:  #  Aumentado de 5 a 6 letras
                        if var_diag in var_user or var_user in var_diag:
                            encontrada = True
                            print(f"       Match substring: '{palabra_usuario}' contiene '{palabra_diag}'")
                            break
                if encontrada:
                    break
            
            if encontrada:
                continue
            
            # Mtodo 3: DISTANCIA DE LEVENSHTEIN (similar) - MS ESTRICTO
            for var_diag in variantes_diag:
                for var_user in variantes_usuario:
                    #  Solo si ambas palabras son LARGAS (7+ letras) para evitar "duel"~"piel"
                    if len(var_diag) >= 7 and len(var_user) >= 7:
                        distancia = distancia_levenshtein(var_diag, var_user)
                        #  Distancia MS ESTRICTA: mximo 20% de la palabra
                        max_distancia = max(1, int(min(len(var_diag), len(var_user)) * 0.2))
                        if distancia <= max_distancia:
                            encontrada = True
                            print(f"       Match Levenshtein: '{palabra_usuario}' ~ '{palabra_diag}' (distancia={distancia}, max={max_distancia})")
                            break
                if encontrada:
                    break
        
        if encontrada and palabra_diag not in palabras_encontradas:
            palabras_encontradas.append(palabra_diag)
    
    porcentaje = len(palabras_encontradas) / len(palabras_diag_filtradas)
    cumple = len(palabras_encontradas) >= palabras_requeridas
    
    print(f"    Diagnstico '{nombre_diagnostico}': {len(palabras_encontradas)}/{len(palabras_diag_filtradas)} palabras = {porcentaje*100:.0f}% (requiere {palabras_requeridas} = {umbral*100:.0f}%) -> {'' if cumple else ''}")
    
    return cumple



def verificar_match_sintoma(nombre_sintoma, palabras_usuario, umbral=0.8):
    """
    Verifica si el sntoma hace match con las palabras del usuario.
    Similar a detectar_diagnostico_por_palabras pero para sntomas.
    
    Reglas:
    - Match exacto (completo) -> siempre True
    - Multi-palabra: requiere 80% de coincidencia
    - Evita matches parciales como "infeccion" -> "infeccion Clamidia"
    
     VERSIN: 2025-11-16 - Match estricto de sntomas (BUGFIX breaks)
    """
    palabras_sintoma = normalizar_texto(nombre_sintoma).split()
    
    # Palabras a ignorar
    palabras_ignorar = {'de', 'del', 'la', 'el', 'los', 'las', 'en', 'por', 'para', 'con', 'y', 'a'}
    palabras_sintoma_filtradas = [p for p in palabras_sintoma if p not in palabras_ignorar]
    
    if not palabras_sintoma_filtradas:
        return False
    
    # Si es una sola palabra, debe ser match exacto
    if len(palabras_sintoma_filtradas) == 1:
        palabra_sintoma = palabras_sintoma_filtradas[0]
        variantes_sintoma = normalizar_palabra_busqueda(palabra_sintoma)
        
        for palabra_usuario in palabras_usuario:
            variantes_usuario = normalizar_palabra_busqueda(palabra_usuario)
            # Match exacto
            for var_sint in variantes_sintoma:
                if var_sint in variantes_usuario:
                    return True
        return False
    
    # Si es multi-palabra, aplicar umbral 80%
    import math
    palabras_requeridas = math.ceil(len(palabras_sintoma_filtradas) * umbral)
    palabras_encontradas = 0

    # Palabras calificadoras que pueden ser opcionales
    calificadores = {'intensa', 'leve', 'aguda', 'cronica', 'severa', 'moderada',
                     'persistente', 'temporal', 'ocasional', 'frecuente', 'constante',
                     'nasal', 'estomacal', 'abdominal', 'muscular', 'articular'}

    #  BUGFIX: Contar TODAS las palabras que hacen match
    for palabra_sintoma in palabras_sintoma_filtradas:
        encontrada_esta_palabra = False
        variantes_sintoma = normalizar_palabra_busqueda(palabra_sintoma)

        for palabra_usuario in palabras_usuario:
            if encontrada_esta_palabra:  # Ya encontramos esta palabra del sntoma
                break

            variantes_usuario = normalizar_palabra_busqueda(palabra_usuario)

            # Match exacto de variantes
            for var_sint in variantes_sintoma:
                for var_user in variantes_usuario:
                    if var_sint == var_user:
                        palabras_encontradas += 1
                        encontrada_esta_palabra = True
                        break
                if encontrada_esta_palabra:
                    break

    # Regla especial: si el sintoma tiene 2 palabras y una es calificador,
    # aceptar match con solo la palabra principal
    if len(palabras_sintoma_filtradas) == 2:
        tiene_calificador = any(p in calificadores for p in palabras_sintoma_filtradas)
        if tiene_calificador and palabras_encontradas >= 1:
            return True

    cumple = palabras_encontradas >= palabras_requeridas
    return cumple



# ============================================
# FUNCIN PRINCIPAL: OBTENER PRODUCTOS
#  VERSIN: 2025-11-08 18:30 - COMPLETA CON TODAS LAS MEJORAS
# ============================================

def buscar_medicamentos_directos(busqueda, conn, precio_min='', precio_max='', permitir_sin_cotizaciones=0):
    """
    Busca medicamentos por nombre/fabricante/componente activo.

    Args:
        busqueda: Texto de bsqueda
        conn: Conexin a BD
        precio_min: Precio mnimo (opcional)
        precio_max: Precio mximo (opcional)
        permitir_sin_cotizaciones: Flag para permitir productos sin cotizaciones

    Returns:
        tuple: (productos, palabras_sin_match, porcentaje_exito)
    """
    # Validar entrada
    if not busqueda or not busqueda.strip():
        return [], [], 0

    # Filtrar stopwords (palabras muy comunes sin significado médico)
    stopwords = {'de', 'del', 'la', 'el', 'los', 'las', 'un', 'una', 'unos', 'unas', 'a', 'con', 'en', 'por', 'para', 'y', 'o', 'que'}
    palabras_busqueda = [p for p in normalizar_texto(busqueda).split() if p not in stopwords]
    if not palabras_busqueda:
        return [], [], 0

    print(f"\nBUSQUEDA DIRECTA: {palabras_busqueda}")

    CAMPOS_BUSQUEDA = 3  # nombre, fabricante, componente_activo

    query = """SELECT DISTINCT p.id as precio_id, p.medicamento_id, p.fabricante_id, p.precio, p.imagen as imagen_precio,
        m.nombre as medicamento_nombre, m.presentacion, m.concentracion, m.imagen as imagen_medicamento,
        m.componente_activo_id, ca.nombre as componente_activo_nombre, f.nombre as fabricante_nombre,
        s.nombre as sintoma_nombre, ms.sintoma_id
        FROM precios p
        INNER JOIN medicamentos m ON p.medicamento_id = m.id
        INNER JOIN fabricantes f ON p.fabricante_id = f.id
        LEFT JOIN medicamentos ca ON m.componente_activo_id = ca.id
        LEFT JOIN medicamento_sintoma ms ON m.id = ms.medicamento_id
        LEFT JOIN sintomas s ON ms.sintoma_id = s.id
        LEFT JOIN (
            SELECT medicamento_id, fabricante_id, COUNT(*) as num_cotizaciones
            FROM precios_competencia
            GROUP BY medicamento_id, fabricante_id
        ) cot ON p.medicamento_id = cot.medicamento_id AND p.fabricante_id = cot.fabricante_id
        WHERE m.activo = '1'"""

    params = []
    condiciones = []

    for palabra in palabras_busqueda:
        for variante in normalizar_palabra_busqueda(palabra):
            # Usar la misma normalización que en test4 (con tildes correctas)
            sql_norm = "LOWER(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE({}, 'á','a'), 'é','e'), 'í','i'), 'ó','o'), 'ú','u'), 'ñ','n'))"

            # 🆕 Estrategia de búsqueda según longitud de palabra:
            # - Palabras CORTAS (≤3 letras): Solo palabra COMPLETA (evita "me" → "Melatonina")
            # - Palabras LARGAS (>3 letras): Palabra completa Y subcadena (permite "aspirin" → "Aspirina")

            for campo in ['m.nombre', 'f.nombre', 'ca.nombre']:
                campo_norm = sql_norm.format(campo)

                if len(variante) <= 3:
                    # Palabras cortas: SOLO palabra completa
                    condiciones.extend([
                        f"{campo_norm} LIKE ?",  # inicio: "me tableta"
                        f"{campo_norm} LIKE ?",  # final: "tableta me"
                        f"{campo_norm} LIKE ?",  # medio: "dolor me forte"
                        f"{campo_norm} = ?"      # exacta: "me"
                    ])
                    params.extend([
                        f'{variante} %',
                        f'% {variante}',
                        f'% {variante} %',
                        f'{variante}'
                    ])
                else:
                    # Palabras largas: Palabra completa + subcadena (más flexible)
                    condiciones.extend([
                        f"{campo_norm} LIKE ?",  # inicio palabra completa
                        f"{campo_norm} LIKE ?",  # final palabra completa
                        f"{campo_norm} LIKE ?",  # medio palabra completa
                        f"{campo_norm} = ?",     # exacta
                        f"{campo_norm} LIKE ?"   # subcadena (permite "aspirin" → "Aspirina")
                    ])
                    params.extend([
                        f'{variante} %',
                        f'% {variante}',
                        f'% {variante} %',
                        f'{variante}',
                        f'%{variante}%'          # búsqueda parcial
                    ])

    if condiciones:
        query += f" AND ({' OR '.join(condiciones)})"

    if precio_min:
        try:
            query += " AND p.precio >= %s"
            params.append(float(precio_min))
        except ValueError:
            pass
    if precio_max:
        try:
            query += " AND p.precio <= %s"
            params.append(float(precio_max))
        except ValueError:
            pass

    query += " AND (%s = 1 OR COALESCE(cot.num_cotizaciones, 0) > 0) AND p.precio > 0"
    params.append(permitir_sin_cotizaciones)

    productos = conn.execute(query, params).fetchall()
    print(f"   Productos encontrados: {len(productos)}")

    palabras_con_match = set()
    medicamentos_por_palabra = {palabra: [] for palabra in palabras_busqueda}  # 🆕 DEBUG

    for p in productos:
        # 🆕 FIX: Normalizar nombres igual que en el query (quitar tildes, minúsculas)
        med_nombre = normalizar_texto(p['medicamento_nombre']) if p['medicamento_nombre'] else ''
        fab_nombre = normalizar_texto(p['fabricante_nombre']) if p['fabricante_nombre'] else ''
        comp_nombre = normalizar_texto(p['componente_activo_nombre']) if p['componente_activo_nombre'] else ''

        for palabra in palabras_busqueda:
            if palabra in med_nombre or palabra in fab_nombre or palabra in comp_nombre:
                palabras_con_match.add(palabra)
                medicamentos_por_palabra[palabra].append(p['medicamento_nombre'])  # 🆕 Guardar nombre original para debug

    # 🆕 MANTENER ORDEN ORIGINAL: usar lista en lugar de list comprehension para preservar el orden
    palabras_sin_match = []
    for p in palabras_busqueda:
        if p not in palabras_con_match:
            palabras_sin_match.append(p)

    porcentaje = (len(palabras_con_match) / len(palabras_busqueda)) * 100 if palabras_busqueda else 0

    # 🆕 DEBUG: Mostrar detalle de matches
    print(f"\n[BUSQUEDA DIRECTA]")
    print(f"   Palabras CON match ({len(palabras_con_match)}): {list(palabras_con_match)}")
    print(f"   Palabras SIN match ({len(palabras_sin_match)}): {palabras_sin_match}")
    print(f"   Porcentaje de exito: {porcentaje:.0f}%")

    for palabra in palabras_busqueda:
        if palabra in palabras_con_match:
            ejemplos = medicamentos_por_palabra[palabra][:3]  # Mostrar max 3 ejemplos
            print(f"      '{palabra}' -> {len(medicamentos_por_palabra[palabra])} meds (ej: {ejemplos})")
        else:
            print(f"      '{palabra}' -> NO encontrado")
    print()

    return productos, palabras_sin_match, porcentaje


@app.route('/api/test_sintoma', methods=['GET'])
def test_sintoma():
    """Endpoint de prueba para verificar síntoma y medicamentos asociados"""
    sintoma = request.args.get('q', 'comezon').strip()

    try:
        conn = get_db_connection()

        # Test 1: Buscar el síntoma en la BD
        sintoma_normalizado = normalizar_texto(sintoma)
        test1 = conn.execute("""
            SELECT id, nombre
            FROM sintomas
            WHERE LOWER(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(
                nombre, 'á','a'), 'é','e'), 'í','i'), 'ó','o'), 'ú','u'), 'ñ','n')
            ) LIKE %s
            LIMIT 10
        """, (f'%{sintoma_normalizado}%',)).fetchall()

        # Test 2: Si encontró síntomas, buscar medicamentos asociados
        test2 = []
        test3 = []  # Con filtro de cotizaciones
        if test1:
            sintoma_ids = [s['id'] for s in test1]
            placeholders = ','.join(['%s'] * len(sintoma_ids))

            # Test 2: TODOS los medicamentos asociados
            test2 = conn.execute(f"""
                SELECT DISTINCT
                    m.id,
                    m.nombre,
                    s.nombre as sintoma_nombre,
                    COUNT(p.id) as num_precios
                FROM medicamentos m
                INNER JOIN medicamento_sintoma ms ON m.id = ms.medicamento_id
                INNER JOIN sintomas s ON ms.sintoma_id = s.id
                LEFT JOIN precios p ON m.id = p.medicamento_id AND p.precio > 0
                WHERE ms.sintoma_id IN ({placeholders})
                AND m.activo = '1'
                GROUP BY m.id, m.nombre, s.nombre
                ORDER BY num_precios DESC
            """, sintoma_ids).fetchall()

            # Test 3: Solo medicamentos CON cotizaciones
            test3 = conn.execute(f"""
                SELECT DISTINCT
                    m.id,
                    m.nombre,
                    p.precio,
                    COALESCE(cot.num_cotizaciones, 0) as num_cotizaciones
                FROM medicamentos m
                INNER JOIN medicamento_sintoma ms ON m.id = ms.medicamento_id
                INNER JOIN precios p ON m.id = p.medicamento_id
                LEFT JOIN (
                    SELECT medicamento_id, fabricante_id, COUNT(*) as num_cotizaciones
                    FROM precios_competencia
                    GROUP BY medicamento_id, fabricante_id
                ) cot ON p.medicamento_id = cot.medicamento_id AND p.fabricante_id = cot.fabricante_id
                WHERE ms.sintoma_id IN ({placeholders})
                AND m.activo = '1'
                AND p.precio > 0
                AND COALESCE(cot.num_cotizaciones, 0) > 0
            """, sintoma_ids).fetchall()

        conn.close()

        return jsonify({
            'busqueda': sintoma,
            'normalizado': sintoma_normalizado,
            'sintomas_encontrados': [dict(s) for s in test1],
            'total_medicamentos_asociados': len(test2),
            'medicamentos_asociados_muestra': [dict(m) for m in test2[:5]],
            'medicamentos_con_cotizaciones': len(test3),
            'con_cotizaciones_muestra': [dict(m) for m in test3[:5]],
            'mensaje': f'De {len(test2)} medicamentos asociados, solo {len(test3)} tienen cotizaciones. Ese es el problema.'
        })
    except Exception as e:
        import traceback
        return jsonify({'error': str(e), 'traceback': traceback.format_exc()}), 500


@app.route('/api/productos', methods=['GET'])
def obtener_productos():
    """
    API para obtener productos con filtros + deteccin inteligente de diagnsticos y sntomas
    
     VERSIN: 2025-11-09 07:30 - FINAL CON SUBSTRING + LEVENSHTEIN
     MEJORAS:
       - Deteccin flexible de plurales en sntomas Y diagnsticos
       - Diagnsticos multi-palabra con substring + distancia Levenshtein
       - Regla 60% para diagnsticos multi-palabra
       - Detecta sntomas faltantes en diagnsticos directos
       - Interfaz mejorada con tipos de deteccin
    """
    from datetime import datetime
    try:
        print(f"\n{'='*70}")
        print(f" API /api/productos - VERSIN: 2025-11-09 07:30 FINAL")
        print(f"   Timestamp: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print(f"   Funcionalidad: Substring + Levenshtein ACTIVOS")
        print(f"{'='*70}\n")
    
        # Parmetros de bsqueda
        busqueda = request.args.get('q', '').strip()
        sintoma_id = request.args.get('sintoma_id', '')
        precio_min = request.args.get('precio_min', '')
        precio_max = request.args.get('precio_max', '')
        busqueda_sintomas = request.args.get('sintomas_busqueda', '').strip()
        categoria_id = request.args.get('categoria_id', '').strip()

        # Parámetros de paginación
        page = request.args.get('page', '1')
        limit = request.args.get('limit', '30')
        try:
            page = int(page)
            limit = int(limit)
            if page < 1:
                page = 1
            if limit < 1 or limit > 100:
                limit = 30
        except ValueError:
            page = 1
            limit = 30

        try:
            conn = get_db_connection()

            # Cargar configuracin de publicacin
            config_row = conn.execute("SELECT permitir_publicar_sin_cotizaciones FROM CONFIGURACION_PRECIOS LIMIT 1").fetchone()
            permitir_sin_cotizaciones = config_row[0] if config_row else 0
        except Exception as e:
            print(f"ERROR en obtener_productos al inicio: {e}")
            import traceback
            traceback.print_exc()
            return jsonify({'ok': False, 'error': str(e), 'productos': []}), 500

        #  PASO 1: Buscar medicamentos directos
        productos_directos = []
        if busqueda:
            productos_directos, palabras_sin_match, porcentaje_exito = buscar_medicamentos_directos(busqueda, conn, precio_min, precio_max, permitir_sin_cotizaciones)
            # 🆕 REGLA: Si hay palabras sin match, buscar TAMBIÉN por síntomas
            # IMPORTANTE: Usar búsqueda ORIGINAL (con stopwords) para detección de síntomas contextuales
            if palabras_sin_match:
                busqueda_sintomas = busqueda  # Usar texto original para preservar contexto (ej: "dolor DE espalda")
                print(f"   Activando busqueda por sintomas (texto original): '{busqueda_sintomas}'")

        # DETECCIN INTELIGENTE: DIAGNSTICOS + SNTOMAS
        sintomas_detectados = []
        sintomas_detectados_ids = []
        diagnosticos_detectados_directo = []
        diagnosticos_detectados_directo_ids = []
        diagnosticos_posibles = {}
        sintomas_faltantes_por_diagnostico = {}
        busqueda_parcial_aplicada = False

        #  PASO 2: Procesar sntomas (si aplica)
        if busqueda_sintomas:
            # ============================================
            # PARTE 1: NORMALIZACIN (SIN CAMBIOS)
            # ============================================
            normalizaciones_diagnosticos = {
                'gripa': 'gripe',
                'resfriado': 'gripe',
                'resfrio': 'gripe',
                'gripal': 'gripe',
                'resfro': 'gripe',
                'catarro': 'gripe',
                'jaqueca': 'migraa',
                'migrania': 'migraa',
                'migrana': 'migraa',
                'cefalea': 'migraa',
                'acidez': 'gastritis',
                'agruras': 'gastritis',
                'reflujo': 'gastritis',
                'alergia': 'alergia estacional',
                'alergico': 'alergia estacional',
                'alergica': 'alergia estacional',
                'rinitis': 'alergia estacional',
                'colitis': 'gastroenteritis',
                'diarrea': 'gastroenteritis',
                'angina': 'amigdalitis',
                'anginas': 'amigdalitis',
                'cistitis': 'infeccin urinaria'
            }
        
            normalizaciones_sintomas = {
                'toso': 'tos', 'tosiendo': 'tos', 'toser': 'tos',
                'duele': 'dolor', 'doliendo': 'dolor', 'doler': 'dolor',
                'arde': 'ardor', 'ardiendo': 'ardor', 'arder': 'ardor',
                'pica': 'picazn', 'picando': 'picazn', 'picar': 'picazn',
                'vomito': 'vmito', 'vomitando': 'vmito', 'vomitar': 'vmito',
                'mareo': 'mareo', 'mareando': 'mareo', 'marear': 'mareo',
                'estornudo': 'estornudo', 'estornudando': 'estornudo', 'estornudar': 'estornudo',
                'orino': 'orinar', 'orinando': 'orinar',
                'sudo': 'sudor', 'sudando': 'sudor', 'sudar': 'sudor'
            }
        
            texto_normalizado = normalizar_texto(busqueda_sintomas)
        
            for forma_verbal, sustantivo in normalizaciones_sintomas.items():
                texto_normalizado = texto_normalizado.replace(forma_verbal, sustantivo)
        
            palabras = texto_normalizado.split()
        
            # ============================================
            #  FILTRO PREVIO: Eliminar palabras intiles
            # ============================================
            palabras_ignorar_busqueda = {
                'es', 'que', 'me', 'la', 'el', 'en', 'de', 'y', 'a', 'un', 'una',
                'mi', 'tu', 'su', 'lo', 'le', 'se', 'si', 'no', 'con', 'por',
                'para', 'como', 'pero', 'muy', 'mas', 'o', 'del', 'al', 'las',
                'los', 'esta', 'este', 'son', 'hay', 'fue', 'ser', 'estar', 'tengo',
                'tiene', 'siento'
            }
        
            # Filtrar palabras ignorables (mnimo 3 letras)
            palabras_filtradas = [p for p in palabras if p not in palabras_ignorar_busqueda and len(p) >= 3]
        
            #  REGLA CRTICA: Verificar si hay palabras significativas (5+ letras)
            palabras_significativas = [p for p in palabras_filtradas if len(p) >= 5]
        
            print(f"   Palabras originales: {palabras}")
            print(f"   Palabras filtradas: {palabras_filtradas}")
            print(f"   Palabras significativas (5+ letras): {palabras_significativas}")
        
            # ============================================
            # PARTE 2: BUSCAR DIAGNSTICOS DIRECTAMENTE ( MEJORADO CON 80% Y FILTROS)
            # ============================================
        
            #  SOLO buscar diagnsticos si hay palabras significativas
            if palabras_significativas:
                print("\n PARTE 2: Buscando diagnsticos directamente...")
            
                # Primero buscar con trigramas/bigramas en el diccionario
                for i in range(len(palabras_filtradas) - 2):
                    trigrama = f"{palabras_filtradas[i]} {palabras_filtradas[i+1]} {palabras_filtradas[i+2]}"
                    if trigrama in normalizaciones_diagnosticos:
                        diagnostico_normalizado = normalizaciones_diagnosticos[trigrama]
                        try:
                            query_diag = """
                                SELECT DISTINCT id, descripcion
                                FROM diagnosticos
                                WHERE LOWER(descripcion) LIKE ?
                                LIMIT 1
                            """
                            resultado = conn.execute(query_diag, [f'%{diagnostico_normalizado}%']).fetchone()
                            if resultado and resultado['id'] not in diagnosticos_detectados_directo_ids:
                                diagnosticos_detectados_directo.append(resultado['descripcion'])
                                diagnosticos_detectados_directo_ids.append(resultado['id'])
                                print(f"    Diagnstico directo (trigrama): {resultado['descripcion']}")
                        except Exception as e:
                            print(f"Error buscando diagnstico trigrama: {e}")
            
                for i in range(len(palabras_filtradas) - 1):
                    bigrama = f"{palabras_filtradas[i]} {palabras_filtradas[i+1]}"
                    if bigrama in normalizaciones_diagnosticos:
                        diagnostico_normalizado = normalizaciones_diagnosticos[bigrama]
                        try:
                            query_diag = """
                                SELECT DISTINCT id, descripcion
                                FROM diagnosticos
                                WHERE LOWER(descripcion) LIKE ?
                                LIMIT 1
                            """
                            resultado = conn.execute(query_diag, [f'%{diagnostico_normalizado}%']).fetchone()
                            if resultado and resultado['id'] not in diagnosticos_detectados_directo_ids:
                                diagnosticos_detectados_directo.append(resultado['descripcion'])
                                diagnosticos_detectados_directo_ids.append(resultado['id'])
                                print(f"    Diagnstico directo (bigrama): {resultado['descripcion']}")
                        except Exception as e:
                            print(f"Error buscando diagnstico bigrama: {e}")
            
                for palabra in palabras_filtradas:
                    if len(palabra) >= 5 and palabra in normalizaciones_diagnosticos:
                        diagnostico_normalizado = normalizaciones_diagnosticos[palabra]
                        try:
                            query_diag = """
                                SELECT DISTINCT id, descripcion
                                FROM diagnosticos
                                WHERE LOWER(descripcion) LIKE ?
                                LIMIT 1
                            """
                            resultado = conn.execute(query_diag, [f'%{diagnostico_normalizado}%']).fetchone()
                            if resultado and resultado['id'] not in diagnosticos_detectados_directo_ids:
                                diagnosticos_detectados_directo.append(resultado['descripcion'])
                                diagnosticos_detectados_directo_ids.append(resultado['id'])
                                print(f"    Diagnstico directo (palabra normalizada): {resultado['descripcion']}")
                        except Exception as e:
                            print(f"Error buscando diagnstico palabra: {e}")
            
                #  BSQUEDA CON REGLA 80% (antes era 60%)
                print("\n    Buscando con normalizacin de plurales y regla 80%...")
                try:
                    # Obtener TODOS los diagnsticos
                    query_todos_diag = "SELECT id, descripcion FROM diagnosticos"
                    todos_diagnosticos = conn.execute(query_todos_diag).fetchall()
                
                    for diag in todos_diagnosticos:
                        if diag['id'] not in diagnosticos_detectados_directo_ids:
                            # Verificar si cumple regla 80% con palabras continuas
                            if detectar_diagnostico_por_palabras(diag['descripcion'], palabras_filtradas, umbral=0.8):
                                diagnosticos_detectados_directo.append(diag['descripcion'])
                                diagnosticos_detectados_directo_ids.append(diag['id'])
                                print(f"    Diagnstico directo (80% match): {diag['descripcion']}")
                except Exception as e:
                    print(f"Error buscando diagnsticos con regla 80%: {e}")
            else:
                print("\n No hay palabras significativas (5+ letras), saltando bsqueda de diagnsticos")
        
            # ============================================
            # PARTE 3: BUSCAR SNTOMAS ( CON PLURALES)
            # ============================================
        
            print("\n PARTE 3: Buscando sntomas...")
        
            sintomas_encontrados = {}
        
            #  PASO 1: BUSCAR PALABRAS EXACTAS CON VARIANTES DE PLURAL/SINGULAR
            for palabra in palabras:
                if len(palabra) >= 3:  # Permitir palabras de 3+ letras (ej: "tos")
                    try:
                        #  Generar variantes (singular/plural)
                        variantes = normalizar_palabra_busqueda(palabra)
                        placeholders = ','.join(['?' for _ in variantes])
                    
                        print(f"    Buscando sntoma exacto con variantes: {variantes}")

                        query_exacta = f"""
                            SELECT DISTINCT id, nombre
                            FROM sintomas
                            WHERE LOWER(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(
                                    nombre, 'á', 'a'), 'é', 'e'), 'í', 'i'), 'ó', 'o'), 'ú', 'u'), 'ñ', 'n')
                                ) IN ({placeholders})
                            LIMIT 1
                        """
                        resultado = conn.execute(query_exacta, variantes).fetchone()
                        if resultado:
                            sintomas_encontrados[resultado['id']] = resultado['nombre']
                            print(f"    Sntoma encontrado: '{resultado['nombre']}' (ID: {resultado['id']})")
                        else:
                            print(f"    No se encontr sntoma con variantes: {variantes}")
                    except Exception as e:
                        print(f"Error buscando sntoma exacto: {e}")
        
            # PASO 2: BUSCAR CON CONTEXTO (lgica existente, SIN CAMBIOS)
            palabras_contextuales = {
                'dolor': ['cabeza', 'espalda', 'pecho', 'abdominal', 'estomago', 'garganta', 
                         'oido', 'ojo', 'ocular', 'muscular', 'articular', 'lumbar', 'cervical',
                         'pierna', 'brazo', 'muela', 'dental', 'rodilla', 'baja', 'cuello'],
                'tos': ['frecuencia', 'seca', 'flema', 'productiva'],
                'orinar': ['frecuencia', 'dolor', 'ardor'],
                'sudor': ['excesivo', 'nocturno', 'frio']
            }
        
            palabras_muy_cortas = ['dolo', 'dol', 'do']
        
            contextos_detectados = {}
            for i, palabra in enumerate(palabras):
                if palabra in palabras_muy_cortas:
                    continue
                
                if palabra in palabras_contextuales:
                    contexto_encontrado = []
                    for j in range(i+1, min(i+4, len(palabras))):
                        if palabras[j] in palabras_contextuales[palabra]:
                            contexto_encontrado.append(palabras[j])
                
                    if contexto_encontrado:
                        contextos_detectados[palabra] = contexto_encontrado
        
            for palabra_clave, contextos in contextos_detectados.items():
                for contexto in contextos:
                    try:
                        query = """
                            SELECT DISTINCT id, nombre, LENGTH(nombre) as len
                            FROM sintomas
                            WHERE (LOWER(nombre) LIKE ? AND LOWER(nombre) LIKE ?)
                               OR LOWER(nombre) LIKE ?
                            ORDER BY len ASC
                            LIMIT 2
                        """
                        frase_completa = f'%{palabra_clave}%{contexto}%'
                        resultados = conn.execute(query, [
                            f'%{palabra_clave}%', 
                            f'%{contexto}%',
                            frase_completa
                        ]).fetchall()
                    
                        for r in resultados:
                            if r['id'] not in sintomas_encontrados:  # No duplicar
                                sintomas_encontrados[r['id']] = r['nombre']
                    except Exception as e:
                        print(f"Error buscando sntoma con contexto: {e}")
        
            # PASO 3: BUSCAR PALABRAS SIMPLES CON LIKE Y VERIFICACIN ESTRICTA
            palabras_simples = [p for p in palabras_filtradas  #  Usar palabras_filtradas
                               if len(p) >= 5
                               and p not in palabras_contextuales.keys()
                               and p not in palabras_muy_cortas
                               and p not in ['mucho', 'poco', 'muy', 'esta', 'tengo', 'siento', 'baja', 'alta']]
        
            for palabra in palabras_simples:
                try:
                    # Generar variantes para buscar con LIKE
                    variantes = normalizar_palabra_busqueda(palabra)
                
                    # Buscar cada variante
                    for variante in variantes:
                        query = """
                            SELECT DISTINCT id, nombre, LENGTH(nombre) as len
                            FROM sintomas
                            WHERE LOWER(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(
                                    nombre, '', 'a'), '', 'e'), '', 'i'), '', 'o'), '', 'u'), '', 'n')
                                ) LIKE ?
                               OR LOWER(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(
                                    descripcion_lower, '', 'a'), '', 'e'), '', 'i'), '', 'o'), '', 'u'), '', 'n')
                                ) LIKE ?
                            ORDER BY len ASC
                            LIMIT 10
                        """
                        resultados = conn.execute(query, [f'%{variante}%', f'%{variante}%']).fetchall()
                    
                        for r in resultados:
                            if r['id'] not in sintomas_encontrados:  # No duplicar
                                #  VERIFICACIN ESTRICTA: Realmente hace match?
                                if verificar_match_sintoma(r['nombre'], palabras_filtradas):
                                    sintomas_encontrados[r['id']] = r['nombre']
                                    print(f"    Sntoma verificado: '{r['nombre']}'")
                                else:
                                    print(f"    Sntoma rechazado (match parcial): '{r['nombre']}'")
                except Exception as e:
                    print(f"Error buscando sntoma simple: {e}")
        
            for sid, nombre in sintomas_encontrados.items():
                sintomas_detectados.append(nombre)
                sintomas_detectados_ids.append(sid)
        
            print(f"\n RESUMEN - Sntomas detectados: {len(sintomas_detectados)}")
            for i, (sid, nombre) in enumerate(zip(sintomas_detectados_ids, sintomas_detectados)):
                print(f"   {i+1}. ID {sid}: {nombre}")
            print("")
        

            # ============================================
            #  BSQUEDA PARCIAL (ltimo recurso - una sola palabra sin resultados)
            # ============================================
            busqueda_parcial_aplicada = False
        
            if len(sintomas_detectados) == 0 and len(palabras_filtradas) == 1:
                palabra_busqueda = palabras_filtradas[0]
                print(f"\n BSQUEDA PARCIAL: No se encontraron sntomas exactos para '{palabra_busqueda}'")
                print(f"   Buscando sntomas que contengan esta palabra...")
            
                try:
                    # Buscar sntomas que contengan la palabra (normalizado)
                    query_parcial = """
                        SELECT DISTINCT id, nombre, LENGTH(nombre) as len
                        FROM sintomas
                        WHERE LOWER(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(
                                nombre, '', 'a'), '', 'e'), '', 'i'), '', 'o'), '', 'u'), '', 'n')
                            ) LIKE ?
                        ORDER BY len ASC
                        LIMIT 10
                    """
                
                    sintomas_parciales = conn.execute(query_parcial, [f'%{palabra_busqueda}%']).fetchall()
                
                    if sintomas_parciales:
                        print(f"    Encontrados {len(sintomas_parciales)} sntomas que contienen '{palabra_busqueda}':")
                        for s in sintomas_parciales:
                            sintomas_detectados.append(s['nombre'])
                            sintomas_detectados_ids.append(s['id'])
                            print(f"      - {s['nombre']}")
                    
                        busqueda_parcial_aplicada = True
                    else:
                        print(f"    No se encontraron sntomas que contengan '{palabra_busqueda}'")
                    
                except Exception as e:
                    print(f"    Error en bsqueda parcial: {e}")


            # ============================================
            # PARTE 4: DIAGNSTICOS POR SNTOMAS (SIN CAMBIOS)
            # ============================================
        
            if sintomas_detectados_ids:
                try:
                    placeholders_ids = ','.join(['?' for _ in sintomas_detectados_ids])
                    query_diagnosticos = f"""
                        SELECT
                            d.id,
                            d.descripcion,
                            COUNT(DISTINCT ds.sintoma_id) as sintomas_coincidentes,
                            (SELECT COUNT(*) FROM diagnostico_sintoma WHERE diagnostico_id = d.id) as sintomas_totales
                        FROM diagnosticos d
                        INNER JOIN diagnostico_sintoma ds ON d.id = ds.diagnostico_id
                        WHERE ds.sintoma_id IN ({placeholders_ids})
                        GROUP BY d.id
                        HAVING COUNT(DISTINCT ds.sintoma_id) > 0
                        ORDER BY COUNT(DISTINCT ds.sintoma_id) DESC
                        LIMIT 10
                    """
                
                    diagnosticos = conn.execute(query_diagnosticos, sintomas_detectados_ids).fetchall()
                
                    for diag in diagnosticos:
                        porcentaje_match = (diag['sintomas_coincidentes'] / diag['sintomas_totales']) * 100
                    
                        print(f"    Diagnstico evaluado: {diag['descripcion']}")
                        print(f"      Coincidencias: {diag['sintomas_coincidentes']}/{diag['sintomas_totales']} = {porcentaje_match:.1f}%")
                        print(f"      Pasa filtro 80%? {porcentaje_match > 80}")
                        print(f"      Ya est en directos? {diag['id'] in diagnosticos_detectados_directo_ids}")
                    
                        #  FILTRO 80%: Solo agregar si porcentaje > 80%
                        if porcentaje_match > 80 and diag['id'] not in diagnosticos_detectados_directo_ids:
                            diagnosticos_posibles[diag['id']] = {
                                'nombre': diag['descripcion'],
                                'coincidencias': diag['sintomas_coincidentes'],
                                'total_sintomas': diag['sintomas_totales'],
                                'porcentaje': porcentaje_match,
                                'tipo': 'por_sintomas'
                            }
                except Exception as e:
                    print(f"Error buscando diagnsticos por sntomas: {e}")
        
            # Agregar diagnsticos detectados directamente (siempre 100%)
            for i, diag_id in enumerate(diagnosticos_detectados_directo_ids):
                try:
                    total_sintomas_query = """
                        SELECT COUNT(*) as total
                        FROM diagnostico_sintoma
                        WHERE diagnostico_id = ?
                    """
                    total_result = conn.execute(total_sintomas_query, [diag_id]).fetchone()
                    total_sintomas = total_result['total'] if total_result else 1
                
                    diagnosticos_posibles[diag_id] = {
                        'nombre': diagnosticos_detectados_directo[i],
                        'coincidencias': total_sintomas,
                        'total_sintomas': total_sintomas,
                        'porcentaje': 100,
                        'tipo': 'directo'
                    }
                
                    #  DETECTAR SNTOMAS FALTANTES PARA DIAGNSTICOS DIRECTOS
                    try:
                        query_sintomas_diag = """
                            SELECT s.id, s.nombre
                            FROM diagnostico_sintoma ds
                            INNER JOIN sintomas s ON ds.sintoma_id = s.id
                            WHERE ds.diagnostico_id = ?
                        """
                        sintomas_del_diag = conn.execute(query_sintomas_diag, [diag_id]).fetchall()
                    
                        faltantes = []
                        for sint in sintomas_del_diag:
                            if sint['id'] not in sintomas_detectados_ids:
                                faltantes.append({
                                    'id': sint['id'],
                                    'nombre': sint['nombre']
                                })
                    
                        if faltantes:
                            sintomas_faltantes_por_diagnostico[diag_id] = faltantes
                            print(f"    Sntomas faltantes para '{diagnosticos_detectados_directo[i]}':")
                            for f in faltantes:
                                print(f"      - {f['nombre']} (ID: {f['id']})")
                    except Exception as e:
                        print(f"Error detectando sntomas faltantes: {e}")
                
                except Exception as e:
                    print(f"Error agregando diagnstico directo: {e}")
    
        # ============================================
        # QUERY DE PRODUCTOS (SIN CAMBIOS)
        # ============================================
    
        query = """
            SELECT DISTINCT
                p.id as precio_id,
                p.medicamento_id,
                p.fabricante_id,
                p.precio,
                p.imagen as imagen_precio,
                m.nombre as medicamento_nombre,
                m.presentacion,
                m.concentracion,
                m.imagen as imagen_medicamento,
                m.componente_activo_id,
                ca.nombre as componente_activo_nombre,
                f.nombre as fabricante_nombre,
                s.nombre as sintoma_nombre,
                ms.sintoma_id
            FROM precios p
            INNER JOIN medicamentos m ON p.medicamento_id = m.id
            INNER JOIN fabricantes f ON p.fabricante_id = f.id
            LEFT JOIN medicamentos ca ON m.componente_activo_id = ca.id
            LEFT JOIN medicamento_sintoma ms ON m.id = ms.medicamento_id
            LEFT JOIN sintomas s ON ms.sintoma_id = s.id
            LEFT JOIN (
                SELECT medicamento_id, fabricante_id, COUNT(*) as num_cotizaciones
                FROM precios_competencia
                GROUP BY medicamento_id, fabricante_id
            ) cot ON p.medicamento_id = cot.medicamento_id AND p.fabricante_id = cot.fabricante_id
            WHERE m.activo = '1'
        """

        params = []

        #  PASO 3: Buscar productos por sntomas (si se detectaron)
        productos_sintomas = []
        if sintomas_detectados_ids:
            todos_sintomas_ids = list(sintomas_detectados_ids)
        
            for diag_id in diagnosticos_detectados_directo_ids:
                try:
                    query_sintomas_diag = """
                        SELECT DISTINCT sintoma_id
                        FROM diagnostico_sintoma
                        WHERE diagnostico_id = %s
                    """
                    sintomas_diag = conn.execute(query_sintomas_diag, [diag_id]).fetchall()
                    for s in sintomas_diag:
                        if s['sintoma_id'] not in todos_sintomas_ids:
                            todos_sintomas_ids.append(s['sintoma_id'])
                except Exception as e:
                    print(f"Error obteniendo sntomas de diagnstico: {e}")
        
            if todos_sintomas_ids:
                placeholders = ','.join(['%s' for _ in todos_sintomas_ids])
                # 🆕 FIX: Incluir información de síntomas con LEFT JOIN para obtener sintomas_ids_por_precio
                query_sintomas = """SELECT
                p.id as precio_id,
                p.medicamento_id,
                p.fabricante_id,
                p.precio,
                p.imagen as imagen_precio,
                m.nombre as medicamento_nombre,
                m.presentacion,
                m.concentracion,
                m.imagen as imagen_medicamento,
                m.componente_activo_id,
                ca.nombre as componente_activo_nombre,
                f.nombre as fabricante_nombre,
                s.nombre as sintoma_nombre,
                s.id as sintoma_id
            FROM precios p
            INNER JOIN medicamentos m ON p.medicamento_id = m.id
            INNER JOIN fabricantes f ON p.fabricante_id = f.id
            LEFT JOIN medicamentos ca ON m.componente_activo_id = ca.id
            LEFT JOIN medicamento_sintoma ms ON m.id = ms.medicamento_id
            LEFT JOIN sintomas s ON ms.sintoma_id = s.id
            LEFT JOIN (
                SELECT medicamento_id, fabricante_id, COUNT(*) as num_cotizaciones
                FROM precios_competencia
                GROUP BY medicamento_id, fabricante_id
            ) cot ON p.medicamento_id = cot.medicamento_id AND p.fabricante_id = cot.fabricante_id
            WHERE m.activo = '1'
            AND m.id IN (SELECT DISTINCT medicamento_id FROM medicamento_sintoma WHERE sintoma_id IN (""" + placeholders + """))"""
                params_sintomas = todos_sintomas_ids

                if precio_min:
                    try:
                        query_sintomas += " AND p.precio >= %s"
                        params_sintomas.append(float(precio_min))
                    except ValueError:
                        pass
                if precio_max:
                    try:
                        query_sintomas += " AND p.precio <= %s"
                        params_sintomas.append(float(precio_max))
                    except ValueError:
                        pass

                query_sintomas += " AND (%s = 1 OR COALESCE(cot.num_cotizaciones, 0) > 0) AND p.precio > 0"
                params_sintomas.append(permitir_sin_cotizaciones)

                productos_sintomas = conn.execute(query_sintomas, params_sintomas).fetchall()
                print(f"    Productos por sntomas: {len(productos_sintomas)}")

        #  PASO 4: Combinar resultados (directos primero, luego sntomas sin duplicados)
        # 🆕 PRIORIDAD DE RESULTADOS:
        #    1. Medicamentos encontrados por búsqueda directa (var_medicamentos_directos)
        #    2. Medicamentos encontrados por síntomas (solo si no están en directos)
        #    La búsqueda por síntomas usó palabras_sin_match en su orden original
        hay_limite = False
        if busqueda or busqueda_sintomas:
            # Combinar siempre: productos_directos primero, luego productos_sintomas sin duplicados
            ids_directos = {p['precio_id'] for p in productos_directos}
            productos_sintomas_unicos = [p for p in productos_sintomas if p['precio_id'] not in ids_directos]
            productos = list(productos_directos) + productos_sintomas_unicos
            print(f"   Combinado: {len(productos_directos)} directos + {len(productos_sintomas_unicos)} sintomas = {len(productos)}")
        else:
            # Sin bsqueda, mostrar primeros 50 (o filtrar por categoría)
            hay_limite = True

            # Filtro por categoría si se proporciona
            if categoria_id:
                query += """
                    AND m.id IN (
                        SELECT medicamento_id FROM medicamento_categoria
                        WHERE categoria_id = %s
                    )
                """
                params.append(int(categoria_id))

            query += " AND (%s = 1 OR COALESCE(cot.num_cotizaciones, 0) > 0) AND p.precio > 0"
            params.append(permitir_sin_cotizaciones)

            # Aplicar paginación: primero obtener precio_ids únicos paginados
            # Esto evita que el LIMIT se aplique a las filas antes del agrupamiento
            subquery_precios = """
                SELECT DISTINCT p.id, m.nombre as med_nombre
                FROM precios p
                INNER JOIN medicamentos m ON p.medicamento_id = m.id
                LEFT JOIN (
                    SELECT medicamento_id, fabricante_id, COUNT(*) as num_cotizaciones
                    FROM precios_competencia
                    GROUP BY medicamento_id, fabricante_id
                ) cot ON p.medicamento_id = cot.medicamento_id AND p.fabricante_id = cot.fabricante_id
                WHERE m.activo = '1'
            """

            subquery_params = []
            if categoria_id:
                subquery_precios += """
                    AND m.id IN (
                        SELECT medicamento_id FROM medicamento_categoria
                        WHERE categoria_id = %s
                    )
                """
                subquery_params.append(int(categoria_id))

            subquery_precios += " AND (%s = 1 OR COALESCE(cot.num_cotizaciones, 0) > 0) AND p.precio > 0"
            subquery_params.append(permitir_sin_cotizaciones)

            # Contar total de productos únicos
            count_query = f"SELECT COUNT(*) FROM ({subquery_precios}) AS subq"
            total_productos = conn.execute(count_query, subquery_params).fetchone()[0]

            # Aplicar LIMIT y OFFSET a los precio_ids únicos
            offset = (page - 1) * limit
            subquery_precios += f" ORDER BY med_nombre LIMIT {limit} OFFSET {offset}"

            # Usar los precio_ids paginados en el query principal (solo los IDs)
            query += f" AND p.id IN (SELECT id FROM ({subquery_precios}) AS precios_paginados)"
            params.extend(subquery_params)
            query += " ORDER BY m.nombre"

            productos = conn.execute(query, params).fetchall()

        total_disponible = len(productos)
        productos_agrupados = {}
        for p in productos:
            clave = p['precio_id']
            if clave not in productos_agrupados:
                productos_agrupados[clave] = {
                    'precio': p,
                    'sintomas': [],
                    'sintomas_ids': []
                }
            if p['sintoma_nombre']:
                productos_agrupados[clave]['sintomas'].append(p['sintoma_nombre'])
                if p['sintoma_id']:
                    productos_agrupados[clave]['sintomas_ids'].append(p['sintoma_id'])
    
        productos = [v['precio'] for v in productos_agrupados.values()]
        sintomas_por_precio = {k: v['sintomas'] for k, v in productos_agrupados.items()}
        sintomas_ids_por_precio = {k: v['sintomas_ids'] for k, v in productos_agrupados.items()}
    
        sintomas_filtrados_por_precio = {}
        if busqueda_sintomas and sintomas_detectados:
            for precio_id, sintomas_lista in sintomas_por_precio.items():
                sintomas_filtrados = [s for s in sintomas_lista if s in sintomas_detectados]
                sintomas_filtrados_por_precio[precio_id] = sintomas_filtrados
        else:
            sintomas_filtrados_por_precio = {precio_id: [] for precio_id in sintomas_por_precio.keys()}

        #  CALCULAR SNTOMAS SOBRANTES
        sintomas_sobrantes = []
        sintomas_sobrantes_ids = []
    
        if diagnosticos_posibles:
            # Obtener sntomas del mejor diagnstico
            mejor_diagnostico = max(diagnosticos_posibles.values(), key=lambda x: x['porcentaje'])
            mejor_diagnostico_id = [k for k, v in diagnosticos_posibles.items() if v == mejor_diagnostico][0]
        
            try:
                query_sintomas_diagnostico = """
                    SELECT DISTINCT sintoma_id
                    FROM diagnostico_sintoma
                    WHERE diagnostico_id = ?
                """
                sintomas_del_diagnostico = conn.execute(query_sintomas_diagnostico, [mejor_diagnostico_id]).fetchall()
                sintomas_del_diagnostico_ids = [s['sintoma_id'] for s in sintomas_del_diagnostico]
            
                # Calcular sobrantes
                for i, sid in enumerate(sintomas_detectados_ids):
                    if sid not in sintomas_del_diagnostico_ids:
                        sintomas_sobrantes_ids.append(sid)
                        sintomas_sobrantes.append(sintomas_detectados[i])
            except Exception as e:
                print(f"Error calculando sntomas sobrantes: {e}")
    
        conn.close()
    
        # ============================================
        # CALCULAR SCORE (SIN CAMBIOS)
        # ============================================

        # Marcar cuáles productos fueron encontrados por búsqueda directa
        ids_productos_directos = {p['precio_id'] for p in productos_directos}

        productos_con_score = []
        for p in productos:
            imagen = p['imagen_precio'] if p['imagen_precio'] else p['imagen_medicamento']
            sintomas_medicamento_ids = sintomas_ids_por_precio.get(p['precio_id'], [])
        
            coincidencias_sintomas = len(set(sintomas_medicamento_ids) & set(sintomas_detectados_ids))

            coincidencias_diagnosticos_directos = 0
            for diag_id in diagnosticos_detectados_directo_ids:
                try:
                    query_sintomas_diag = """
                        SELECT DISTINCT sintoma_id
                        FROM diagnostico_sintoma
                        WHERE diagnostico_id = %s
                    """
                    sintomas_diag = conn.execute(query_sintomas_diag, [diag_id]).fetchall()
                    sintomas_diag_ids = [s['sintoma_id'] for s in sintomas_diag]

                    match = len(set(sintomas_medicamento_ids) & set(sintomas_diag_ids))
                    if match > 0:
                        coincidencias_diagnosticos_directos += match
                except Exception as e:
                    print(f"Error calculando coincidencias diagnstico: {e}")
        
            score = 0
            mejor_diagnostico = None
            tipo_deteccion = None
        
            if coincidencias_diagnosticos_directos > 0:
                score = 100 + (coincidencias_diagnosticos_directos * 10)
                for diag_id, diag_info in diagnosticos_posibles.items():
                    if diag_info.get('tipo') == 'directo':
                        mejor_diagnostico = diag_info['nombre']
                        tipo_deteccion = 'directo'
                        break
            elif diagnosticos_posibles:
                mejor_porcentaje = 0
                for diag_id, diag_info in diagnosticos_posibles.items():
                    if diag_info.get('tipo') == 'por_sintomas':
                        if coincidencias_sintomas >= diag_info['coincidencias'] * 0.5:
                            if diag_info['porcentaje'] > mejor_porcentaje:
                                mejor_porcentaje = diag_info['porcentaje']
                                mejor_diagnostico = diag_info['nombre']
                                tipo_deteccion = 'por_sintomas'
                                score = (coincidencias_sintomas * 10) + 20
            else:
                score = coincidencias_sintomas * 10
        
            # Determinar si fue encontrado por búsqueda directa
            es_directo = p['precio_id'] in ids_productos_directos

            # 🆕 FILTRO: Solo incluir si es directo O tiene coincidencias de síntomas
            # EXCEPCIÓN: Si no hay búsqueda (carga inicial), incluir TODOS
            hay_busqueda = bool(busqueda or busqueda_sintomas)
            if not hay_busqueda or es_directo or coincidencias_sintomas > 0 or coincidencias_diagnosticos_directos > 0:
                productos_con_score.append({
                    'precio_id': p['precio_id'],
                    'medicamento_id': p['medicamento_id'],
                    'fabricante_id': p['fabricante_id'],
                    'nombre': p['medicamento_nombre'],
                    'presentacion': p['presentacion'] or '',
                    'concentracion': p['concentracion'] or '',
                    'fabricante': p['fabricante_nombre'],
                    'precio': p['precio'],
                    'imagen': imagen,
                    'componente_activo': p['componente_activo_nombre'] if p['componente_activo_nombre'] else None,
                    'sintomas_filtrados': sintomas_filtrados_por_precio.get(p['precio_id'], []),
                    'sintomas_totales': sintomas_por_precio.get(p['precio_id'], []),
                    'score': score,
                    'diagnostico_detectado': mejor_diagnostico,
                    'tipo_deteccion': tipo_deteccion,
                    'coincidencias': coincidencias_sintomas,
                    'es_directo': es_directo  # Marca si fue encontrado por búsqueda directa
                })
    
        # ============================================
        #  ORDENAMIENTO INTELIGENTE
        # ============================================
    
        if busqueda_sintomas and (sintomas_detectados_ids or diagnosticos_detectados_directo_ids):
        
            # DETECTAR TIPO DE BSQUEDA
            es_busqueda_diagnostico = len(diagnosticos_detectados_directo_ids) > 0
        
            if es_busqueda_diagnostico:
                # 
                # REGLA 2: BSQUEDA POR DIAGNSTICO
                # Prioridad: MS cobertura del diagnstico
                # 
            
                print(f"\n ORDENAMIENTO: Bsqueda por DIAGNSTICO")
            
                # Obtener sntomas del diagnstico principal
                diagnostico_principal_id = diagnosticos_detectados_directo_ids[0]
            
                try:
                    conn_ordenamiento = get_db_connection()  #  Nueva conexin
                
                    query_sintomas_diag = """
                        SELECT DISTINCT sintoma_id
                        FROM diagnostico_sintoma
                        WHERE diagnostico_id = %s
                    """
                    sintomas_del_diagnostico = conn_ordenamiento.execute(query_sintomas_diag, [diagnostico_principal_id]).fetchall()
                    sintomas_diag_ids = [s['sintoma_id'] for s in sintomas_del_diagnostico]
                    total_sintomas_diagnostico = len(sintomas_diag_ids)
                
                    print(f"   Diagnstico tiene {total_sintomas_diagnostico} sntomas")
                
                    # Calcular cobertura de cada producto
                    for producto in productos_con_score:
                        sintomas_producto = sintomas_ids_por_precio.get(producto['precio_id'], [])
                    
                        # Contar cuntos sntomas del diagnstico cubre este producto
                        cobertura = len([s for s in sintomas_producto if s in sintomas_diag_ids])
                    
                        #  Contar sntomas EXTRA (que NO estn en el diagnstico)
                        sintomas_extra = len([s for s in sintomas_producto if s not in sintomas_diag_ids])
                    
                        producto['cobertura_diagnostico'] = cobertura
                        producto['sintomas_extra'] = sintomas_extra
                        producto['sintomas_totales_count'] = len(sintomas_producto)
                        producto['porcentaje_cobertura'] = (cobertura / total_sintomas_diagnostico * 100) if total_sintomas_diagnostico > 0 else 0
                    
                        print(f"      {producto['nombre'][:30]:30} -> Cubre {cobertura}/{total_sintomas_diagnostico} ({producto['porcentaje_cobertura']:.0f}%) | Extras: {sintomas_extra} | Total: {len(sintomas_producto)}")
                
                    #  Ordenar: DIRECTOS primero, luego MENOS extras, luego MÁS cobertura
                    productos_con_score.sort(key=lambda x: (not x['es_directo'], x['sintomas_extra'], -x['cobertura_diagnostico']))
                
                    print(f"    Orden: Menos extras -> Ms cobertura del diagnstico")
                
                    conn_ordenamiento.close()  #  Cerrar conexin
                
                except Exception as e:
                    print(f"    Error calculando cobertura: {e}")
                    # Ordenar: DIRECTOS primero, luego por score
                    productos_con_score.sort(key=lambda x: (not x['es_directo'], -x['score']))
        
            else:
                # 
                # REGLA 1: BSQUEDA POR SNTOMAS
                # Prioridad: MENOS sntomas totales (especfico)
                # 
            
                print(f"\n ORDENAMIENTO: Bsqueda por SNTOMAS")
            
                for producto in productos_con_score:
                    sintomas_totales = len(sintomas_ids_por_precio.get(producto['precio_id'], []))
                
                    # Score inverso: menos sntomas = ms score
                    # Frmula: 1000 / sntomas_totales
                    if sintomas_totales > 0:
                        producto['especificidad_score'] = 1000 / sintomas_totales
                    else:
                        producto['especificidad_score'] = 0
                
                    producto['sintomas_totales_count'] = sintomas_totales
                
                    print(f"      {producto['nombre'][:30]:30} -> {sintomas_totales} sntomas (score: {producto['especificidad_score']:.1f})")

                # Ordenar: DIRECTOS primero, luego MENOS síntomas (más específico)
                productos_con_score.sort(key=lambda x: (not x['es_directo'], -x['especificidad_score'], -x['coincidencias']))
            
                print(f"    Orden: Ms especfico -> Ms genrico")
            
                # AHORA aplicar el ordenamiento por sntomas (alternar)
                # PASO 0: Separar productos DIRECTOS de productos por síntomas
                productos_directos_final = [p for p in productos_con_score if p['es_directo']]
                productos_sintomas_final = [p for p in productos_con_score if not p['es_directo']]

                # PASO 1: Agrupar por sntoma principal (SOLO productos de síntomas)
                productos_por_sintoma = {}

                for producto in productos_sintomas_final:
                    sintomas_del_producto = sintomas_ids_por_precio.get(producto['precio_id'], [])
                    sintomas_coincidentes = [s for s in sintomas_detectados_ids if s in sintomas_del_producto]

                    if sintomas_coincidentes:
                        sintoma_principal = sintomas_coincidentes[0]

                        if sintoma_principal not in productos_por_sintoma:
                            productos_por_sintoma[sintoma_principal] = []

                        productos_por_sintoma[sintoma_principal].append(producto)

                # PASO 2: Extraer el MEJOR de cada sntoma (ya ordenados por especificidad)
                mejores_por_sintoma = []

                for sintoma_id in sintomas_detectados_ids:
                    if sintoma_id in productos_por_sintoma and productos_por_sintoma[sintoma_id]:
                        mejor = productos_por_sintoma[sintoma_id][0]
                        mejores_por_sintoma.append(mejor)
                        productos_por_sintoma[sintoma_id] = productos_por_sintoma[sintoma_id][1:]

                # PASO 3: Alternar el RESTO
                resto_productos = []
                tiene_productos = True

                while tiene_productos:
                    tiene_productos = False
                    for sintoma_id in sintomas_detectados_ids:
                        if sintoma_id in productos_por_sintoma and productos_por_sintoma[sintoma_id]:
                            resto_productos.append(productos_por_sintoma[sintoma_id].pop(0))
                            tiene_productos = True

                # PASO 4: Resultado final - DIRECTOS PRIMERO, luego síntomas
                productos_con_score = productos_directos_final + mejores_por_sintoma + resto_productos

                print(f"    Productos DIRECTOS: {len(productos_directos_final)}")
                print(f"    Mejores por sntoma: {len(mejores_por_sintoma)}")
                print(f"    Resto alternado: {len(resto_productos)}")
    
        else:
            # Sin síntomas ni diagnósticos: DIRECTOS primero, luego por score
            productos_con_score.sort(key=lambda x: (not x['es_directo'], -x['score']))

    
        diagnosticos_response = []
        for diag_id, diag_info in diagnosticos_posibles.items():
            diag_response = {
                **diag_info,
                'es_directo': diag_info.get('tipo') == 'directo'
            }
        
            #  Agregar sntomas faltantes si es diagnstico directo
            if diag_id in sintomas_faltantes_por_diagnostico:
                diag_response['sintomas_faltantes'] = sintomas_faltantes_por_diagnostico[diag_id]
        
            diagnosticos_response.append(diag_response)
    
        # Calcular información de paginación
        total_count = total_productos if 'total_productos' in locals() else len(productos_con_score)
        total_pages = (total_count + limit - 1) // limit if limit > 0 else 1

        return jsonify({
            'ok': True,
            'productos': productos_con_score,
            'total': len(productos_con_score),
            'total_disponible': total_disponible,  # Total de productos sin LIMIT
            'hay_limite': hay_limite,  # Indica si se aplic LIMIT 50
            'sintomas_detectados': sintomas_detectados,
            'diagnosticos_posibles': diagnosticos_response,
            'diagnosticos_directos': diagnosticos_detectados_directo,
            'sintomas_sobrantes': sintomas_sobrantes,  #  Sntomas que no estn en el diagnstico
            'busqueda_parcial': busqueda_parcial_aplicada,  #  Flag para mostrar mensaje especial en frontend
            # Información de paginación
            'pagination': {
                'page': page,
                'limit': limit,
                'total_items': total_count,
                'total_pages': total_pages
            }
        })
    except Exception as e:
        print(f"\nERROR CRITICO en /api/productos: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'ok': False, 'error': str(e), 'productos': []}), 500


# -------------------------------------------------------------------
# --- ZONA 4: FLUJO DE REGISTRO CONVERSACIONAL (ETAPAS 1 a 5) ---
# -------------------------------------------------------------------

@app.route('/verificar_progreso_paciente')
def verificar_progreso_paciente():
    """Verifica qu datos faltan en el perfil del paciente y redirige a la etapa pendiente."""
    device_id = session.get('dispositivo_id')
    
    if session.get('rol') != 'Paciente' or not device_id:
        flash("Acceso no autorizado para completar el registro.", 'danger')
        return redirect(url_for('index'))

    conn = get_db_connection()
    usuario = conn.execute("SELECT edad, peso_aprox, genero, estado_organos FROM usuarios WHERE dispositivo_id = ?", (device_id,)).fetchone()
    conn.close()

    if not usuario:
        flash("Error: No se encontr el registro del usuario.", 'danger')
        return redirect(url_for('logout')) 

    if usuario['edad'] is None:
        return redirect(url_for('etapa2_edad')) 

    if usuario['peso_aprox'] is None:
        return redirect(url_for('etapa3_peso'))
        
    if usuario['genero'] is None:
        return redirect(url_for('etapa4_genero'))
    
    if usuario['estado_organos'] is None:
        return redirect(url_for('etapa5_organos'))

    return redirect(url_for('paciente_area'))

@app.route('/etapa1_nuevo_registro', methods=['GET', 'POST'])
def etapa1_nuevo_registro():
    """Muestra el formulario de registro de nombre, con plantillas distintas por rol."""

    #  Log de sesin al entrar
    app.logger.debug(f" Sesin al entrar en etapa1_nuevo_registro: {dict(session)}")

    #  Forzar persistencia del rol si viene desde /admin
    if not session.get('rol_temporal'):
        if request.referrer and '/admin' in request.referrer:
            session['rol_temporal'] = 'Administrador'
            app.logger.debug(" Rol temporal asignado automticamente como 'Administrador' (referrer /admin)")
        else:
            session['rol_temporal'] = 'Paciente'
            app.logger.debug(" Rol temporal asignado automticamente como 'Paciente' (referrer sin /admin)")
    else:
        app.logger.debug(f" Rol temporal ya existente en sesin: {session['rol_temporal']}")

    # Determinar rol y plantilla
    rol_temp = session.get('rol_temporal')
    template_name = '1_registro_admin.html' if rol_temp == 'Administrador' else '1_registro_paciente.html'

    #  Log de verificacin antes de mostrar plantilla
    app.logger.debug(f" Rol temporal activo: {rol_temp}, Plantilla: {template_name}")
    app.logger.debug(f" Sesin antes de procesar mtodo {request.method}: {dict(session)}")

    if request.method == 'POST':
        app.logger.debug(f" POST recibido en etapa1_nuevo_registro: {request.form}")  #  Verifica que llegue el nombre
        nombre = request.form.get('nombre')

        if not nombre:
            flash("El nombre es requerido.", 'warning')
            app.logger.warning(" Intento de registro sin nombre")
            return render_template(template_name, rol=rol_temp)

        #  Log previo a la redireccin
        app.logger.debug(f" Redirigiendo a etapa1_registro_completo con nombre={nombre}")
        app.logger.debug(f" Sesin justo antes de redirigir: {dict(session)}")

        return redirect(url_for('etapa1_registro_completo', nombre=nombre))
        
    # Si es GET, simplemente renderizamos la plantilla
    app.logger.debug(f" Renderizando plantilla: {template_name} con rol={rol_temp}")
    app.logger.debug(f" Sesin al final del GET: {dict(session)}")

    return render_template(template_name, rol=rol_temp)



@app.route('/etapa1_registro_completo', methods=['GET', 'POST'])
def etapa1_registro_completo():
    """Guarda el usuario en la base de datos y establece la sesin de acceso."""
    
    #  Log inicial
    app.logger.debug(f" Llamada a etapa1_registro_completo - Mtodo: {request.method}")
    app.logger.debug(f" Estado completo de la sesin al entrar: {dict(session)}")

    # Obtener datos
    nombre = request.form.get('nombre') if request.method == 'POST' else request.args.get('nombre')
    device_id = session.get('dispositivo_id')
    rol = session.get('rol_temporal')

    app.logger.debug(f" Datos recibidos -> nombre={nombre}, device_id={device_id}, rol={rol}")

    # Validacin bsica
    if not all([nombre, device_id, rol]):
        app.logger.error(f" Datos faltantes -> nombre={nombre}, device_id={device_id}, rol={rol}")
        flash("Error en el registro. Informacin incompleta.", 'danger')
        return redirect(url_for('etapa1_nuevo_registro'))

    # Guardar en base de datos
    conn = get_db_connection()
    try:
        # Verificar si ya existe el dispositivo
        existente = conn.execute(
            "SELECT id FROM usuarios WHERE dispositivo_id = ?", (device_id,)
        ).fetchone()

        if existente:
            usuario_id = existente['id']
            app.logger.warning(f" Usuario con device_id {device_id} ya existente (ID={usuario_id}).")
        else:
            conn.execute("""
                INSERT INTO usuarios (dispositivo_id, nombre, fecha_registro, rol)
                VALUES (?, ?, ?, ?)
            """, (device_id, nombre, datetime.now().isoformat(), rol))
            conn.commit()

            usuario_id = conn.execute(
                "SELECT id FROM usuarios WHERE dispositivo_id = ?", (device_id,)
            ).fetchone()['id']

            app.logger.debug(f" Usuario insertado con ID {usuario_id}")

        # Verificacin del registro
        usuario_verif = conn.execute(
            "SELECT id, nombre, rol, fecha_registro FROM usuarios WHERE id = ?", 
            (usuario_id,)
        ).fetchone()
        app.logger.debug(
            f" Verificacin DB -> ID: {usuario_verif['id']}, Nombre: {usuario_verif['nombre']}, Rol: {usuario_verif['rol']}"
        )

    except Exception as e:
        conn.close()
        app.logger.error(f" Error durante el registro: {e}")
        flash("Ocurri un error guardando el registro.", 'danger')
        return redirect(url_for('etapa1_nuevo_registro'))
    finally:
        conn.close()

    # Actualizar sesin
    session['rol'] = rol
    session['nombre'] = nombre
    session['usuario_id'] = usuario_id
    session.pop('rol_temporal', None)

    app.logger.debug(f" Sesin actualizada tras registro: {dict(session)}")

    flash(f"Bienvenido/a, {nombre}! Tu registro como {rol} ha sido exitoso.", 'success')

    # Redirigir segn el rol
    try:
        if rol == 'Administrador':
            app.logger.debug(" Redirigiendo al rea de administrador (admin_area)")
            return redirect(url_for('admin_area'))
        else:
            app.logger.debug(" Redirigiendo al flujo de paciente (verificar_progreso_paciente)")
            return redirect(url_for('verificar_progreso_paciente'))
    except Exception as e:
        app.logger.error(f" Error redirigiendo despus del registro: {e}")
        flash("Registro completado, pero no se pudo redirigir automticamente.", 'warning')
        return redirect(url_for('index'))


@app.route('/etapa2_edad', methods=['GET', 'POST'])
def etapa2_edad():
    """Pregunta por el rango de edad y guarda el dato en la DB."""
    if session.get('rol') != 'Paciente':
        return redirect(url_for('index'))
        
    opciones = [
        ('18-30 aos', '18-30'), 
        ('31-50 aos', '31-50'), 
        ('51-70 aos', '51-70'), 
        ('Ms de 70 aos', '+70')
    ]

    if request.method == 'POST':
        rango_edad = request.form.get('rango_edad')
        device_id = session.get('dispositivo_id')
        
        if not rango_edad or not device_id:
            flash("Seleccin invlida.", 'warning')
            return redirect(url_for('etapa2_edad'))

        conn = get_db_connection()
        conn.execute("UPDATE usuarios SET edad = ? WHERE dispositivo_id = ?", 
                      (rango_edad, device_id))
        conn.commit()
        conn.close()

        return redirect(url_for('verificar_progreso_paciente'))
        
    return render_template('2_etapa_conversacional.html', 
                           pregunta="Para empezar, en qu rango de edad te encuentras?",
                           opciones=opciones,
                           nombre_campo="rango_edad")

@app.route('/etapa3_peso', methods=['GET', 'POST'])
def etapa3_peso():
    """Pregunta por el rango de peso y guarda el dato en la DB."""
    if session.get('rol') != 'Paciente':
        return redirect(url_for('index'))
        
    opciones = [
        ('Menos de 60 kg', '0-60'), 
        ('60 - 80 kg', '60-80'), 
        ('81 - 100 kg', '81-100'), 
        ('Ms de 100 kg', '+100')
    ]

    if request.method == 'POST':
        rango_peso = request.form.get('rango_peso')
        device_id = session.get('dispositivo_id')
        
        if not rango_peso or not device_id:
            flash("Seleccin invlida.", 'warning')
            return redirect(url_for('etapa3_peso'))

        conn = get_db_connection()
        conn.execute("UPDATE usuarios SET peso_aprox = ? WHERE dispositivo_id = ?", 
                      (rango_peso, device_id))
        conn.commit()
        conn.close()

        return redirect(url_for('verificar_progreso_paciente'))
        
    return render_template('2_etapa_conversacional.html', 
                           pregunta="Para un clculo preciso, cul es tu rango de peso aproximado?",
                           opciones=opciones,
                           nombre_campo="rango_peso")

@app.route('/etapa4_genero', methods=['GET', 'POST'])
def etapa4_genero():
    """Pregunta por el gnero y guarda el dato en la DB."""
    if session.get('rol') != 'Paciente':
        return redirect(url_for('index'))
        
    opciones = [
        ('Mujer', 'Femenino'), 
        ('Hombre', 'Masculino'), 
        ('Otro/No especificar', 'Otro')
    ]

    if request.method == 'POST':
        genero = request.form.get('genero')
        device_id = session.get('dispositivo_id')
        
        if not genero or not device_id:
            flash("Seleccin invlida.", 'warning')
            return redirect(url_for('etapa4_genero'))

        conn = get_db_connection()
        conn.execute("UPDATE usuarios SET genero = ? WHERE dispositivo_id = ?", 
                      (genero, device_id))
        conn.commit()
        conn.close()

        return redirect(url_for('verificar_progreso_paciente'))
        
    return render_template('2_etapa_conversacional.html', 
                           pregunta=f"{session.get('nombre')}, cul es tu gnero?",
                           opciones=opciones,
                           nombre_campo="genero")

@app.route('/etapa5_organos', methods=['GET', 'POST'])
def etapa5_organos():
    """Pregunta por el estado de rganos y guarda el dato en la DB."""
    if session.get('rol') != 'Paciente':
        return redirect(url_for('index'))
        
    opciones = [
        ('Sanos (Sin problemas conocidos)', 'Sanos'), 
        ('Problemas leves (Insuficiencia leve, grasa)', 'Leve'), 
        ('Problemas moderados a graves (Dilisis, Cirrosis)', 'Grave')
    ]

    if request.method == 'POST':
        estado = request.form.get('estado_organos')
        device_id = session.get('dispositivo_id')
        
        if not estado or not device_id:
            flash("Seleccin invlida.", 'warning')
            return redirect(url_for('etapa5_organos'))

        conn = get_db_connection()
        conn.execute("UPDATE usuarios SET estado_organos = ? WHERE dispositivo_id = ?", 
                      (estado, device_id))
        conn.commit()
        conn.close()

        return redirect(url_for('verificar_progreso_paciente'))
        
    return render_template('2_etapa_conversacional.html', 
                           pregunta="Finalmente, cul es el estado de salud de tus rganos (Hgado/Rin)?",
                           opciones=opciones,
                           nombre_campo="estado_organos")

# -------------------------------------------------------------------
# --- ZONA 5: REA DE PACIENTE Y LGICA DE SUGERENCIAS ---
# -------------------------------------------------------------------

@app.route('/paciente_area')
def paciente_area():
    """Muestra la interfaz principal de consulta de sntomas."""
    if session.get('rol') != 'Paciente':
        flash("Acceso no autorizado.", 'danger')
        return redirect(url_for('index'))
    
    device_id = session.get('dispositivo_id')
    conn = get_db_connection()
    usuario = conn.execute("SELECT estado_organos FROM usuarios WHERE dispositivo_id = ?", (device_id,)).fetchone()

    # 1. Verifica si el registro est completo
    if usuario['estado_organos'] is None:
        flash("An debes completar tu registro clnico.", 'warning')
        return redirect(url_for('verificar_progreso_paciente'))

    # 2. Obtiene la base de datos de sntomas
    sintomas_db = conn.execute("SELECT id, nombre, descripcion_lower FROM sintomas").fetchall()
    conn.close()

    sintomas_json = json.dumps([dict(s) for s in sintomas_db])

    return render_template('consulta_sintomas.html', 
                           sintomas_data=sintomas_json)

@app.route('/mostrar_sugerencias', methods=['POST'])
def mostrar_sugerencias():
    """
    Ruta que recibe la lista de sntomas, consulta la base de datos, 
    filtra los medicamentos y renderiza los resultados.
    """
    if session.get('rol') != 'Paciente':
        flash("Acceso no autorizado.", 'danger')
        return redirect(url_for('index'))
    
    sintomas_seleccionados_ids_str = request.form.getlist('sintomas_id')
    
    if not sintomas_seleccionados_ids_str:
        flash("No seleccionaste ningn sntoma. Por favor, intntalo de nuevo.", 'warning')
        return redirect(url_for('paciente_area'))

    placeholders = ','.join(['?'] * len(sintomas_seleccionados_ids_str))
    
    conn = get_db_connection()
    
    # 2. Consulta de medicamentos usando la tabla de asociacin MEDICAMENTO_SINTOMA
    # Nota: Tu tabla de medicamentos en el data_initializer NO tiene stock_actual. 
    # Lo he quitado de la consulta para evitar errores, pero si lo necesitas, debes agregarlo en la inicializacin.
    sql_query = f"""
    SELECT 
        m.id, m.nombre, m.presentacion, m.concentracion, 
        m.uso, m.imagen, m.stock_actual
    FROM 
        medicamentos m
    JOIN 
        MEDICAMENTO_SINTOMA ms ON m.id = ms.medicamento_id
    WHERE 
        ms.sintoma_id IN ({placeholders})
    GROUP BY
        m.id, m.nombre, m.presentacion, m.concentracion, m.uso, m.imagen
    """
    
    sugerencias_db = conn.execute(sql_query, sintomas_seleccionados_ids_str).fetchall()
    conn.close()
    
    medicamentos_sugeridos = [dict(row) for row in sugerencias_db]

    # --- Simulacin de Diagnstico Principal (Opcional) ---
    diagnostico = None
    if len(sintomas_seleccionados_ids_str) >= 2 and '1' in sintomas_seleccionados_ids_str and '2' in sintomas_seleccionados_ids_str:
        diagnostico = "Sndrome General de Resfriado/Gripe Leve"
    
    return render_template('sugerencias_resultado.html', 
                           medicamentos=medicamentos_sugeridos,
                           diagnostico_principal=diagnostico,
                           sintomas_seleccionados_ids=sintomas_seleccionados_ids_str 
                          )


# ========================================
# REQUERIMIENTOS - API JSON
# ========================================
@app.route('/api/requerimientos', methods=['GET'])
@admin_required
def obtener_requerimientos():
    """Devuelve todos los requerimientos en JSON"""
    try:
        conn = get_db_connection()
        requerimientos = conn.execute("""
            SELECT id, descripcion, modulo, prioridad, estado, fecha_creacion
            FROM requerimientos
            ORDER BY id DESC
        """).fetchall()
        conn.close()
        return jsonify({
            'ok': True,
            'requerimientos': [dict(r) for r in requerimientos],
            'total': len(requerimientos)
        })
    except Exception as e:
        print(f"Error obteniendo requerimientos: {e}")
        return jsonify({'ok': False, 'error': str(e)}), 500


@app.route('/api/requerimientos', methods=['POST'])
@admin_required
def crear_requerimiento():
    """Crea un nuevo requerimiento"""
    try:
        data = request.get_json()
        descripcion = data.get('descripcion', '').strip()
        modulo = data.get('modulo', '').strip()
        prioridad = data.get('prioridad', '').strip()
        estado = data.get('estado', 'Planificacin').strip()
        
        if not all([descripcion, modulo, prioridad]):
            return jsonify({'ok': False, 'error': 'Faltan campos requeridos'}), 400
        
        conn = get_db_connection()
        cursor = conn.execute("""
            INSERT INTO requerimientos (descripcion, modulo, prioridad, estado)
            VALUES (?, ?, ?, ?)
        """, (descripcion, modulo, prioridad, estado))
        conn.commit()
        requerimiento_id = cursor.lastrowid
        conn.close()
        
        return jsonify({
            'ok': True,
            'mensaje': 'Requerimiento creado',
            'id': requerimiento_id
        }), 201
    except Exception as e:
        print(f"Error creando requerimiento: {e}")
        return jsonify({'ok': False, 'error': str(e)}), 500


@app.route('/api/requerimientos/<int:requerimiento_id>', methods=['PUT'])
@admin_required
def actualizar_requerimiento(requerimiento_id):
    """Actualiza un requerimiento"""
    try:
        data = request.get_json()
        descripcion = data.get('descripcion', '').strip()
        modulo = data.get('modulo', '').strip()
        prioridad = data.get('prioridad', '').strip()
        estado = data.get('estado', '').strip()
        
        conn = get_db_connection()
        conn.execute("""
            UPDATE requerimientos
            SET descripcion = ?, modulo = ?, prioridad = ?, estado = ?
            WHERE id = ?
        """, (descripcion, modulo, prioridad, estado, requerimiento_id))
        conn.commit()
        conn.close()
        
        return jsonify({'ok': True, 'mensaje': 'Requerimiento actualizado'})
    except Exception as e:
        print(f"Error actualizando requerimiento: {e}")
        return jsonify({'ok': False, 'error': str(e)}), 500


@app.route('/admin/requerimientos')
@admin_required
def admin_requerimientos():
    """Pgina de gestin de requerimientos"""
    return render_template('admin_requerimientos.html')



@app.route('/api/requerimientos/buscar_codigo', methods=['GET'])
@admin_required
def buscar_codigo_seccion():
    """Busca una seccin especfica en un archivo HTML/JS"""
    try:
        archivo = request.args.get('archivo', '').strip()
        identificador = request.args.get('identificador', '').strip()
        
        if not archivo or not identificador:
            return jsonify({'ok': False, 'error': 'Faltan parmetros'}), 400
        
        # Construir ruta del archivo
        ruta_archivo = os.path.join(app.root_path, 'templates', archivo)
        
        if not os.path.exists(ruta_archivo):
            return jsonify({'ok': False, 'error': f'Archivo no encontrado: {archivo}'}), 404
        
        # Leer archivo
        with open(ruta_archivo, 'r', encoding='utf-8') as f:
            contenido = f.read()
        
        lineas = contenido.split('\n')
        
        # BUSCAR POR FUNCIN JAVASCRIPT
        patron_func = f'(async\\s+)?function\\s+{re.escape(identificador)}\\s*\\('
        for i, linea in enumerate(lineas):
            if re.search(patron_func, linea, re.IGNORECASE):
                # Encontrada la funcin, extraer completa
                inicio = i
                fin = extraer_funcion_completa(lineas, i)
                codigo = '\n'.join(lineas[inicio:fin+1])
                return jsonify({
                    'ok': True,
                    'tipo': 'Funcin JavaScript',
                    'identificador': identificador,
                    'codigo': codigo,
                    'linea': i + 1
                })
        
        # BUSCAR POR ID HTML
        patron_id = f'id=["\']?{re.escape(identificador)}["\']?'
        for i, linea in enumerate(lineas):
            if re.search(patron_id, linea, re.IGNORECASE):
                inicio = max(0, i - 2)
                fin = min(len(lineas), i + 10)
                codigo = '\n'.join(lineas[inicio:fin])
                return jsonify({
                    'ok': True,
                    'tipo': 'ID HTML',
                    'identificador': identificador,
                    'codigo': codigo,
                    'linea': i + 1
                })
        
        # BUSCAR POR CLASE CSS
        patron_clase = f'class=["\']?[^"\']*{re.escape(identificador)}[^"\']*["\']?'
        for i, linea in enumerate(lineas):
            if re.search(patron_clase, linea, re.IGNORECASE):
                inicio = max(0, i - 2)
                fin = min(len(lineas), i + 10)
                codigo = '\n'.join(lineas[inicio:fin])
                return jsonify({
                    'ok': True,
                    'tipo': 'Clase CSS',
                    'identificador': identificador,
                    'codigo': codigo,
                    'linea': i + 1
                })
        
        return jsonify({'ok': False, 'error': f'No se encontr: {identificador}'}), 404
        
    except Exception as e:
        print(f"Error buscando cdigo: {e}")
        return jsonify({'ok': False, 'error': str(e)}), 500

def extraer_funcion_completa(lineas, inicio):
    """Extrae una funcin completa contando llaves"""
    contador_llaves = 0
    dentro_funcion = False
    
    for i in range(inicio, len(lineas)):
        linea = lineas[i]
        
        # Contar llaves (ignorar strings)
        en_string = False
        escape = False
        for char in linea:
            if escape:
                escape = False
                continue
            if char == '\\':
                escape = True
                continue
            if char == '"' or char == "'":
                en_string = not en_string
                continue
            if not en_string:
                if char == '{':
                    contador_llaves += 1
                    dentro_funcion = True
                elif char == '}':
                    contador_llaves -= 1
        
        # Si cerramos todas las llaves, fin de funcin
        if dentro_funcion and contador_llaves == 0:
            return i
    
    return len(lineas) - 1



@app.route('/api/requerimientos/<int:requerimiento_id>/referencias', methods=['GET'])
@admin_required
def obtener_referencias(requerimiento_id):
    """Obtiene todas las referencias de un requerimiento"""
    try:
        conn = get_db_connection()
        referencias = conn.execute("""
            SELECT id, archivo_relacionado, seccion_identificador, descripcion_referencia, estado
            FROM requerimiento_referencias
            WHERE requerimiento_id = ?
            ORDER BY id DESC
        """, (requerimiento_id,)).fetchall()
        conn.close()
        return jsonify({
            'ok': True,
            'referencias': [dict(r) for r in referencias],
            'total': len(referencias)
        })
    except Exception as e:
        print(f"Error obteniendo referencias: {e}")
        return jsonify({'ok': False, 'error': str(e)}), 500

@app.route('/api/requerimientos/<int:requerimiento_id>/referencias', methods=['POST'])
@admin_required
def agregar_referencia(requerimiento_id):
    """Agrega una referencia a un requerimiento"""
    try:
        data = request.get_json()
        archivo = data.get('archivo_relacionado', '').strip()
        identificador = data.get('seccion_identificador', '').strip()
        descripcion = data.get('descripcion_referencia', '').strip()
        estado = data.get('estado', 'Pendiente').strip()
        
        if not archivo or not identificador:
            return jsonify({'ok': False, 'error': 'Faltan campos requeridos'}), 400
        
        conn = get_db_connection()
        cursor = conn.execute("""
            INSERT INTO requerimiento_referencias 
            (requerimiento_id, archivo_relacionado, seccion_identificador, descripcion_referencia, estado)
            VALUES (?, ?, ?, ?, ?)
        """, (requerimiento_id, archivo, identificador, descripcion, estado))
        conn.commit()
        ref_id = cursor.lastrowid
        conn.close()
        
        return jsonify({
            'ok': True,
            'mensaje': 'Referencia agregada',
            'id': ref_id
        }), 201
    except Exception as e:
        print(f"Error agregando referencia: {e}")
        return jsonify({'ok': False, 'error': str(e)}), 500


@app.route('/api/requerimientos/<int:requerimiento_id>/referencias/<int:ref_id>/estado', methods=['PUT'])
@admin_required
def actualizar_estado_referencia(requerimiento_id, ref_id):
    """Actualiza el estado de una referencia"""
    try:
        data = request.get_json()
        nuevo_estado = data.get('estado', '').strip()
        
        if not nuevo_estado:
            return jsonify({'ok': False, 'error': 'Estado requerido'}), 400
        
        conn = get_db_connection()
        conn.execute("""
            UPDATE requerimiento_referencias
            SET estado = ?
            WHERE id = ? AND requerimiento_id = ?
        """, (nuevo_estado, ref_id, requerimiento_id))
        conn.commit()
        conn.close()
        
        return jsonify({'ok': True, 'mensaje': 'Estado actualizado'})
    except Exception as e:
        print(f"Error actualizando estado: {e}")
        return jsonify({'ok': False, 'error': str(e)}), 500


@app.route('/api/requerimientos/<int:requerimiento_id>/referencias/<int:ref_id>', methods=['DELETE'])
@admin_required
def eliminar_referencia(requerimiento_id, ref_id):
    """Elimina una referencia de un requerimiento"""
    try:
        conn = get_db_connection()
        conn.execute("""
            DELETE FROM requerimiento_referencias
            WHERE id = ? AND requerimiento_id = ?
        """, (ref_id, requerimiento_id))
        conn.commit()
        conn.close()
        
        return jsonify({'ok': True, 'mensaje': 'Referencia eliminada'})
    except Exception as e:
        print(f"Error eliminando referencia: {e}")
        return jsonify({'ok': False, 'error': str(e)}), 500


@app.route('/api/requerimientos/extraer_identificadores', methods=['GET'])
@admin_required
def extraer_identificadores():
    """Extrae todos los identificadores de un archivo HTML/JS"""
    try:
        archivo = request.args.get('archivo', '').strip()
        
        if not archivo:
            return jsonify({'ok': False, 'error': 'Archivo requerido'}), 400
        
        ruta_archivo = os.path.join(app.root_path, 'templates', archivo)
        
        if not os.path.exists(ruta_archivo):
            return jsonify({'ok': False, 'error': f'Archivo no encontrado'}), 404
        
        with open(ruta_archivo, 'r', encoding='utf-8') as f:
            contenido = f.read()
        
        # EXTRAER FUNCIONES JAVASCRIPT
        funciones = re.findall(r'(async\s+)?function\s+(\w+)\s*\(', contenido, re.IGNORECASE)
        funciones_lista = list(set([f[1] for f in funciones]))
        
        # EXTRAER IDS HTML
        ids = re.findall(r'id=["\']([^"\']+)["\']', contenido)
        ids_lista = list(set(ids))
        
        # EXTRAER CLASES CSS
        clases = re.findall(r'class=["\']([^"\']+)["\']', contenido)
        clases_set = set()
        for clase_str in clases:
            clases_set.update(clase_str.split())
        clases_lista = list(clases_set)
        
        # EXTRAER VARIABLES
        variables = re.findall(r'(?:let|const|var)\s+(\w+)\s*=', contenido)
        variables_lista = list(set(variables))
        
        return jsonify({
            'ok': True,
            'archivo': archivo,
            'funciones': sorted(funciones_lista),
            'ids': sorted(ids_lista),
            'clases': sorted(clases_lista),
            'variables': sorted(variables_lista),
            'totales': {
                'funciones': len(funciones_lista),
                'ids': len(ids_lista),
                'clases': len(clases_lista),
                'variables': len(variables_lista)
            }
        })
        
    except Exception as e:
        print(f"Error extrayendo identificadores: {e}")
        return jsonify({'ok': False, 'error': str(e)}), 500


@app.route('/api/archivos', methods=['GET'])
@admin_required
def listar_archivos():
    """Lista todos los archivos registrados"""
    try:
        conn = get_db_connection()
        archivos = conn.execute("""
            SELECT id, nombre_archivo, descripcion, ruta, fecha_creacion, fecha_actualizacion
            FROM archivos
            ORDER BY fecha_creacion DESC
        """).fetchall()
        conn.close()
        
        return jsonify({
            'ok': True,
            'archivos': [dict(a) for a in archivos],
            'total': len(archivos)
        })
    except Exception as e:
        print(f"Error listando archivos: {e}")
        return jsonify({'ok': False, 'error': str(e)}), 500

@app.route('/api/archivos', methods=['POST'])
@admin_required
def agregar_archivo():
    """Agrega un archivo a la tabla"""
    try:
        data = request.get_json()
        nombre = data.get('nombre_archivo', '').strip()
        descripcion = data.get('descripcion', '').strip()
        ruta = data.get('ruta', '').strip()
        
        if not nombre:
            return jsonify({'ok': False, 'error': 'Nombre de archivo requerido'}), 400
        
        conn = get_db_connection()
        cursor = conn.execute("""
            INSERT INTO archivos (nombre_archivo, descripcion, ruta)
            VALUES (?, ?, ?)
        """, (nombre, descripcion, ruta))
        conn.commit()
        archivo_id = cursor.lastrowid
        conn.close()
        
        return jsonify({
            'ok': True,
            'mensaje': 'Archivo agregado',
            'id': archivo_id
        }), 201
    except Exception as e:
        print(f"Error agregando archivo: {e}")
        return jsonify({'ok': False, 'error': str(e)}), 500

@app.route('/api/archivos/<int:archivo_id>', methods=['DELETE'])
@admin_required
def eliminar_archivo(archivo_id):
    """Elimina un archivo de la tabla"""
    try:
        conn = get_db_connection()
        conn.execute("DELETE FROM archivos WHERE id = ?", (archivo_id,))
        conn.commit()
        conn.close()
        
        return jsonify({'ok': True, 'mensaje': 'Archivo eliminado'})
    except Exception as e:
        print(f"Error eliminando archivo: {e}")
        return jsonify({'ok': False, 'error': str(e)}), 500



@app.route('/api/archivos/poblar', methods=['POST'])
@admin_required
def poblar_archivos():
    """Escanea templates y agrega archivos automticamente"""
    try:
        extensiones = request.get_json().get('extensiones', ['.html', '.js', '.py', '.css']) if request.is_json else ['.html', '.js', '.py', '.css']
        templates_path = os.path.join(app.root_path, 'templates')
        conexion = get_db_connection()
        
        archivos_agregados = 0
        archivos_existentes = 0
        
        # Escanear todos los archivos con extensiones permitidas
        for archivo in os.listdir(templates_path):
            if any(archivo.endswith(ext) for ext in extensiones):
                # Verificar si ya existe
                existe = conexion.execute(
                    "SELECT id FROM archivos WHERE nombre_archivo = ?",
                    (archivo,)
                ).fetchone()
                
                if not existe:
                    conexion.execute("""
                        INSERT INTO archivos (nombre_archivo, descripcion, ruta)
                        VALUES (?, ?, ?)
                    """, (archivo, f'Archivo: {archivo}', f'/templates/{archivo}'))
                    archivos_agregados += 1
                else:
                    archivos_existentes += 1
        
        conexion.commit()
        conexion.close()
        
        return jsonify({
            'ok': True,
            'mensaje': f'Poblado completado',
            'agregados': archivos_agregados,
            'existentes': archivos_existentes,
            'extensiones': extensiones
        })
        
    except Exception as e:
        print(f"Error poblando archivos: {e}")
        return jsonify({'ok': False, 'error': str(e)}), 500


# -------------------------------------------------------------------
# --- ZONA 6: REA DE ADMINISTRADOR Y UTILERAS ---
# -------------------------------------------------------------------

def require_role(target_role):
    """Decorador simple para proteger las rutas."""
    def decorator(f):
        def wrapper(*args, **kwargs):
            if 'rol' not in session or session['rol'] != target_role:
                flash(f"Acceso denegado. Se requiere el rol '{target_role}'.", 'danger')
                return redirect(url_for('index'))
            return f(*args, **kwargs)
        wrapper.__name__ = f.__name__
        return wrapper
    return decorator



@app.route('/area_admin')
@admin_required
def admin_area():
    """Ruta principal para el Administrador. Muestra el men."""
    #  Renderizamos la plantilla admin_menu.html
    return render_template('admin_menu.html', 
                           nombre=session['nombre'],
                           device_id=session['dispositivo_id'])

# --- RUTAS DEL MEN ADMINISTRADOR ---
@app.route('/admin_menu')
@admin_required
def admin_menu():
    """Muestra el men principal de administracin."""
    return render_template('admin_menu.html')

@app.route('/admin/categorias')
@admin_required
def admin_categorias():
    """Página de gestión de categorías"""
    return render_template('admin_categorias.html')

@app.route('/admin/nuevo_admin')
@admin_required
def registro_admin_form_protegido():
    """Muestra el men del administrador con un mensaje de accin (para que la plantilla sea reutilizable)."""
    return render_template('admin_menu.html',
                           nombre=session['nombre'],
                           device_id=session['dispositivo_id'],
                           mensaje_accion="Aqu se gestionar el registro de nuevos administradores.")

@app.route('/admin/medicamentos')
@admin_required
def lista_medicamentos():
    """Muestra la lista mejorada de medicamentos."""
    return render_template('lista_medicamentos_mejorada.html')

@app.route('/admin/gestor_medicamentos_top')
@admin_required
def gestor_medicamentos_top():
    """Pgina para gestionar los medicamentos top vendidos"""
    conn = get_db_connection()
    # Obtener sntomas existentes
    sintomas = conn.execute("SELECT id, nombre FROM sintomas ORDER BY nombre").fetchall()
    conn.close()
    
    sintomas_list = [dict(s) for s in sintomas]
    
    return render_template('gestor_medicamentos_top.html', sintomas=sintomas_list)


@app.route('/admin/medicamentos/verificar_top', methods=['POST'])
@admin_required
def verificar_medicamentos_top():
    """Verifica si los medicamentos TOP existen en la BD y retorna sus IDs nicos"""
    try:
        data = request.get_json()
        medicamentos = data.get('medicamentos', [])
        conn = get_db_connection()
        resultados = []
        
        for med in medicamentos:
            #  CAMBIO CLAVE #1: Usar 'descripcion' que S viene del frontend
            descripcion_completa = med.get('descripcion', '').strip()
            laboratorio_nombre = med.get('laboratorio', '').strip()
            
            if not descripcion_completa:
                print(f" Medicamento sin descripcin: {med}")
                continue
            
            print(f" Buscando: {descripcion_completa} | Laboratorio: {laboratorio_nombre}")
            
            #  CAMBIO CLAVE #2: Buscar por NOMBRE EXACTO (no LIKE)
            medicamento_bd = conn.execute(
                "SELECT id, nombre, presentacion FROM medicamentos WHERE LOWER(nombre) = LOWER(?)",
                (descripcion_completa,)  # Bsqueda exacta
            ).fetchone()
            
            if medicamento_bd:
                # Existe en BD
                presentacion_bd = medicamento_bd['presentacion'].lower() if medicamento_bd['presentacion'] else ''
                presentacion_frontend = med.get('presentacion', '').lower()
                misma_presentacion = presentacion_frontend == presentacion_bd
                
                print(f" ENCONTRADO - ID: {medicamento_bd['id']}")
                
                resultados.append({
                    'nombre': descripcion_completa,  #  CAMBIO #3: Usar descripcion completa
                    'presentacion': med.get('presentacion'),
                    'concentracion': med.get('concentracion'),
                    'laboratorio': laboratorio_nombre,
                    'existe': True,
                    'misma_presentacion': misma_presentacion,
                    'id': medicamento_bd['id'],  #  ID NICO Y CORRECTO
                    'presentacion_bd': medicamento_bd['presentacion']
                })
            else:
                # No existe en BD
                print(f" NO ENCONTRADO")
                
                resultados.append({
                    'nombre': descripcion_completa,  #  CAMBIO #3: Usar descripcion completa
                    'presentacion': med.get('presentacion'),
                    'concentracion': med.get('concentracion'),
                    'laboratorio': laboratorio_nombre,
                    'existe': False,
                    'misma_presentacion': False,
                    'id': None
                })
        
        conn.close()
        
        return jsonify({
            'ok': True,
            'resultados': resultados
        })
        
    except Exception as e:
        print(f" Error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'ok': False, 'error': str(e)}), 500


@app.route('/admin/medicamentos/buscar_por_sintoma', methods=['POST'])
@admin_required
def buscar_medicamentos_por_sintoma():
    """Busca medicamentos que traten un sntoma especfico"""
    try:
        data = request.get_json()
        sintoma = data.get('sintoma', '').strip()
        
        if not sintoma:
            return jsonify({'ok': False, 'error': 'Sntoma requerido'}), 400
        
        conn = get_db_connection()
        
        # Buscar si el sntoma existe en la BD
        sintoma_normalizado = normalizar_texto(sintoma)
        sintoma_bd = conn.execute(
            """SELECT id FROM sintomas WHERE LOWER(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(
                nombre, '', 'a'), '', 'e'), '', 'i'), '', 'o'), '', 'u'), '', 'n')
            ) LIKE ?""",
            (f'%{sintoma_normalizado}%',)
        ).fetchone()
        
        resultados = []
        
        if sintoma_bd:
            # Sntoma existe: buscar medicamentos asociados
            medicamentos = conn.execute("""
                SELECT m.id, m.nombre, m.presentacion, m.concentracion
                FROM medicamentos m
                INNER JOIN medicamento_sintoma ms ON m.id = ms.medicamento_id
                WHERE ms.sintoma_id = ?
                ORDER BY m.nombre
            """, (sintoma_bd['id'],)).fetchall()
        else:
            # Sntoma no existe: buscar por bsqueda en Google (placeholder)
            # Por ahora devolvemos vaco, en futuro integraramos bsqueda en Google
            medicamentos = []
        
        # Procesar resultados
        for med in medicamentos:
            resultados.append({
                'nombre': med['nombre'],
                'presentacion': med['presentacion'] or 'N/A',
                'concentracion': med['concentracion'] or 'N/A',
                'laboratorio': 'Por definir',
                'existe': True,
                'misma_presentacion': True,  # Ya est en BD, se considera OK
                'id': med['id']
            })
        
        conn.close()
        
        return jsonify({
            'ok': True,
            'resultados': resultados,
            'sintoma_encontrado': sintoma_bd is not None
        })
    
    except Exception as e:
        print(f" Error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'ok': False, 'error': str(e)}), 500


@app.route('/admin/medicamentos/lista_json')
@admin_required
def lista_medicamentos_json():
    """Devuelve JSON con medicamentos + indicadores de completitud + PRIMERA IMAGEN"""
    try:
        filtro = request.args.get('filtro', 'todos')
        buscar = request.args.get('buscar', '').strip().lower()
        conn = get_db_connection()
        query = """
        SELECT
        m.id,
        m.nombre,
        m.presentacion,
        m.concentracion,
        m.stock_actual,
        m.imagen,
        m.componente_activo_id,
        (SELECT nombre FROM medicamentos WHERE id = m.componente_activo_id) as componente_activo_nombre,
        (SELECT COUNT(*) FROM precios WHERE medicamento_id = m.id AND precio > 0) as cantidad_precios,
        (SELECT COUNT(*) FROM precios WHERE medicamento_id = m.id AND imagen IS NOT NULL AND imagen != '') as tiene_imagen_oferta,
        (SELECT COUNT(*) FROM medicamento_sintoma WHERE medicamento_id = m.id) as cantidad_sintomas,
        (SELECT STRING_AGG(DISTINCT f.nombre, ',')
        FROM precios p
        INNER JOIN fabricantes f ON p.fabricante_id = f.id
        WHERE p.medicamento_id = m.id AND p.precio > 0) as fabricantes_str,
        (SELECT STRING_AGG(DISTINCT s.nombre, ',')
        FROM medicamento_sintoma ms
        INNER JOIN sintomas s ON ms.sintoma_id = s.id
        WHERE ms.medicamento_id = m.id) as sintomas_str_list,
        (SELECT p.imagen FROM precios p WHERE p.medicamento_id = m.id AND p.imagen IS NOT NULL AND p.imagen != '' LIMIT 1) as primera_imagen_precio
        FROM medicamentos m
        WHERE 1=1
        """

        params = []
        if buscar:
            query += " AND LOWER(m.nombre) LIKE %s"
            params.append(f'%{buscar}%')

        query += " ORDER BY m.nombre ASC"
        medicamentos = conn.execute(query, params).fetchall()
        
        resultado = []
        for med in medicamentos:
            med_dict = dict(med)
            tiene_precio = med_dict['cantidad_precios'] > 0
            tiene_sintomas = med_dict['cantidad_sintomas'] > 0
            tiene_imagen = med_dict['tiene_imagen_oferta'] > 0
            
            med_dict['indicador_precio'] = '' if tiene_precio else ''
            med_dict['indicador_sintomas'] = '' if tiene_sintomas else ''
            med_dict['indicador_imagen'] = '' if tiene_imagen else ''
            med_dict['es_completo'] = tiene_precio and tiene_sintomas and tiene_imagen
            med_dict['completitud'] = {
                'precio': tiene_precio,
                'sintomas': tiene_sintomas,
                'imagen': tiene_imagen
            }
            
            if filtro == 'sin_precio' and tiene_precio:
                continue
            elif filtro == 'sin_sintomas' and tiene_sintomas:
                continue
            elif filtro == 'sin_imagen' and tiene_imagen:
                continue
            elif filtro == 'incompletos' and med_dict['es_completo']:
                continue
            
            resultado.append(med_dict)
        
        conn.close()
        
        total = len(resultado)
        completos = sum(1 for m in resultado if m['es_completo'])
        sin_precio = sum(1 for m in resultado if not m['completitud']['precio'])
        sin_sintomas = sum(1 for m in resultado if not m['completitud']['sintomas'])
        sin_imagen = sum(1 for m in resultado if not m['completitud']['imagen'])
        
        return jsonify({
            'ok': True,
            'medicamentos': resultado,
            'estadisticas': {
                'total': total,
                'completos': completos,
                'sin_precio': sin_precio,
                'sin_sintomas': sin_sintomas,
                'sin_imagen': sin_imagen
            },
            'filtro_activo': filtro,
            'busqueda_activa': buscar
        })
    
    except Exception as e:
        print(f" Error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'ok': False, 'error': str(e)}), 500


def allowed_file(filename: str) -> bool:
    """Devuelve True si la extensin del archivo est permitida."""
    return bool(filename and '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXT)


def hash_image_content(file_stream) -> str:
    """
    Calcula un hash SHA1 del contenido del stream para nombre nico.
    Asegura volver al inicio del stream si es posible.
    """
    try:
        # Guardar posicin actual
        pos = file_stream.tell()
    except Exception:
        pos = None

    try:
        file_stream.seek(0)
    except Exception:
        pass

    data = file_stream.read()
    h = hashlib.sha1(data).hexdigest()

    # Volver al inicio para que quien llame pueda leer de nuevo
    try:
        file_stream.seek(0)
    except Exception:
        pass

    # Restaurar posicin anterior si exista
    try:
        if pos is not None:
            file_stream.seek(pos)
    except Exception:
        pass

    return h

# Ruta auxiliar (opcional) para servir uploads si prefieres no usar url_for('static', ...)
# Puedes usar en plantillas: url_for('uploaded_file', filename=med.imagen)
@app.route('/uploads/<path:filename>')
def uploaded_file(filename):
    folder = app.config['UPLOAD_FOLDER']
    app.logger.debug(f"Servir archivo desde uploads: {filename}")
    return send_from_directory(folder, filename)



@app.route('/admin/medicamentos/editar/<medicamento_id>', methods=['GET', 'POST'])
@admin_required
def editar_medicamento_admin(medicamento_id):
    """Permite editar o crear un medicamento, incluyendo la imagen."""
    
    es_nuevo = (medicamento_id == 'nuevo')
    conn = get_db_connection()
    
    # Detectar si es nuevo
    if es_nuevo:
        nombre_precar = request.args.get('nombre', '')
        laboratorio_precar = request.args.get('laboratorio', '')
        
        medicamento_dict = {
            'id': None,
            'nombre': nombre_precar,
            'presentacion': '',
            'concentracion': '',
            'codigo_atc_puro': '',
            'descripcion_tecnica_atc': '',
            'imagen': None,
            'uso': '',
            'stock_actual': 0,
            'activo': '1',  # Por defecto activo
            'laboratorio_sugerido': laboratorio_precar  #  Solo pasar, no guardar
        }
        
        if request.method == 'GET':
            conn.close()
            return render_template('editar_medicamento.html', medicamento=medicamento_dict, es_nuevo=True)
    else:
        # Si no es nuevo, obtener de BD
        medicamento = conn.execute("SELECT * FROM medicamentos WHERE id = %s", (medicamento_id,)).fetchone()

        if not medicamento:
            conn.close()
            flash("Medicamento no encontrado.", "danger")
            return redirect(url_for('lista_medicamentos'))

        medicamento_dict = dict(medicamento)

    if request.method == 'POST':
        nombre = request.form['nombre']
        presentacion = request.form.get('presentacion', '').strip()  # Opcional
        concentracion = request.form.get('concentracion', '').strip()  # Opcional
        codigo_atc_puro = request.form.get('codigo_atc_puro', '').upper().strip()
        descripcion_tecnica_atc = request.form.get('descripcion_tecnica_atc', '')
        # Campo activo (checkbox)
        activo = '1' if request.form.get('activo') == '1' else '0'
        #  Componente activo
        componente_activo_id = request.form.get('componente_activo_id', '')
        if componente_activo_id and componente_activo_id.strip():
            componente_activo_id = int(componente_activo_id)
        else:
            componente_activo_id = None

        # Mantenemos el nombre de la imagen actual por defecto
        imagen_filename = medicamento_dict.get('imagen')

        # --- Lgica de Subida de Nueva Imagen ---
        if 'imagen' in request.files:
            file = request.files['imagen']
            if file.filename != '' and allowed_file(file.filename):
                ext = file.filename.rsplit('.', 1)[1].lower()
                filename_to_save = f"{nombre}.{ext}"
                filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename_to_save)
                file.stream.seek(0)
                file.save(filepath)
                imagen_filename = filename_to_save
        
        try:
            if es_nuevo:
                # INSERTAR nuevo medicamento
                cursor = conn.execute("""
                INSERT INTO medicamentos (nombre, presentacion, concentracion, imagen, codigo_atc_puro, descripcion_tecnica_atc, uso, stock_actual, componente_activo_id, activo)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                RETURNING id
                """, (nombre, presentacion, concentracion, imagen_filename, codigo_atc_puro, descripcion_tecnica_atc, '', 0, componente_activo_id, activo))
                nuevo_id = cursor.fetchone()[0]
                conn.commit()
                flash("Medicamento creado exitosamente.", "success")
                conn.close()
                return redirect(url_for('editar_medicamento_admin', medicamento_id=nuevo_id))
            else:
                # ACTUALIZAR medicamento existente
                conn.execute("""
                UPDATE medicamentos SET
                nombre = %s, presentacion = %s, concentracion = %s, imagen = %s,
                codigo_atc_puro = %s, descripcion_tecnica_atc = %s, componente_activo_id = %s, activo = %s
                WHERE id = %s
                """, (nombre, presentacion, concentracion, imagen_filename,
                codigo_atc_puro, descripcion_tecnica_atc, componente_activo_id, activo, medicamento_id))
                
                conn.commit()
                flash("Medicamento actualizado exitosamente.", "success")
                conn.close()
                return redirect(url_for('editar_medicamento_admin', medicamento_id=medicamento_id))            
       
        except sqlite3.Error as e:
            flash(f"Error al guardar el medicamento: {e}", "danger")
            conn.rollback()
            conn.close()
        
        # Si hay error, mostrar formulario nuevamente
        return render_template('editar_medicamento.html', medicamento=medicamento_dict, es_nuevo=es_nuevo)

    # Si es GET, mostramos el formulario de edicin
    conn.close()
    return render_template('editar_medicamento.html', medicamento=medicamento_dict, es_nuevo=es_nuevo)



@app.route('/admin/precios/<int:medicamento_id>', methods=['GET'])
def obtener_precios_medicamento(medicamento_id):
    """
    Devuelve todos los precios registrados para un medicamento especfico,
    incluyendo fabricante, valor, imagen y fecha.
    """
    try:
        conn = get_db_connection()
        precios_rows = conn.execute("""
            SELECT p.id,
                p.medicamento_id,
                p.fabricante_id,
                f.nombre AS fabricante,
                p.precio,
                p.imagen,
                p.fecha_actualizacion
            FROM precios p
            LEFT JOIN fabricantes f ON p.fabricante_id = f.id
            WHERE p.medicamento_id = %s
            ORDER BY p.fecha_actualizacion DESC
        """, (medicamento_id,)).fetchall()
        precios = [dict(row) for row in precios_rows]
        conn.close()
        
        return jsonify({"ok": True, "precios": precios})
        
    except Exception as e:
        print(f"Error al obtener precios: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route('/admin/medicamentos/<int:medicamento_id>/es_componente_activo')
def verificar_si_es_componente_activo(medicamento_id):
    """Verifica si este medicamento es componente activo de otros"""
    try:
        conn = get_db_connection()
        
        # Buscar medicamentos que lo usan como componente_activo_id
        medicamentos = conn.execute("""
            SELECT id, nombre 
            FROM medicamentos 
            WHERE componente_activo_id = ?
            ORDER BY nombre
        """, (medicamento_id,)).fetchall()
        
        conn.close()
        
        resultados = [dict(m) for m in medicamentos]
        
        return jsonify({
            'ok': True,
            'es_componente_activo': len(resultados) > 0,
            'medicamentos_que_lo_usan': resultados
        })
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500


@app.route('/admin/medicamentos/componentes_activos/lista')
def obtener_componentes_activos():
    """Devuelve lista de componentes activos puros (sin componente_activo_id)"""
    try:
        medicamento_id = request.args.get('medicamento_id', type=int)
        
        conn = get_db_connection()
        
        # Solo medicamentos donde componente_activo_id IS NULL
        # EXCLUIR el medicamento actual
        if medicamento_id:
            componentes = conn.execute("""
                SELECT id, nombre 
                FROM medicamentos 
                WHERE componente_activo_id IS NULL 
                AND id != ?
                ORDER BY LOWER(nombre)
            """, (medicamento_id,)).fetchall()
        else:
            componentes = conn.execute("""
                SELECT id, nombre 
                FROM medicamentos 
                WHERE componente_activo_id IS NULL
                ORDER BY LOWER(nombre)
            """).fetchall()
        
        conn.close()
        
        resultados = [dict(c) for c in componentes]
        
        return jsonify({
            'ok': True,
            'componentes_activos': resultados
        })
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500
    

@app.route('/admin/fabricantes', methods=['GET'])
def listar_fabricantes():
    """
    Devuelve la lista de todos los fabricantes.
    """
    try:
        conn = get_db_connection()
        fabricantes_rows = conn.execute("SELECT id, nombre FROM fabricantes ORDER BY nombre").fetchall()
        fabricantes = [dict(row) for row in fabricantes_rows]
        conn.close()

        return jsonify({"ok": True, "fabricantes": fabricantes})
    except Exception as e:
        print(f"Error al listar fabricantes: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500



@app.route('/admin/fabricantes/by_name', methods=['POST'])
def obtener_o_crear_fabricante():
    """
    Entrada JSON: {"nombre": "Genfar"}
    Retorna: {"ok": True, "fabricante": {"id": 3, "nombre": "Genfar"}}
    Si ya existe, devuelve el existente; si no, lo crea y devuelve.
    """
    try:
        data = request.get_json(force=True)
        nombre = (data.get('nombre') or "").strip()
        if not nombre:
            return jsonify({"ok": False, "error": "nombre requerido"}), 400

        conn = get_db_connection()

        # Buscar si ya existe (insensible a maysculas/minsculas)
        row = conn.execute(
            "SELECT id, nombre FROM fabricantes WHERE lower(nombre)=%s",
            (nombre.lower(),)
        ).fetchone()

        if row:
            fab = dict(row)
        else:
            cur = conn.execute("INSERT INTO fabricantes (nombre) VALUES (%s) RETURNING id", (nombre,))
            fab_id = cur.fetchone()[0]
            conn.commit()
            row = conn.execute("SELECT id, nombre FROM fabricantes WHERE id=%s", (fab_id,)).fetchone()
            fab = dict(row)

        conn.close()
        return jsonify({"ok": True, "fabricante": fab})

    except Exception as e:
        print(f"Error al crear/obtener fabricante: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route('/admin/precios/<int:precio_id>', methods=['DELETE'])
@admin_required
def eliminar_precio(precio_id):
    """Elimina un precio de la BD"""
    try:
        conn = get_db_connection()
        conn.execute('DELETE FROM precios WHERE id = ?', (precio_id,))
        conn.commit()
        conn.close()
        
        return jsonify({'ok': True, 'mensaje': 'Precio eliminado'})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500


@app.route('/admin/precios/guardar', methods=['POST'])
def guardar_precio():
    """
    Guarda o actualiza un precio de medicamento.
    Entrada JSON:
    {
        "id": 5,                # opcional (si existe, se actualiza)
        "medicamento_id": 123,
        "fabricante_id": 2,
        "precio": 5400.0
    }
    """
    try:
        data = request.get_json(force=True)
        med_id = data.get("medicamento_id")
        fab_id = data.get("fabricante_id")
        precio = data.get("precio")
        precio_id = data.get("id")

        if not (med_id and fab_id and precio is not None):
            return jsonify({"ok": False, "error": "Datos incompletos"}), 400

        conn = get_db_connection()
        fecha = datetime.now().strftime("%Y-%m-%d")

        if precio_id:  #  actualizar existente
            conn.execute("""
                UPDATE precios
                SET fabricante_id = %s, precio = %s, fecha_actualizacion = %s
                WHERE id = %s
            """, (fab_id, precio, fecha, precio_id))
            conn.commit()
            nuevo_id = precio_id
        else:  #  insertar nuevo
            cur = conn.execute("""
                INSERT INTO precios (medicamento_id, fabricante_id, precio, fecha_actualizacion)
                VALUES (%s, %s, %s, %s)
                RETURNING id
            """, (med_id, fab_id, precio, fecha))
            nuevo_id = cur.fetchone()[0]
            conn.commit()

        conn.close()

        return jsonify({
            "ok": True, 
            "mensaje": "Precio guardado correctamente",
            "precio_id": nuevo_id  #  RETORNA EL ID
        })

    except Exception as e:
        print(f"Error al guardar precio: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500

@app.route('/admin/precios/<int:precio_id>/imagen', methods=['POST'])
@admin_required
def subir_imagen_precio(precio_id):
    """Sube/actualiza imagen para un precio especfico"""
    print(f" INICIO - Subir imagen para precio_id: {precio_id}")
    
    try:
        if 'imagen' not in request.files:
            print(" No se envi imagen")
            return jsonify({'ok': False, 'error': 'No se envi imagen'}), 400
        
        file = request.files['imagen']
        if not file or file.filename == '':
            print(" Archivo vaco")
            return jsonify({'ok': False, 'error': 'Archivo vaco'}), 400
        
        print(f" Archivo recibido: {file.filename}")
        
        # Obtener info del precio
        conn = get_db_connection()
        precio = conn.execute("""
            SELECT p.medicamento_id, m.nombre as med_nombre, f.nombre as fab_nombre
            FROM precios p
            INNER JOIN medicamentos m ON p.medicamento_id = m.id
            INNER JOIN fabricantes f ON p.fabricante_id = f.id
            WHERE p.id = ?
        """, (precio_id,)).fetchone()
        
        if not precio:
            conn.close()
            print(f" Precio {precio_id} no encontrado")
            return jsonify({'ok': False, 'error': 'Precio no encontrado'}), 404
        
        print(f" Precio encontrado: {precio['med_nombre']} - {precio['fab_nombre']}")
        
        # ===  NUEVA LGICA SEGURA PARA NOMBRE DE ARCHIVO ===
        import hashlib
        from werkzeug.utils import secure_filename

        # Usar solo un hash del nombre + fabricante + timestamp para evitar conflictos y lmites
        nombre_base = f"{precio['med_nombre']}_{precio['fab_nombre']}_{precio_id}".encode('utf-8')
        hash_corto = hashlib.md5(nombre_base).hexdigest()[:12]  # 12 caracteres = suficiente unicidad
        ext = file.filename.rsplit('.', 1)[1].lower() if '.' in file.filename else 'jpg'
        if ext not in ALLOWED_EXT:
            conn.close()
            print(f" Extensin no permitida: {ext}")
            return jsonify({'ok': False, 'error': f'Extensin no permitida: {ext}'}), 400

        # Nombre final: corto, seguro, sin tildes, sin espacios, sin lmite de longitud
        filename_to_save = f"med_{precio_id}_{hash_corto}.{ext}"
        # =====================================================

        print(f" Nombre generado: {filename_to_save}")
        
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename_to_save)
        
        # Guardar archivo
        file.stream.seek(0)
        file.save(filepath)
        print(f" Archivo guardado en: {filepath}")
        
        # Actualizar BD
        print(f" Actualizando BD - precio_id: {precio_id}, imagen: {filename_to_save}")
        cursor = conn.execute("UPDATE precios SET imagen = ? WHERE id = ?", 
                     (filename_to_save, precio_id))
        print(f" Filas afectadas: {cursor.rowcount}")
        
        conn.commit()
        print(" COMMIT exitoso")
        
        # Verificar
        verificar = conn.execute("SELECT imagen FROM precios WHERE id = ?", (precio_id,)).fetchone()
        print(f" Verificacin - imagen en BD: {verificar['imagen'] if verificar else 'NULL'}")
        
        conn.close()
        
        return jsonify({
            'ok': True, 
            'imagen': filename_to_save,
            'mensaje': 'Imagen actualizada correctamente'
        }), 200
        
    except Exception as e:
        print(f" ERROR: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'ok': False, 'error': str(e)}), 500


@app.route("/admin/medicamentos/nuevo", methods=['GET', 'POST'])
@admin_required   # o @admin_required si usas ese decorador; usa el que tengas definido
def nuevo_medicamento():
    """Ruta para agregar un nuevo medicamento, con subida de imagen y stock inicial."""
    if request.method == 'POST':
        nombre = request.form.get('nombre', '').strip()
        presentacion = request.form.get('presentacion', '').strip()
        concentracion = request.form.get('concentracion', '').strip()
        fabricante_nombre = request.form.get('fabricante_nombre', '').strip()
        codigo_atc_puro = request.form.get('codigo_atc_puro', '').upper().strip()
        descripcion_tecnica_atc = request.form.get('descripcion_tecnica_atc', '').strip()
        cantidad_inicial_raw = request.form.get('cantidad_inicial', '0')

        # Validar cantidad inicial
        try:
            cantidad_inicial = int(cantidad_inicial_raw)
            if cantidad_inicial < 0:
                raise ValueError()
        except ValueError:
            flash("La cantidad inicial debe ser un nmero entero vlido (>=0).", "danger")
            return redirect(url_for('nuevo_medicamento'))

        # Validaciones mnimas
        if not nombre or not fabricante_nombre:
            flash("Nombre y fabricante son requeridos.", "warning")
            return redirect(url_for('nuevo_medicamento'))

        # Manejo de imagen
        imagen_filename = None
        file = request.files.get('imagen')
        if file and file.filename:
            if allowed_file(file.filename):
                # Nombre seguro y basado en hash del contenido
                try:
                    file_hash = hash_image_content(file.stream)
                except Exception as e:
                    app.logger.error(f"Error hashing image stream: {e}")
                    file_hash = None

                ext = secure_filename(file.filename).rsplit('.', 1)[1].lower() if '.' in file.filename else ''
                if file_hash and ext:
                    filename_to_save = f"{file_hash}.{ext}"
                    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename_to_save)

                    # Guardar solo si no existe
                    if not os.path.exists(filepath):
                        try:
                            file.stream.seek(0)
                        except Exception:
                            pass
                        file.save(filepath)
                        app.logger.debug(f" Imagen guardada en: {filepath}")
                    else:
                        app.logger.debug(f" Imagen ya existe: {filepath}")

                    imagen_filename = filename_to_save
                else:
                    app.logger.debug(" No se pudo generar filename para la imagen; se omitir.")
            else:
                app.logger.debug(f" Archivo con extensin no permitida: {file.filename}")
                flash("Tipo de archivo no permitido. Usa png/jpg/jpeg/gif.", "warning")
                return redirect(url_for('nuevo_medicamento'))

        conn = get_db_connection()
        try:
            # Fabricante: insertar si no existe
            fabricante = conn.execute("SELECT id FROM fabricantes WHERE nombre = ?", (fabricante_nombre,)).fetchone()
            if fabricante is None:
                cur = conn.execute("INSERT INTO fabricantes (nombre) VALUES (?)", (fabricante_nombre,))
                fabricante_id = cur.lastrowid
            else:
                fabricante_id = fabricante['id']

            # Insertar medicamento (stock_actual inicializar en 0; se actualizar ms abajo)
            cur = conn.execute(
                "INSERT INTO medicamentos (nombre, presentacion, concentracion, codigo_atc_puro, descripcion_tecnica_atc, uso, imagen, stock_actual) VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
                (nombre, presentacion, concentracion, codigo_atc_puro, descripcion_tecnica_atc, '', imagen_filename, 0)
            )
            medicamento_id = cur.lastrowid

            # Si hay cantidad inicial, registrar existencia y actualizar stock_actual
            if cantidad_inicial > 0:
                conn.execute(
                    "INSERT INTO existencias (medicamento_id, fabricante_id, tipo_movimiento, cantidad, fecha) VALUES (?, ?, ?, ?, ?)",
                    (medicamento_id, fabricante_id, 'entrada', cantidad_inicial, datetime.now().isoformat())
                )
                conn.execute("UPDATE medicamentos SET stock_actual = stock_actual + ? WHERE id = ?", (cantidad_inicial, medicamento_id))

            conn.commit()
            flash(f"Medicamento '{nombre}' creado exitosamente.", "success")
            app.logger.debug(f" Medicamento creado ID={medicamento_id}, imagen={imagen_filename}, stock={cantidad_inicial}")
        except Exception as e:
            conn.rollback()
            app.logger.error(f" Error al guardar medicamento o existencia: {e}")
            flash("Ocurri un error guardando el medicamento.", "danger")
            return redirect(url_for('nuevo_medicamento'))
        finally:
            conn.close()

        return redirect(url_for('lista_medicamentos'))

    # GET -> render form
    return render_template('nuevo_medicamento.html')


@app.route('/admin/medicamentos/guardar_imagen_pegada/<int:medicamento_id>', methods=['POST'])
@require_role('Administrador')
def guardar_imagen_pegada(medicamento_id):
    """Recibe imagen pegada, la guarda y retorna la respuesta."""
    try:
        # Obtener la imagen del request (viene como archivo)
        if 'imagen' not in request.files:
            return {'error': 'No se envi imagen'}, 400
        
        file = request.files['imagen']
        
        if not file or file.filename == '':
            return {'error': 'Archivo vaco'}, 400
        
        # Obtener el nombre del medicamento
        conn = get_db_connection()
        medicamento = conn.execute("SELECT nombre FROM medicamentos WHERE id = ?", (medicamento_id,)).fetchone()
        conn.close()
        
        if not medicamento:
            return {'error': 'Medicamento no encontrado'}, 404
        
        nombre = medicamento['nombre']
        
        # Obtener extensin
        ext = file.filename.rsplit('.', 1)[1].lower() if '.' in file.filename else 'jpg'
        
        if ext not in ALLOWED_EXT:
            return {'error': f'Extensin no permitida: {ext}'}, 400
        
        # Generar nombre del archivo
        filename_to_save = f"{nombre}.{ext}"
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename_to_save)
        
        # Guardar archivo
        file.stream.seek(0)
        file.save(filepath)
        
        # Actualizar BD
        conn = get_db_connection()
        conn.execute("UPDATE medicamentos SET imagen = ? WHERE id = ?", 
                    (filename_to_save, medicamento_id))
        conn.commit()
        conn.close()
        
        return {'success': True, 'imagen': filename_to_save}, 200
        
    except Exception as e:
        print(f"Error guardando imagen pegada: {e}")
        return {'error': str(e)}, 500


@app.route('/api/medicamentos/<int:medicamento_id>/validar_eliminacion', methods=['GET'])
@admin_required
def validar_eliminacion_medicamento(medicamento_id):
    """Valida si un medicamento puede ser eliminado"""
    try:
        conn = get_db_connection()
        
        # Validacin 1: Existencias
        existencias = conn.execute(
            "SELECT COUNT(*) as total FROM existencias WHERE medicamento_id=%s",
            (medicamento_id,)
        ).fetchone()

        if existencias and existencias['total'] > 0:
            conn.close()
            return jsonify({
                'ok': False,
                'error': f'No se puede eliminar: Existen {existencias["total"]} registro(s) en existencias'
            })

        # Validacin 2: Componente activo (otros medicamentos lo usan)
        medicamentos_dependientes = conn.execute(
            "SELECT COUNT(*) as total FROM medicamentos WHERE componente_activo_id=%s",
            (medicamento_id,)
        ).fetchone()
        
        if medicamentos_dependientes and medicamentos_dependientes['total'] > 0:
            conn.close()
            return jsonify({
                'ok': False,
                'error': f'No se puede eliminar: Este medicamento es componente activo de {medicamentos_dependientes["total"]} otro(s) medicamento(s)'
            })
        
        #  NUEVO: Mostrar cuntos registros relacionados se eliminarn
        registros_relacionados = {}
        
        tablas_dependientes = [
            ('MEDICAMENTO_SINTOMA', 'medicamento_id'),
            ('DIAGNOSTICO_MEDICAMENTO', 'medicamento_id'),
            ('RECETAS', 'medicamento_id'),
            ('PRECIOS', 'medicamento_id'),
            ('COMPONENTES_ACTIVOS_SUGERENCIAS', 'medicamento_id'),
            ('PRECIOS_COMPETENCIA', 'medicamento_id'),
            ('SUGERIR_SINTOMAS', 'medicamento_id')
        ]
        
        for tabla, columna in tablas_dependientes:
            count = conn.execute(
                f"SELECT COUNT(*) as total FROM {tabla} WHERE {columna}=%s",
                (medicamento_id,)
            ).fetchone()
            
            if count and count['total'] > 0:
                registros_relacionados[tabla] = count['total']
        
        conn.close()
        
        return jsonify({
            'ok': True,
            'registros_relacionados': registros_relacionados
        })
        
    except Exception as e:
        if conn:
            conn.close()
        return jsonify({'ok': False, 'error': str(e)}), 500




@app.route("/medicamentos/actualizar-nombre/<int:medicamento_id>", methods=["POST"])
@admin_required
def actualizar_nombre_medicamento(medicamento_id):
    """Actualiza el nombre de un medicamento"""
    conn = None
    try:
        conn = get_db_connection()
        data = request.get_json()
        nuevo_nombre = data.get('nombre', '').strip()

        if not nuevo_nombre:
            return jsonify({'ok': False, 'error': 'El nombre no puede estar vaco'}), 400

        # Actualizar en tabla MEDICAMENTOS
        conn.execute(
            "UPDATE MEDICAMENTOS SET nombre = ? WHERE id = ?",
            (nuevo_nombre, medicamento_id)
        )
        conn.commit()

        return jsonify({'ok': True})

    except Exception as e:
        if conn:
            conn.rollback()
        print(f"Error al actualizar nombre: {e}")
        traceback.print_exc()
        return jsonify({'ok': False, 'error': str(e)}), 500
    finally:
        if conn:
            conn.close()

@app.route("/admin/medicamentos/actualizar-fabricante", methods=["POST"])
@admin_required
def actualizar_fabricante_medicamento():
    """Actualiza el fabricante de un medicamento (cambia el precio de un fabricante a otro)"""
    conn = None
    try:
        conn = get_db_connection()
        data = request.get_json()
        medicamento_id = data.get('medicamento_id')
        fabricante_id_antiguo = data.get('fabricante_id_antiguo')
        fabricante_id_nuevo = data.get('fabricante_id_nuevo')

        if not medicamento_id or not fabricante_id_nuevo:
            return jsonify({'ok': False, 'error': 'Datos incompletos'}), 400

        # Verificar si ya existe un precio para este medicamento con el nuevo fabricante
        existe = conn.execute(
            "SELECT COUNT(*) as count FROM precios WHERE medicamento_id = %s AND fabricante_id = %s",
            (medicamento_id, fabricante_id_nuevo)
        ).fetchone()

        if existe['count'] > 0:
            return jsonify({'ok': False, 'error': 'Ya existe un precio para este medicamento con el fabricante seleccionado'}), 400

        # Actualizar el fabricante en la tabla PRECIOS
        conn.execute(
            "UPDATE precios SET fabricante_id = %s WHERE medicamento_id = %s AND fabricante_id = %s",
            (fabricante_id_nuevo, medicamento_id, fabricante_id_antiguo)
        )

        # Actualizar también en precios_competencia si existen
        conn.execute(
            "UPDATE precios_competencia SET fabricante_id = %s WHERE medicamento_id = %s AND fabricante_id = %s",
            (fabricante_id_nuevo, medicamento_id, fabricante_id_antiguo)
        )

        # Actualizar en cotizaciones para que no queden huérfanas
        conn.execute(
            "UPDATE cotizaciones SET fabricante_id = %s WHERE medicamento_id = %s AND fabricante_id = %s",
            (fabricante_id_nuevo, medicamento_id, fabricante_id_antiguo)
        )

        conn.commit()

        return jsonify({'ok': True})

    except Exception as e:
        if conn:
            conn.rollback()
        print(f"Error al actualizar fabricante: {e}")
        traceback.print_exc()
        return jsonify({'ok': False, 'error': str(e)}), 500
    finally:
        if conn:
            conn.close()


@app.route("/admin/cotizaciones/huerfanas", methods=["GET"])
@admin_required
def obtener_cotizaciones_huerfanas():
    """
    Busca cotizaciones huérfanas (donde el fabricante_id no coincide con ningún precio activo).
    Retorna lista de cotizaciones que necesitan corrección.
    """
    conn = None
    try:
        conn = get_db_connection()

        # Buscar cotizaciones donde el fabricante_id no existe en la tabla precios
        # para ese medicamento_id
        # Query compatible con PostgreSQL
        query = """
            SELECT
                c.id as cotizacion_id,
                c.medicamento_id,
                c.fabricante_id as fabricante_id_cotizacion,
                c.precio as precio_cotizado,
                c.fecha_cotizacion,
                c.tercero_id,
                m.nombre as medicamento_nombre,
                f_cotiz.nombre as fabricante_cotizacion,
                t.nombre as tercero_nombre,
                STRING_AGG(DISTINCT CAST(f_valido.id AS TEXT) || ':' || f_valido.nombre, ',') as fabricantes_validos
            FROM cotizaciones c
            INNER JOIN medicamentos m ON c.medicamento_id = m.id
            LEFT JOIN fabricantes f_cotiz ON c.fabricante_id = f_cotiz.id
            LEFT JOIN terceros t ON c.tercero_id = t.id
            LEFT JOIN precios p ON c.medicamento_id = p.medicamento_id
            LEFT JOIN fabricantes f_valido ON p.fabricante_id = f_valido.id
            WHERE NOT EXISTS (
                SELECT 1 FROM precios
                WHERE precios.medicamento_id = c.medicamento_id
                AND precios.fabricante_id = c.fabricante_id
            )
            GROUP BY c.id, c.medicamento_id, c.fabricante_id, c.precio, c.fecha_cotizacion,
                     c.tercero_id, m.nombre, f_cotiz.nombre, t.nombre
            ORDER BY c.fecha_cotizacion DESC
        """

        cotizaciones_huerfanas = conn.execute(query).fetchall()

        resultados = []
        for cot in cotizaciones_huerfanas:
            fabricantes_validos = []
            if cot['fabricantes_validos']:
                for fab in cot['fabricantes_validos'].split(','):
                    if ':' in fab:
                        fab_id, fab_nombre = fab.split(':', 1)
                        fabricantes_validos.append({
                            'id': int(fab_id),
                            'nombre': fab_nombre
                        })

            resultados.append({
                'cotizacion_id': cot['cotizacion_id'],
                'medicamento_id': cot['medicamento_id'],
                'medicamento_nombre': cot['medicamento_nombre'],
                'fabricante_id_cotizacion': cot['fabricante_id_cotizacion'],
                'fabricante_cotizacion': cot['fabricante_cotizacion'],
                'precio_cotizado': cot['precio_cotizado'],
                'fecha_cotizacion': cot['fecha_cotizacion'],
                'tercero_id': cot['tercero_id'],
                'tercero_nombre': cot['tercero_nombre'],
                'fabricantes_validos': fabricantes_validos
            })

        conn.close()

        return jsonify({
            'ok': True,
            'cotizaciones_huerfanas': resultados,
            'total': len(resultados)
        })

    except Exception as e:
        if conn:
            conn.close()
        print(f"Error al buscar cotizaciones huérfanas: {e}")
        traceback.print_exc()
        return jsonify({'ok': False, 'error': str(e)}), 500


@app.route("/medicamentos/eliminar/<int:medicamento_id>", methods=["POST"])
@admin_required
def eliminar_medicamento(medicamento_id):
    """
    Ruta POST para eliminar un medicamento (completo o solo un fabricante).
    
    Lgica:
    - Si tiene 1 solo fabricante -> elimina medicamento completo
    - Si tiene mltiples fabricantes -> requiere fabricante_id para eliminar solo esa combinacin
    """
    conn = None
    try:
        conn = get_db_connection()
        
        #  Leer fabricante_id del body (opcional)
        data = request.get_json() or {}
        fabricante_id = data.get('fabricante_id')
        forzar_completo = data.get('forzar_completo', False)
        
        #  Contar cuntos fabricantes tiene este medicamento
        fabricantes = conn.execute(
            "SELECT COUNT(DISTINCT fabricante_id) as total FROM precios WHERE medicamento_id=%s",
            (medicamento_id,)
        ).fetchone()
        
        total_fabricantes = fabricantes['total'] if fabricantes else 0
        
        #  CASO 1: Usuario fuerza eliminacin completa
        if forzar_completo:
            return eliminar_medicamento_completo(conn, medicamento_id)
        
        #  CASO 2: Mltiples fabricantes
        if total_fabricantes > 1:
            if not fabricante_id:
                return jsonify({
                    'ok': False,
                    'error': f'Este medicamento tiene {total_fabricantes} fabricantes. Debe especificar cul eliminar.',
                    'requiere_fabricante': True,
                    'total_fabricantes': total_fabricantes
                }), 400
            
            # Eliminar solo la combinacin medicamento + fabricante especfico
            return eliminar_medicamento_fabricante(conn, medicamento_id, fabricante_id)
        
        #  CASO 2: Solo 1 fabricante o ninguno -> eliminar medicamento completo
        return eliminar_medicamento_completo(conn, medicamento_id)
        
    except (sqlite3.Error, Exception) as e:
        if conn:
            try:
                conn.rollback()
            except:
                pass
        
        print(f" Error al eliminar medicamento: {e}")
        return jsonify({
            'ok': False,
            'error': f'Error: {str(e)}'
        }), 500
    
    finally:
        if conn:
            try:
                conn.close()
            except:
                pass


def eliminar_medicamento_fabricante(conn, medicamento_id, fabricante_id):
    """Elimina solo la combinacin medicamento + fabricante especfico"""
    
    #  VALIDACIN: Verificar si existen registros en existencias para esta combinacin
    existencias = conn.execute(
        "SELECT COUNT(*) as total FROM existencias WHERE medicamento_id=%s AND fabricante_id=%s",
        (medicamento_id, fabricante_id)
    ).fetchone()

    if existencias and existencias['total'] > 0:
        return jsonify({
            'ok': False,
            'error': f'No se puede eliminar: Existen {existencias["total"]} registro(s) en existencias para este medicamento con este fabricante'
        }), 400

    #  Iniciar transaccin
    conn.execute("BEGIN TRANSACTION")

    #  Obtener nombre del medicamento para el mensaje
    med = conn.execute("SELECT nombre FROM medicamentos WHERE id=%s", (medicamento_id,)).fetchone()
    nombre_med = med['nombre'] if med else f"ID {medicamento_id}"

    #  Eliminar de PRECIOS
    conn.execute(
        "DELETE FROM precios WHERE medicamento_id=%s AND fabricante_id=%s",
        (medicamento_id, fabricante_id)
    )
    print(f" Eliminado precio de medicamento {medicamento_id} con fabricante {fabricante_id}")

    #  Eliminar de PRECIOS_COMPETENCIA
    result = conn.execute(
        "DELETE FROM precios_competencia WHERE medicamento_id=%s AND fabricante_id=%s",
        (medicamento_id, fabricante_id)
    )
    eliminados_comp = result.rowcount
    if eliminados_comp > 0:
        print(f" Eliminados {eliminados_comp} precio(s) de competencia")
    
    #  Confirmar transaccin
    conn.commit()
    
    return jsonify({
        'ok': True,
        'mensaje': f'Eliminado precio de "{nombre_med}" con fabricante ID {fabricante_id}',
        'tipo': 'fabricante'
    })


def eliminar_medicamento_completo(conn, medicamento_id):
    """Elimina el medicamento completo y todas sus dependencias"""
    
    #  VALIDACIN: Verificar si existen registros en existencias
    existencias = conn.execute(
        "SELECT COUNT(*) as total FROM existencias WHERE medicamento_id=%s",
        (medicamento_id,)
    ).fetchone()

    if existencias and existencias['total'] > 0:
        return jsonify({
            'ok': False,
            'error': f'No se puede eliminar: Existen {existencias["total"]} registro(s) en existencias para este medicamento'
        }), 400

    #  Iniciar transaccin
    conn.execute("BEGIN TRANSACTION")

    #  Obtener imagen antes de eliminar
    cur = conn.execute("SELECT imagen FROM medicamentos WHERE id=%s", (medicamento_id,)).fetchone()
    imagen_filename = cur['imagen'] if cur and 'imagen' in cur else None

    #  ELIMINAR DE TODAS LAS TABLAS RELACIONADAS (en orden)
    tablas_a_limpiar = [
        ('MEDICAMENTO_SINTOMA', 'medicamento_id'),
        ('DIAGNOSTICO_MEDICAMENTO', 'medicamento_id'),
        ('RECETAS', 'medicamento_id'),
        ('PRECIOS', 'medicamento_id'),
        ('COMPONENTES_ACTIVOS_SUGERENCIAS', 'medicamento_id'),
        ('PRECIOS_COMPETENCIA', 'medicamento_id'),
        ('SUGERIR_SINTOMAS', 'medicamento_id')
    ]

    total_eliminados = 0

    for tabla, columna in tablas_a_limpiar:
        result = conn.execute(f"DELETE FROM {tabla} WHERE {columna}=%s", (medicamento_id,))
        eliminados = result.rowcount
        if eliminados > 0:
            total_eliminados += eliminados
            print(f" Eliminados {eliminados} registro(s) de {tabla}")

    #  Eliminar el medicamento principal
    conn.execute("DELETE FROM medicamentos WHERE id=%s", (medicamento_id,))
    print(f" Medicamento ID {medicamento_id} eliminado")

    #  Eliminar imagen si existe y no la usan otros medicamentos
    if imagen_filename:
        file_path = os.path.join(app.config["UPLOAD_FOLDER"], imagen_filename)
        if os.path.exists(file_path):
            count_refs = conn.execute(
                "SELECT COUNT(*) FROM medicamentos WHERE imagen=%s",
                (imagen_filename,)
            ).fetchone()[0]
            
            if count_refs == 0:
                os.remove(file_path)
                print(f" Imagen {imagen_filename} eliminada del disco")
            else:
                print(f" Imagen {imagen_filename} mantenida: usada por {count_refs} medicamento(s)")
    
    #  Confirmar transaccin
    conn.commit()
    
    return jsonify({
        'ok': True,
        'mensaje': f'Medicamento ID {medicamento_id} y {total_eliminados} registro(s) relacionado(s) eliminados correctamente',
        'tipo': 'completo'
    })

@app.route('/admin/diagnosticos')
@admin_required
def lista_diagnosticos():
    """Muestra el men del administrador con un mensaje de accin (para que la plantilla sea reutilizable)."""
    return render_template('lista_diagnosticos.html', 
                           nombre=session['nombre'],
                           device_id=session['dispositivo_id'],
                           mensaje_accion="Aqu se gestionar la lgica de Diagnsticos.")

@app.route('/admin/sintomas')
@admin_required
def lista_sintomas():
    """Muestra el men del administrador con un mensaje de accin (para que la plantilla sea reutilizable)."""
    return render_template('lista_sintomas.html',
                           nombre=session['nombre'],
                           device_id=session['dispositivo_id'],
                           mensaje_accion="Aqu se gestionarn los Sntomas y sus relaciones.")

# --- FIN RUTAS DEL MEN ADMINISTRADOR ---



@app.route('/admin/sintomas/fusionar', methods=['POST'])
@admin_required
def fusionar_sintomas():
    """Fusiona varios sntomas en uno solo, migrando todos los medicamentos y diagnsticos"""
    conn = None
    try:
        data = request.get_json()
        sintoma_principal_id = data.get('sintoma_principal_id')
        sintomas_secundarios_ids = data.get('sintomas_secundarios_ids', [])
        
        if not sintoma_principal_id or not sintomas_secundarios_ids:
            return jsonify({'ok': False, 'error': 'Datos incompletos'}), 400
        
        print(f" Iniciando fusin de sntomas:")
        print(f"   Principal: {sintoma_principal_id}")
        print(f"   Secundarios: {sintomas_secundarios_ids}")
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Iniciar transaccin
        conn.execute("BEGIN TRANSACTION")
        
        # ========================================
        # MIGRAR MEDICAMENTOS
        # ========================================
        # Obtener todos los medicamentos nicos de los sntomas secundarios
        medicamentos_ids = set()
        for sintoma_id in sintomas_secundarios_ids:
            cursor.execute("""
                SELECT medicamento_id FROM medicamento_sintoma
                WHERE sintoma_id = ?
            """, (sintoma_id,))
            
            for row in cursor.fetchall():
                medicamentos_ids.add(row['medicamento_id'])
        
        print(f"    Medicamentos encontrados en secundarios: {len(medicamentos_ids)}")
        
        # Obtener medicamentos del principal
        cursor.execute("""
            SELECT medicamento_id FROM medicamento_sintoma
            WHERE sintoma_id = ?
        """, (sintoma_principal_id,))
        
        medicamentos_principal = {row['medicamento_id'] for row in cursor.fetchall()}
        print(f"    Medicamentos ya en principal: {len(medicamentos_principal)}")
        
        # Agregar solo medicamentos que no estn ya en el principal
        medicamentos_nuevos = medicamentos_ids - medicamentos_principal
        print(f"    Medicamentos a agregar: {len(medicamentos_nuevos)}")
        
        for med_id in medicamentos_nuevos:
            cursor.execute("""
                INSERT INTO medicamento_sintoma (medicamento_id, sintoma_id)
                VALUES (?, ?)
            """, (med_id, sintoma_principal_id))
        
        # ========================================
        #  MIGRAR DIAGNSTICOS
        # ========================================
        # Obtener todos los diagnsticos nicos de los sntomas secundarios
        diagnosticos_ids = set()
        for sintoma_id in sintomas_secundarios_ids:
            cursor.execute("""
                SELECT diagnostico_id FROM diagnostico_sintoma
                WHERE sintoma_id = ?
            """, (sintoma_id,))
            
            for row in cursor.fetchall():
                diagnosticos_ids.add(row['diagnostico_id'])
        
        print(f"    Diagnsticos encontrados en secundarios: {len(diagnosticos_ids)}")
        
        # Obtener diagnsticos del principal
        cursor.execute("""
            SELECT diagnostico_id FROM diagnostico_sintoma
            WHERE sintoma_id = ?
        """, (sintoma_principal_id,))
        
        diagnosticos_principal = {row['diagnostico_id'] for row in cursor.fetchall()}
        print(f"    Diagnsticos ya en principal: {len(diagnosticos_principal)}")
        
        # Agregar solo diagnsticos que no estn ya en el principal
        diagnosticos_nuevos = diagnosticos_ids - diagnosticos_principal
        print(f"    Diagnsticos a agregar: {len(diagnosticos_nuevos)}")
        
        for diag_id in diagnosticos_nuevos:
            cursor.execute("""
                INSERT INTO diagnostico_sintoma (diagnostico_id, sintoma_id)
                VALUES (?, ?)
            """, (diag_id, sintoma_principal_id))
        
        # ========================================
        # LIMPIAR ASOCIACIONES DE SECUNDARIOS
        # ========================================
        # Eliminar asociaciones de medicamentos
        for sintoma_id in sintomas_secundarios_ids:
            cursor.execute("""
                DELETE FROM medicamento_sintoma
                WHERE sintoma_id = ?
            """, (sintoma_id,))
        
        #  Eliminar asociaciones de diagnsticos
        for sintoma_id in sintomas_secundarios_ids:
            cursor.execute("""
                DELETE FROM diagnostico_sintoma
                WHERE sintoma_id = ?
            """, (sintoma_id,))
        
        # Eliminar los sntomas secundarios
        for sintoma_id in sintomas_secundarios_ids:
            cursor.execute("""
                DELETE FROM sintomas
                WHERE id = ?
            """, (sintoma_id,))
        
        conn.commit()
        conn.close()
        
        print(f"    Fusin completada exitosamente")
        print(f"    Resumen: +{len(medicamentos_nuevos)} medicamentos, +{len(diagnosticos_nuevos)} diagnsticos")
        
        return jsonify({
            'ok': True,
            'mensaje': f' Sntomas fusionados exitosamente. Se agregaron {len(medicamentos_nuevos)} medicamentos y {len(diagnosticos_nuevos)} diagnsticos.'
        })
        
    except Exception as e:
        if conn:
            try:
                conn.rollback()
                conn.close()
            except:
                pass
        
        print(f"    ERROR en fusin: {str(e)}")
        import traceback
        traceback.print_exc()
        
        return jsonify({'ok': False, 'error': str(e)}), 500


@app.route('/logout')
def logout():
    session.clear()
    flash("Has cerrado la sesin.", 'info')
    return redirect(url_for('index'))


def calcular_dosis_automatica(medicamento_id, peso_paciente=70):
    """
    Funcin PLACEHOLDER: Simula el clculo de la dosis y el tiempo mximo.
    
    NOTA: En un proyecto real, esto usara la concentracin del medicamento,
    la edad/peso del paciente y tablas mdicas.
    """
    # Usamos los IDs como strings porque vienen de request.args.get()
    if medicamento_id == '1': # Ejemplo para Ibuprofeno
        dosis = "400 mg"
        frecuencia = "Cada 8 horas"
        duracion_dias = 3
    elif medicamento_id == '2': # Ejemplo para Paracetamol
        dosis = "500 mg"
        frecuencia = "Cada 6 horas"
        duracion_dias = 5
    else:
        dosis = "Dosis Estndar"
        frecuencia = "Segn indicacin"
        duracion_dias = 7
        
    return dosis, frecuencia, duracion_dias


def guardar_receta(paciente_id, medicamento_id, dosis, frecuencia, duracion_dias, sintomas_str):
    """
    Funcin PLACEHOLDER: Simula el guardado de la receta en la tabla 'recetas'.
    """
    conn = get_db_connection()
    #  CORRECCIN CLAVE: La columna fecha_emision usa SQL function datetime('now')
    # Por lo tanto, el nmero de placeholders (?) debe coincidir con la tupla (6 vs 6).
    try:
        conn.execute("""
            INSERT INTO recetas 
                (paciente_id, medicamento_id, dosis, frecuencia, duracion_dias, fecha_emision, sintomas_base)
            VALUES (?, ?, ?, ?, ?, CURRENT_TIMESTAMP, ?) 
            """, 
            (paciente_id, medicamento_id, dosis, frecuencia, duracion_dias, sintomas_str)
        )
        conn.commit()
        receta_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
    except Exception as e:
        print(f"Error al guardar la receta: {e}")
        receta_id = None
    finally:
        conn.close()
        
    return receta_id


@app.route('/nueva_receta', methods=['GET'])
#  Mantenemos el rol de Administrador por si es el nico que puede 'emitir' la receta.
# Si el paciente puede auto-recetarse, cambia a require_role('Paciente')
@admin_required
def nueva_receta():
    """
    Genera automticamente la receta con dosis calculada y la guarda en la DB.
    """
    # 1. Obtener datos de la URL y del usuario
    paciente_id = session.get('usuario_id') # ID del usuario logueado
    
    if not paciente_id:
        flash("Error: Debes estar logueado para generar una receta.", 'danger')
        # Cambiamos 'login' por 'index' para que pase por el before_request
        return redirect(url_for('index')) 
        
    medicamento_id = request.args.get('medicamento_id')
    medicamento_nombre = request.args.get('medicamento_nombre')
    medicamento_presentacion = request.args.get('medicamento_presentacion')
    sintomas_ids_str = request.args.get('sintomas_ids')
    
    if not medicamento_id:
        flash("Error: No se seleccion un medicamento.", 'danger')
        return redirect(url_for('paciente_area'))

    # 2. Calcular dosis automticamente (usando la funcin placeholder)
    dosis_calc, frecuencia_calc, duracion_calc = calcular_dosis_automatica(medicamento_id)
    
    # 3. Guardar la receta en la base de datos
    receta_id = guardar_receta(
        paciente_id, 
        medicamento_id, 
        dosis_calc, 
        frecuencia_calc, 
        duracion_calc, 
        sintomas_ids_str
    )

    if receta_id:
        # Aqu puedes reducir el stock si es necesario (lgica futura)
        
        flash(f"Receta #{receta_id} generada automticamente y guardada.", 'success')
        
        # 4. Mostrar el resultado al usuario (renderizar una nueva plantilla)
        return render_template('receta_final.html',
                               receta_id=receta_id,
                               med_nombre=medicamento_nombre,
                               med_presentacion=medicamento_presentacion,
                               dosis=dosis_calc,
                               frecuencia=frecuencia_calc,
                               duracion=duracion_calc
                              )
    else:
        flash("Error al guardar la receta. Consulte el log.", 'danger')
        return redirect(url_for('paciente_area'))




@app.route('/admin/medicamentos/siguiente_incompleto')
@admin_required
def siguiente_medicamento_incompleto():
    """Obtiene el prximo medicamento que falta completar"""
    try:
        conn = get_db_connection()
        
        # Buscar medicamento incompleto
        medicamento = conn.execute("""
            SELECT m.id
            FROM medicamentos m
            WHERE 
                (SELECT COUNT(*) FROM precios WHERE medicamento_id = m.id AND precio > 0) = 0
                OR (SELECT COUNT(*) FROM medicamento_sintoma WHERE medicamento_id = m.id) = 0
                OR m.imagen IS NULL OR m.imagen = ''
            ORDER BY m.id ASC
            LIMIT 1
        """).fetchone()
        
        conn.close()
        
        if medicamento:
            return jsonify({'ok': True, 'medicamento_id': medicamento['id']})
        else:
            return jsonify({'ok': False, 'error': 'No hay medicamentos incompletos'})
    
    except Exception as e:
        print(f"Error: {e}")
        return jsonify({'ok': False, 'error': str(e)}), 500




# ========================================
# ENDPOINTS PARA GESTIN DE SNTOMAS - SQLite
# Copiar y pegar en tu archivo 1_medicamentos.py
# Usa tu funcin existente get_db_connection()
# ========================================

# No necesitas importar nada nuevo, ya tienes:
# - requests, BeautifulSoup, sqlite3, jsonify, request



# ========================================
# ASOCIAR MEDICAMENTO A SNTOMA
# ========================================
@app.route('/admin/sintomas/<int:sintoma_id>/asociar-medicamento', methods=['POST'])
@admin_required
def asociar_medicamento_sintoma(sintoma_id):
    """Asocia un medicamento a un sntoma"""
    try:
        data = request.get_json()
        medicamento_id = data.get('medicamento_id')
        
        if not medicamento_id:
            return jsonify({'ok': False, 'error': 'Falta el ID del medicamento'}), 400
        
        print(f" Asociando medicamento {medicamento_id} a sntoma {sintoma_id}")
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Verificar que el sntoma existe
        cursor.execute("SELECT nombre FROM sintomas WHERE id = ?", (sintoma_id,))
        sintoma = cursor.fetchone()
        if not sintoma:
            conn.close()
            return jsonify({'ok': False, 'error': 'Sntoma no encontrado'}), 404
        
        # Verificar que el medicamento existe
        cursor.execute("SELECT nombre FROM medicamentos WHERE id = ?", (medicamento_id,))
        medicamento = cursor.fetchone()
        if not medicamento:
            conn.close()
            return jsonify({'ok': False, 'error': 'Medicamento no encontrado'}), 404
        
        # Verificar si ya est asociado
        cursor.execute("""
            SELECT COUNT(*) as count FROM medicamento_sintoma
            WHERE medicamento_id = ? AND sintoma_id = ?
        """, (medicamento_id, sintoma_id))
        
        if cursor.fetchone()['count'] > 0:
            conn.close()
            return jsonify({'ok': False, 'error': 'El medicamento ya est asociado a este sntoma'}), 400
        
        # Crear la asociacin
        cursor.execute("""
            INSERT INTO medicamento_sintoma (medicamento_id, sintoma_id)
            VALUES (?, ?)
        """, (medicamento_id, sintoma_id))
        
        conn.commit()
        conn.close()
        
        print(f"    Medicamento '{medicamento['nombre']}' asociado a sntoma '{sintoma['nombre']}'")
        return jsonify({
            'ok': True,
            'mensaje': f"Medicamento '{medicamento['nombre']}' asociado correctamente"
        })
        
    except Exception as e:
        print(f"    ERROR: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'ok': False, 'error': str(e)}), 500


# ========================================
# DESASOCIAR MEDICAMENTO DE SNTOMA
# ========================================
@app.route('/admin/sintomas/<int:sintoma_id>/desasociar-medicamento', methods=['POST'])
@admin_required
def desasociar_medicamento_sintoma(sintoma_id):
    """Desasocia un medicamento de un sntoma"""
    try:
        data = request.get_json()
        medicamento_id = data.get('medicamento_id')
        
        if not medicamento_id:
            return jsonify({'ok': False, 'error': 'Falta el ID del medicamento'}), 400
        
        print(f" Desasociando medicamento {medicamento_id} de sntoma {sintoma_id}")
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Obtener nombres para el mensaje
        cursor.execute("SELECT nombre FROM sintomas WHERE id = ?", (sintoma_id,))
        sintoma = cursor.fetchone()
        
        cursor.execute("SELECT nombre FROM medicamentos WHERE id = ?", (medicamento_id,))
        medicamento = cursor.fetchone()
        
        # Verificar si existe la asociacin
        cursor.execute("""
            SELECT COUNT(*) as count FROM medicamento_sintoma
            WHERE medicamento_id = ? AND sintoma_id = ?
        """, (medicamento_id, sintoma_id))
        
        if cursor.fetchone()['count'] == 0:
            conn.close()
            return jsonify({'ok': False, 'error': 'El medicamento no est asociado a este sntoma'}), 400
        
        # Eliminar la asociacin
        cursor.execute("""
            DELETE FROM medicamento_sintoma
            WHERE medicamento_id = ? AND sintoma_id = ?
        """, (medicamento_id, sintoma_id))
        
        conn.commit()
        conn.close()
        
        nombre_med = medicamento['nombre'] if medicamento else 'Desconocido'
        nombre_sint = sintoma['nombre'] if sintoma else 'Desconocido'
        
        print(f"    Medicamento '{nombre_med}' desasociado de sntoma '{nombre_sint}'")
        return jsonify({
            'ok': True,
            'mensaje': f"Medicamento '{nombre_med}' desasociado correctamente"
        })
        
    except Exception as e:
        print(f"    ERROR: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'ok': False, 'error': str(e)}), 500



# ========================================
# NUEVA: OBTENER SNTOMAS CON MEDICAMENTOS
# ========================================
# ========================================
# OBTENER SNTOMAS CON MEDICAMENTOS Y DIAGNSTICOS
# ========================================
@app.route('/admin/sintomas/completo/json', methods=['GET'])
@admin_required
def obtener_sintomas_completo():
    """Devuelve todos los sntomas con sus medicamentos y diagnsticos asociados"""
    try:
        print(" FUNCIN obtener_sintomas_completo() INICIADA")
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Obtener todos los sntomas
        cursor.execute("""
            SELECT id, nombre, descripcion_lower 
            FROM sintomas 
            ORDER BY nombre
        """)
        sintomas = cursor.fetchall()
        
        resultado = []
        
        for s in sintomas:
            # 1. Medicamentos asociados
            cursor.execute("""
                SELECT m.id, m.nombre, m.presentacion
                FROM medicamento_sintoma ms
                INNER JOIN medicamentos m ON ms.medicamento_id = m.id
                WHERE ms.sintoma_id = ?
                ORDER BY m.nombre
            """, (s['id'],))
            medicamentos = cursor.fetchall()

            # 2. Diagnsticos asociados (usando la tabla diagnostico_sintoma + diagnosticos.descripcion)
            cursor.execute("""
                SELECT d.id, d.descripcion
                FROM diagnostico_sintoma ds
                INNER JOIN diagnosticos d ON ds.diagnostico_id = d.id
                WHERE ds.sintoma_id = ?
                ORDER BY d.descripcion
            """, (s['id'],))
            diagnosticos = cursor.fetchall()

            resultado.append({
                'id': s['id'],
                'nombre': s['nombre'],
                'sinonimos': s['descripcion_lower'] or '',
                'medicamentos': [
                    {
                        'id': m['id'],
                        'nombre': m['nombre'],
                        'presentacion': m['presentacion']
                    }
                    for m in medicamentos
                ],
                'diagnosticos': [
                    {
                        'id': d['id'],
                        'nombre': d['descripcion']  # Usamos 'descripcion' como nombre legible
                    }
                    for d in diagnosticos
                ]
            })
        
        conn.close()
        print(f"    Devolviendo {len(resultado)} sntomas con medicamentos y diagnsticos")
        return jsonify({'ok': True, 'sintomas': resultado})
    
    except Exception as e:
        print(f"    ERROR: {str(e)}")
        return jsonify({'ok': False, 'error': str(e)}), 500




# ========================================
# 1. OBTENER TODOS LOS SNTOMAS
# ========================================
@app.route('/admin/sintomas/json', methods=['GET'])
#@admin_required#
def obtener_sintomas():
    print(" FUNCIN obtener_sintomas() INICIADA")
    """Devuelve todos los sntomas disponibles en la BD"""
    try:
        print("    Conectando a BD...")
        conn = get_db_connection()
        sintomas = conn.execute("""
            SELECT id, nombre, descripcion_lower
            FROM sintomas
            ORDER BY nombre
        """).fetchall()
        print(f"    Sntomas obtenidos: {len(sintomas)}")

        conn.close()
        
        resultado = [
            {
                'id': s['id'],
                'nombre': s['nombre'],
                'sinonimos': s['descripcion_lower'] or ''
            }
            for s in sintomas
        ]

        print(f"    Devolviendo {len(resultado)} sntomas")
        return jsonify({'ok': True, 'sintomas': resultado})
    
    except Exception as e:
        print(f"    ERROR: {str(e)}")
        return jsonify({'ok': False, 'error': str(e)}), 500


# ========================================
# 2. OBTENER SNTOMAS DE UN MEDICAMENTO
# ========================================
@app.route('/admin/medicamentos/<int:medicamento_id>/sintomas', methods=['GET'])
#@admin_required#
def obtener_sintomas_medicamento(medicamento_id):
    print(f" FUNCIN obtener_sintomas_medicamento() INICIADA - ID: {medicamento_id}")

    """Devuelve los sntomas asociados a un medicamento especfico"""
    try:
        print(f"    Buscando sntomas para medicamento {medicamento_id}...")
        conn = get_db_connection()
        sintomas = conn.execute("""
            SELECT ms.sintoma_id, s.nombre, s.descripcion_lower
            FROM medicamento_sintoma ms
            INNER JOIN sintomas s ON ms.sintoma_id = s.id
            WHERE ms.medicamento_id = %s
            ORDER BY s.nombre
        """, (medicamento_id,)).fetchall()

        print(f"    Sntomas encontrados: {len(sintomas)}")
        conn.close()
        
        resultado = [
            {
                'sintoma_id': s['sintoma_id'],
                'nombre': s['nombre'],
                'sinonimos': s['descripcion_lower'] or ''
            }
            for s in sintomas
        ]

        print(f"    Devolviendo {len(resultado)} sntomas asociados")
        return jsonify({'ok': True, 'sintomas': resultado})
    
    except Exception as e:
        print(f"    ERROR: {str(e)}")
        return jsonify({'ok': False, 'error': str(e)}), 500



@app.route('/admin/medicamentos/<int:medicamento_id>/hermanos_componente')
@admin_required
def obtener_hermanos_componente(medicamento_id):
    """
    Devuelve medicamentos hermanos (mismo componente activo) con sus sntomas.
    Puede recibir ?componente_activo_id=X para usar un valor diferente al guardado
    """
    try:
        conn = get_db_connection()
        
        #  Obtener el componente_activo_id: primero del parmetro, luego de BD
        componente_id = request.args.get('componente_activo_id', type=int)
        
        if not componente_id:
            # Si no vino en query string, obtener de BD
            med_actual = conn.execute(
                "SELECT componente_activo_id FROM medicamentos WHERE id = ?",
                (medicamento_id,)
            ).fetchone()
            componente_id = med_actual['componente_activo_id'] if med_actual else None
        
        if not componente_id:
            conn.close()
            return jsonify({
                'ok': True,
                'componente_id': None,
                'componente_nombre': None,
                'medicamentos_hermanos': [],
                'sintomas_actuales': []
            })
        
        # 2. Obtener nombre del componente
        componente = conn.execute(
            "SELECT nombre FROM medicamentos WHERE id = ?",
            (componente_id,)
        ).fetchone()
        
        componente_nombre = componente['nombre'] if componente else None
        
        # 3. Obtener medicamentos hermanos (que comparten el mismo componente_activo_id)
        hermanos = conn.execute("""
            SELECT id, nombre
            FROM medicamentos
            WHERE componente_activo_id = ?
            AND id != ?
            ORDER BY nombre
        """, (componente_id, medicamento_id)).fetchall()
        
        # 4. Para cada hermano, obtener sus sntomas
        medicamentos_hermanos = []
        for hermano in hermanos:
            sintomas = conn.execute("""
                SELECT ms.sintoma_id, s.nombre
                FROM medicamento_sintoma ms
                INNER JOIN sintomas s ON ms.sintoma_id = s.id
                WHERE ms.medicamento_id = ?
                ORDER BY s.nombre
            """, (hermano['id'],)).fetchall()
            
            medicamentos_hermanos.append({
                'id': hermano['id'],
                'nombre': hermano['nombre'],
                'sintomas': [{'id': s['sintoma_id'], 'nombre': s['nombre']} for s in sintomas],
                'cantidad_sintomas': len(sintomas)
            })
        
        # 5. Obtener sntomas actuales del medicamento
        sintomas_actuales_ids = conn.execute("""
            SELECT sintoma_id
            FROM medicamento_sintoma
            WHERE medicamento_id = ?
        """, (medicamento_id,)).fetchall()
        
        sintomas_actuales = [s['sintoma_id'] for s in sintomas_actuales_ids]
        
        conn.close()
        
        return jsonify({
            'ok': True,
            'componente_id': componente_id,
            'componente_nombre': componente_nombre,
            'medicamentos_hermanos': medicamentos_hermanos,
            'sintomas_actuales': sintomas_actuales
        })
        
    except Exception as e:
        print(f" Error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'ok': False, 'error': str(e)}), 500


# ========================================
# 3. GUARDAR SNTOMAS DE UN MEDICAMENTO
# ========================================
@app.route('/admin/medicamentos/<int:medicamento_id>/sintomas', methods=['POST'])
#@admin_required#
def guardar_sintomas_medicamento(medicamento_id):
    print(f" FUNCIN guardar_sintomas_medicamento() INICIADA - ID: {medicamento_id}")
    """
    Guarda los sntomas asociados a un medicamento.
    Recibe: {"sintomas": [1, 2, 3, 4]}
    """
    try:
        data = request.get_json()
        print(f"    Datos recibidos: {data}")

        sintomas_ids = data.get('sintomas', [])
        print(f"    IDs de sntomas a guardar: {sintomas_ids}")
        print(f"    Total de sntomas: {len(sintomas_ids)}")

        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Eliminar relaciones anteriores
        print(f"    Eliminando relaciones anteriores...")
        cursor.execute("""
            DELETE FROM medicamento_sintoma 
            WHERE medicamento_id = ?
        """, (medicamento_id,))
        deleted = cursor.rowcount  #  AGREGAR ESTA LNEA
        print(f"    Relaciones eliminadas: {deleted}")

        # Insertar nuevas relaciones
        if sintomas_ids:
            print(f"    Insertando {len(sintomas_ids)} nuevas relaciones...")
            for sintoma_id in sintomas_ids:
                print(f"       Insertando medicamento={medicamento_id}, sintoma={sintoma_id}")
    
                cursor.execute("""
                    INSERT OR IGNORE INTO medicamento_sintoma (medicamento_id, sintoma_id)
                    VALUES (?, ?)
                """, (medicamento_id, sintoma_id))
        
        conn.commit()
        print(f"    COMMIT exitoso!")
        conn.close()
        
        return jsonify({
            'ok': True, 
            'mensaje': f'Se guardaron {len(sintomas_ids)} sntomas correctamente'
        })
    
    except Exception as e:
        print(f"    ERROR: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'ok': False, 'error': str(e)}), 500


# ========================================
# ACTUALIZAR UN SNTOMA EXISTENTE
# ========================================
@app.route('/admin/sintomas/<int:sintoma_id>/actualizar', methods=['POST'])
@admin_required
def actualizar_sintoma(sintoma_id):
    """Actualiza el nombre y descripcin de un sntoma existente"""
    try:
        data = request.get_json()
        nombre = data.get('nombre', '').strip()
        descripcion = data.get('descripcion', '').strip()
        
        if not nombre:
            return jsonify({'ok': False, 'error': 'El nombre es obligatorio'}), 400
        
        print(f" Actualizando sntoma ID: {sintoma_id}")
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Verificar que el sntoma existe
        cursor.execute("SELECT id FROM sintomas WHERE id = ?", (sintoma_id,))
        if not cursor.fetchone():
            conn.close()
            return jsonify({'ok': False, 'error': 'Sntoma no encontrado'}), 404
        
        # Verificar que no exista otro sntoma con el mismo nombre
        cursor.execute("""
            SELECT id FROM sintomas 
            WHERE LOWER(nombre) = LOWER(?) AND id != ?
        """, (nombre, sintoma_id))
        
        duplicado = cursor.fetchone()
        if duplicado:
            conn.close()
            return jsonify({'ok': False, 'error': f'Ya existe otro sntoma con el nombre "{nombre}"'}), 400
        
        # Actualizar
        cursor.execute("""
            UPDATE sintomas 
            SET nombre = ?, descripcion_lower = ?
            WHERE id = ?
        """, (nombre, descripcion, sintoma_id))
        
        conn.commit()
        conn.close()
        
        print(f"    Sntoma {sintoma_id} actualizado correctamente")
        return jsonify({
            'ok': True,
            'mensaje': f'Sntoma "{nombre}" actualizado exitosamente'
        })
        
    except Exception as e:
        print(f"    ERROR: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'ok': False, 'error': str(e)}), 500


# ========================================
# ELIMINAR UN SNTOMA
# ========================================
@app.route('/admin/sintomas/<int:sintoma_id>/eliminar', methods=['POST'])
@admin_required
def eliminar_sintoma(sintoma_id):
    """Elimina un sntoma y TODAS sus relaciones (medicamentos y diagnsticos)"""
    conn = None
    try:
        print(f" Eliminando sntoma ID: {sintoma_id}")
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        #  Verificar que el sntoma existe
        cursor.execute("SELECT nombre FROM sintomas WHERE id = ?", (sintoma_id,))
        sintoma = cursor.fetchone()
        
        if not sintoma:
            conn.close()
            return jsonify({'ok': False, 'error': 'Sntoma no encontrado'}), 404
        
        nombre_sintoma = sintoma['nombre']
        
        #  Iniciar transaccin
        conn.execute("BEGIN TRANSACTION")
        
        #  Contar relaciones existentes
        cursor.execute("""
            SELECT COUNT(*) as total 
            FROM medicamento_sintoma 
            WHERE sintoma_id = ?
        """, (sintoma_id,))
        total_medicamentos = cursor.fetchone()['total']
        
        cursor.execute("""
            SELECT COUNT(*) as total 
            FROM diagnostico_sintoma 
            WHERE sintoma_id = ?
        """, (sintoma_id,))
        total_diagnosticos = cursor.fetchone()['total']
        
        print(f"    Medicamentos asociados: {total_medicamentos}")
        print(f"    Diagnsticos asociados: {total_diagnosticos}")
        
        #  Eliminar relaciones con medicamentos
        cursor.execute("""
            DELETE FROM medicamento_sintoma 
            WHERE sintoma_id = ?
        """, (sintoma_id,))
        eliminados_med = cursor.rowcount
        if eliminados_med > 0:
            print(f"    Eliminadas {eliminados_med} relaciones con medicamentos")
        
        #  Eliminar relaciones con diagnsticos
        cursor.execute("""
            DELETE FROM diagnostico_sintoma 
            WHERE sintoma_id = ?
        """, (sintoma_id,))
        eliminados_diag = cursor.rowcount
        if eliminados_diag > 0:
            print(f"    Eliminadas {eliminados_diag} relaciones con diagnsticos")
        
        #  Eliminar el sntoma
        cursor.execute("DELETE FROM sintomas WHERE id = ?", (sintoma_id,))
        print(f"    Sntoma '{nombre_sintoma}' eliminado")
        
        #  Confirmar transaccin
        conn.commit()
        
        print(f"    Sntoma '{nombre_sintoma}' eliminado correctamente")
        return jsonify({
            'ok': True,
            'mensaje': f'Sntoma "{nombre_sintoma}" eliminado exitosamente',
            'medicamentos_desasociados': eliminados_med,
            'diagnosticos_desasociados': eliminados_diag
        })
        
    except Exception as e:
        if conn:
            try:
                conn.rollback()
                print(f"    Rollback ejecutado por error")
            except:
                pass
        
        print(f"    ERROR: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'ok': False, 'error': str(e)}), 500
    
    finally:
        if conn:
            try:
                conn.close()
            except:
                pass


#  FUNCIN PARA LIMPIAR NOMBRE
    def limpiar_nombre_medicamento(texto):
        """Limpia el nombre removiendo presentacin, unidades, formas, etc"""
        import re
        texto = texto.lower()
        
        palabras_remover = [
            'inyeccin', 'inyeccion', 'comprimido', 'comprimidos', 'cpsula', 'cpsulas',
            'capsula', 'capsulas', 'gotas', 'gota', 'ampolla', 'ampollas', 'jarabe',
            'solucin', 'solucion', 'suspensin', 'suspension', 'crema', 'ungento',
            'unguento', 'gel', 'polvo', 'tableta', 'pastilla', 'tabletas', 'pastillas',
            'spray', 'aerosol', 'locin', 'pomada', 'supositorios', 'supositorio',
            'parche', 'parches', 'paquete', 'pack', 'caja', 'frasco', 'botella'
        ]
        
        for palabra in palabras_remover:
            texto = re.sub(r'\b' + palabra + r'\b', '', texto)
        
        texto = re.sub(r'\b\d+\s*(mg|ml|g|gr|%|ui|mcg|g)\b', '', texto)
        texto = re.sub(r'\b(paquete|pack)\s+[x]\s+\d+\b', '', texto)
        texto = re.sub(r'\s+', ' ', texto).strip()
        
        return texto if texto else "medicamento"


#  PARA SNTOMAS (remueve TODO)
def limpiar_nombre_para_sintomas(texto):
    """Limpia removiendo presentacin, unidades, formas - PARA SNTOMAS"""
    import re
    texto = texto.lower()
    
    palabras_remover = [
        'inyeccin', 'inyeccion', 'comprimido', 'comprimidos', 'cpsula', 'cpsulas',
        'capsula', 'capsulas', 'gotas', 'gota', 'ampolla', 'ampollas', 'jarabe',
        'solucin', 'solucion', 'suspensin', 'suspension', 'crema', 'ungento',
        'unguento', 'gel', 'polvo', 'tableta', 'pastilla', 'tabletas', 'pastillas',
        'spray', 'aerosol', 'locin', 'pomada', 'supositorios', 'supositorio',
        'parche', 'parches', 'paquete', 'pack', 'caja', 'frasco', 'botella'
    ]
    
    for palabra in palabras_remover:
        texto = re.sub(r'\b' + palabra + r'\b', '', texto)
    
    #  REMUEVE TAMBIN concentracin
    texto = re.sub(r'\b\d+\s*(mg|ml|g|gr|%|ui|mcg|g)\b', '', texto)
    texto = re.sub(r'\/?\s*(mg|ml|g|gr|%|ui|mcg|g)\b', '', texto)  # Remover /ml, mg sin nmero
    texto = re.sub(r'\b(paquete|pack)\s+[x]\s+\d+\b', '', texto)
    texto = re.sub(r'\s+', ' ', texto).strip()
    
    return texto if texto else "medicamento"


# ========================================
# 4. BUSCAR SNTOMAS EN LNEA
# ========================================
@app.route('/admin/buscar_sintomas_medicamento', methods=['GET'])
@admin_required
def buscar_sintomas_medicamento():
    """
    Busca sntomas usando Google Custom Search API.
    Detecta sntomas EXISTENTES y sntomas NUEVOS con filtros inteligentes.
    """
    nombre = request.args.get('nombre', '')
    medicamento_id = request.args.get('medicamento_id', '')
    if not nombre:
        return jsonify({'ok': False, 'error': 'Falta el nombre del medicamento'})
    
    #  LIMPIAR NOMBRE INMEDIATAMENTE
    nombre = limpiar_nombre_para_sintomas(nombre)
    print(f" Buscando sntomas para: {nombre}")
    try:
        import re
        
        #  OBTENER COMPONENTE ACTIVO (SI EXISTE)
        nombre_componente_activo = None
        if medicamento_id:
            conn = get_db_connection()
            cursor = conn.cursor()
            cursor.execute("""
                SELECT m.nombre 
                FROM medicamentos m 
                WHERE m.id = (
                    SELECT componente_activo_id 
                    FROM medicamentos 
                    WHERE id = ?
                )
            """, (medicamento_id,))
            resultado = cursor.fetchone()
            conn.close()
            if resultado:
                nombre_componente_activo = resultado['nombre'].split()[0]  # Simplificar
                print(f"  Componente activo encontrado: {nombre_componente_activo}")
        
        # Google Custom Search API
        GOOGLE_API_KEY = 'AIzaSyCiAtNFl95bJJFuqiNsiYynBS3LuDisq9g'
        SEARCH_ENGINE_ID = '40c8305664a9147e9'
        
        #  TEXTO COMBINADO
        texto_completo = ""
        
        # BSQUEDA 1: Con el nombre del medicamento
        print(f" BSQUEDA 1: Consultando '{nombre}'...")
        query = f"{nombre} para qu sirve indicaciones sntomas trata alivia"
        url_api = f"https://www.googleapis.com/customsearch/v1?key={GOOGLE_API_KEY}&cx={SEARCH_ENGINE_ID}&q={query}&num=10"
        resp = requests.get(url_api, timeout=10)
        if resp.status_code == 200:
            data = resp.json()
            items = data.get("items", [])
            print(f"    Resultados: {len(items)}")
            for item in items:
                texto_completo += f" {item.get('title', '')} {item.get('snippet', '')}"
        else:
            print(f"    Error API: {resp.status_code}")
        
        #  BSQUEDA 2: Con el componente activo (SI EXISTE)
        if nombre_componente_activo:
            print(f" BSQUEDA 2: Consultando '{nombre_componente_activo}'...")
            query = f"{nombre_componente_activo} para qu sirve indicaciones sntomas trata alivia"
            url_api = f"https://www.googleapis.com/customsearch/v1?key={GOOGLE_API_KEY}&cx={SEARCH_ENGINE_ID}&q={query}&num=10"
            resp = requests.get(url_api, timeout=10)
            if resp.status_code == 200:
                data = resp.json()
                items = data.get("items", [])
                print(f"    Resultados: {len(items)}")
                for item in items:
                    texto_completo += f" {item.get('title', '')} {item.get('snippet', '')}"
            else:
                print(f"    Error API: {resp.status_code}")
        
        if not texto_completo.strip():
            return jsonify({'ok': True, 'sintomas_encontrados': [], 'sintomas_nuevos_sugeridos': [], 'total': 0})
        
        print(f"  Texto combinado: {len(texto_completo)} caracteres")
        
        # Cargar sntomas BD
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT id, nombre, descripcion_lower FROM sintomas")
        sintomas_bd = cursor.fetchall()
        conn.close()
        
        # ==========================================
        # PARTE 1: BUSCAR SNTOMAS EXISTENTES
        # ==========================================
        sintomas_encontrados = []
        texto_lower = texto_completo.lower()
        
        for sintoma in sintomas_bd:
            sid = sintoma['id']
            nombre_sintoma = sintoma['nombre']
            descripcion = sintoma['descripcion_lower'] or ''
            
            patron = r'\b' + re.escape(nombre_sintoma.lower()) + r'\b'
            
            if re.search(patron, texto_lower):
                sintomas_encontrados.append({
                    'id': sid,
                    'nombre': nombre_sintoma,
                    'confianza': 'alta',
                    'motivo': f'Encontrado: "{nombre_sintoma}"',
                    'tipo': 'existente'
                })
                print(f"    {nombre_sintoma}")
                continue
            
            # Buscar en sinnimos
            if descripcion:
                terminos = [t.strip() for t in descripcion.split(',') if len(t.strip()) > 3]
                for termino in terminos:
                    patron_t = r'\b' + re.escape(termino.lower()) + r'\b'
                    if re.search(patron_t, texto_lower):
                        sintomas_encontrados.append({
                            'id': sid,
                            'nombre': nombre_sintoma,
                            'confianza': 'media',
                            'motivo': f'Trmino: "{termino}"',
                            'tipo': 'existente'
                        })
                        print(f"    {nombre_sintoma} (trmino: {termino})")
                        break
        
        # ==========================================
        # PARTE 2: DETECTAR SNTOMAS NUEVOS
        # ==========================================
        print(f"    Buscando sntomas nuevos...")
        
        # Lista negra de palabras que NO son sntomas
        palabras_prohibidas = {
            'adultos', 'adulto', 'nios', 'nio', 'hijo', 'hijos', 'paciente', 'pacientes',
            'persona', 'personas', 'indicaciones', 'tratamiento', 'medicamento', 'medicina',
            'dosis', 'tomar', 'usar', 'recibir', 'depende', 'muchos', 'algunos', 'estos',
            'esos', 'aquellos', 'cir', 'antes', 'despus', 'durante', 'siempre', 'nunca',
            'sntomas', 'sintomas', 'causa', 'causas', 'cuidado', 'alivio', 'aliviar'
        }
        
        # Patrones MS ESPECFICOS para sntomas mdicos
        patrones_sintomas = [
            # Patrn: "alivia/trata/reduce + sntoma"
            r'(?:alivia|trata|reduce|calma|combate)\s+(?:el|la)?\s*([a-z]{4,20}(?:\s+[a-z]{4,20})?)',
            # Patrn: "dolor/fiebre/inflamacin + especificador"
            r'\b(dolor|fiebre|inflamacin|picazn|nuseas|vmitos|tos|mareos|diarrea)\s+(?:de|en)?\s*([a-z\s]{0,20})',
            # Patrn: "para la/el + sntoma"
            r'para\s+(?:el|la)\s+([a-z]{4,20}(?:\s+[a-z]{4,20})?)\b',
        ]
        
        candidatos_raw = set()
        for patron in patrones_sintomas:
            matches = re.finditer(patron, texto_lower, re.IGNORECASE)
            for match in matches:
                # Obtener el grupo capturado
                if match.lastindex and match.lastindex >= 1:
                    candidato = match.group(1).strip()
                    # Si hay un segundo grupo (ej: "dolor de cabeza")
                    if match.lastindex >= 2 and match.group(2):
                        candidato = f"{candidato} {match.group(2)}".strip()
                    
                    # Limpiar
                    candidato = re.sub(r'\b(el|la|los|las|de|del|en|con|por|para|y|o|su|tu|mi)\b', '', candidato)
                    candidato = re.sub(r'\s+', ' ', candidato).strip()
                    
                    # Validaciones bsicas
                    if candidato and 4 <= len(candidato) <= 35:
                        # Contar palabras
                        palabras = candidato.split()
                        if 1 <= len(palabras) <= 3:  # Mximo 3 palabras
                            candidatos_raw.add(candidato)
        
        print(f"    Candidatos detectados: {len(candidatos_raw)}")
        
        sintomas_nuevos = []
        
        for candidato in candidatos_raw:
            # OPCIN 1: LIMPIAR palabras prohibidas (no descartar todo)
            palabras_candidato = candidato.lower().split()
            palabras_limpias = [p for p in palabras_candidato if p not in palabras_prohibidas]
            candidato_limpio = ' '.join(palabras_limpias)
            
            # Si despus de limpiar queda algo vlido
            if candidato_limpio and len(candidato_limpio) >= 4:
                # Verificar que empiece con letra
                if not candidato_limpio[0].isalpha():
                    continue
                
                # Capitalizar correctamente
                nombre_cap = ' '.join(w.capitalize() for w in candidato_limpio.split())
                
                # Evitar duplicados en la misma lista
                if not any(s['nombre_sugerido'].lower() == candidato_limpio for s in sintomas_nuevos):
                    sintomas_nuevos.append({
                        'nombre_sugerido': nombre_cap,
                        'descripcion_sugerida': candidato_limpio,
                        'tipo': 'nuevo',
                        'confianza': 'sugerencia'
                    })
                    
                    if candidato != candidato_limpio:
                        print(f"    Nuevo: {nombre_cap} (limpiado de: '{candidato}')")
                    else:
                        print(f"    Nuevo: {nombre_cap}")
            else:
                if len(palabras_candidato) > 0:
                    print(f"    Descartado: '{candidato}' (solo palabras prohibidas)")
        
        # Limitar a 8 sugerencias nuevas
        sintomas_nuevos = sintomas_nuevos[:8]
        
        # ==========================================
        # PROCESAR RESULTADOS
        # ==========================================
        # Eliminar duplicados en existentes
        sintomas_unicos = {}
        for s in sintomas_encontrados:
            if s['id'] not in sintomas_unicos:
                sintomas_unicos[s['id']] = s
            elif s['confianza'] == 'alta':
                sintomas_unicos[s['id']] = s
        
        resultado_existentes = list(sintomas_unicos.values())
        resultado_existentes.sort(key=lambda x: 0 if x['confianza'] == 'alta' else 1)
        
        print(f"    Existentes: {len(resultado_existentes)}, Nuevos: {len(sintomas_nuevos)}")
        
        # ==========================================
        # FILTRAR INDICACIONES RECHAZADAS
        # ==========================================
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT indicacion_nombre FROM indicaciones_rechazadas")
        rechazadas = cursor.fetchall()
        conn.close()
        
        # Convertir a set de nombres en lowercase para comparacin exacta
        rechazadas_set = {r['indicacion_nombre'].lower() for r in rechazadas}
        
        # Filtrar sntomas nuevos (comparacin EXACTA)
        sintomas_nuevos_filtrados = [
            s for s in sintomas_nuevos 
            if s['descripcion_sugerida'].lower() not in rechazadas_set
        ]
        
        print(f"    Rechazadas totales: {len(rechazadas_set)}")
        print(f"    Nuevos antes: {len(sintomas_nuevos)}, despus: {len(sintomas_nuevos_filtrados)}")
        
        return jsonify({
            'ok': True,
            'medicamento': nombre,
            'sintomas_encontrados': resultado_existentes,
            'sintomas_nuevos_sugeridos': sintomas_nuevos_filtrados,
            'total': len(resultado_existentes) + len(sintomas_nuevos_filtrados),
            'fuente': 'Google Custom Search API'
        })
    
    except Exception as e:
        print(f"    ERROR: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'ok': False, 'error': f'Error: {str(e)}'})
        
# ========================================
# NOTA ADICIONAL: Si tu tabla SINTOMAS usa descripcion_lower
# y quieres ver mejor los sinnimos, considera actualizar
# los datos para que descripcion_lower contenga los trminos
# alternativos separados por comas
# ========================================


# ========================================
# 5. RECHAZAR INDICACIN (NUEVA)
# ========================================
@app.route('/admin/indicaciones/rechazar', methods=['POST'])
@admin_required
def rechazar_indicacion():
    """
    Rechaza una indicacin para que no vuelva a sugerirse en NINGN medicamento.
    Guarda en tabla indicaciones_rechazadas.
    """
    try:
        data = request.get_json()
        indicacion_nombre = data.get('indicacion_nombre', '').strip().lower()
        
        if not indicacion_nombre:
            return jsonify({'ok': False, 'error': 'Falta nombre de indicacin'})
        
        print(f" Rechazando indicacin: '{indicacion_nombre}'")
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Verificar si ya est rechazada
        cursor.execute(
            "SELECT id FROM indicaciones_rechazadas WHERE indicacion_nombre = ?",
            (indicacion_nombre,)
        )
        existe = cursor.fetchone()
        
        if existe:
            conn.close()
            return jsonify({'ok': True, 'mensaje': 'Ya estaba rechazada'})
        
        # Obtener admin_id (del usuario logueado)
        admin_id = session.get('usuario_id', 1)  # Ajusta segn tu variable de sesin
        
        # Insertar en tabla
        cursor.execute(
            """
            INSERT INTO indicaciones_rechazadas (indicacion_nombre, admin_id)
            VALUES (?, ?)
            """,
            (indicacion_nombre, admin_id)
        )
        conn.commit()
        conn.close()
        
        print(f" Indicacin rechazada guardada")
        return jsonify({'ok': True, 'mensaje': 'Indicacin rechazada exitosamente'})
    
    except Exception as e:
        print(f" Error rechazando: {str(e)}")
        return jsonify({'ok': False, 'error': str(e)})



@app.route('/admin/sintomas/crear', methods=['POST'])
@admin_required
def crear_sintoma():
    """
    Crea un sntoma nuevo y lo asocia a un medicamento.
    """
    try:
        import time
        
        data = request.get_json()
        nombre = data.get('nombre', '').strip()
        descripcion = data.get('descripcion', '').strip()
        medicamento_id = data.get('medicamento_id')
        
        if not nombre:
            return jsonify({'ok': False, 'error': 'Falta el nombre del sntoma'}), 400
        
        print(f" Procesando sntoma: {nombre}")
        
        maxReintentos = 5
        for intento in range(1, maxReintentos + 1):
            conn = None
            try:
                conn = get_db_connection()
                cursor = conn.cursor()
                
                print(f"    Intento {intento}/{maxReintentos}")
                
                # Verificar si ya existe
                cursor.execute("""
                    SELECT id FROM sintomas 
                    WHERE LOWER(nombre) = LOWER(?)
                """, (nombre,))
                
                existente = cursor.fetchone()
                
                if existente:
                    sintoma_id = existente['id']
                    print(f"    Sntoma ya existe con ID: {sintoma_id}")
                    accion = 'encontrado'
                else:
                    cursor.execute("""
                        INSERT INTO sintomas (nombre, descripcion_lower)
                        VALUES (?, ?)
                    """, (nombre, descripcion if descripcion else nombre))
                    
                    sintoma_id = cursor.lastrowid
                    print(f"    Sntoma CREADO con ID: {sintoma_id}")
                    accion = 'creado'
                
                # Asociar al medicamento
                if medicamento_id:
                    cursor.execute("""
                        SELECT COUNT(*) as count FROM medicamento_sintoma
                        WHERE medicamento_id = ? AND sintoma_id = ?
                    """, (medicamento_id, sintoma_id))
                    
                    relacion = cursor.fetchone()
                    relacion_existe = relacion['count'] > 0 if relacion else False
                    
                    if not relacion_existe:
                        cursor.execute("""
                            INSERT INTO medicamento_sintoma (medicamento_id, sintoma_id)
                            VALUES (?, ?)
                        """, (medicamento_id, sintoma_id))
                        print(f"    Asociado al medicamento")
                
                conn.commit()
                conn.close()
                print(f"    XITO en intento {intento}")
                                
                return jsonify({
                    'ok': True,
                    'sintoma_id': sintoma_id,
                    'accion': accion,
                    'ya_existia': existente is not None,
                    'mensaje': f' Sntoma "{nombre}" creado exitosamente'
                })
                
            except sqlite3.OperationalError as e:
                if conn:
                    try:
                        conn.close()
                    except:
                        pass
                
                print(f"    Error intento {intento}: {e}")
                
                if 'locked' in str(e) and intento < maxReintentos:
                    espera = intento * 0.5
                    print(f"    Esperando {espera}s antes de reintentar...")
                    time.sleep(espera)
                    continue
                else:
                    raise
                    
            except Exception as e:
                if conn:
                    try:
                        conn.close()
                    except:
                        pass
                print(f"    Error: {e}")
                raise
        
        return jsonify({'ok': False, 'error': 'BD bloqueada despus de reintentos'}), 500
        
    except Exception as e:
        print(f"    ERROR FINAL: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'ok': False, 'error': str(e)}), 500



# ==========================================
# RUTAS PARA GESTIÓN DE DIAGNÓSTICOS
# ==========================================

from flask import Blueprint, render_template, request, jsonify, flash, redirect, url_for

# Crear blueprint (o agregar a tu app principal)
# admin_bp = Blueprint('admin', __name__)

# ==========================================
# 1. PGINA PRINCIPAL DE DIAGNSTICOS
# ==========================================
@app.route('/admin/diagnosticos')
def admin_diagnosticos():
    """Renderiza la pgina de gestin de diagnsticos"""
    return render_template('lista_diagnosticos.html')

# ==========================================
# RUTAS ACTUALIZADAS PARA GESTIÓN DE DIAGNÓSTICOS
# (Con campo descripcion_lower)
# ==========================================

# ==========================================
# 2. LISTAR TODOS LOS DIAGNÓSTICOS (JSON)
# ==========================================
@app.route('/admin/diagnosticos/json', methods=['GET'])
def listar_diagnosticos_json():
    """Devuelve todos los diagnósticos con conteo de síntomas"""
    conn = get_db_connection()
    
    query = """
        SELECT 
            d.id,
            d.descripcion as nombre,
            d.descripcion_lower as sinonimos,
            COUNT(ds.sintoma_id) as total_sintomas
        FROM diagnosticos d
        LEFT JOIN diagnostico_sintoma ds ON d.id = ds.diagnostico_id
        GROUP BY d.id
        ORDER BY d.descripcion
    """
    
    diagnosticos = conn.execute(query).fetchall()
    conn.close()
    
    return jsonify({
        'ok': True,
        'diagnosticos': [dict(d) for d in diagnosticos]
    })

# ==========================================
# 3. OBTENER UN DIAGNSTICO ESPECFICO (JSON)
# ==========================================
@app.route('/admin/diagnosticos/<int:diagnostico_id>/json', methods=['GET'])
def obtener_diagnostico_json(diagnostico_id):
    """Devuelve un diagnstico con sus sntomas asociados"""
    conn = get_db_connection()
    
    # Obtener diagnstico
    diagnostico = conn.execute(
        'SELECT id, descripcion as nombre, descripcion_lower as sinonimos FROM diagnosticos WHERE id = ?',
        (diagnostico_id,)
    ).fetchone()
    
    if not diagnostico:
        conn.close()
        return jsonify({'ok': False, 'error': 'Diagnstico no encontrado'}), 404
    
    # Obtener sntomas asociados
    sintomas = conn.execute("""
        SELECT s.id, s.nombre, s.descripcion_lower as sinonimos
        FROM sintomas s
        INNER JOIN diagnostico_sintoma ds ON s.id = ds.sintoma_id
        WHERE ds.diagnostico_id = ?
        ORDER BY s.nombre
    """, (diagnostico_id,)).fetchall()
    
    conn.close()
    
    diagnostico_dict = dict(diagnostico)
    diagnostico_dict['sintomas'] = [dict(s) for s in sintomas]
    
    return jsonify({
        'ok': True,
        'diagnostico': diagnostico_dict
    })

# ==========================================
# 4. CREAR DIAGNSTICO
# ==========================================
@app.route('/admin/diagnosticos/crear', methods=['POST'])
def crear_diagnostico():
    """Crea un nuevo diagnstico"""
    data = request.get_json()
    nombre = data.get('nombre', '').strip()
    descripcion = data.get('descripcion', '').strip()
    
    if not nombre:
        return jsonify({'ok': False, 'error': 'El nombre es requerido'}), 400
    
    conn = get_db_connection()
    
    # Verificar si ya existe
    existe = conn.execute(
        'SELECT id FROM diagnosticos WHERE LOWER(descripcion) = LOWER(?)',
        (nombre,)
    ).fetchone()
    
    if existe:
        conn.close()
        return jsonify({'ok': False, 'error': 'Ya existe un diagnstico con ese nombre'}), 400
    
    # Crear diagnstico
    try:
        conn.execute(
            'INSERT INTO diagnosticos (descripcion, descripcion_lower) VALUES (?, ?)',
            (nombre, descripcion.lower() if descripcion else '')
        )
        conn.commit()
        conn.close()
        
        return jsonify({
            'ok': True,
            'mensaje': f'Diagnstico "{nombre}" creado exitosamente'
        })
    except Exception as e:
        conn.close()
        return jsonify({'ok': False, 'error': f'Error al crear diagnstico: {str(e)}'}), 500

# ==========================================
# 5. ACTUALIZAR DIAGNSTICO
# ==========================================
@app.route('/admin/diagnosticos/<int:diagnostico_id>/actualizar', methods=['POST'])
def actualizar_diagnostico(diagnostico_id):
    """Actualiza un diagnstico existente"""
    data = request.get_json()
    nombre = data.get('nombre', '').strip()
    descripcion = data.get('descripcion', '').strip()
    
    if not nombre:
        return jsonify({'ok': False, 'error': 'El nombre es requerido'}), 400
    
    conn = get_db_connection()
    
    # Verificar que existe
    existe = conn.execute(
        'SELECT id FROM diagnosticos WHERE id = ?',
        (diagnostico_id,)
    ).fetchone()
    
    if not existe:
        conn.close()
        return jsonify({'ok': False, 'error': 'Diagnstico no encontrado'}), 404
    
    # Verificar duplicados (excepto el actual)
    duplicado = conn.execute(
        'SELECT id FROM diagnosticos WHERE LOWER(descripcion) = LOWER(?) AND id != ?',
        (nombre, diagnostico_id)
    ).fetchone()
    
    if duplicado:
        conn.close()
        return jsonify({'ok': False, 'error': 'Ya existe otro diagnstico con ese nombre'}), 400
    
    # Actualizar
    try:
        conn.execute(
            'UPDATE diagnosticos SET descripcion = ?, descripcion_lower = ? WHERE id = ?',
            (nombre, descripcion.lower() if descripcion else '', diagnostico_id)
        )
        conn.commit()
        conn.close()
        
        return jsonify({
            'ok': True,
            'mensaje': f'Diagnstico "{nombre}" actualizado exitosamente'
        })
    except Exception as e:
        conn.close()
        return jsonify({'ok': False, 'error': f'Error al actualizar diagnstico: {str(e)}'}), 500

# ==========================================
# 6. ELIMINAR DIAGNSTICO
# ==========================================
@app.route('/admin/diagnosticos/<int:diagnostico_id>/eliminar', methods=['POST'])
def eliminar_diagnostico(diagnostico_id):
    """Elimina un diagnstico y TODAS sus relaciones (sntomas y medicamentos)"""
    conn = None
    try:
        conn = get_db_connection()
        
        #  Verificar que existe
        diagnostico = conn.execute(
            'SELECT descripcion FROM diagnosticos WHERE id = ?',
            (diagnostico_id,)
        ).fetchone()
        
        if not diagnostico:
            conn.close()
            return jsonify({'ok': False, 'error': 'Diagnstico no encontrado'}), 404
        
        nombre_diagnostico = diagnostico['descripcion']
        
        #  Iniciar transaccin
        conn.execute("BEGIN TRANSACTION")
        
        #  Contar relaciones existentes
        cursor = conn.cursor()
        
        cursor.execute("""
            SELECT COUNT(*) as total 
            FROM diagnostico_sintoma 
            WHERE diagnostico_id = ?
        """, (diagnostico_id,))
        total_sintomas = cursor.fetchone()['total']
        
        cursor.execute("""
            SELECT COUNT(*) as total 
            FROM diagnostico_medicamento 
            WHERE diagnostico_id = ?
        """, (diagnostico_id,))
        total_medicamentos = cursor.fetchone()['total']
        
        print(f" Eliminando diagnstico ID {diagnostico_id}: '{nombre_diagnostico}'")
        print(f"    Sntomas asociados: {total_sintomas}")
        print(f"    Medicamentos asociados: {total_medicamentos}")
        
        #  Eliminar relaciones con sntomas
        conn.execute(
            'DELETE FROM diagnostico_sintoma WHERE diagnostico_id = ?',
            (diagnostico_id,)
        )
        eliminados_sint = conn.total_changes
        if total_sintomas > 0:
            print(f"    Eliminadas {total_sintomas} relaciones con sntomas")
        
        #  Eliminar relaciones con medicamentos
        conn.execute(
            'DELETE FROM diagnostico_medicamento WHERE diagnostico_id = ?',
            (diagnostico_id,)
        )
        if total_medicamentos > 0:
            print(f"    Eliminadas {total_medicamentos} relaciones con medicamentos")
        
        #  Eliminar diagnstico
        conn.execute(
            'DELETE FROM diagnosticos WHERE id = ?',
            (diagnostico_id,)
        )
        print(f"    Diagnstico '{nombre_diagnostico}' eliminado")
        
        #  Confirmar transaccin
        conn.commit()
        
        print(f"    Diagnstico eliminado correctamente")
        return jsonify({
            'ok': True,
            'mensaje': f'Diagnstico "{nombre_diagnostico}" eliminado exitosamente',
            'sintomas_desasociados': total_sintomas,
            'medicamentos_desasociados': total_medicamentos
        })
        
    except Exception as e:
        if conn:
            try:
                conn.rollback()
                print(f"    Rollback ejecutado por error")
            except:
                pass
        
        print(f"    ERROR: {str(e)}")
        return jsonify({'ok': False, 'error': f'Error al eliminar diagnstico: {str(e)}'}), 500
    
    finally:
        if conn:
            try:
                conn.close()
            except:
                pass

# ==========================================
# 7. AGREGAR SNTOMA AL DIAGNSTICO
# ==========================================
@app.route('/admin/diagnosticos/<int:diagnostico_id>/sintomas/agregar', methods=['POST'])
def agregar_sintoma_diagnostico(diagnostico_id):
    """Asocia un sntoma a un diagnstico"""
    data = request.get_json()
    sintoma_id = data.get('sintoma_id')
    
    if not sintoma_id:
        return jsonify({'ok': False, 'error': 'ID de sntoma requerido'}), 400
    
    conn = get_db_connection()
    
    # Verificar que ambos existen
    diagnostico = conn.execute('SELECT id FROM diagnosticos WHERE id = ?', (diagnostico_id,)).fetchone()
    sintoma = conn.execute('SELECT id, nombre FROM sintomas WHERE id = ?', (sintoma_id,)).fetchone()
    
    if not diagnostico or not sintoma:
        conn.close()
        return jsonify({'ok': False, 'error': 'Diagnstico o sntoma no encontrado'}), 404
    
    # Verificar si ya est asociado
    existe = conn.execute(
        'SELECT 1 FROM diagnostico_sintoma WHERE diagnostico_id = ? AND sintoma_id = ?',
        (diagnostico_id, sintoma_id)
    ).fetchone()
    
    if existe:
        conn.close()
        return jsonify({'ok': False, 'error': 'El sntoma ya est asociado a este diagnstico'}), 400
    
    # Agregar relacin
    try:
        conn.execute(
            'INSERT INTO diagnostico_sintoma (diagnostico_id, sintoma_id) VALUES (?, ?)',
            (diagnostico_id, sintoma_id)
        )
        conn.commit()
        conn.close()
        
        return jsonify({
            'ok': True,
            'mensaje': f'Sntoma "{sintoma["nombre"]}" agregado exitosamente'
        })
    except Exception as e:
        conn.close()
        return jsonify({'ok': False, 'error': f'Error al agregar sntoma: {str(e)}'}), 500

# ==========================================
# 8. QUITAR SNTOMA DEL DIAGNSTICO
# ==========================================
@app.route('/admin/diagnosticos/<int:diagnostico_id>/sintomas/quitar', methods=['POST'])
def quitar_sintoma_diagnostico(diagnostico_id):
    """Desasocia un sntoma de un diagnstico"""
    data = request.get_json()
    sintoma_id = data.get('sintoma_id')
    
    if not sintoma_id:
        return jsonify({'ok': False, 'error': 'ID de sntoma requerido'}), 400
    
    conn = get_db_connection()
    
    # Obtener nombre del sntoma
    sintoma = conn.execute('SELECT nombre FROM sintomas WHERE id = ?', (sintoma_id,)).fetchone()
    
    if not sintoma:
        conn.close()
        return jsonify({'ok': False, 'error': 'Sntoma no encontrado'}), 404
    
    # Eliminar relacin
    try:
        conn.execute(
            'DELETE FROM diagnostico_sintoma WHERE diagnostico_id = ? AND sintoma_id = ?',
            (diagnostico_id, sintoma_id)
        )
        conn.commit()
        conn.close()
        
        return jsonify({
            'ok': True,
            'mensaje': f'Sntoma "{sintoma["nombre"]}" quitado exitosamente'
        })
    except Exception as e:
        conn.close()
        return jsonify({'ok': False, 'error': f'Error al quitar sntoma: {str(e)}'}), 500

# ==========================================
# 9. BUSCAR SNTOMAS (para el buscador)
# ==========================================
@app.route('/admin/sintomas/buscar', methods=['GET'])
def buscar_sintomas():
    """Busca sntomas por nombre o sinnimos"""
    termino = request.args.get('q', '').strip()
    
    if len(termino) < 2:
        return jsonify({'ok': False, 'error': 'Trmino muy corto'}), 400
    
    conn = get_db_connection()
    
    query = """
        SELECT id, nombre, descripcion_lower as sinonimos
        FROM sintomas
        WHERE LOWER(nombre) LIKE ? OR LOWER(descripcion_lower) LIKE ?
        ORDER BY nombre
        LIMIT 20
    """
    
    sintomas = conn.execute(query, (f'%{termino.lower()}%', f'%{termino.lower()}%')).fetchall()
    conn.close()
    
    return jsonify({
        'ok': True,
        'sintomas': [dict(s) for s in sintomas]
    })





@app.route('/buscar_medicamentos')
def buscar_medicamentos():
    """Devuelve coincidencias de medicamentos segn el texto buscado."""
    nombre = request.args.get('nombre', '').strip()

    if not nombre:
        return jsonify({"ok": False, "error": "Debe proporcionar un nombre", "medicamentos": []})

    conn = sqlite3.connect('medicamentos.db')
    c = conn.cursor()
    try:
        c.execute("""
            SELECT id, nombre
            FROM MEDICAMENTOS
            WHERE lower(nombre) LIKE ?
            ORDER BY nombre ASC
            LIMIT 10
        """, (f"%{nombre.lower()}%",))
        resultados = [{"id": r[0], "nombre": r[1]} for r in c.fetchall()]
        conn.close()

        return jsonify({"ok": True, "medicamentos": resultados})
    except Exception as e:
        conn.close()
        return jsonify({"ok": False, "error": str(e), "medicamentos": []})




# --- Crear medicamento rpido (sin recargar la pgina) ---
@app.route('/crear_medicamento_rapido', methods=['POST'])
def crear_medicamento_rapido():
    data = request.get_json()
    nombre = (data.get('nombre') or '').strip()
    medicamento_id = data.get('medicamento_id')  #  ID del medicamento actual

    if not nombre:
        return jsonify({"ok": False, "error": "El nombre no puede estar vacío"}), 400

    conn = sqlite3.connect('medicamentos.db')
    c = conn.cursor()
    try:
        # Buscar si ya existe un medicamento con ese nombre
        c.execute("SELECT id, nombre FROM MEDICAMENTOS WHERE lower(nombre) = lower(?)", (nombre,))
        existente = c.fetchone()
        
        if existente:
            componente_id, med_nombre = existente
        else:
            # Crear nuevo registro
            c.execute("INSERT INTO MEDICAMENTOS (nombre, activo) VALUES (?, 1)", (nombre,))
            conn.commit()
            componente_id = c.lastrowid
            med_nombre = nombre

        #  ASOCIAR como componente activo del medicamento actual
        if medicamento_id:
            c.execute("""
                UPDATE MEDICAMENTOS 
                SET componente_activo_id = ? 
                WHERE id = ?
            """, (componente_id, medicamento_id))
            conn.commit()
            print(f" Medicamento {medicamento_id} ahora tiene componente activo {componente_id}")

        conn.close()
        return jsonify({
            "ok": True,
            "id": componente_id,
            "nombre": med_nombre,
            "exists": bool(existente)
        })
    except Exception as e:
        conn.rollback()
        conn.close()
        return jsonify({"ok": False, "error": str(e)}), 500


# ============================================================
# FUNCION AUXILIAR - SIMPLIFICAR NOMBRE
# ============================================================
def simplificar_nombre_medicamento(nombre):
    """Simplifica el nombre eliminando descriptores innecesarios."""
    import re
    
    # Palabras a eliminar
    palabras_innecesarias = [
        'analgsico', 'analgesico', 'antiinflamatorio', 'antibitico', 'antibiotico',
        'antifngico', 'antifungico', 'antiviral', 'antihistamnico', 'antihistaminico',
        'antipirtico', 'antipiretico', 'antiespasmdico', 'antiespasmdico',
        'tpico', 'topico', 'oral', 'nasal', 'oftlmico', 'oftalmica',
        'parenteral', 'solucin', 'solucion', 'suspensin', 'suspension'
    ]
    
    nombre_limpio = nombre.lower()
    
    for palabra in palabras_innecesarias:
        nombre_limpio = re.sub(r'\b' + palabra + r'\b', '', nombre_limpio, flags=re.IGNORECASE)
    
    nombre_limpio = re.sub(r'\s+', ' ', nombre_limpio).strip()
    
    return nombre_limpio



# ============================================================
# BSQUEDA DE PRECIOS EN LNEA - CON FABRICANTES
# ============================================================
from collections import defaultdict
@app.route('/admin/buscar_precios', methods=['GET'])
def buscar_precios():
    """Busca precios validando nombre, concentracin, volumen y cantidad."""
    import re
    from statistics import median
    from collections import defaultdict
    
    print(" INICIO - Bsqueda de precios")
    
    nombre_original = request.args.get('nombre', '').strip()
    nombre = simplificar_nombre_medicamento(nombre_original)
    print(f" Medicamento original: '{nombre_original}'")
    print(f" Medicamento simplificado: '{nombre}'")
    
    if not nombre:
        return jsonify({"ok": False, "error": "Debe proporcionar un nombre"}), 400

    try:
        conn = get_db_connection()
        fabricantes_rows = conn.execute("SELECT nombre FROM fabricantes").fetchall()
        fabricantes_db = [row['nombre'].strip().upper() for row in fabricantes_rows]
        conn.close()

        print(f" Fabricantes en BD: {len(fabricantes_db)}")
        
        palabras_ignorar = {'tableta', 'capsula', 'frasco', 'caja', 'envase', 'blister', 'sobre', 
                           'solucion', 'jarabe', 'suspension', 'crema', 'gel', 'pomada', 'tubo',
                           'ampolla', 'inyectable', 'oral', 'topica', 'spray', 'gotas', 'unguento'}
        
        palabras_nombre = []
        for palabra in nombre.lower().split():
            palabra_limpia = re.sub(r'[^a-z]', '', palabra)
            if (len(palabra_limpia) >= 4 and 
                palabra_limpia not in palabras_ignorar and
                not re.match(r'^\d+$', palabra_limpia)):
                palabras_nombre.append(palabra_limpia)
            if len(palabras_nombre) >= 2:
                break
        
        if palabras_nombre:
            print(f" Palabras clave a buscar: {palabras_nombre}")
        
        patron_concentracion = re.compile(r'(\d+(?:[.,]\d+)?)\s*(mg|mcg|ui|%)\b', re.IGNORECASE)
        match_conc = patron_concentracion.search(nombre)
        
        concentracion_buscada = None
        if match_conc:
            valor = match_conc.group(1).replace(',', '.')
            unidad = match_conc.group(2).lower()
            concentracion_buscada = f"{valor}{unidad}"
            print(f" Concentracin detectada: {concentracion_buscada}")
        
        patron_volumen = re.compile(r'\b(\d+(?:[.,]\d+)?)\s*(ml|l|cc)\b', re.IGNORECASE)
        patron_peso = re.compile(r'\b(\d+(?:[.,]\d+)?)\s*(g|kg)\b', re.IGNORECASE)
        
        match_vol = patron_volumen.search(nombre)
        if not match_vol:
            match_vol = patron_peso.search(nombre)
        
        volumen_buscado = None
        volumen_min = None
        volumen_max = None
        unidad_vol = None
        valor_vol = None
        if match_vol:
            valor_vol = float(match_vol.group(1).replace(',', '.'))
            unidad_vol = match_vol.group(2).lower()
            volumen_buscado = f"{valor_vol}{unidad_vol}"
            
            volumen_min = valor_vol * 0.8
            volumen_max = valor_vol * 1.2
            print(f" Volumen/Peso detectado: {volumen_buscado} (tolerancia: {volumen_min:.0f}-{volumen_max:.0f}{unidad_vol})")
        
        patron_cantidad = re.compile(r'(?:caja|frasco|envase|blister|sobre)?\s*x\s*(\d+)', re.IGNORECASE)
        match_cant = patron_cantidad.search(nombre)
        
        if match_cant:
            cantidad_exacta = int(match_cant.group(1))
            precio_minimo_caja = cantidad_exacta * 80
            tiene_cantidad = True
            print(f" Presentacin: x{cantidad_exacta} unidades (precio mnimo: ${precio_minimo_caja:,})")
        else:
            cantidad_exacta = None
            precio_minimo_caja = 1000
            tiene_cantidad = False
            print(f" Presentacin: unitaria (precio mnimo: ${precio_minimo_caja:,})")
        
        GOOGLE_API_KEY = 'AIzaSyCiAtNFl95bJJFuqiNsiYynBS3LuDisq9g'
        SEARCH_ENGINE_ID = '40c8305664a9147e9'
        
        sitios = [
            "farmatodo.com.co",
            "larebajavirtual.com",
            "cruzverde.com.co",
            "drogueriascafam.com.co",
            "locatelcolombia.com",
            "drogueriascolsubsidio.com.co"
        ]
        
        query = f"{nombre} precio " + " OR ".join([f"site:{s}" for s in sitios])
        url_api = f"https://www.googleapis.com/customsearch/v1?key={GOOGLE_API_KEY}&cx={SEARCH_ENGINE_ID}&q={query}&num=10"
        
        print(" Consultando Google API...")
        resp = requests.get(url_api, timeout=15)
        
        if resp.status_code != 200:
            print(f" Error API: {resp.status_code}")
            return jsonify({"ok": False, "error": f"Error Google: {resp.status_code}"}), 500

        data = resp.json()
        items = data.get("items", [])
        print(f" Resultados de Google: {len(items)}")
        
        if not items:
            return jsonify({"ok": False, "error": "No se encontraron resultados"}), 404

        patron_precio = re.compile(
            r'(?:Precio[:\s.]+)?\$\s*(\d{1,3}(?:[.,]\d{3})+)|'
            r'\$\s*(\d{4,7})\b|'
            r'(\d{1,3}(?:[.,]\d{3})+)\s*(?:COP|cop|pesos?)',
            re.IGNORECASE
        )
        
        patron_fabricante = re.compile(
            r'\(([A-Z][A-Z\s&.]+)\)|'
            r'(?:de|marca)\s+([A-Z][A-Z\s&.]+)',
            re.IGNORECASE
        )
        
        resultados_raw = []
        
        for item in items:
            link = item.get("link", "")
            titulo = item.get("title", "")
            snippet = item.get("snippet", "")
            texto_completo = f"{titulo} {snippet}".lower()
            #  DEBUG: Ver qu extrae del snippet
            print(f"\n TTULO: {titulo}")
            print(f" SNIPPET: {snippet[:200]}")

            #  VALIDACIN: Rechazar si el ttulo NO contiene palabras clave
            if palabras_nombre:
                titulo_lower = titulo.lower()
                contiene_palabra_clave = any(palabra in titulo_lower for palabra in palabras_nombre)
                
                if not contiene_palabra_clave:
                    print(f"  DESCARTADO POR TTULO: {titulo[:60]}... (no tiene: {palabras_nombre})")
                    continue
            
            if palabras_nombre:
                texto_busqueda_palabras = re.sub(r'[^a-z\s]', '', texto_completo)
                encontro_palabra = False
                for palabra in palabras_nombre:
                    if palabra in texto_busqueda_palabras:
                        encontro_palabra = True
                        break
                
                if not encontro_palabra:
                    print(f"   DESCARTADO: {titulo[:60]}... (no contiene palabras clave: {palabras_nombre})")
                    continue
            
            if concentracion_buscada:
                match_item = patron_concentracion.search(texto_completo)
                
                if match_item:
                    valor_item = match_item.group(1).replace(',', '.')
                    unidad_item = match_item.group(2).lower()
                    concentracion_item = f"{valor_item}{unidad_item}"
                    
                    if concentracion_item != concentracion_buscada:
                        print(f"   DESCARTADO: {titulo[:60]}... (tiene {concentracion_item} vs {concentracion_buscada})")
                        continue
                else:
                    print(f"   DESCARTADO: {titulo[:60]}... (sin concentracin clara)")
                    continue
            
            if volumen_buscado:
                match_vol_item = patron_volumen.search(texto_completo)
                if not match_vol_item:
                    match_vol_item = patron_peso.search(texto_completo)
                
                if match_vol_item:
                    valor_vol_item = float(match_vol_item.group(1).replace(',', '.'))
                    unidad_vol_item = match_vol_item.group(2).lower()
                    
                    if unidad_vol_item == unidad_vol:
                        if volumen_min <= valor_vol_item <= volumen_max:
                            volumen_item = f"{valor_vol_item}{unidad_vol_item}"
                            if valor_vol_item != valor_vol:
                                print(f"     Volumen cercano aceptado: {volumen_item}")
                        else:
                            print(f"   DESCARTADO: {titulo[:60]}... (tiene {valor_vol_item}{unidad_vol_item} fuera de rango {volumen_min:.0f}-{volumen_max:.0f}{unidad_vol})")
                            continue
                    else:
                        print(f"   DESCARTADO: {titulo[:60]}... (unidad diferente: {unidad_vol_item} vs {unidad_vol})")
                        continue
                else:
                    print(f"   DESCARTADO: {titulo[:60]}... (sin volumen/peso claro)")
                    continue
            
            if tiene_cantidad:
                match_cantidad_item = patron_cantidad.search(texto_completo)
                
                if not match_cantidad_item:
                    print(f"   DESCARTADO: {titulo[:60]}... (sin presentacin x cantidad)")
                    continue
                
                cantidad_item = int(match_cantidad_item.group(1))
                
                if cantidad_item != cantidad_exacta:
                    print(f"   DESCARTADO: {titulo[:60]}... (tiene x{cantidad_item} vs x{cantidad_exacta})")
                    continue
            
            link_lower = link.lower()
            sitio = "Otros"
            if "farmatodo" in link_lower:
                sitio = "Farmatodo"
            elif "larebaja" in link_lower:
                sitio = "La Rebaja"
            elif "cruzverde" in link_lower:
                sitio = "Cruz Verde"
            elif "cafam" in link_lower:
                sitio = "Cafam"
            elif "locatel" in link_lower:
                sitio = "Locatel"
            elif "colsubsidio" in link_lower:
                sitio = "Colsubsidio"
            
            fabricante = "Sin especificar"
            texto_busqueda = f"{titulo} {snippet}".upper()
            
            for fab_bd in fabricantes_db:
                if fab_bd in texto_busqueda:
                    fabricante = fab_bd.title()
                    print(f"     Fabricante identificado: {fabricante}")
                    break
            
            if fabricante == "Sin especificar":
                match_fab = patron_fabricante.search(titulo)
                if match_fab:
                    fabricante = (match_fab.group(1) or match_fab.group(2) or "").strip().title()
                
                if fabricante == "Sin especificar":
                    match_fab = patron_fabricante.search(snippet)
                    if match_fab:
                        fabricante = (match_fab.group(1) or match_fab.group(2) or "").strip().title()
            
            texto_completo_precio = f"{titulo} {snippet}"
            coincidencias = patron_precio.findall(texto_completo_precio)
            precios_encontrados = []

            # Extraer TODOS los precios
            precios_candidatos = []
            for match in coincidencias:
                precio_str = next((m for m in match if m), '')
                if not precio_str:
                    continue
                limpio = re.sub(r'[.,\s]', '', precio_str)
                try:
                    valor = int(limpio)
                    if 100 <= valor <= 500000:  # Rango amplio para capturar ambos
                        precios_candidatos.append(valor)
                except ValueError:
                    continue

            #  Si tiene cantidad, identificar cul es de caja
            if tiene_cantidad and cantidad_exacta and len(precios_candidatos) >= 2:
                precios_candidatos.sort()
                precios_validados = []
                
                for precio in precios_candidatos:
                    # Verificar si es mltiplo aproximado de otros
                    es_multiplicado = False
                    for otro_precio in precios_candidatos:
                        if precio != otro_precio:
                            ratio = precio / otro_precio
                            # Si la razn es cercana a cantidad_exacta (con tolerancia 10%)
                            if abs(ratio - cantidad_exacta) < cantidad_exacta * 0.1:
                                es_multiplicado = True
                                print(f"  {precio:,} = {otro_precio:,}  {ratio:.1f} (multiplicado, correcto)")
                                break
                    
                    if es_multiplicado or precio >= precio_minimo_caja:
                        precios_validados.append(precio)
                    else:
                        print(f"  {precio:,} descartado (parece unitario, no multiplicado)")
                
                precios_encontrados = precios_validados
            else:
                # Sin cantidad, aceptar todos los que pasen validacin
                #  Si tiene cantidad pero solo encuentra 1 precio bajo, multiplicar
                if tiene_cantidad and cantidad_exacta and len(precios_candidatos) == 1:
                    unico_precio = precios_candidatos[0]
                    if 500 <= unico_precio <= 5000:  # Rango tpico unitario
                        print(f"  Precio nico detectado: ${unico_precio:,} -> multiplicando por {cantidad_exacta}")
                        precios_candidatos[0] = unico_precio * cantidad_exacta

                precios_encontrados = [p for p in precios_candidatos 
                                    if precio_minimo_caja <= p <= 500000]


            if precios_encontrados:
                #  Si tiene cantidad, usar el MAYOR (es la caja)
                if tiene_cantidad and cantidad_exacta:
                    precio = max(precios_encontrados)
                else:
                    precio = min(precios_encontrados)
                
                resultados_raw.append({
                    "fuente": sitio,
                    "fabricante": fabricante,
                    "precio": precio,
                    "link": link,
                    "titulo": titulo
                })
                
                detalles = []
                if tiene_cantidad:
                    detalles.append(f"x{cantidad_exacta}")
                if volumen_buscado:
                    detalles.append(volumen_buscado)
                
                detalle_str = f" ({', '.join(detalles)})" if detalles else ""
                print(f"   {sitio} - {fabricante}: ${precio:,}{detalle_str}")
        
        if not resultados_raw:
            print(" No se encontraron precios")
            return jsonify({"ok": False, "error": "No se encontraron precios para este medicamento"}), 404
        
        por_fabricante = defaultdict(list)
        for r in resultados_raw:
            por_fabricante[r["fabricante"]].append(r)
        
        fabricantes_info = []
        for fab, lista in por_fabricante.items():
            precios = [item["precio"] for item in lista]
            mediana = int(median(precios))
            
            fabricantes_info.append({
                "fabricante": fab,
                "mediana": mediana,
                "detalles": lista,
                "total_precios": len(precios)
            })
        
        print(f" SUCCESS! Fabricantes encontrados: {len(fabricantes_info)}")
        
        presentacion_parts = []
        if tiene_cantidad:
            presentacion_parts.append(f"x{cantidad_exacta}")
        if volumen_buscado:
            presentacion_parts.append(volumen_buscado)
        if not presentacion_parts:
            presentacion_parts.append("unitaria")
        
        presentacion_texto = " ".join(presentacion_parts)
        
        return jsonify({
            "ok": True,
            "medicamento": nombre_original,
            "presentacion": presentacion_texto,
            "fabricantes": fabricantes_info,
            "total_fabricantes": len(fabricantes_info)
        })

    except Exception as e:
        print(f" Error general: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"ok": False, "error": "Error al procesar la bsqueda"}), 500


# ========================================
# ENDPOINTS PARA GESTIN DE MEDICAMENTOS TOP
# ========================================

# 1 OBTENER LISTA DE MEDICAMENTOS TOP
@app.route('/admin/medicamentos_top/lista')
@admin_required
def obtener_lista_medicamentos_top():
    """Obtiene todos los medicamentos TOP de la BD con info de existencia"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        medicamentos = cursor.execute("""
            SELECT id, componente_activo, descripcion, laboratorio, orden, activo
            FROM medicamentos_top
            WHERE activo = '1'
            ORDER BY orden ASC
        """).fetchall()
        
        resultado = []
        for med in medicamentos:
            # Verificar si existe en la tabla medicamentos
            existe_med = cursor.execute(
                "SELECT id FROM medicamentos WHERE LOWER(nombre) = LOWER(?)",
                (med['descripcion'],)
            ).fetchone()
            
            resultado.append({
                'id': med['id'],
                'componente_activo': med['componente_activo'],
                'descripcion': med['descripcion'],
                'laboratorio': med['laboratorio'],
                'orden': med['orden'],
                'activo': med['activo'],
                'existe_en_medicamentos': existe_med is not None,
                'medicamento_id': existe_med['id'] if existe_med else None
            })
        
        conn.close()
        
        return jsonify({
            'ok': True,
            'medicamentos': resultado,
            'total': len(resultado)
        })
    except Exception as e:
        print(f" Error obteniendo lista: {e}")
        return jsonify({'ok': False, 'error': str(e)}), 500


# 2 ELIMINAR MEDICAMENTO DE TOP
@app.route('/admin/medicamentos_top/<int:medicamento_id>', methods=['DELETE'])
@admin_required
def eliminar_medicamento_top(medicamento_id):
    """Elimina un medicamento de la lista TOP (soft delete)"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Verificar que existe
        existe = cursor.execute(
            "SELECT id FROM medicamentos_top WHERE id = ?",
            (medicamento_id,)
        ).fetchone()
        
        if not existe:
            conn.close()
            return jsonify({
                'ok': False,
                'error': f'Medicamento TOP ID {medicamento_id} no encontrado'
            }), 404
        
        # Eliminar (soft delete)
        cursor.execute(
            "UPDATE medicamentos_top SET activo = 0 WHERE id = ?",
            (medicamento_id,)
        )
        conn.commit()
        
        # Obtener el nombre para confirmar
        medicamento = cursor.execute(
            "SELECT descripcion FROM medicamentos_top WHERE id = ?",
            (medicamento_id,)
        ).fetchone()
        
        conn.close()
        
        return jsonify({
            'ok': True,
            'mensaje': f' Medicamento "{medicamento["descripcion"]}" eliminado de TOP',
            'medicamento_id': medicamento_id
        })
    except Exception as e:
        print(f" Error eliminando medicamento TOP: {e}")
        return jsonify({'ok': False, 'error': str(e)}), 500


# 3 AGREGAR MEDICAMENTO A TOP
@app.route('/admin/medicamentos_top/agregar', methods=['POST'])
@admin_required
def agregar_medicamento_top():
    """
    Agrega un nuevo medicamento a la lista TOP.
    Entrada JSON:
    {
        "componente_activo": "Ibuprofeno",
        "descripcion": "Ibuprofeno 400 mg Tableta Caja x 20",
        "laboratorio": "Genrico (Procaps)"
    }
    """
    try:
        data = request.get_json()
        
        componente_activo = data.get('componente_activo', '').strip()
        descripcion = data.get('descripcion', '').strip()
        laboratorio = data.get('laboratorio', '').strip()
        
        if not all([componente_activo, descripcion, laboratorio]):
            return jsonify({
                'ok': False,
                'error': 'Faltan campos: componente_activo, descripcion, laboratorio'
            }), 400
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Verificar que no exista duplicado
        duplicado = cursor.execute(
            "SELECT id FROM medicamentos_top WHERE LOWER(descripcion) = LOWER(?)",
            (descripcion,)
        ).fetchone()
        
        if duplicado:
            conn.close()
            return jsonify({
                'ok': False,
                'error': f'Este medicamento ya existe en la lista TOP'
            }), 400
        
        # Obtener el nuevo orden (mximo actual + 1)
        max_orden = cursor.execute(
            "SELECT MAX(orden) as max_orden FROM medicamentos_top"
        ).fetchone()
        
        nuevo_orden = (max_orden['max_orden'] or 0) + 1
        
        # Insertar
        cursor.execute("""
            INSERT INTO medicamentos_top 
            (componente_activo, descripcion, laboratorio, orden, activo)
            VALUES (?, ?, ?, ?, 1)
        """, (componente_activo, descripcion, laboratorio, nuevo_orden))
        
        conn.commit()
        
        nuevo_id = cursor.lastrowid
        
        conn.close()
        
        return jsonify({
            'ok': True,
            'mensaje': f' Medicamento "{descripcion}" agregado a TOP',
            'medicamento_id': nuevo_id,
            'orden': nuevo_orden
        }), 201
        
    except Exception as e:
        print(f" Error agregando medicamento TOP: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'ok': False, 'error': str(e)}), 500






# ============================================
# ANLISIS DE COMPONENTES ACTIVOS
# ============================================

@app.route('/admin/componentes_activos/analizar', methods=['GET'])
@admin_required
def analizar_componentes_activos():
    """Obtiene medicamentos candidatos para anlisis de componente activo"""
    conn = get_db_connection()
    
    #  LIMPIAR TABLA TEMPORAL ANTES DE ANALIZAR (auto-limpieza)
    conn.execute('DELETE FROM componentes_activos_sugerencias')
    conn.commit()
    
    # Medicamentos SIN componente_activo_id
    # EXCLUYENDO los que SON componentes activos de otros
    candidatos = conn.execute('''
        SELECT m.id, m.nombre
        FROM medicamentos m
        WHERE m.componente_activo_id IS NULL
        AND m.id NOT IN (
            SELECT DISTINCT componente_activo_id 
            FROM medicamentos 
            WHERE componente_activo_id IS NOT NULL
        )
        ORDER BY m.nombre
    ''').fetchall()
    
    conn.close()
    
    return jsonify({
        'total': len(candidatos),
        'medicamentos': [dict(row) for row in candidatos]
    })


@app.route('/admin/componentes_activos/guardar_sugerencia', methods=['POST'])
@admin_required
def guardar_sugerencia():
    """Guarda una sugerencia de componente activo en tabla temporal"""
    data = request.json
    conn = get_db_connection()
    
    conn.execute('''
        INSERT INTO componentes_activos_sugerencias 
        (medicamento_id, nombre_medicamento, nombre_base_extraido, 
         componente_activo_sugerido, componente_activo_id_sugerido, 
         necesita_crear_nuevo, confianza, fuente_validacion, observaciones)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
    ''', (
        data['medicamento_id'],
        data['nombre_medicamento'],
        data.get('nombre_base_extraido'),
        data['componente_activo_sugerido'],
        data.get('componente_activo_id_sugerido'),
        data.get('necesita_crear_nuevo', False),
        data.get('confianza', 'media'),
        data.get('fuente_validacion'),
        data.get('observaciones')
    ))
    
    conn.commit()
    conn.close()
    
    return jsonify({'success': True})


@app.route('/admin/componentes_activos/sugerencias', methods=['GET'])
@admin_required
def ver_sugerencias():
    """Lista todas las sugerencias"""
    conn = get_db_connection()
    
    sugerencias = conn.execute('''
        SELECT * FROM componentes_activos_sugerencias
        ORDER BY estado, confianza DESC, fecha_analisis DESC
    ''').fetchall()
    
    conn.close()
    
    return jsonify({
        'sugerencias': [dict(row) for row in sugerencias]
    })



@app.route('/admin/componentes_activos')
@admin_required
def admin_componentes_activos():
    """Pgina de anlisis de componentes activos"""
    return render_template('analizar_componentes.html')




@app.route('/admin/componentes_activos/analizar_medicamento', methods=['POST'])
@admin_required
def analizar_medicamento_ia():
    """Analiza un medicamento usando bsqueda web para determinar su componente activo"""
    data = request.json
    medicamento_id = data['medicamento_id']
    nombre_medicamento = data['nombre_medicamento']
    
    conn = get_db_connection()
    
    # 1. Extraer nombre base (sin concentracin ni presentacin)
    nombre_base = extraer_nombre_base(nombre_medicamento)
    
    # 2. Buscar en BD si ya existe un medicamento con ese nombre base
    componente_existente = conn.execute(
        'SELECT id, nombre FROM medicamentos WHERE LOWER(nombre) = LOWER(?)',
        (nombre_base,)
    ).fetchone()
    
    # 3. Preparar respuesta base
    resultado = {
        'medicamento_id': medicamento_id,
        'nombre_medicamento': nombre_medicamento,
        'nombre_base_extraido': nombre_base,
        'componente_activo_sugerido': nombre_base,
        'componente_activo_id_sugerido': None,
        'necesita_crear_nuevo': False,
        'confianza': 'media',
        'fuente_validacion': '',
        'observaciones': ''
    }
    
    # 4. Si existe en BD, usarlo
    if componente_existente:
        resultado['componente_activo_id_sugerido'] = componente_existente['id']
        resultado['componente_activo_sugerido'] = componente_existente['nombre']
        resultado['confianza'] = 'alta'
        resultado['observaciones'] = f'Encontrado en BD: {componente_existente["nombre"]}'
    else:
        # 5. No existe - necesita crear nuevo
        resultado['necesita_crear_nuevo'] = True
        resultado['componente_activo_sugerido'] = nombre_base
        resultado['observaciones'] = f'Crear nuevo componente: {nombre_base}'
    
    conn.close()
    
    return jsonify(resultado)


def extraer_nombre_base(nombre_completo):
    """Extrae el componente activo del medicamento"""
    import re
    
    nombre = nombre_completo.strip()
    nombre_lower = nombre.lower()
    
    # 1. FILTRAR NO-MEDICAMENTOS
    no_medicamentos = [
        'aguja', 'jeringa', 'catter', 'sonda', 'bomba', 'bolsa', 'termmetro',
        'tensimetro', 'glucmetro', 'nebulizador', 'espaciador',
        'guante', 'mascarilla', 'alcohol', 'algodn', 'gasa', 'venda', 'esparadrapo',
        'lanceta', 'tira reactiva', 'suero', 'agua destilada',
        'almohada', 'almohadilla', 'faja', 'rodillera'
    ]
    
    for item in no_medicamentos:
        if item in nombre_lower:
            return None
    
    # 2. CASOS SIN COMPONENTE ESPECFICO
    if any(x in nombre_lower for x in ['combinaciones', 'varias', 'varios', 'citotxicos', 'mltiple']):
        return None
    
    # 3. COMPUESTOS QUMICOS COMPLETOS
    compuestos_patron = r'((?:bicarbonato|carbonato|cloruro|sulfato|fosfato|acetato|hidrxido|xido)\s+de\s+\w+(?:\s*\+\s*\w+)?)'
    match_compuesto = re.search(compuestos_patron, nombre_lower)
    if match_compuesto:
        compuesto = match_compuesto.group(1).strip()
        if '+' in compuesto:
            partes = [p.strip().title() for p in compuesto.split('+')]
            return ' + '.join(partes)
        return compuesto.title()
    
    # 4. LIMPIAR PASO A PASO
    nombre_limpio = nombre
    
    # A. Clasificadores al inicio
    clasificadores = [
        'analgsico', 'anestsico', 'antibitico', 'antiinflamatorio', 'antimictico',
        'antifngico', 'antiemtico', 'antisptico', 'anticido', 'anticonceptivo',
        'antirretroviral', 'antineoplsico', 'antispticos'
    ]
    for clasificador in clasificadores:
        nombre_limpio = re.sub(rf'^\s*{clasificador}\s+', '', nombre_limpio, flags=re.IGNORECASE)
    
    # B. Eliminar "autoinyector" al inicio
    nombre_limpio = re.sub(r'^\s*autoinyector\s+', '', nombre_limpio, flags=re.IGNORECASE)
    
    # C. Concentraciones
    nombre_limpio = re.sub(r'\d+\.?\d*\s*%', '', nombre_limpio)
    nombre_limpio = re.sub(r'\d+\.?\d*\s*(mg|mcg|g|g|ml|UI|U|mEq|L|Gx\d+Mm)/?\d*', '', nombre_limpio, flags=re.IGNORECASE)
    
    # D. Formas farmacuticas
    formas = [
        'tableta', 'tabletas', 'capsula', 'capsulas', 'cpsula', 'cpsulas',
        'jarabe', 'crema', 'gel', 'pomada', 'ungento',
        'suspension', 'suspensin', 'solucion', 'solucin',
        'ampolla', 'ampollas', 'inyectable', 'inyeccin',
        'gotas', 'gota', 'spray', 'inhalador', 'supositorio', 'supositorios',
        'parche', 'sobre', 'sobres', 'polvo', 'granulado',
        'colirio', 'dosis', 'comprimidos?', 'efervescente'
    ]
    for forma in formas:
        nombre_limpio = re.sub(rf'\b{forma}s?\b', '', nombre_limpio, flags=re.IGNORECASE)
    
    # E. Vas y descriptores
    descriptores = [
        'tpico', 'tpica', 'oral', 'oftlmico', 'oftlmica', 'tico', 'tica',
        'nasal', 'rectal', 'vaginal', 'dental', 'bucal',
        'local', 'hormonal', 'combinado', 'porttil', 'peditrico', 'reutilizable'
    ]
    for desc in descriptores:
        nombre_limpio = re.sub(rf'\b{desc}\b', '', nombre_limpio, flags=re.IGNORECASE)
    
    # F. Cantidades y envases (incluyendo "tubo", "frasco")
    nombre_limpio = re.sub(r'\s*(?:caja|frasco|tubo|sobre|ampolla|blister|envase)\s*(?:x|de|con)?\s*\d*.*$', '', nombre_limpio, flags=re.IGNORECASE)
    nombre_limpio = re.sub(r'\s*x\s*\d+.*$', '', nombre_limpio, flags=re.IGNORECASE)
    
    # G. Nmeros sueltos
    nombre_limpio = re.sub(r'\b\d+\b', '', nombre_limpio)
    
    # H. Palabras finales sobrantes: "tubo", "frasco" (por si quedaron)
    palabras_finales = ['tubo', 'frasco', 'envase', 'caja', 'sobre', 'autoinyector']
    for palabra in palabras_finales:
        nombre_limpio = re.sub(rf'\b{palabra}\b', '', nombre_limpio, flags=re.IGNORECASE)
    
    # I. Limpiar "en/con/de" al final
    nombre_limpio = re.sub(r'\s+(?:en|con|de)\s*$', '', nombre_limpio, flags=re.IGNORECASE)
    
    # J. Limpiar espacios y caracteres especiales
    nombre_limpio = re.sub(r'\s+', ' ', nombre_limpio).strip()
    nombre_limpio = re.sub(r'[/\(\)]+$', '', nombre_limpio).strip()
    
    # K. Eliminar preposiciones sobrantes
    nombre_limpio = re.sub(r'^\s*(?:de|con|para|en)\s+', '', nombre_limpio, flags=re.IGNORECASE)
    nombre_limpio = re.sub(r'\s+(?:de|con|para|en)\s*$', '', nombre_limpio, flags=re.IGNORECASE)
    
    # L. Validar
    if not nombre_limpio or len(nombre_limpio) < 3:
        return None
    
    # M. Capitalizar
    if '+' in nombre_limpio:
        partes = [p.strip().capitalize() for p in nombre_limpio.split('+')]
        return ' + '.join(partes)
    
    palabras = nombre_limpio.split()
    nombre_final = ' '.join(word.capitalize() for word in palabras)
    
    return nombre_final



@app.route('/admin/componentes_activos/importar_corregidos', methods=['POST'])
@admin_required
def importar_corregidos():
    """Importa JSON corregido manualmente y actualiza la tabla temporal"""
    data = request.json
    datos_corregidos = data.get('datos', [])
    
    if not datos_corregidos:
        return jsonify({'success': False, 'error': 'No hay datos para importar'})
    
    conn = get_db_connection()
    actualizados = 0
    creados = 0
    
    try:
        for item in datos_corregidos:
            medicamento_id = item['medicamento_id']
            componente_final = item['componente_activo_final']
            confianza = item.get('confianza', 'alta')
            observacion = item.get('observacion', '')
            
            # Buscar si el componente activo ya existe en BD
            componente_existente = conn.execute(
                'SELECT id FROM medicamentos WHERE LOWER(nombre) = LOWER(?)',
                (componente_final,)
            ).fetchone()
            
            componente_activo_id = componente_existente['id'] if componente_existente else None
            necesita_crear = not componente_existente
            
            # Actualizar o insertar en tabla temporal
            existe = conn.execute(
                'SELECT id FROM componentes_activos_sugerencias WHERE medicamento_id = ?',
                (medicamento_id,)
            ).fetchone()
            
            if existe:
                # Actualizar
                conn.execute('''
                    UPDATE componentes_activos_sugerencias 
                    SET componente_activo_sugerido = ?,
                        componente_activo_id_sugerido = ?,
                        necesita_crear_nuevo = ?,
                        confianza = ?,
                        observaciones = ?,
                        fuente_validacion = 'Anlisis manual con IA'
                    WHERE medicamento_id = ?
                ''', (componente_final, componente_activo_id, necesita_crear, confianza, observacion, medicamento_id))
                actualizados += 1
            else:
                # Insertar nuevo (por si acaso)
                med = conn.execute('SELECT nombre FROM medicamentos WHERE id = ?', (medicamento_id,)).fetchone()
                if med:
                    conn.execute('''
                        INSERT INTO componentes_activos_sugerencias 
                        (medicamento_id, nombre_medicamento, nombre_base_extraido, 
                         componente_activo_sugerido, componente_activo_id_sugerido, 
                         necesita_crear_nuevo, confianza, fuente_validacion, observaciones)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                    ''', (medicamento_id, med['nombre'], '', componente_final, componente_activo_id, 
                          necesita_crear, confianza, 'Anlisis manual con IA', observacion))
                    creados += 1
        
        conn.commit()
        conn.close()
        
        return jsonify({
            'success': True,
            'actualizados': actualizados,
            'creados': creados
        })
        
    except Exception as e:
        conn.close()
        return jsonify({'success': False, 'error': str(e)})





@app.route('/admin/componentes_activos/guardar_estados', methods=['POST'])
@admin_required
def guardar_estados():
    """Guarda los estados (pendiente/aprobado/rechazado) en la BD"""
    data = request.json
    estados = data.get('estados', [])
    
    if not estados:
        return jsonify({'success': False, 'error': 'No hay estados para guardar'})
    
    conn = get_db_connection()
    actualizados = 0
    
    try:
        for item in estados:
            sugerencia_id = item['id']
            nuevo_estado = item['estado']
            
            conn.execute('''
                UPDATE componentes_activos_sugerencias 
                SET estado = ?
                WHERE id = ?
            ''', (nuevo_estado, sugerencia_id))
            actualizados += 1
        
        conn.commit()
        conn.close()
        
        return jsonify({
            'success': True,
            'actualizados': actualizados
        })
        
    except Exception as e:
        conn.close()
        return jsonify({'success': False, 'error': str(e)})






@app.route('/admin/componentes_activos/aplicar_aprobados', methods=['POST'])
@admin_required
def aplicar_aprobados():
    """Aplica los cambios aprobados: crea componentes activos y asigna relaciones"""
    conn = get_db_connection()
    
    try:
        # 1. Obtener sugerencias aprobadas
        aprobados = conn.execute('''
            SELECT * FROM componentes_activos_sugerencias
            WHERE estado = 'aprobado'
            ORDER BY componente_activo_sugerido
        ''').fetchall()
        
        if not aprobados:
            conn.close()
            return jsonify({'success': False, 'error': 'No hay sugerencias aprobadas'})
        
        componentes_creados = 0
        asignaciones_realizadas = 0
        
        # 2. Procesar cada sugerencia aprobada
        for sug in aprobados:
            medicamento_id = sug['medicamento_id']
            componente_nombre = sug['componente_activo_sugerido']
            necesita_crear = sug['necesita_crear_nuevo']
            componente_id_existente = sug['componente_activo_id_sugerido']
            
            componente_activo_id = None
            
            # 3. Si necesita crear, crear el componente activo base
            if necesita_crear:
                # Verificar si ya existe (por si acaso)
                existe = conn.execute(
                    'SELECT id FROM medicamentos WHERE LOWER(nombre) = LOWER(?)',
                    (componente_nombre,)
                ).fetchone()
                
                if existe:
                    componente_activo_id = existe['id']
                else:
                    # Crear nuevo componente activo base (sin concentracin, genrico)
                    conn.execute('''
                        INSERT INTO medicamentos (nombre)
                        VALUES (?)
                    ''', (componente_nombre,))
                    componente_activo_id = conn.execute('SELECT last_insert_rowid()').fetchone()[0]
                    componentes_creados += 1
            else:
                # Ya existe, usar el ID
                componente_activo_id = componente_id_existente
            
            # 4. Asignar componente_activo_id al medicamento
            if componente_activo_id:
                conn.execute('''
                    UPDATE medicamentos
                    SET componente_activo_id = ?
                    WHERE id = ?
                ''', (componente_activo_id, medicamento_id))
                asignaciones_realizadas += 1
        
        conn.commit()
        
        # 5. Limpiar tabla temporal de aprobados
        conn.execute("DELETE FROM componentes_activos_sugerencias WHERE estado = 'aprobado'")
        conn.commit()
        
        conn.close()
        
        return jsonify({
            'success': True,
            'componentes_creados': componentes_creados,
            'asignaciones_realizadas': asignaciones_realizadas
        })
        
    except Exception as e:
        conn.close()
        return jsonify({'success': False, 'error': str(e)})



@app.route('/admin/actualizar_precios', methods=['GET', 'POST'])
def actualizar_precios():
    """Pantalla para actualizar precios comparando con competencia."""
    if request.method == 'POST':
        accion = request.form.get('accion')
        
        conn = sqlite3.connect('medicamentos.db')
        cursor = conn.cursor()
        
        # Accin: Actualizar nombre del medicamento
        if accion == 'actualizar_nombre':
            medicamento_id = request.form.get('medicamento_id')
            nuevo_nombre = request.form.get('nuevo_nombre')
            
            cursor.execute("""
                UPDATE medicamentos 
                SET nombre = ?
                WHERE id = ?
            """, (nuevo_nombre, medicamento_id))
            
            conn.commit()
            conn.close()
            return jsonify({"ok": True})
        
        # Accin: Guardar precios
        elif accion == 'guardar_precios':
            medicamento_id = request.form.get('medicamento_id')
            precio_id = request.form.get('precio_id')
            fabricante_id = request.form.get('fabricante_id')
            precio_sugerido = request.form.get('precio_sugerido')
            
            # Actualizar precio y fabricante en PRECIOS
            cursor.execute("""
                UPDATE PRECIOS 
                SET precio = ?, fabricante_id = ?, fecha_actualizacion = DATE('now')
                WHERE id = ?
            """, (precio_sugerido, fabricante_id, precio_id))
            
            print(f" Actualizado precio_id={precio_id}: precio={precio_sugerido}, fabricante={fabricante_id}")
            
            # Borrar precios antiguos de competencia para este medicamento
            cursor.execute("""
                DELETE FROM precios_competencia 
                WHERE medicamento_id = ?
            """, (medicamento_id,))
            
            # Guardar nuevos precios de competencia
            competidores = cursor.execute("""
                SELECT t.id 
                FROM terceros t
                JOIN terceros_competidores tc ON t.id = tc.tercero_id
            """).fetchall()
            
            for (comp_id,) in competidores:
                precio_comp = request.form.get(f'precio_competidor_{comp_id}')
                if precio_comp:
                    cursor.execute("""
                        INSERT INTO precios_competencia 
                        (medicamento_id, competidor_id, precio, fecha_actualizacion)
                        VALUES (?, ?, ?, DATE('now'))
                    """, (medicamento_id, comp_id, precio_comp))
                    print(f"   Guardado precio competidor {comp_id}: ${precio_comp}")
            
            conn.commit()
            conn.close()
            return jsonify({"ok": True})
    
    # GET - Mostrar formulario
    conn = sqlite3.connect('medicamentos.db')
    cursor = conn.cursor()
    
    #    OBTENER CONFIGURACIN GLOBAL (NUEVO)
    config_row = cursor.execute("SELECT * FROM configuracion_precios WHERE id = 1").fetchone()
    if config_row:
        # Asumiendo que la tabla tiene columnas: id, descuento_competencia, recargo_escaso, redondeo_superior
        config = {
            'descuento_competencia': config_row[1],
            'recargo_escaso': config_row[2],
            'redondeo_superior': config_row[3],
            'ganancia_min_escaso': config_row[4],  # <-- Nuevo
            'ganancia_max_escaso': config_row[5],   # <-- Nuevo
            'base_escaso': config_row[6],
            'usar_precio': config_row[7] if len(config_row) > 7 else config_row[6]  # Usar nuevo campo si existe
        }
    else:
        # Valores por defecto si no existe la configuracin
        config = {
            'descuento_competencia': 200,
            'recargo_escaso': 30,
            'redondeo_superior': 100,
            'ganancia_min_escaso': 2000,  # <-- Nuevo
            'ganancia_max_escaso': 10000,  # <-- Nuevo
            'base_escaso': 'minimo'
        }
    
    # Obtener medicamentos con precio 0
    medicamentos = cursor.execute("""
        SELECT 
            m.id, 
            m.nombre, 
            f.id, 
            f.nombre, 
            p.precio, 
            p.id
        FROM medicamentos m
        JOIN PRECIOS p ON m.id = p.medicamento_id
        JOIN FABRICANTES f ON p.fabricante_id = f.id
        WHERE p.precio = 0
        ORDER BY 
            EXISTS (
                SELECT 1 
                FROM MEDICAMENTO_SINTOMA ms 
                WHERE ms.medicamento_id = m.id
            ) DESC,
            m.nombre ASC
        LIMIT 10
    """).fetchall()

    # Obtener todos los fabricantes
    fabricantes = cursor.execute("SELECT id, nombre FROM FABRICANTES ORDER BY nombre").fetchall()
    
    # Obtener competidores
    competidores = cursor.execute("""
        SELECT t.id, t.nombre 
        FROM terceros t
        JOIN terceros_competidores tc ON t.id = tc.tercero_id
        ORDER BY t.nombre
    """).fetchall()
    
    conn.close()
    
    return render_template('actualizar_precios.html', 
                         medicamentos=medicamentos,
                         fabricantes=fabricantes,
                         competidores=competidores,
                         config=config)  #    PASAR CONFIG AL TEMPLATE



@app.route('/admin/guardar-configuracion', methods=['POST'])
def guardar_configuracion():
    try:
        data = request.get_json()
        conn = get_db_connection()
        conn.execute("""
            UPDATE configuracion_precios
            SET
                descuento_competencia = ?,
                recargo_escaso = ?,
                redondeo_superior = ?,
                ganancia_min_escaso = ?,
                ganancia_max_escaso = ?,
                base_escaso = ?,
                usar_precio = ?
            WHERE id = 1
        """, (
            data['descuento_competencia'],
            data['recargo_escaso'],
            data['redondeo_superior'],
            data['ganancia_min_escaso'],
            data['ganancia_max_escaso'],
            data.get('base_escaso', 'minimo'),
            data.get('usar_precio', data.get('base_escaso', 'minimo'))  # Sincronizar ambos campos
        ))
        conn.commit()
        conn.close()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500



@app.route('/admin/precios/guardar-fabricante', methods=['POST'])
def guardar_fabricante():
    """Guarda o actualiza el fabricante asociado a un medicamento en la tabla precios."""
    data = request.get_json()
    med_id = data.get('medicamento_id')
    fab_id = data.get('fabricante_id')
    
    if not med_id or not fab_id:
        return jsonify({'ok': False, 'error': 'Faltan datos'}), 400

    db = get_db_connection()
    # Verificar si ya existe un registro en precios
    precio = db.execute('SELECT id FROM precios WHERE medicamento_id = ?', (med_id,)).fetchone()
    if precio:
        # Actualizar fabricante
        db.execute('UPDATE precios SET fabricante_id = ? WHERE id = ?', (fab_id, precio['id']))
    else:
        # Crear nuevo registro (sin precio an)
        db.execute('INSERT INTO precios (medicamento_id, fabricante_id, precio) VALUES (?, ?, 0)', (med_id, fab_id))
    db.commit()
    db.close()
    return jsonify({'ok': True})



@app.route('/admin/precios/guardar-precio-competencia', methods=['POST'])
def guardar_precio_competencia():
    """Guarda un precio individual de competencia en la tabla precios_competencia."""
    try:
        data = request.get_json()
        med_id = data.get('medicamento_id')
        fab_id = data.get('fabricante_id')
        comp_id = data.get('competidor_id')
        precio_val = data.get('precio')
        url = data.get('url')  # Nuevo campo opcional

        if not med_id or not comp_id or precio_val is None:
            return jsonify({'ok': False, 'error': 'Faltan datos'}), 400

        conn = get_db_connection()
        cursor = conn.cursor()

        # Verificar si ya existe (ahora incluye fabricante_id)
        if fab_id:
            cursor.execute(
                'SELECT id FROM precios_competencia WHERE medicamento_id = ? AND fabricante_id = ? AND competidor_id = ?',
                (med_id, fab_id, comp_id)
            )
        else:
            cursor.execute(
                'SELECT id FROM precios_competencia WHERE medicamento_id = ? AND competidor_id = ?',
                (med_id, comp_id)
            )
        existe = cursor.fetchone()

        if existe:
            # Actualizar precio y URL si existe
            if url:
                cursor.execute(
                    'UPDATE precios_competencia SET precio = ?, url = ?, fecha_actualizacion = datetime("now") WHERE id = ?',
                    (precio_val, url, existe[0])
                )
            else:
                cursor.execute(
                    'UPDATE precios_competencia SET precio = ?, fecha_actualizacion = datetime("now") WHERE id = ?',
                    (precio_val, existe[0])
                )
        else:
            # Insertar nuevo registro
            if fab_id:
                if url:
                    cursor.execute(
                        'INSERT INTO precios_competencia (medicamento_id, fabricante_id, competidor_id, precio, url, fecha_actualizacion) VALUES (?, ?, ?, ?, ?, datetime("now"))',
                        (med_id, fab_id, comp_id, precio_val, url)
                    )
                else:
                    cursor.execute(
                        'INSERT INTO precios_competencia (medicamento_id, fabricante_id, competidor_id, precio, fecha_actualizacion) VALUES (?, ?, ?, ?, datetime("now"))',
                        (med_id, fab_id, comp_id, precio_val)
                    )
            else:
                if url:
                    cursor.execute(
                        'INSERT INTO precios_competencia (medicamento_id, competidor_id, precio, url, fecha_actualizacion) VALUES (?, ?, ?, ?, datetime("now"))',
                        (med_id, comp_id, precio_val, url)
                    )
                else:
                    cursor.execute(
                        'INSERT INTO precios_competencia (medicamento_id, competidor_id, precio, fecha_actualizacion) VALUES (?, ?, ?, datetime("now"))',
                        (med_id, comp_id, precio_val)
                    )

        conn.commit()
        conn.close()
        return jsonify({'ok': True})

    except Exception as e:
        print(f"Error al guardar precio competencia: {e}")
        return jsonify({'ok': False, 'error': str(e)}), 500





# --- RUTAS PARA LA PRUEBA DE CONSULTA DE SNTOMAS ---

@app.route('/api/sintomas/buscar', methods=['GET'])
def buscar_sintomas_api():
    """
    Busca sntomas en la base de datos que coincidan con un trmino parcial.
    Utiliza la lgica existente de carga de sntomas.
    Asume que get_db_connection est disponible globalmente en app.py.
    """
    # No importamos get_db_connection aqu, se asume disponible globalmente
    termino = request.args.get('q', '').strip().lower()
    if not termino:
        return jsonify({'ok': True, 'sintomas': []})

    try:
        # Usamos directamente get_db_connection asumiendo que est definida en app.py
        conn = get_db_connection()
        # Buscar coincidencias en el nombre o descripcion_lower, similar a otras partes
        # Usamos LOWER para bsqueda insensible a maysculas
        cursor = conn.execute("""
            SELECT id, nombre
            FROM sintomas
            WHERE LOWER(nombre) LIKE ? OR LOWER(descripcion_lower) LIKE ?
            ORDER BY nombre
            LIMIT 10
        """, (f'%{termino}%', f'%{termino}%'))
        resultados = cursor.fetchall()
        conn.close()

        # Convertir Row objects a diccionarios
        sintomas = [{'id': row['id'], 'nombre': row['nombre']} for row in resultados]
        return jsonify({'ok': True, 'sintomas': sintomas})
    except Exception as e:
        print(f"Error buscando sntomas: {e}")
        import traceback
        traceback.print_exc() # Agregar traceback para depurar
        return jsonify({'ok': False, 'error': str(e)}), 500



@app.route('/api/sintomas/crear', methods=['POST'])
def crear_sintoma_api():
    data = request.get_json()
    nombre = data.get('nombre', '').strip()

    if not nombre:
        return jsonify({'ok': False, 'error': 'El nombre del sntoma es obligatorio.'}), 400

    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        #  Usa solo LOWER(nombre) en la consulta, SIN aplicar .lower() en Python
        cursor.execute("SELECT id, nombre FROM sintomas WHERE LOWER(nombre) = LOWER(?)", (nombre,))
        existente = cursor.fetchone()

        if existente:
            conn.close()
            return jsonify({
                'ok': True,
                'mensaje': 'Sntoma ya existente.',
                'id': existente['id'],
                'nombre': existente['nombre']  # Usa el nombre tal como est en BD
            })

        #  Inserta el nombre original, y su versin normalizada
        nombre_lower = nombre.lower().strip()
        cursor.execute(
            "INSERT INTO sintomas (nombre, descripcion_lower) VALUES (?, ?)",
            (nombre, nombre_lower)
        )
        conn.commit()
        nuevo_id = cursor.lastrowid
        conn.close()

        print(f" Sntoma creado: {nombre} (ID: {nuevo_id})")
        return jsonify({
            'ok': True,
            'mensaje': 'Sntoma creado exitosamente.',
            'id': nuevo_id,
            'nombre': nombre
        })
    except Exception as e:
        print(f" Error creando sntoma: {e}")
        traceback.print_exc()  #  Esto imprime el traceback completo en la consola del servidor

        if 'conn' in locals():
            conn.rollback()
            conn.close()
        return jsonify({'ok': False, 'error': 'Error interno al crear el sntoma.'}), 500



@app.route('/api/diagnosticos/buscar_simple', methods=['POST'])
def buscar_diagnosticos_simple_api():
    """
    Versin corregida: Recibe IDs de sntomas seleccionados por el usuario.
    Muestra diagnsticos que requieran *todos* los sntomas (por ID) ingresados.
    Calcula y muestra los sntomas faltantes para completar cada diagnstico.
    """
    data = request.get_json()
    # Ahora esperamos recibir 'ids_sintomas_usuario' en lugar de 'nombres_sintomas'
    ids_sintomas_usuario_lista = data.get('ids_sintomas_usuario', []) # Cambiado el nombre del campo
    ids_sintomas_usuario = set(ids_sintomas_usuario_lista) # Convertir a conjunto para operaciones eficientes

    if not ids_sintomas_usuario:
        # Devolver lista vaca si no hay sntomas
        print(" No se proporcionaron IDs de sntomas.")
        return jsonify({'ok': True, 'diagnosticos': [], 'mensaje': 'No se proporcionaron sntomas.'})

    print(f" IDs de sntomas ingresados por el usuario (directamente del frontend): {ids_sintomas_usuario}")

    try:
        # --- PASO 1: Obtener TODOS los diagnsticos y sus sntomas requeridos desde la tabla intermedia ---
        conn = get_db_connection()
        query_all_diags = """
            SELECT ds.diagnostico_id, d.descripcion, ds.sintoma_id
            FROM diagnostico_sintoma ds
            JOIN diagnosticos d ON ds.diagnostico_id = d.id
            ORDER BY ds.diagnostico_id, ds.sintoma_id
        """
        cursor_all = conn.execute(query_all_diags)
        filas_diag_sintoma = cursor_all.fetchall()
        conn.close()

        # --- PASO 2: Agrupar sntomas por diagnstico ---
        diag_sintomas_map = {}
        for fila in filas_diag_sintoma:
            diag_id = fila['diagnostico_id']
            descripcion = fila['descripcion']
            sintoma_id = fila['sintoma_id']
            if diag_id not in diag_sintomas_map:
                diag_sintomas_map[diag_id] = {'descripcion': descripcion, 'sintomas_requeridos_ids': set()}
            diag_sintomas_map[diag_id]['sintomas_requeridos_ids'].add(sintoma_id)

        print(f" Diagnosticos cargados desde la tabla 'diagnostico_sintoma': {len(diag_sintomas_map)}")

        # --- PASO 3: Filtrar diagnsticos. ---
        # Un diagnstico es vlido si *todos* los IDs de sntomas que el usuario ingres
        # estn *dentro* de los sntomas que *requiere* el diagnstico.
        # Es decir, el conjunto de IDs de sntomas del usuario (ids_sintomas_usuario)
        # debe ser un SUBCONJUNTO del conjunto de sntomas requeridos por el diagnstico.
        diagnosticos_filtrados = []
        for diag_id, info in diag_sintomas_map.items():
            ids_requeridos_diag = info['sintomas_requeridos_ids']
            descripcion_diag = info['descripcion']

            # Verificamos si todos los IDs de sntomas del usuario estn en los requeridos por este diagnstico
            if ids_sintomas_usuario.issubset(ids_requeridos_diag):
                print(f"    Diag {diag_id} ('{descripcion_diag}') contiene todos los IDs de sintomas del usuario.")
                # Calculamos cuntos sntomas faltan para completar este diagnstico
                ids_faltantes_para_diag = ids_requeridos_diag - ids_sintomas_usuario
                cantidad_faltantes = len(ids_faltantes_para_diag)

                # Obtenemos los nombres de los sntomas faltantes para mostrarlos
                nombres_faltantes = []
                if ids_faltantes_para_diag:
                    placeholders_faltantes = ','.join(['?' for _ in ids_faltantes_para_diag])
                    query_nombres_falt = f"SELECT nombre, id FROM sintomas WHERE id IN ({placeholders_faltantes})"
                    conn = get_db_connection()
                    cursor_nombres = conn.execute(query_nombres_falt, list(ids_faltantes_para_diag))
                    for row_nombre in cursor_nombres.fetchall():
                        nombres_faltantes.append({'id': row_nombre['id'], 'nombre': row_nombre['nombre']})
                    conn.close()

                # Aadimos el diagnstico a la lista final
                diagnosticos_filtrados.append({
                    'id': diag_id,
                    'descripcion': descripcion_diag,
                    'sintomas_faltantes': nombres_faltantes,
                    'cantidad_faltantes': cantidad_faltantes
                })
            # else:
            #     print(f"    Diag {diag_id} ('{descripcion_diag}') NO contiene todos los IDs de sintomas del usuario.")

        # --- PASO 4: Ordenar los diagnstomas filtrados por la cantidad de sntomas faltantes (menor primero) ---
        diagnosticos_filtrados.sort(key=lambda x: x['cantidad_faltantes'])

        print(f" Diagnosticos finales (filtrados y ordenados): {len(diagnosticos_filtrados)} encontrados.")
        for d in diagnosticos_filtrados:
            print(f"  - {d['descripcion']} (ID: {d['id']}) - Faltan: {d['cantidad_faltantes']}")

        # Devolver la lista de diagnsticos filtrados
        return jsonify({'ok': True, 'diagnosticos': diagnosticos_filtrados})

    except Exception as e:
        print(f" Error en buscar_diagnosticos_simple_api: {e}")
        import traceback
        traceback.print_exc()
        # Devolver lista vaca en caso de error
        return jsonify({'ok': False, 'error': str(e), 'diagnosticos': []}), 500




@app.route('/api/medicamentos/por_sintomas_ids', methods=['POST'])
def obtener_medicamentos_por_sintomas_ids():
    """
    API para obtener medicamentos sugeridos basados en una lista de IDs de sntomas.
    Reutiliza parte de la lgica de seleccin de medicamentos de la funcin original
    'obtener_productos', pero adaptada para recibir directamente IDs de sntomas.
    Incorpora un algoritmo greedy para ordenar los medicamentos sugeridos de forma
    que se cubran los sntomas objetivo de manera eficiente y menos redundante posible
    antes de la lnea divisoria de cobertura completa.
    """
    from datetime import datetime
    print(f"\n{'='*70}")
    print(f" API /api/medicamentos/por_sintomas_ids - Iniciando")
    print(f"   Timestamp: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"{'='*70}\n")

    data = request.get_json()
    ids_sintomas_usuario = data.get('ids_sintomas_usuario', [])
    # Opcional: Parmetros de filtro
    precio_min = data.get('precio_min', '')
    precio_max = data.get('precio_max', '')

    if not ids_sintomas_usuario:
        print("    No se proporcionaron IDs de sntomas.")
        return jsonify({
            'ok': True,
            'productos': [],
            'total': 0,
            'mensaje': 'No se proporcionaron IDs de sntomas.'
        })

    print(f"    IDs de sntomas recibidos: {ids_sintomas_usuario}")

    try:
        conn = get_db_connection()

        # --- PARTE CLAVE: Adaptacin de la lgica de 'obtener_productos' ---
        # Simulamos las variables que 'obtener_productos' calcula a partir del texto de bsqueda
        # pero ahora las tenemos como entrada (ids_sintomas_usuario).
        sintomas_objetivo_ids = set(ids_sintomas_usuario) # Usamos un set para bsquedas rpidas
        # Obtener nombres de los sntomas para mensajes o referencias
        placeholders_ids = ','.join(['?' for _ in ids_sintomas_usuario])
        query_nombres_sint = f"""
            SELECT id, nombre FROM sintomas WHERE id IN ({placeholders_ids})
        """
        resultados_sint = conn.execute(query_nombres_sint, ids_sintomas_usuario).fetchall()
        sintomas_detectados = [row['nombre'] for row in resultados_sint]
        diagnosticos_posibles = {} # En esta ruta, no se generan diagnsticos posibles a partir de IDs de sntomas
                                   # (esto se hace en la otra parte de la consulta).
                                   # Se podra integrar si se pasan tambin IDs de diagnsticos detectados.

        print(f"    Sntomas detectados (nombres): {sintomas_detectados}")

        # QUERY DE PRODUCTOS (similar a la de 'obtener_productos', pero filtrando por IDs de sntomas)
        # Ahora incluimos los sntomas asociados al medicamento
        query = """
            SELECT DISTINCT
                p.id as precio_id,
                p.medicamento_id,
                p.fabricante_id,
                p.precio,
                p.imagen as imagen_precio,
                m.nombre as medicamento_nombre,
                m.presentacion,
                m.concentracion,
                m.imagen as imagen_medicamento,
                m.componente_activo_id,
                ca.nombre as componente_activo_nombre,
                f.nombre as fabricante_nombre,
                -- Nuevos campos para los sntomas del medicamento
                ms.sintoma_id as sintoma_asociado_id,
                s.nombre as sintoma_asociado_nombre
            FROM precios p
            INNER JOIN medicamentos m ON p.medicamento_id = m.id
            INNER JOIN fabricantes f ON p.fabricante_id = f.id
            LEFT JOIN medicamentos ca ON m.componente_activo_id = ca.id
            -- JOIN para obtener los sntomas asociados al medicamento
            LEFT JOIN medicamento_sintoma ms ON m.id = ms.medicamento_id
            LEFT JOIN sintomas s ON ms.sintoma_id = s.id
            WHERE m.activo = '1'
        """

        params = []

        # Filtro por IDs de sntomas
        if sintomas_objetivo_ids: # <-- Verificar que no est vaco antes de aplicar el filtro
             # Filtrar precios cuyos medicamentos estn asociados a los sntomas ingresados
             query += f"""
                 AND m.id IN (
                     SELECT DISTINCT medicamento_id
                     FROM medicamento_sintoma
                     WHERE sintoma_id IN ({placeholders_ids})
                 )
             """
             params.extend(ids_sintomas_usuario)

        if precio_min:
            try:
                query += " AND p.precio >= ?"
                params.append(float(precio_min))
            except ValueError:
                pass # Si no es un nmero vlido, ignorar el filtro

        if precio_max:
            try:
                query += " AND p.precio <= ?"
                params.append(float(precio_max))
            except ValueError:
                pass # Si no es un nmero vlido, ignorar el filtro

        query += " AND p.precio > 0"
        query += " ORDER BY m.nombre, f.nombre, p.id" # Orden inicial para agrupacin

        productos_raw = conn.execute(query, params).fetchall()

        # --- AGRUPACIN DE RESULTADOS POR PRECIO_ID ---
        # Agrupamos las filas crudas por precio_id, acumulando los sntomas asociados
        productos_agrupados = {}
        for row in productos_raw:
            precio_id = row['precio_id']
            if precio_id not in productos_agrupados:
                productos_agrupados[precio_id] = {
                    'precio': {
                        'precio_id': row['precio_id'],
                        'medicamento_id': row['medicamento_id'],
                        'fabricante_id': row['fabricante_id'],
                        'precio': row['precio'],
                        'imagen_precio': row['imagen_precio'],
                        'medicamento_nombre': row['medicamento_nombre'],
                        'presentacion': row['presentacion'],
                        'concentracion': row['concentracion'],
                        'imagen_medicamento': row['imagen_medicamento'],
                        'componente_activo_id': row['componente_activo_id'],
                        'componente_activo_nombre': row['componente_activo_nombre'],
                        'fabricante_nombre': row['fabricante_nombre']
                    },
                    'sintomas_asociados': [] # Lista para almacenar {'id': X, 'nombre': 'Y'}
                }
            # Agregar sntoma asociado si existe
            if row['sintoma_asociado_id']:
                productos_agrupados[precio_id]['sintomas_asociados'].append({
                    'id': row['sintoma_asociado_id'],
                    'nombre': row['sintoma_asociado_nombre']
                })

        # Extraer la lista de productos ya agrupados
        productos = [v['precio'] for v in productos_agrupados.values()]
        # Extraer la lista de sntomas asociados por precio_id
        sintomas_asociados_por_precio = {k: v['sintomas_asociados'] for k, v in productos_agrupados.items()}

        # --- CLCULO DE DETALLES INICIALES PARA EL ALGORITMO GREEDY ---
        productos_con_detalle = []
        for p in productos:
            precio_id = p['precio_id']
            sintomas_medicamento = sintomas_asociados_por_precio.get(precio_id, [])
            ids_sintomas_medicamento = {s['id'] for s in sintomas_medicamento}

            # Calcular coincidencias directas con los IDs objetivo
            ids_coincidentes = sintomas_objetivo_ids.intersection(ids_sintomas_medicamento)
            coincidencias = len(ids_coincidentes)

            # Calcular sntomas sobrantes (asociados al medicamento pero no objetivo)
            ids_sobrantes = ids_sintomas_medicamento - sintomas_objetivo_ids
            sintomas_sobrantes = [s for s in sintomas_medicamento if s['id'] in ids_sobrantes]

            # Calcular especificidad
            sintomas_totales_count = len(ids_sintomas_medicamento)
            especificidad_score = 1000 / sintomas_totales_count if sintomas_totales_count > 0 else 0

            productos_con_detalle.append({
                'precio_id': p['precio_id'],
                'medicamento_id': p['medicamento_id'],
                'fabricante_id': p['fabricante_id'],
                'nombre': p['medicamento_nombre'],
                'presentacion': p['presentacion'] or '',
                'concentracion': p['concentracion'] or '',
                'fabricante': p['fabricante_nombre'],
                'precio': p['precio'],
                'imagen': p['imagen_precio'] if p['imagen_precio'] else p['imagen_medicamento'],
                'componente_activo': p['componente_activo_nombre'] if p['componente_activo_nombre'] else None,
                'sintomas_totales': [s['nombre'] for s in sintomas_medicamento], # Lista de nombres
                'sintomas_ids_asociados': [s['id'] for s in sintomas_medicamento], # Lista de IDs
                'ids_coincidentes': ids_coincidentes, # Conjunto de IDs coincidentes
                'sintomas_coincidentes_nombres': [s['nombre'] for s in sintomas_medicamento if s['id'] in ids_coincidentes],
                'ids_sobrantes': ids_sobrantes, # Conjunto de IDs sobrantes
                'sintomas_sobrantes_nombres': [s['nombre'] for s in sintomas_medicamento if s['id'] in ids_sobrantes],
                'coincidencias': coincidencias,
                'sintomas_totales_count': sintomas_totales_count,
                'especificidad_score': especificidad_score
            })

        # --- ALGORITMO GREEDY PARA SELECCIN Y ORDENAMIENTO ---
        print(f"\n APLICANDO ALGORITMO GREEDY...")
        print(f"   Sntomas objetivo (IDs): {sintomas_objetivo_ids}")
        sintomas_cubiertos_ids = set()
        productos_ordenados_greedy = []
        productos_restantes = productos_con_detalle[:] # Copia de la lista

        while productos_restantes and len(sintomas_cubiertos_ids) < len(sintomas_objetivo_ids):
            mejor_producto = None
            mejor_valor = -float('inf') # Usamos -inf para asegurar que cualquier valor positivo sea mejor
            indice_mejor = -1

            print(f"   Iteracin: Sntomas cubiertos: {sintomas_cubiertos_ids}, Objetivo: {sintomas_objetivo_ids}")

            for i, prod in enumerate(productos_restantes):
                # Verificar que 'ids_coincidentes' exista y sea un conjunto
                ids_coincidentes_prod = prod.get('ids_coincidentes', set())
                if not isinstance(ids_coincidentes_prod, set):
                     print(f"      Producto {prod['nombre']} tiene 'ids_coincidentes' invlido: {ids_coincidentes_prod}, tipo: {type(ids_coincidentes_prod)}. Saltando.")
                     continue

                # Calcular cuntos *nuevos* sntomas objetivo cubrira este medicamento
                nuevos_sintomas = ids_coincidentes_prod - sintomas_cubiertos_ids
                valor_nuevos = len(nuevos_sintomas)

                if valor_nuevos == 0:
                    # Si no aporta nuevos sntomas objetivo, su valor es muy bajo
                    valor_actual = -1000
                else:
                    # Valorar positivamente la cobertura de nuevos sntomas objetivo
                    # y negativamente los sntomas sobrantes (generalidad)
                    # y ligeramente positivamente la especificidad
                    ids_sobrantes_prod = prod.get('ids_sobrantes', set())
                    if not isinstance(ids_sobrantes_prod, set):
                         print(f"      Producto {prod['nombre']} tiene 'ids_sobrantes' invlido: {ids_sobrantes_prod}, tipo: {type(ids_sobrantes_prod)}. Saltando.")
                         continue
                    penalizacion_sobrantes = len(ids_sobrantes_prod)
                    valor_actual = (valor_nuevos * 100) - (penalizacion_sobrantes * 10) + (prod['especificidad_score'] * 0.1)

                print(f"     - Producto {prod['nombre'][:20]}: Nuevos: {len(nuevos_sintomas)}, Sobrantes: {len(ids_sobrantes_prod)}, Valor: {valor_actual:.2f}")

                if valor_actual > mejor_valor:
                    mejor_valor = valor_actual
                    mejor_producto = prod
                    indice_mejor = i

            if mejor_producto is None:
                # Si no hay ningn producto que aporte nuevos sntomas objetivo, salir del bucle
                print("  No se encontr un medicamento que aportara nuevos sntomas objetivo. Deteniendo seleccin greedy.")
                break

            # Marcar los sntomas objetivo de este producto como cubiertos
            ids_coincidentes_final = mejor_producto.get('ids_coincidentes', set())
            if isinstance(ids_coincidentes_final, set):
                 sintomas_cubiertos_ids.update(ids_coincidentes_final)
            else:
                 print(f"      Producto {mejor_producto['nombre']} tiene 'ids_coincidentes' invlido al intentar actualizar cubiertos: {ids_coincidentes_final}.")
                 # Opcional: Salir o manejar el error
                 break

            productos_ordenados_greedy.append(mejor_producto)
            productos_restantes.pop(indice_mejor) # Quitar el producto seleccionado
            print(f"        Seleccionado: {mejor_producto['nombre'][:30]:30} (Aporta: {len(ids_coincidentes_final)}, Nuevos: {len(ids_coincidentes_final - (sintomas_cubiertos_ids - ids_coincidentes_final))})")


        # Agregar los productos restantes (que no aportaron nuevos sntomas objetivo o se aaden despus)
        productos_ordenados_greedy.extend(productos_restantes)
        print(f"    Seleccin Greedy completada. Total productos: {len(productos_ordenados_greedy)}")

        # --- CLCULO DE SCORE FINAL (opcional, basado en el orden greedy) ---
        for i, p in enumerate(productos_ordenados_greedy):
             # El score puede reflejar el orden greedy o mantener la lgica anterior
             # Por ejemplo, basado en coincidencias y especificidad, o simplemente el orden
             # Vamos con una mtrica que combine la informacin til
             ids_sobrantes_p = p.get('ids_sobrantes', set())
             penalizacion = len(ids_sobrantes_p) if isinstance(ids_sobrantes_p, set) else 0
             p['score'] = (p['coincidencias'] * 100) - (penalizacion * 10) + p['especificidad_score']
             # Opcional: Aadir un campo para indicar si es parte de la "receta eficiente" o no
             # p['es_parte_de_receta_eficiente'] = i < len(productos_seleccionados_greedy) # Calculable si se desea

        productos_con_score = productos_ordenados_greedy

        # --- CONVERSIN DE 'sets' A 'lists' PARA SERIALIZACIN JSON ---
        # Flask no puede serializar objetos 'set' a JSON.
        # Convertimos los sets relevantes a listas antes de devolverlos.
        for producto in productos_con_score:
            if isinstance(producto.get('ids_coincidentes'), set):
                producto['ids_coincidentes'] = list(producto['ids_coincidentes'])
            if isinstance(producto.get('ids_sobrantes'), set):
                producto['ids_sobrantes'] = list(producto['ids_sobrantes'])
            # Asegurarse de que 'sintomas_ids_asociados' tambin sea una lista si se calcula como set en otro lado
            # (Aunque en este cdigo se construye como lista, es bueno verificar si se modifica ms adelante)
            # if isinstance(producto.get('sintomas_ids_asociados'), set):
            #     producto['sintomas_ids_asociados'] = list(producto['sintomas_ids_asociados'])


        conn.close()

        print(f"    Medicamentos encontrados y ordenados (Greedy): {len(productos_con_score)}")
        return jsonify({
            'ok': True,
            'productos': productos_con_score,
            'total': len(productos_con_score),
            'sintomas_ingresados_ids': list(sintomas_objetivo_ids), # Convertir set a list
            'sintomas_ingresados_nombres': sintomas_detectados
        })

    except Exception as e:
        print(f" Error en obtener_medicamentos_por_sintomas_ids: {e}")
        import traceback
        traceback.print_exc() # Imprimir el traceback completo para depuracin
        return jsonify({'ok': False, 'error': str(e), 'productos': [], 'total': 0}), 500





@app.route('/api/medicamentos/economicos', methods=['POST'])
def obtener_medicamentos_economicos():
    """
    API para obtener medicamentos sugeridos basados en una lista de IDs de sntomas,
    priorizando la minimizacin del costo total para cubrir todos los sntomas objetivo.
    """
    from datetime import datetime
    print(f"\n{'='*70}")
    print(f" API /api/medicamentos/economicos - Iniciando")
    print(f"   Timestamp: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"{'='*70}\n")

    data = request.get_json()
    ids_sintomas_usuario = data.get('ids_sintomas_usuario', [])
    # Opcional: Parmetros de filtro
    precio_min = data.get('precio_min', '')
    precio_max = data.get('precio_max', '')

    if not ids_sintomas_usuario:
        print("    No se proporcionaron IDs de sntomas.")
        return jsonify({
            'ok': True,
            'productos': [],
            'total': 0,
            'mensaje': 'No se proporcionaron IDs de sntomas.'
        })

    print(f"    IDs de sntomas recibidos: {ids_sintomas_usuario}")

    try:
        conn = get_db_connection()

        # --- PARTE CLAVE: Adaptacin de la lgica de 'obtener_productos' ---
        # Simulamos las variables que 'obtener_productos' calcula a partir del texto de bsqueda
        # pero ahora las tenemos como entrada (ids_sintomas_usuario).
        sintomas_objetivo_ids = set(ids_sintomas_usuario) # Usamos un set para bsquedas rpidas
        # Obtener nombres de los sntomas para mensajes o referencias
        placeholders_ids = ','.join(['?' for _ in ids_sintomas_usuario])
        query_nombres_sint = f"""
            SELECT id, nombre FROM sintomas WHERE id IN ({placeholders_ids})
        """
        resultados_sint = conn.execute(query_nombres_sint, ids_sintomas_usuario).fetchall()
        sintomas_detectados = [row['nombre'] for row in resultados_sint]
        diagnosticos_posibles = {} # En esta ruta, no se generan diagnsticos posibles a partir de IDs de sntomas
                                   # (esto se hace en la otra parte de la consulta).
                                   # Se podra integrar si se pasan tambin IDs de diagnsticos detectados.

        print(f"    Sntomas detectados (nombres): {sintomas_detectados}")

        # QUERY DE PRODUCTOS (similar a la de 'obtener_productos', pero filtrando por IDs de sntomas)
        # Ahora incluimos los sntomas asociados al medicamento
        query = """
            SELECT DISTINCT
                p.id as precio_id,
                p.medicamento_id,
                p.fabricante_id,
                p.precio,
                p.imagen as imagen_precio,
                m.nombre as medicamento_nombre,
                m.presentacion,
                m.concentracion,
                m.imagen as imagen_medicamento,
                m.componente_activo_id,
                ca.nombre as componente_activo_nombre,
                f.nombre as fabricante_nombre,
                -- Nuevos campos para los sntomas del medicamento
                ms.sintoma_id as sintoma_asociado_id,
                s.nombre as sintoma_asociado_nombre
            FROM precios p
            INNER JOIN medicamentos m ON p.medicamento_id = m.id
            INNER JOIN fabricantes f ON p.fabricante_id = f.id
            LEFT JOIN medicamentos ca ON m.componente_activo_id = ca.id
            -- JOIN para obtener los sntomas asociados al medicamento
            LEFT JOIN medicamento_sintoma ms ON m.id = ms.medicamento_id
            LEFT JOIN sintomas s ON ms.sintoma_id = s.id
            WHERE m.activo = '1'
        """

        params = []

        # Filtro por IDs de sntomas
        if sintomas_objetivo_ids:
             # Filtrar precios cuyos medicamentos estn asociados a los sntomas ingresados
             query += f"""
                 AND m.id IN (
                     SELECT DISTINCT medicamento_id
                     FROM medicamento_sintoma
                     WHERE sintoma_id IN ({placeholders_ids})
                 )
             """
             params.extend(ids_sintomas_usuario)

        if precio_min:
            try:
                query += " AND p.precio >= ?"
                params.append(float(precio_min))
            except ValueError:
                pass # Si no es un nmero vlido, ignorar el filtro

        if precio_max:
            try:
                query += " AND p.precio <= ?"
                params.append(float(precio_max))
            except ValueError:
                pass # Si no es un nmero vlido, ignorar el filtro

        query += " AND p.precio > 0"
        query += " ORDER BY m.nombre, f.nombre, p.id" # Orden inicial para agrupacin

        productos_raw = conn.execute(query, params).fetchall()

        # --- AGRUPACIN DE RESULTADOS POR PRECIO_ID ---
        # Agrupamos las filas crudas por precio_id, acumulando los sntomas asociados
        productos_agrupados = {}
        for row in productos_raw:
            precio_id = row['precio_id']
            if precio_id not in productos_agrupados:
                productos_agrupados[precio_id] = {
                    'precio': {
                        'precio_id': row['precio_id'],
                        'medicamento_id': row['medicamento_id'],
                        'fabricante_id': row['fabricante_id'],
                        'precio': row['precio'],
                        'imagen_precio': row['imagen_precio'],
                        'medicamento_nombre': row['medicamento_nombre'],
                        'presentacion': row['presentacion'],
                        'concentracion': row['concentracion'],
                        'imagen_medicamento': row['imagen_medicamento'],
                        'componente_activo_id': row['componente_activo_id'],
                        'componente_activo_nombre': row['componente_activo_nombre'],
                        'fabricante_nombre': row['fabricante_nombre']
                    },
                    'sintomas_asociados': [] # Lista para almacenar {'id': X, 'nombre': 'Y'}
                }
            # Agregar sntoma asociado si existe
            if row['sintoma_asociado_id']:
                productos_agrupados[precio_id]['sintomas_asociados'].append({
                    'id': row['sintoma_asociado_id'],
                    'nombre': row['sintoma_asociado_nombre']
                })

        # Extraer la lista de productos ya agrupados
        productos = [v['precio'] for v in productos_agrupados.values()]
        # Extraer la lista de sntomas asociados por precio_id
        sintomas_asociados_por_precio = {k: v['sintomas_asociados'] for k, v in productos_agrupados.items()}

        # --- CLCULO DE DETALLES INICIALES ---
        productos_con_detalle = []
        for p in productos:
            precio_id = p['precio_id']
            sintomas_medicamento = sintomas_asociados_por_precio.get(precio_id, [])
            ids_sintomas_medicamento = {s['id'] for s in sintomas_medicamento}

            # Calcular coincidencias directas con los IDs objetivo
            ids_coincidentes = sintomas_objetivo_ids.intersection(ids_sintomas_medicamento)
            coincidencias = len(ids_coincidentes)

            # Calcular sntomas sobrantes (asociados al medicamento pero no ingresados por el usuario)
            ids_sobrantes = ids_sintomas_medicamento - sintomas_objetivo_ids
            sintomas_sobrantes = [s for s in sintomas_medicamento if s['id'] in ids_sobrantes]

            # Calcular especificidad
            sintomas_totales_count = len(ids_sintomas_medicamento)
            especificidad_score = 1000 / sintomas_totales_count if sintomas_totales_count > 0 else 0

            productos_con_detalle.append({
                'precio_id': p['precio_id'],
                'medicamento_id': p['medicamento_id'],
                'fabricante_id': p['fabricante_id'],
                'nombre': p['medicamento_nombre'],
                'presentacion': p['presentacion'] or '',
                'concentracion': p['concentracion'] or '',
                'fabricante': p['fabricante_nombre'],
                'precio': p['precio'],
                'imagen': p['imagen_precio'] if p['imagen_precio'] else p['imagen_medicamento'],
                'componente_activo': p['componente_activo_nombre'] if p['componente_activo_nombre'] else None,
                'sintomas_totales': [s['nombre'] for s in sintomas_medicamento], # Lista de nombres
                'sintomas_ids_asociados': [s['id'] for s in sintomas_medicamento], # Lista de IDs
                'ids_coincidentes': list(ids_coincidentes), # IDs que coinciden (como lista para JSON)
                'sintomas_coincidentes_nombres': [s['nombre'] for s in sintomas_medicamento if s['id'] in ids_coincidentes], # Nombres que coinciden
                'ids_sobrantes': list(ids_sobrantes), # IDs sobrantes (como lista para JSON)
                'sintomas_sobrantes_nombres': [s['nombre'] for s in sintomas_medicamento if s['id'] in ids_sobrantes], # Nombres sobrantes
                'coincidencias': coincidencias,
                'sintomas_totales_count': sintomas_totales_count,
                'especificidad_score': especificidad_score
            })

        # --- ALGORITMO DE SELECCIN ECONMICA ---
        # Objetivo: Encontrar una combinacin de medicamentos que cubra todos los sintomas_objetivo_ids
        # al costo total ms bajo posible.
        # Usaremos un enfoque greedy basado en el precio y la eficiencia de cobertura.

        print(f"\n APLICANDO ALGORITMO ECONMICO...")
        print(f"   Sntomas objetivo (IDs): {sintomas_objetivo_ids}")

        # Inicializar conjunto de sntomas cubiertos
        sintomas_cubiertos_ids = set()
        productos_seleccionados = []
        productos_restantes = productos_con_detalle[:]

        # Iterar hasta cubrir todos los sntomas objetivo o agotar opciones
        while productos_restantes and len(sintomas_cubiertos_ids) < len(sintomas_objetivo_ids):
            mejor_producto = None
            mejor_valor = float('inf') # Minimizar el valor (precio ponderado)
            indice_mejor = -1

            for i, prod in enumerate(productos_restantes):
                # Calcular cuntos *nuevos* sntomas objetivo cubrira este medicamento
                ids_nuevos = set(prod['ids_coincidentes']) - sintomas_cubiertos_ids
                nuevos_sintomas = len(ids_nuevos)

                if nuevos_sintomas == 0:
                    # Si no aporta nuevos sntomas objetivo, su valor es muy alto (no lo queremos)
                    valor_actual = float('inf')
                else:
                    # Valorar el medicamento por su precio, ponderado por cuntos nuevos sntomas aporta
                    # (menos sntomas nuevos = ms caro por sntoma nuevo)
                    # Tambien se puede ponderar por especificidad si se desea
                    valor_actual = prod['precio'] / nuevos_sintomas # Precio por nuevo sntoma cubierto

                if valor_actual < mejor_valor:
                    mejor_valor = valor_actual
                    mejor_producto = prod
                    indice_mejor = i

            if mejor_producto is None:
                # Si no hay ningn producto que aporte nuevos sntomas objetivo, salir del bucle
                print("  No se encontr un medicamento que aportara nuevos sntomas objetivo. Deteniendo seleccin econmica.")
                break

            # Marcar los sntomas de este producto como cubiertos
            for sid in mejor_producto['ids_coincidentes']:
                sintomas_cubiertos_ids.add(sid)

            # Aadir el producto a la lista de seleccionados
            productos_seleccionados.append(mejor_producto)
            # Quitar el producto de la lista de restantes
            productos_restantes.pop(indice_mejor)
            print(f"        Seleccionado: {mejor_producto['nombre'][:30]:30} (Precio: ${mejor_producto['precio']}, Nuevos Sntomas: {len(set(mejor_producto['ids_coincidentes']) - (sintomas_cubiertos_ids - set(mejor_producto['ids_coincidentes'])))})")

        # Agregar los productos restantes si se desea mostrarlos tambin (aunque no sean parte de la "receta econmica")
        # productos_seleccionados.extend(productos_restantes) # Descomentar si se quiere mostrar todo

        productos_economicos = productos_seleccionados

        # --- CLCULO DE SCORE FINAL (opcional, basado en el orden econmico) ---
        for i, p in enumerate(productos_economicos):
             # El score puede reflejar el orden econmico o mantener la lgica anterior
             # Por ejemplo, basado en coincidencias y especificidad, o simplemente el orden
             # Vamos con una mtrica que combine la informacin til
             ids_sobrantes_p = set(p.get('ids_sobrantes', []))
             penalizacion = len(ids_sobrantes_p)
             # El score aqu podra ser inversamente proporcional al precio y proporcional a la cobertura
             # Por simplicidad, lo dejamos como el precio original o basado en el orden econmico
             p['score'] = 1 / (p['precio'] + 1) # Ejemplo: score basado en inverso del precio
             # Opcional: Aadir un campo para indicar si es parte de la "receta eficiente" o no
             # p['es_parte_de_receta_economica'] = i < len(productos_seleccionados) # Calculable si se desea

        conn.close()

        print(f"    Medicamentos encontrados y ordenados (Econmico): {len(productos_economicos)}")
        return jsonify({
            'ok': True,
            'productos': productos_economicos,
            'total': len(productos_economicos),
            'sintomas_ingresados_ids': list(sintomas_objetivo_ids),
            'sintomas_ingresados_nombres': sintomas_detectados
        })

    except Exception as e:
        print(f" Error en obtener_medicamentos_economicos: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'ok': False, 'error': str(e), 'productos': [], 'total': 0}), 500

@app.route('/prueba_consulta')
def prueba_consulta():
    return render_template('prueba_sintomas.html')

# --- FIN RUTAS PARA PRUEBA ---


# ============================================
# RUTAS PARA COMPRAS A PROVEEDORES
# Agregar estas rutas al final de 1_medicamentos.py
# (antes del if __name__ == '__main__':)
# ============================================

@app.route('/api/compras/obtener_lista')
@admin_required
def obtener_lista_compras():
    """
    Obtiene la lista de productos pendientes agrupados por droguera
    """
    try:
        print(" DEBUG: Iniciando obtener_lista_compras")
        conn = get_db_connection()

        # 1. Obtener todos los productos pendientes con sus cantidades AGRUPADAS
        query_productos = """
            SELECT
                m.id as medicamento_id,
                m.nombre as medicamento,
                e.fabricante_id,
                f.nombre as fabricante,
                SUM(e.cantidad) as cantidad_total,
                STRING_AGG(CAST(e.id AS TEXT), ',') as existencias_ids
            FROM existencias e
            INNER JOIN pedidos p ON e.pedido_id = p.id
            INNER JOIN medicamentos m ON e.medicamento_id = m.id
            LEFT JOIN fabricantes f ON e.fabricante_id = f.id
            WHERE
                e.tipo_movimiento = 'salida'
                AND (e.estado = 'pendiente' OR e.estado IS NULL)
                AND p.estado = 'pendiente'
            GROUP BY m.id, m.nombre, e.fabricante_id, f.nombre
        """

        print(" DEBUG: Ejecutando query productos pendientes")
        productos_pendientes = conn.execute(query_productos).fetchall()
        print(f" DEBUG: Encontrados {len(productos_pendientes)} productos pendientes")

        # 2. Para cada producto, encontrar el MEJOR precio (ms bajo) de todos los proveedores
        proveedores = {}
        sin_precios = []

        for producto in productos_pendientes:
            medicamento_id = producto[0]  # medicamento_id
            fabricante_id = producto[2]   # fabricante_id

            # Buscar el mejor precio para este producto (solo cotizaciones activas)
            query_mejor_precio = """
                SELECT
                    t.id as drogueria_id,
                    t.nombre as drogueria,
                    pc.precio as precio_unitario,
                    pc.url as url
                FROM precios_competencia pc
                INNER JOIN terceros t ON pc.competidor_id = t.id
                WHERE pc.medicamento_id = ? AND pc.fabricante_id = ?
                  AND pc.activo = TRUE
                  AND (pc.inactivo_hasta IS NULL OR pc.inactivo_hasta < CURRENT_TIMESTAMP)
                ORDER BY pc.precio ASC
                LIMIT 1
            """

            mejor_proveedor = conn.execute(query_mejor_precio, (medicamento_id, fabricante_id)).fetchone()

            if mejor_proveedor:
                # Hay precio de competencia - asignar al proveedor con mejor precio
                drogueria_id = mejor_proveedor[0]    # drogueria_id
                precio_unitario = mejor_proveedor[2] # precio_unitario
                url = mejor_proveedor[3]             # url
                cantidad_total = producto[4]         # cantidad_total
                subtotal = cantidad_total * precio_unitario

                # Crear proveedor si no existe
                if drogueria_id not in proveedores:
                    proveedores[drogueria_id] = {
                        'id': drogueria_id,
                        'nombre': mejor_proveedor[1],  # drogueria nombre
                        'total': 0,
                        'productos': []
                    }

                # Agregar producto al proveedor
                nombre_completo = producto[1]  # medicamento
                if producto[3]:                # fabricante
                    nombre_completo += f" - {producto[3]}"

                proveedores[drogueria_id]['productos'].append({
                    'existencias_ids': producto[5],  # existencias_ids - IDs separados por comas
                    'medicamento_id': medicamento_id,
                    'medicamento': nombre_completo,
                    'url': url,  # URL de la cotizacin
                    'cantidad': cantidad_total,
                    'precio_unitario': precio_unitario,
                    'subtotal': subtotal,
                    'fabricante_id': fabricante_id
                })

                proveedores[drogueria_id]['total'] += subtotal
            else:
                # No hay precio de competencia - agregar a lista de sin precios
                nombre_completo = producto[1]  # medicamento
                if producto[3]:                # fabricante
                    nombre_completo += f" - {producto[3]}"

                sin_precios.append({
                    'medicamento_id': medicamento_id,
                    'medicamento': nombre_completo,
                    'cantidad': producto[4],  # cantidad_total
                    'existencias_ids': producto[5]  # existencias_ids
                })

        conn.close()

        print(f" DEBUG: Proveedores: {len(proveedores)}, Sin precios: {len(sin_precios)}")
        return jsonify({
            'success': True,
            'proveedores': list(proveedores.values()),
            'sin_precios': sin_precios
        })

    except Exception as e:
        print(f" ERROR en obtener_lista_compras: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/compras/gestionar_cotizacion', methods=['POST'])
@admin_required
def gestionar_cotizacion():
    """
    Gestiona la cotizacin cuando un proveedor no tiene el producto
    Opciones: inactivar temporalmente, eliminar, o buscar siguiente proveedor
    """
    try:
        data = request.json
        medicamento_id = data.get('medicamento_id')
        fabricante_id = data.get('fabricante_id')
        proveedor_actual_id = data.get('proveedor_id')
        accion = data.get('accion')  # 'temporal', 'eliminar', 'ninguna'
        dias_inactivo = data.get('dias_inactivo', 0)  # 1, 7, 30

        conn = get_db_connection()

        # Aplicar accin segn lo elegido
        if accion == 'temporal':
            # Inactivar temporalmente
            conn.execute("""
                UPDATE precios_competencia
                SET inactivo_hasta = CURRENT_TIMESTAMP + INTERVAL '%s days'
                WHERE medicamento_id = ? AND fabricante_id = ? AND competidor_id = ?
            """ % dias_inactivo, [medicamento_id, fabricante_id, proveedor_actual_id])

        elif accion == 'eliminar':
            # Eliminar cotizacin permanentemente
            conn.execute("""
                DELETE FROM precios_competencia
                WHERE medicamento_id = ? AND fabricante_id = ? AND competidor_id = ?
            """, [medicamento_id, fabricante_id, proveedor_actual_id])

        # Buscar siguiente mejor proveedor (solo si la accin fue temporal o eliminar)
        siguiente_proveedor = None
        if accion in ['temporal', 'eliminar']:
            query_siguiente = """
                SELECT
                    t.id as drogueria_id,
                    t.nombre as drogueria,
                    pc.precio as precio_unitario
                FROM precios_competencia pc
                INNER JOIN terceros t ON pc.competidor_id = t.id
                WHERE pc.medicamento_id = ? AND pc.fabricante_id = ?
                  AND pc.activo = TRUE
                  AND (pc.inactivo_hasta IS NULL OR pc.inactivo_hasta < CURRENT_TIMESTAMP)
                  AND pc.competidor_id != ?
                ORDER BY pc.precio ASC
                LIMIT 1
            """
            resultado = conn.execute(query_siguiente, [medicamento_id, fabricante_id, proveedor_actual_id]).fetchone()

            if resultado:
                siguiente_proveedor = {
                    'id': resultado[0],
                    'nombre': resultado[1],
                    'precio': resultado[2]
                }

        conn.commit()
        conn.close()

        return jsonify({
            'success': True,
            'siguiente_proveedor': siguiente_proveedor,
            'mensaje': 'Cotizacin actualizada'
        })

    except Exception as e:
        print(f" ERROR en gestionar_cotizacion: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/compras/registrar', methods=['POST'])
@admin_required
def registrar_compra():
    """
    Registra la compra de productos a un proveedor
    Formato de productos: array de objetos con {existencia_id, precio_compra}
    """
    try:
        data = request.json
        drogueria_id = data.get('drogueria_id')
        numero_documento = data.get('numero_documento')
        productos = data.get('productos', [])

        if not drogueria_id or not numero_documento or not productos:
            return jsonify({'success': False, 'message': 'Datos incompletos'}), 400

        conn = get_db_connection()

        # Crear un diccionario de precios por existencia_id
        precios_map = {}
        existencias_ids = []

        for prod in productos:
            if isinstance(prod, dict):
                # Nuevo formato: {existencia_id, precio_compra}
                existencias_ids.append(prod['existencia_id'])
                precios_map[prod['existencia_id']] = prod.get('precio_compra', 0)
            else:
                # Formato legacy: solo IDs
                existencias_ids.append(prod)
                precios_map[prod] = 0

        # 1. Obtener informacin de los productos a comprar
        placeholders = ','.join(['%s'] * len(existencias_ids))
        query = f"""
            SELECT
                e.id,
                e.medicamento_id,
                e.fabricante_id,
                e.cantidad,
                e.pedido_id
            FROM existencias e
            WHERE e.id IN ({placeholders})
        """

        existencias = conn.execute(query, existencias_ids).fetchall()

        # 2. Crear ENTRADAS en inventario (compras al proveedor)
        for item in existencias:
            precio_compra = precios_map.get(item[0], 0)  # item[0] es id

            # Obtener siguiente ID para existencias
            cursor_seq = conn.execute("SELECT COALESCE(MAX(id), 0) + 1 FROM existencias")
            next_existencia_id = cursor_seq.fetchone()[0]

            # Insertar entrada con costo_unitario
            conn.execute("""
                INSERT INTO existencias
                (id, medicamento_id, fabricante_id, tipo_movimiento, cantidad, fecha, id_tercero, numero_documento, pedido_id, costo_unitario)
                VALUES (?, ?, ?, 'entrada', ?, CURRENT_TIMESTAMP, ?, ?, NULL, ?)
            """, [
                next_existencia_id,
                item[1],  # medicamento_id
                item[2],  # fabricante_id
                item[3],  # cantidad
                drogueria_id,
                numero_documento,
                precio_compra
            ])

            # Actualizar stock y costo_unitario en tabla precios con promedio ponderado
            if precio_compra > 0:
                # Obtener stock y costo actual
                precio_actual = conn.execute("""
                    SELECT stock_fabricante, costo_unitario
                    FROM precios
                    WHERE medicamento_id = ? AND fabricante_id = ?
                """, [item[1], item[2]]).fetchone()

                if precio_actual:
                    stock_actual = precio_actual[0] or 0
                    costo_actual = precio_actual[1] or 0

                    # Calcular promedio ponderado
                    costo_total_anterior = stock_actual * costo_actual
                    costo_entrada = item[3] * precio_compra
                    nuevo_costo_total = costo_total_anterior + costo_entrada
                    nueva_cantidad = stock_actual + item[3]
                    nuevo_costo_unitario = nuevo_costo_total / nueva_cantidad if nueva_cantidad > 0 else precio_compra

                    # Actualizar stock y costo en precios
                    conn.execute("""
                        UPDATE precios
                        SET stock_fabricante = ?, costo_unitario = ?
                        WHERE medicamento_id = ? AND fabricante_id = ?
                    """, [nueva_cantidad, nuevo_costo_unitario, item[1], item[2]])
                else:
                    # Si no existe en precios, insertar
                    conn.execute("""
                        INSERT INTO precios (medicamento_id, fabricante_id, stock_fabricante, costo_unitario, precio)
                        VALUES (?, ?, ?, ?, 0)
                    """, [item[1], item[2], item[3], precio_compra])

                # Actualizar precio en precios_competencia
                existe = conn.execute("""
                    SELECT id FROM precios_competencia
                    WHERE medicamento_id = ? AND fabricante_id = ? AND competidor_id = ?
                """, [item[1], item[2], drogueria_id]).fetchone()

                if existe:
                    conn.execute("""
                        UPDATE precios_competencia
                        SET precio = ?, fecha_actualizacion = CURRENT_TIMESTAMP
                        WHERE medicamento_id = ? AND fabricante_id = ? AND competidor_id = ?
                    """, [precio_compra, item[1], item[2], drogueria_id])
                else:
                    conn.execute("""
                        INSERT INTO precios_competencia
                        (medicamento_id, fabricante_id, competidor_id, precio, fecha_actualizacion)
                        VALUES (?, ?, ?, ?, CURRENT_TIMESTAMP)
                    """, [item[1], item[2], drogueria_id, precio_compra])

        # 3. Actualizar SALIDAS a estado 'comprado'
        placeholders_update = ','.join(['%s'] * len(existencias_ids))
        query_actualizar = f"""
            UPDATE existencias
            SET estado = 'comprado'
            WHERE id IN ({placeholders_update})
        """
        conn.execute(query_actualizar, existencias_ids)

        # 4. Verificar y actualizar estados de pedidos
        pedidos_afectados = list(set([item[4] for item in existencias]))  # item[4] es pedido_id

        for pedido_id in pedidos_afectados:
            actualizar_estado_pedido(conn, pedido_id)

        conn.commit()
        conn.close()

        return jsonify({
            'success': True,
            'message': 'Compra registrada exitosamente',
            'pedidos_actualizados': pedidos_afectados
        })

    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


def actualizar_estado_pedido(conn, pedido_id):
    """
    Actualiza el estado de un pedido segn el estado de sus productos
    """
    # Verificar si TODOS los productos del pedido estn comprados
    resultado = conn.execute("""
        SELECT COUNT(*) as total,
               SUM(CASE WHEN estado = 'comprado' THEN 1 ELSE 0 END) as comprados
        FROM existencias
        WHERE pedido_id = ? AND tipo_movimiento = 'salida'
    """, [pedido_id]).fetchone()

    if resultado and resultado[0] > 0 and resultado[0] == resultado[1]:  # total y comprados
        # Todos los productos estn comprados -> Cambiar pedido a "en_camino"
        conn.execute("""
            UPDATE pedidos
            SET estado = 'en_camino'
            WHERE id = ?
        """, [pedido_id])

# --- FIN RUTAS PARA COMPRAS A PROVEEDORES ---


@app.route('/api/run-migration-now')
@admin_required
def run_migration_endpoint():
    """Endpoint temporal para ejecutar migraciones manualmente"""
    try:
        conn = get_db_connection()
        mensajes = []
        mensajes.append("Ejecutando migraciones...")

        # Migracin 1
        conn.execute("ALTER TABLE existencias ADD COLUMN IF NOT EXISTS costo_unitario DECIMAL(10,2) DEFAULT 0")
        mensajes.append(" existencias.costo_unitario")

        # Migracin 2
        conn.execute("ALTER TABLE precios ADD COLUMN IF NOT EXISTS costo_unitario DECIMAL(10,2) DEFAULT 0")
        mensajes.append(" precios.costo_unitario")

        # Migracin 3
        conn.execute("ALTER TABLE precios_competencia ADD COLUMN IF NOT EXISTS activo BOOLEAN DEFAULT TRUE")
        mensajes.append(" precios_competencia.activo")

        # Migracin 4
        conn.execute("ALTER TABLE precios_competencia ADD COLUMN IF NOT EXISTS inactivo_hasta TIMESTAMP")
        mensajes.append(" precios_competencia.inactivo_hasta")

        # Migracin 5: Arreglar secuencia del id en precios_competencia
        try:
            # Crear secuencia si no existe
            conn.execute("""
                CREATE SEQUENCE IF NOT EXISTS precios_competencia_id_seq
                OWNED BY precios_competencia.id
            """)
            mensajes.append(" Secuencia precios_competencia_id_seq creada")

            # Obtener el valor mximo actual del id
            max_id_row = conn.execute("SELECT COALESCE(MAX(id), 0) as max_id FROM precios_competencia").fetchone()
            max_id = max_id_row[0] if max_id_row else 0

            # Setear el valor de la secuencia
            conn.execute(f"SELECT setval('precios_competencia_id_seq', {max_id + 1})")
            mensajes.append(f" Secuencia inicializada en {max_id + 1}")

            # Asignar la secuencia como default al campo id
            conn.execute("""
                ALTER TABLE precios_competencia
                ALTER COLUMN id SET DEFAULT nextval('precios_competencia_id_seq')
            """)
            mensajes.append(" DEFAULT asignado a precios_competencia.id")

        except Exception as e:
            mensajes.append(f" Error en secuencia: {str(e)}")

        # Migracin 6: Migrar datos del pastillero desde SQLite a PostgreSQL
        try:
            import sqlite3
            import os

            sqlite_path = os.path.join(os.path.dirname(__file__), 'medicamentos.db')

            if os.path.exists(sqlite_path):
                mensajes.append(" Iniciando migracin de pastillero desde SQLite...")

                # Conectar a SQLite
                sqlite_conn = sqlite3.connect(sqlite_path)
                sqlite_conn.row_factory = sqlite3.Row
                sqlite_cursor = sqlite_conn.cursor()

                # Leer datos del pastillero
                sqlite_cursor.execute("""
                    SELECT usuario_id, medicamento_id, nombre, cantidad, unidad
                    FROM pastillero_usuarios
                """)

                pastillero_rows = sqlite_cursor.fetchall()
                count = 0

                if pastillero_rows:
                    for row in pastillero_rows:
                        # Verificar si ya existe para evitar duplicados
                        exists = conn.execute("""
                            SELECT id FROM pastillero_usuarios
                            WHERE usuario_id = ? AND medicamento_id IS NOT DISTINCT FROM ?
                            AND nombre = ?
                        """, (row['usuario_id'], row['medicamento_id'], row['nombre'])).fetchone()

                        if not exists:
                            conn.execute("""
                                INSERT INTO pastillero_usuarios (usuario_id, medicamento_id, nombre, cantidad, unidad)
                                VALUES (?, ?, ?, ?, ?)
                            """, (row['usuario_id'], row['medicamento_id'], row['nombre'],
                                  row['cantidad'], row['unidad']))
                            count += 1

                    mensajes.append(f" Migrados {count} medicamentos al pastillero (de {len(pastillero_rows)} encontrados)")
                else:
                    mensajes.append(" No hay medicamentos en el pastillero de SQLite")

                sqlite_conn.close()
            else:
                mensajes.append(" Archivo medicamentos.db no encontrado, saltando migracin de pastillero")

        except Exception as e:
            mensajes.append(f" Error en migracin de pastillero: {str(e)}")

        conn.commit()
        conn.close()

        mensajes.append("<br><strong>MIGRACIONES COMPLETADAS EXITOSAMENTE</strong>")
        return "<br>".join(mensajes)

    except Exception as e:
        import traceback
        return f"ERROR: {str(e)}<br><pre>{traceback.format_exc()}</pre>", 500


@app.route('/api/existencias/browse')
@admin_required
def browse_existencias():
    try:
        conn = get_db_connection()
        cursor = conn.execute("""
            SELECT e.id, e.medicamento_id, m.nombre, e.fabricante_id, e.tipo_movimiento,
                   e.cantidad, e.fecha, e.id_tercero, e.pedido_id, e.estado, e.numero_documento, e.costo_unitario
            FROM existencias e
            LEFT JOIN medicamentos m ON e.medicamento_id = m.id
            ORDER BY e.id DESC
            LIMIT 50
        """)
        rows = cursor.fetchall()
        total_cursor = conn.execute("SELECT COUNT(*) FROM existencias")
        total = total_cursor.fetchone()[0]
        conn.close()

        html = '<table border="1"><tr><th>ID</th><th>Med ID</th><th>Medicamento</th><th>Fab ID</th><th>Tipo</th><th>Cant</th><th>Fecha</th><th>Tercero</th><th>Pedido</th><th>Estado</th><th>Doc</th><th>Costo Unit</th></tr>'
        for row in rows:
            html += f'<tr><td>{row[0]}</td><td>{row[1]}</td><td>{row[2]}</td><td>{row[3]}</td><td>{row[4]}</td><td>{row[5]}</td><td>{row[6]}</td><td>{row[7]}</td><td>{row[8]}</td><td>{row[9]}</td><td>{row[10]}</td><td>{row[11]}</td></tr>'
        html += f'</table><p>Total: {total}</p>'
        return html
    except Exception as e:
        return f'Error: {str(e)}', 500


# --- RUTAS PARA PRecios actualizador con listado ---

def calcular_precio_sugerido(precio_base, config):
    if precio_base <= 0:
        return 0
    recargo = precio_base * (config['recargo_escaso'] / 100.0)
    ganancia = max(
        config['ganancia_min_escaso'],
        min(config['ganancia_max_escaso'], recargo)
    )
    sugerido = precio_base + ganancia
    redondeo = config['redondeo_superior']
    if redondeo > 0:
        sugerido = ((int(sugerido) + redondeo - 1) // redondeo) * redondeo
    return int(sugerido)

@app.route('/admin/precios-dinamicos')
@admin_required
def precios_dinamicos():
    return render_template('precios_dinamicos.html')

@app.route('/admin/fusionar-fabricantes')
@admin_required
def fusionar_fabricantes_page():
    return render_template('fusionar_fabricantes.html')

@app.route('/admin/precios-dinamicos/data')
@admin_required
def precios_dinamicos_data():
    import time
    inicio = time.time()
    try:
        db = get_db_connection()

        # Reactivar automticamente cotizaciones cuya fecha de inactividad ya expir
        db.execute("""
            UPDATE precios_competencia
            SET inactivo_hasta = NULL, activo = TRUE
            WHERE inactivo_hasta IS NOT NULL AND inactivo_hasta < CURRENT_TIMESTAMP
        """)
        db.commit()

        config_row = db.execute("SELECT * FROM CONFIGURACION_PRECIOS LIMIT 1").fetchone()
        if config_row:
            config = dict(config_row)
        else:
            config = {
                'usar_precio': 'minimo',
                'recargo_escaso': 30,
                'ganancia_min_escaso': 2000,
                'ganancia_max_escaso': 10000,
                'redondeo_superior': 100
            }

        query = """
        SELECT
            m.id as medicamento_id,
            m.nombre,
            m.activo,
            f.id as fabricante_id,
            f.nombre as fabricante_nombre,
            COALESCE(p.precio, 0) AS precio_actual,
            CASE WHEN p.imagen IS NOT NULL AND p.imagen != '' THEN 'S' ELSE 'N' END AS tiene_imagen,
            CASE
                WHEN EXISTS (
                    SELECT 1 FROM existencias e
                    WHERE e.medicamento_id = m.id
                    AND e.fabricante_id = f.id
                    AND e.estado = 'pendiente'
                ) THEN 'Pendiente' ELSE '' END AS estado_existencia,
            t.nombre AS competencia_nombre,
            pc.precio AS competencia_precio,
            pc.fecha_actualizacion AS competencia_fecha,
            '' AS origen,
            (SELECT COUNT(*) FROM precios_competencia WHERE medicamento_id = m.id AND fabricante_id = f.id AND url IS NOT NULL AND url != '') AS cotizaciones_con_url,
            (SELECT COUNT(*) FROM precios_competencia WHERE medicamento_id = m.id AND fabricante_id = f.id) AS cotizaciones_total
        FROM precios p
        LEFT JOIN medicamentos m ON p.medicamento_id = m.id
        LEFT JOIN fabricantes f ON p.fabricante_id = f.id
        LEFT JOIN (
            SELECT
                medicamento_id,
                fabricante_id,
                competidor_id,
                precio,
                fecha_actualizacion,
                ROW_NUMBER() OVER (
                    PARTITION BY medicamento_id, fabricante_id
                    ORDER BY precio ASC, fecha_actualizacion DESC
                ) as rn
            FROM precios_competencia
        ) pc ON m.id = pc.medicamento_id AND f.id = pc.fabricante_id AND pc.rn = 1
        LEFT JOIN terceros t ON pc.competidor_id = t.id

        UNION ALL

        SELECT
            NULL as medicamento_id,
            pu.nombre,
            NULL as activo,
            NULL as fabricante_id,
            '' as fabricante_nombre,
            0 AS precio_actual,
            'N' AS tiene_imagen,
            '' AS estado_existencia,
            '' AS competencia_nombre,
            NULL AS competencia_precio,
            '' AS competencia_fecha,
            'PASTILLERO' AS origen,
            0 AS cotizaciones_con_url,
            0 AS cotizaciones_total
        FROM pastillero_usuarios pu
        WHERE pu.medicamento_id IS NULL OR pu.medicamento_id = 0
        GROUP BY pu.nombre

        ORDER BY nombre, fabricante_nombre
        """
        medicamentos = [dict(row) for row in db.execute(query).fetchall()]

        primer_med = medicamentos[0] if medicamentos else None
        competencias_primer = []
        precio_sugerido = 0

        if primer_med:
            competencias_primer = [dict(row) for row in db.execute("""
                SELECT pc.id, t.nombre, pc.precio, pc.fecha_actualizacion, t.id as tercero_id,
                       t.telefono, t.direccion
                FROM precios_competencia pc
                JOIN terceros t ON pc.competidor_id = t.id
                WHERE pc.medicamento_id = ? AND pc.fabricante_id = ?
                ORDER BY pc.precio ASC
            """, (primer_med['medicamento_id'], primer_med['fabricante_id'])).fetchall()]

            if competencias_primer:
                precio_base = min([c['precio'] for c in competencias_primer], default=0)
                if precio_base == 0:
                    precio_base = primer_med['precio_actual']
            else:
                precio_base = primer_med['precio_actual']

            precio_sugerido = calcular_precio_sugerido(precio_base, config)

        db.close()

        tiempo_total = time.time() - inicio
        print(f" /admin/precios-dinamicos/data ejecutado en {tiempo_total:.2f}s - {len(medicamentos)} medicamentos")

        return jsonify({
            'config': config,
            'medicamentos': medicamentos,
            'primer_med_id': primer_med['medicamento_id'] if primer_med else None,
            'primer_fab_id': primer_med['fabricante_id'] if primer_med else None,
            'competencias_primer': competencias_primer,
            'precio_sugerido': precio_sugerido
        })
    except Exception as e:
        print(f" Error en /admin/precios-dinamicos/data: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e), 'medicamentos': [], 'config': {}}), 500

@app.route('/admin/fabricantes/buscar', methods=['GET'])
@admin_required
def buscar_fabricantes():
    query = request.args.get('q', '').strip()
    db = get_db_connection()

    if query:
        fabricantes = db.execute(
            "SELECT id, nombre FROM FABRICANTES WHERE LOWER(nombre) LIKE LOWER(?) ORDER BY LOWER(nombre) LIMIT 20",
            (f'%{query}%',)
        ).fetchall()
    else:
        fabricantes = db.execute(
            "SELECT id, nombre FROM FABRICANTES ORDER BY nombre LIMIT 20"
        ).fetchall()

    db.close()
    return jsonify({'fabricantes': [dict(f) for f in fabricantes]})

@app.route('/admin/fabricantes/crear', methods=['POST'])
@admin_required
def crear_fabricante():
    data = request.get_json()
    nombre = data.get('nombre', '').strip()

    if not nombre:
        return jsonify({'error': 'Nombre requerido'}), 400

    conn = get_db_connection()

    # Verificar si ya existe
    existe = conn.execute("SELECT id FROM fabricantes WHERE nombre = %s", (nombre,)).fetchone()
    if existe:
        conn.close()
        return jsonify({'id': existe['id'], 'nombre': nombre, 'existia': True})

    # Crear nuevo con RETURNING
    cursor = conn.execute("INSERT INTO fabricantes (nombre) VALUES (%s) RETURNING id", (nombre,))
    next_id = cursor.fetchone()[0]
    conn.commit()

    conn.close()

    return jsonify({'id': next_id, 'nombre': nombre, 'existia': False})

@app.route('/admin/huerfano/fusionar', methods=['POST'])
def fusionar_huerfano():
    data = request.get_json()
    nombre_huerfano = data.get('nombre_huerfano', '').strip()
    medicamento_id = data.get('medicamento_id')

    if not nombre_huerfano or not medicamento_id:
        return jsonify({'error': 'Datos incompletos'}), 400

    db = get_db_connection()

    try:
        # Actualizar todos los registros del pastillero con ese nombre
        db.execute("""
            UPDATE pastillero_usuarios
            SET medicamento_id = %s
            WHERE nombre = %s AND (medicamento_id IS NULL OR medicamento_id = 0)
        """, (medicamento_id, nombre_huerfano))

        db.commit()
        db.close()

        return jsonify({'ok': True, 'mensaje': 'Medicamento fusionado correctamente'})

    except Exception as e:
        db.close()
        return jsonify({'error': str(e)}), 500

@app.route('/admin/huerfano/crear', methods=['POST'])
def crear_desde_huerfano():
    data = request.get_json()
    nombre_huerfano = data.get('nombre_huerfano', '').strip()
    nombre_nuevo = data.get('nombre_nuevo', '').strip()
    fabricante_id = data.get('fabricante_id')
    fabricante_nombre = data.get('fabricante_nombre', '').strip()

    if not nombre_huerfano or not nombre_nuevo:
        return jsonify({'error': 'Nombre requerido'}), 400

    db = get_db_connection()

    try:
        # Si no hay fabricante_id, crear el fabricante
        if not fabricante_id and fabricante_nombre:
            existe = db.execute("SELECT id FROM FABRICANTES WHERE nombre = %s", (fabricante_nombre,)).fetchone()
            if existe:
                fabricante_id = existe['id']
            else:
                cursor = db.execute("INSERT INTO FABRICANTES (nombre) VALUES (%s) RETURNING id", (fabricante_nombre,))
                fabricante_id = cursor.fetchone()[0]
                db.commit()

        if not fabricante_id:
            db.close()
            return jsonify({'error': 'Fabricante requerido'}), 400

        # Crear medicamento
        cursor = db.execute("INSERT INTO MEDICAMENTOS (nombre) VALUES (%s) RETURNING id", (nombre_nuevo,))
        medicamento_id = cursor.fetchone()[0]
        db.commit()

        # Crear registro en precios con precio 0
        db.execute("""
            INSERT INTO precios (medicamento_id, fabricante_id, precio, fecha_actualizacion)
            VALUES (%s, %s, 0, CURRENT_TIMESTAMP)
        """, (medicamento_id, fabricante_id))
        db.commit()

        # Actualizar pastillero
        db.execute("""
            UPDATE pastillero_usuarios
            SET medicamento_id = %s
            WHERE nombre = %s AND (medicamento_id IS NULL OR medicamento_id = 0)
        """, (medicamento_id, nombre_huerfano))
        db.commit()

        db.close()

        return jsonify({
            'ok': True,
            'medicamento_id': medicamento_id,
            'fabricante_id': fabricante_id,
            'mensaje': 'Medicamento creado correctamente'
        })

    except Exception as e:
        db.close()
        return jsonify({'error': str(e)}), 500


@app.route('/admin/huerfano/crear-solo-medicamento', methods=['POST'])
def crear_solo_medicamento():
    """Crea medicamento en MEDICAMENTOS sin fabricante ni precio"""
    data = request.get_json()
    nombre_huerfano = data.get('nombre_huerfano', '').strip()
    nombre_nuevo = data.get('nombre_nuevo', '').strip()

    if not nombre_huerfano or not nombre_nuevo:
        return jsonify({'error': 'Nombre requerido'}), 400

    db = get_db_connection()
    try:
        # Crear medicamento
        cursor = db.execute("INSERT INTO MEDICAMENTOS (nombre) VALUES (?)", (nombre_nuevo,))
        medicamento_id = cursor.lastrowid
        db.commit()

        # Actualizar pastillero_usuarios con el nuevo medicamento_id
        db.execute("""
            UPDATE pastillero_usuarios
            SET medicamento_id = ?
            WHERE nombre = ? AND (medicamento_id IS NULL OR medicamento_id = 0)
        """, (medicamento_id, nombre_huerfano))
        db.commit()
        db.close()

        return jsonify({'ok': True, 'medicamento_id': medicamento_id})
    except Exception as e:
        db.close()
        return jsonify({'error': str(e)}), 500


@app.route('/admin/huerfano/finalizar', methods=['POST'])
def finalizar_huerfano():
    """Crea registro en PRECIOS y finaliza el proceso"""
    data = request.get_json()
    medicamento_id = data.get('medicamento_id')
    fabricante_id = data.get('fabricante_id')
    nombre_huerfano = data.get('nombre_huerfano', '').strip()

    if not medicamento_id or not fabricante_id:
        return jsonify({'error': 'Datos incompletos'}), 400

    db = get_db_connection()
    try:
        # Crear registro en precios con precio inicial en 0
        db.execute("""
            INSERT INTO precios (medicamento_id, fabricante_id, precio, fecha_actualizacion)
            VALUES (?, ?, 0, CURRENT_TIMESTAMP)
        """, (medicamento_id, fabricante_id))
        db.commit()
        db.close()

        return jsonify({'ok': True, 'mensaje': 'Medicamento finalizado correctamente'})
    except Exception as e:
        db.close()
        return jsonify({'error': str(e)}), 500


@app.route('/admin/huerfano/autocompletar', methods=['POST'])
def autocompletar_huerfano():
    """Usa Google Custom Search para autocompletar nombre y fabricante"""
    import re

    data = request.get_json()
    nombre = data.get('nombre', '').strip()

    if not nombre:
        return jsonify({'error': 'Nombre requerido'}), 400

    try:
        # Google Custom Search API
        GOOGLE_API_KEY = 'AIzaSyCiAtNFl95bJJFuqiNsiYynBS3LuDisq9g'
        SEARCH_ENGINE_ID = '40c8305664a9147e9'

        query = f"en colombia quien fabrica {nombre}"
        url_api = f"https://www.googleapis.com/customsearch/v1?key={GOOGLE_API_KEY}&cx={SEARCH_ENGINE_ID}&q={query}&num=5"

        resp = requests.get(url_api, timeout=10)

        if resp.status_code != 200:
            return jsonify({'error': 'Error en Google Search API'}), 500

        data_search = resp.json()
        items = data_search.get("items", [])

        if not items:
            return jsonify({'error': 'No se encontraron resultados'}), 404

        # Intentar extraer info del primer resultado
        primer_item = items[0]
        titulo = primer_item.get('title', '')
        snippet = primer_item.get('snippet', '')

        # Buscar fabricante en el ttulo y snippet
        fabricantes_posibles = []

        # Patrones comunes de fabricantes
        patrones_fabricante = [
            r'\|\s*([A-Z][a-zA-Z\s&]+?)\s*\|',  # |Interpharma|
            r'Fabricante[:\s]+([A-Z][a-zA-Z\s&]+)',  # Fabricante: Interpharma
            r'Laboratorio[:\s]+([A-Z][a-zA-Z\s&]+)',  # Laboratorio: Genfar
        ]

        for patron in patrones_fabricante:
            match = re.search(patron, titulo + ' ' + snippet)
            if match:
                fabricantes_posibles.append(match.group(1).strip())

        # Buscar nombre del medicamento mejorado
        nombre_completo = nombre.upper()

        # Intentar extraer presentacin del ttulo
        match_presentacion = re.search(r'(TABLETAS?|CAPSULAS?|SOBRES?|ML|MG|G)\s*(X\s*\d+)?', titulo, re.IGNORECASE)
        if match_presentacion:
            # Extraer parte relevante del ttulo
            nombre_completo = titulo.split('|')[0].strip().upper()

        resultado = {
            'nombre_completo': nombre_completo,
            'fabricante': fabricantes_posibles[0] if fabricantes_posibles else '',
            'titulo_original': titulo,
            'snippet': snippet,
            'confianza': 'alta' if fabricantes_posibles else 'baja'
        }

        return jsonify(resultado)

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/admin/competencia/obtener-dominio-tercero/<int:tercero_id>', methods=['GET'])
def obtener_dominio_tercero(tercero_id):
    """Obtiene el dominio tpico de un tercero basado en sus URLs histricas"""
    from urllib.parse import urlparse
    from collections import Counter

    db = get_db_connection()
    try:
        # Obtener todas las URLs de este tercero
        urls = db.execute("""
            SELECT url
            FROM precios_competencia
            WHERE competidor_id = ? AND url IS NOT NULL AND LENGTH(url) > 0
        """, (tercero_id,)).fetchall()

        if not urls:
            db.close()
            return jsonify({'dominio': None, 'mensaje': 'Sin URLs histricas'})

        # Extraer dominios
        dominios = []
        for row in urls:
            url = row['url']
            try:
                parsed = urlparse(url)
                dominio = parsed.netloc.lower()
                # Remover 'www.' si existe
                if dominio.startswith('www.'):
                    dominio = dominio[4:]
                if dominio:
                    dominios.append(dominio)
            except:
                continue

        db.close()

        if not dominios:
            return jsonify({'dominio': None, 'mensaje': 'No se pudieron extraer dominios'})

        # Encontrar el dominio ms comn
        contador = Counter(dominios)
        dominio_comun = contador.most_common(1)[0][0]
        total_urls = len(dominios)
        frecuencia = contador[dominio_comun]

        return jsonify({
            'dominio': dominio_comun,
            'frecuencia': frecuencia,
            'total': total_urls,
            'porcentaje': round((frecuencia / total_urls) * 100, 1)
        })

    except Exception as e:
        db.close()
        return jsonify({'error': str(e)}), 500


@app.route('/admin/competencia/auditar-urls', methods=['GET'])
def auditar_urls():
    """Audita todas las URLs guardadas para detectar inconsistencias"""
    from urllib.parse import urlparse
    from collections import Counter

    db = get_db_connection()
    try:
        # Obtener todos los terceros con URLs
        terceros = db.execute("""
            SELECT DISTINCT t.id, t.nombre
            FROM terceros t
            JOIN precios_competencia pc ON pc.competidor_id = t.id
            WHERE pc.url IS NOT NULL AND LENGTH(pc.url) > 0
            ORDER BY t.nombre
        """).fetchall()

        problemas = []

        for tercero in terceros:
            tercero_id = tercero['id']
            tercero_nombre = tercero['nombre']

            # Obtener todas las URLs de este tercero
            urls_tercero = db.execute("""
                SELECT pc.id, pc.url, m.nombre as medicamento_nombre
                FROM precios_competencia pc
                JOIN MEDICAMENTOS m ON pc.medicamento_id = m.id
                WHERE pc.competidor_id = ? AND pc.url IS NOT NULL AND LENGTH(pc.url) > 0
            """, (tercero_id,)).fetchall()

            if not urls_tercero:
                continue

            # Extraer dominios
            dominios = []
            urls_con_dominio = []

            for row in urls_tercero:
                url = row['url']
                try:
                    parsed = urlparse(url)
                    dominio = parsed.netloc.lower()
                    if dominio.startswith('www.'):
                        dominio = dominio[4:]

                    if dominio:
                        dominios.append(dominio)
                        urls_con_dominio.append({
                            'id': row['id'],
                            'url': url,
                            'dominio': dominio,
                            'medicamento': row['medicamento_nombre']
                        })
                    else:
                        # URL sin dominio vlido (basura)
                        problemas.append({
                            'tipo': 'url_invalida',
                            'tercero_id': tercero_id,
                            'tercero_nombre': tercero_nombre,
                            'cotizacion_id': row['id'],
                            'url': url,
                            'medicamento': row['medicamento_nombre'],
                            'mensaje': 'URL sin dominio vlido (posible basura)'
                        })
                except Exception as e:
                    # Error al parsear URL (basura)
                    problemas.append({
                        'tipo': 'url_invalida',
                        'tercero_id': tercero_id,
                        'tercero_nombre': tercero_nombre,
                        'cotizacion_id': row['id'],
                        'url': url,
                        'medicamento': row['medicamento_nombre'],
                        'mensaje': f'URL invlida: {str(e)}'
                    })

            # Si no hay dominios vlidos, continuar
            if not dominios:
                continue

            # Encontrar el dominio ms comn (dominio esperado)
            contador = Counter(dominios)
            dominio_esperado = contador.most_common(1)[0][0]
            frecuencia_esperado = contador[dominio_esperado]

            # Detectar URLs con dominio diferente al esperado
            for item in urls_con_dominio:
                if item['dominio'] != dominio_esperado:
                    problemas.append({
                        'tipo': 'dominio_incorrecto',
                        'tercero_id': tercero_id,
                        'tercero_nombre': tercero_nombre,
                        'cotizacion_id': item['id'],
                        'url': item['url'],
                        'medicamento': item['medicamento'],
                        'dominio_encontrado': item['dominio'],
                        'dominio_esperado': dominio_esperado,
                        'mensaje': f"Dominio '{item['dominio']}' no coincide con el esperado '{dominio_esperado}'"
                    })

        db.close()

        return jsonify({
            'total_problemas': len(problemas),
            'problemas': problemas
        })

    except Exception as e:
        db.close()
        return jsonify({'error': str(e)}), 500


@app.route('/admin/competencia/conteo-por-tercero', methods=['GET'])
def conteo_cotizaciones_por_tercero():
    """Obtiene el conteo total de cotizaciones por cada tercero"""
    db = get_db_connection()
    try:
        # Contar cotizaciones por tercero
        resultado = db.execute("""
            SELECT
                t.nombre as tercero_nombre,
                COUNT(pc.id) as total_cotizaciones
            FROM terceros t
            LEFT JOIN precios_competencia pc ON pc.competidor_id = t.id
            GROUP BY t.id, t.nombre
            HAVING total_cotizaciones > 0
            ORDER BY total_cotizaciones DESC
        """).fetchall()

        conteo = {}
        for row in resultado:
            conteo[row['tercero_nombre']] = row['total_cotizaciones']

        db.close()
        return jsonify(conteo)

    except Exception as e:
        db.close()
        return jsonify({'error': str(e)}), 500


@app.route('/admin/competencia/auditoria-dispersion', methods=['GET'])
def auditoria_dispersion():
    """Audita medicamentos con alta dispersin de precios para detectar posibles errores"""
    umbral = request.args.get('umbral', default=30, type=float)

    db = get_db_connection()
    try:
        # Obtener PRODUCTOS (medicamento + fabricante) con mltiples cotizaciones y calcular dispersin
        query = """
            SELECT
                pc.medicamento_id,
                pc.fabricante_id,
                m.nombre as medicamento_nombre,
                f.nombre as fabricante_nombre,
                COUNT(pc.id) as num_cotizaciones,
                MIN(pc.precio) as precio_min,
                MAX(pc.precio) as precio_max,
                AVG(pc.precio) as precio_promedio,
                STRING_AGG(t.nombre || ': $' || pc.precio, ' | ') as "detalle_precios_str"
            FROM precios_competencia pc
            INNER JOIN MEDICAMENTOS m ON pc.medicamento_id = m.id
            LEFT JOIN FABRICANTES f ON pc.fabricante_id = f.id
            LEFT JOIN terceros t ON t.id = pc.competidor_id
            WHERE pc.precio > 0
            GROUP BY pc.medicamento_id, pc.fabricante_id, m.nombre, f.nombre
            HAVING num_cotizaciones >= 2
            ORDER BY num_cotizaciones DESC
        """

        resultados = db.execute(query).fetchall()

        medicamentos_con_dispersion = []

        for row in resultados:
            precio_min = row['precio_min']
            precio_max = row['precio_max']

            # Calcular dispersin porcentual
            if precio_min > 0:
                dispersion_pct = ((precio_max - precio_min) / precio_min) * 100
            else:
                dispersion_pct = 0

            # Filtrar por umbral
            if dispersion_pct >= umbral:
                # Nombre completo del producto (medicamento + fabricante)
                nombre_completo = f"{row['medicamento_nombre']} - {row['fabricante_nombre'] or 'Sin fabricante'}"

                medicamentos_con_dispersion.append({
                    'medicamento_id': row['medicamento_id'],
                    'fabricante_id': row['fabricante_id'],
                    'medicamento_nombre': nombre_completo,
                    'num_cotizaciones': row['num_cotizaciones'],
                    'precio_min': round(precio_min, 2),
                    'precio_max': round(precio_max, 2),
                    'precio_promedio': round(row['precio_promedio'], 2),
                    'dispersion_pct': round(dispersion_pct, 1),
                    'detalle_precios': row['detalle_precios_str']
                })

        db.close()

        return jsonify({
            'umbral_aplicado': umbral,
            'total_encontrados': len(medicamentos_con_dispersion),
            'medicamentos': medicamentos_con_dispersion
        })

    except Exception as e:
        db.close()
        return jsonify({'error': str(e)}), 500


@app.route('/admin/terceros/buscar-o-crear', methods=['POST'])
def buscar_o_crear_tercero():
    data = request.get_json()
    nombre = data.get('nombre', '').strip()
    if not nombre:
        return jsonify({'error': 'Nombre requerido'}), 400

    db = get_db_connection()
    # Buscar tercero existente (insensible a maysculas/minsculas)
    # PostgreSQL: usar LOWER() en lugar de COLLATE NOCASE
    tercero = db.execute("SELECT id FROM terceros WHERE LOWER(nombre) = LOWER(%s)", (nombre,)).fetchone()
    if tercero:
        # Actualizar fecha_actualizacion al usarlo
        db.execute(
            "UPDATE terceros SET fecha_actualizacion = %s WHERE id = %s",
            (datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S'), tercero['id'])
        )
        db.commit()
        db.close()
        return jsonify({'id': tercero['id']})

    # Crear nuevo tercero
    cursor = db.execute(
        "INSERT INTO terceros (nombre, fecha_creacion, fecha_actualizacion) VALUES (%s, %s, %s) RETURNING id",
        (nombre, datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S'), datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S'))
    )
    nuevo_id = cursor.fetchone()[0]
    db.commit()
    db.close()
    return jsonify({'id': nuevo_id})


@app.route('/admin/fix-terceros-sequence', methods=['GET'])
@admin_required
def fix_terceros_sequence():
    """Endpoint temporal para crear la secuencia de terceros en PostgreSQL"""
    try:
        db = get_db_connection()

        # Crear secuencia si no existe
        db.execute('CREATE SEQUENCE IF NOT EXISTS terceros_id_seq')

        # Configurar columna para usar la secuencia
        db.execute("ALTER TABLE terceros ALTER COLUMN id SET DEFAULT nextval('terceros_id_seq')")

        # Obtener MAX id actual
        result = db.execute('SELECT COALESCE(MAX(id), 0) FROM terceros').fetchone()
        max_id = result[0] if result else 0

        # Sincronizar secuencia
        db.execute(f"SELECT setval('terceros_id_seq', {max_id})")

        db.commit()
        db.close()

        return jsonify({
            'success': True,
            'message': f'Secuencia terceros_id_seq creada y sincronizada. MAX ID: {max_id}'
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/admin/fix-alertas-sequence', methods=['GET'])
@admin_required
def fix_alertas_sequence():
    """Endpoint temporal para crear la secuencia de alertas_admin en PostgreSQL"""
    try:
        db = get_db_connection()

        # Crear secuencia si no existe
        db.execute('CREATE SEQUENCE IF NOT EXISTS alertas_admin_id_seq')

        # Configurar columna para usar la secuencia
        db.execute("ALTER TABLE alertas_admin ALTER COLUMN id SET DEFAULT nextval('alertas_admin_id_seq')")

        # Obtener MAX id actual
        result = db.execute('SELECT COALESCE(MAX(id), 0) FROM alertas_admin').fetchone()
        max_id = result[0] if result else 0

        # Sincronizar secuencia
        db.execute(f"SELECT setval('alertas_admin_id_seq', {max_id})")

        db.commit()
        db.close()

        return jsonify({
            'success': True,
            'message': f'Secuencia alertas_admin_id_seq creada y sincronizada. MAX ID: {max_id}'
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/admin/crear-tabla-tokens-vinculacion', methods=['GET'])
@admin_required
def crear_tabla_tokens_vinculacion():
    """Endpoint temporal para crear la tabla de tokens de vinculación"""
    try:
        db = get_db_connection()

        # Crear tabla de tokens de vinculación
        db.execute('''
            CREATE TABLE IF NOT EXISTS tokens_vinculacion (
                token TEXT PRIMARY KEY,
                usuario_id INTEGER NOT NULL,
                nombre TEXT,
                telefono TEXT,
                fecha_creacion TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                usado BOOLEAN DEFAULT FALSE,
                fecha_uso TIMESTAMP
            )
        ''')

        db.commit()
        db.close()

        return jsonify({
            'success': True,
            'message': 'Tabla tokens_vinculacion creada exitosamente'
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/admin/diagnostico-pastillero', methods=['GET'])
@admin_required
def diagnostico_pastillero():
    """Endpoint para diagnosticar estado del pastillero y recordatorios"""
    try:
        db = get_db_connection()

        ahora = datetime.now()

        # Obtener todos los medicamentos del pastillero con recordatorio activo
        pastillero = db.execute('''
            SELECT
                p.id,
                p.usuario_id,
                p.nombre,
                p.cantidad,
                p.horas_entre_tomas,
                p.proxima_toma,
                p.recordatorio_activo,
                t.nombre as usuario_nombre,
                t.telegram_chat_id
            FROM pastillero_usuarios p
            INNER JOIN terceros t ON p.usuario_id = t.id
            WHERE p.recordatorio_activo = TRUE
            ORDER BY p.proxima_toma
        ''').fetchall()

        resultado = {
            'hora_servidor': ahora.strftime('%Y-%m-%d %H:%M:%S'),
            'total_recordatorios_activos': len(pastillero),
            'medicamentos': []
        }

        for med in pastillero:
            proxima_toma = med['proxima_toma']
            if proxima_toma:
                if isinstance(proxima_toma, str):
                    from dateutil import parser
                    proxima_toma = parser.parse(proxima_toma)

                diferencia = (proxima_toma - ahora).total_seconds() / 60  # en minutos
                estado = 'VENCIDO' if diferencia < 0 else 'PENDIENTE'

                resultado['medicamentos'].append({
                    'id': med['id'],
                    'nombre': med['nombre'],
                    'usuario': med['usuario_nombre'],
                    'telegram_chat_id': med['telegram_chat_id'],
                    'cantidad': med['cantidad'],
                    'horas_entre_tomas': med['horas_entre_tomas'],
                    'proxima_toma': proxima_toma.strftime('%Y-%m-%d %H:%M:%S'),
                    'diferencia_minutos': round(diferencia, 1),
                    'estado': estado
                })

        db.close()

        return jsonify(resultado)

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/admin/actualizar-telegram-chat-id', methods=['POST'])
@admin_required
def actualizar_telegram_chat_id():
    """Endpoint para actualizar manualmente el telegram_chat_id de un usuario"""
    try:
        data = request.get_json()
        usuario_id = data.get('usuario_id')
        chat_id = data.get('chat_id')

        if not usuario_id or not chat_id:
            return jsonify({'ok': False, 'error': 'Faltan usuario_id o chat_id'}), 400

        db = get_db_connection()

        # Obtener nombre del usuario
        usuario = db.execute('SELECT nombre FROM terceros WHERE id = %s', (usuario_id,)).fetchone()

        if not usuario:
            db.close()
            return jsonify({'ok': False, 'error': 'Usuario no encontrado'}), 404

        # Actualizar telegram_chat_id
        db.execute('''
            UPDATE terceros
            SET telegram_chat_id = %s
            WHERE id = %s
        ''', (str(chat_id), usuario_id))

        db.commit()
        db.close()

        return jsonify({
            'ok': True,
            'mensaje': f'Chat ID actualizado para {usuario["nombre"]}',
            'usuario': usuario['nombre'],
            'chat_id': chat_id
        })

    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500


@app.route('/admin/migrar-sistema-pastilleros', methods=['POST'])
@admin_required
def migrar_sistema_pastilleros():
    """Endpoint para migrar al nuevo sistema de pastilleros compartidos"""
    try:
        db = get_db_connection()

        resultado = {
            'ok': True,
            'pasos': [],
            'errores': []
        }

        # 1. Crear tabla pastilleros
        resultado['pasos'].append('Creando tabla pastilleros...')
        db.execute('''
            CREATE TABLE IF NOT EXISTS pastilleros (
                id SERIAL PRIMARY KEY,
                nombre VARCHAR(100) NOT NULL,
                creado_por_usuario_id INTEGER NOT NULL REFERENCES terceros(id) ON DELETE CASCADE,
                fecha_creacion TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')

        # 2. Crear tabla relaciones_pastillero
        resultado['pasos'].append('Creando tabla relaciones_pastillero...')
        db.execute('''
            CREATE TABLE IF NOT EXISTS relaciones_pastillero (
                id SERIAL PRIMARY KEY,
                pastillero_id INTEGER NOT NULL REFERENCES pastilleros(id) ON DELETE CASCADE,
                usuario_id INTEGER NOT NULL REFERENCES terceros(id) ON DELETE CASCADE,
                tipo VARCHAR(20) NOT NULL CHECK (tipo IN ('propietario', 'miembro', 'autorizado')),
                fecha_agregado TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(pastillero_id, usuario_id)
            )
        ''')

        # 3. Crear tabla mensajes
        resultado['pasos'].append('Creando tabla mensajes...')
        db.execute('''
            CREATE TABLE IF NOT EXISTS mensajes (
                id SERIAL PRIMARY KEY,
                remitente_id INTEGER NOT NULL REFERENCES terceros(id) ON DELETE CASCADE,
                destinatario_id INTEGER NOT NULL REFERENCES terceros(id) ON DELETE CASCADE,
                mensaje TEXT NOT NULL,
                tipo VARCHAR(30) NOT NULL DEFAULT 'texto' CHECK (tipo IN ('texto', 'invitacion_compartir', 'solicitud_acceso')),
                pastillero_id INTEGER REFERENCES pastilleros(id) ON DELETE CASCADE,
                estado VARCHAR(20) NOT NULL DEFAULT 'pendiente' CHECK (estado IN ('pendiente', 'aceptado', 'rechazado', 'leido')),
                fecha TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')

        # 4. Obtener usuarios con medicamentos
        resultado['pasos'].append('Obteniendo usuarios con medicamentos...')
        usuarios_con_pastillero = db.execute('''
            SELECT DISTINCT usuario_id, t.nombre
            FROM pastillero_usuarios p
            INNER JOIN terceros t ON p.usuario_id = t.id
        ''').fetchall()

        resultado['usuarios_encontrados'] = len(usuarios_con_pastillero)

        # 5. Crear pastilleros para cada usuario
        resultado['pasos'].append(f'Creando pastilleros para {len(usuarios_con_pastillero)} usuarios...')
        pastilleros_creados = 0

        for usuario in usuarios_con_pastillero:
            usuario_id = usuario['usuario_id']
            nombre_usuario = usuario['nombre']

            # Crear pastillero
            pastillero_id = db.execute('''
                INSERT INTO pastilleros (nombre, creado_por_usuario_id)
                VALUES (%s, %s)
                RETURNING id
            ''', (f"Mi pastillero", usuario_id)).fetchone()['id']

            # Crear relación de propietario
            db.execute('''
                INSERT INTO relaciones_pastillero (pastillero_id, usuario_id, tipo)
                VALUES (%s, %s, 'propietario')
            ''', (pastillero_id, usuario_id))

            pastilleros_creados += 1

        resultado['pastilleros_creados'] = pastilleros_creados

        # 6. Agregar columna pastillero_id si no existe
        resultado['pasos'].append('Agregando columna pastillero_id...')
        try:
            db.execute('''
                ALTER TABLE pastillero_usuarios
                ADD COLUMN IF NOT EXISTS pastillero_id INTEGER REFERENCES pastilleros(id) ON DELETE CASCADE
            ''')
        except Exception as e:
            # La columna ya existe, continuar
            pass

        # 7. Migrar medicamentos
        resultado['pasos'].append('Migrando medicamentos al nuevo sistema...')
        db.execute('''
            UPDATE pastillero_usuarios pu
            SET pastillero_id = p.id
            FROM pastilleros p
            WHERE pu.usuario_id = p.creado_por_usuario_id
            AND pu.pastillero_id IS NULL
        ''')

        medicamentos_migrados = db.execute('SELECT COUNT(*) as count FROM pastillero_usuarios WHERE pastillero_id IS NOT NULL').fetchone()['count']
        resultado['medicamentos_migrados'] = medicamentos_migrados

        # 8. Hacer pastillero_id NOT NULL
        resultado['pasos'].append('Configurando restricciones finales...')
        try:
            db.execute('''
                ALTER TABLE pastillero_usuarios
                ALTER COLUMN pastillero_id SET NOT NULL
            ''')
        except Exception as e:
            # Ya está configurado, continuar
            pass

        db.commit()
        db.close()

        resultado['mensaje'] = 'Migración completada exitosamente'
        return jsonify(resultado)

    except Exception as e:
        import traceback
        return jsonify({
            'ok': False,
            'error': str(e),
            'traceback': traceback.format_exc()
        }), 500


@app.route('/admin/terceros/guardar-campo', methods=['POST'])
def guardar_campo_tercero():
    data = request.get_json()
    tercero_id = data.get('tercero_id')
    campo = data.get('campo')
    valor = data.get('valor', '')

    if not tercero_id or campo not in ['telefono', 'direccion']:
        return jsonify({'ok': False})

    db = get_db_connection()
    db.execute(
        f"UPDATE terceros SET {campo} = ?, fecha_actualizacion = ? WHERE id = ?",
        (valor, datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S'), tercero_id)
    )
    db.commit()
    return jsonify({'ok': True})

@app.route('/admin/precios_competencia/guardar-dinamico', methods=['POST'])
def guardar_precio_competencia_dinamico():
    data = request.get_json()
    med_id = data.get('medicamento_id')
    fab_id = data.get('fabricante_id')
    comp_id = data.get('competidor_id')
    precio = data.get('precio')
    url = data.get('url')  # URL opcional

    if not med_id or not fab_id or not comp_id or not precio or float(precio) <= 0:
        return jsonify({'ok': False})

    db = get_db_connection()
    fecha_actual = datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')

    if url:
        db.execute("""
            INSERT INTO precios_competencia
            (medicamento_id, fabricante_id, competidor_id, precio, url, fecha_actualizacion, activo)
            VALUES (?, ?, ?, ?, ?, ?, TRUE)
        """, (med_id, fab_id, comp_id, float(precio), url, fecha_actual))
    else:
        db.execute("""
            INSERT INTO precios_competencia
            (medicamento_id, fabricante_id, competidor_id, precio, fecha_actualizacion, activo)
            VALUES (?, ?, ?, ?, ?, TRUE)
        """, (med_id, fab_id, comp_id, float(precio), fecha_actual))

    # Actualizar fecha_actualizacion del tercero para que aparezca en "recientemente usado"
    db.execute("""
        UPDATE terceros
        SET fecha_actualizacion = ?
        WHERE id = ?
    """, (fecha_actual, comp_id))

    db.commit()

    # Recalcular precio segn polticas automticamente
    precio_nuevo = calcular_precio_segun_politica(med_id, fab_id, db)

    if precio_nuevo is not None:
        # Actualizar precio en tabla precios
        db.execute("""
            UPDATE precios
            SET precio = ?
            WHERE medicamento_id = ? AND fabricante_id = ?
        """, (precio_nuevo, med_id, fab_id))
        db.commit()

    # Contar cotizaciones totales para este producto
    cotizaciones_total = db.execute("""
        SELECT COUNT(*) as total
        FROM precios_competencia
        WHERE medicamento_id = ? AND fabricante_id = ?
    """, (med_id, fab_id)).fetchone()['total']

    db.close()

    return jsonify({
        'ok': True,
        'precio_calculado': precio_nuevo,
        'cotizaciones_total': cotizaciones_total
    })

@app.route('/admin/precios_competencia/listar')
def listar_precios_competencia():
    med_id = request.args.get('medicamento_id', type=int)
    fab_id = request.args.get('fabricante_id', type=int)

    if not med_id or not fab_id:
        return jsonify({'competencias': []})

    db = get_db_connection()
    rows = db.execute("""
        SELECT pc.id, t.nombre, pc.precio, pc.fecha_actualizacion,
               t.id as tercero_id, t.telefono, t.direccion, pc.url,
               pc.activo, pc.inactivo_hasta
        FROM precios_competencia pc
        JOIN terceros t ON pc.competidor_id = t.id
        WHERE pc.medicamento_id = ? AND pc.fabricante_id = ?
        ORDER BY pc.activo DESC, pc.precio ASC
    """, (med_id, fab_id)).fetchall()
    db.close()

    return jsonify({'competencias': [dict(row) for row in rows]})


@app.route('/api/cotizacion/reactivar', methods=['POST'])
@admin_required
def reactivar_cotizacion():
    """Reactivar una cotizacin que estaba marcada como inactiva temporalmente"""
    try:
        data = request.json
        cotizacion_id = data.get('cotizacion_id')

        if not cotizacion_id:
            return jsonify({'success': False, 'message': 'ID de cotizacin requerido'}), 400

        conn = get_db_connection()

        # Reactivar: poner inactivo_hasta = NULL y activo = TRUE
        conn.execute("""
            UPDATE precios_competencia
            SET inactivo_hasta = NULL, activo = TRUE
            WHERE id = ?
        """, [cotizacion_id])

        conn.commit()
        conn.close()

        return jsonify({'success': True, 'message': 'Cotizacin reactivada'})

    except Exception as e:
        print(f" ERROR en reactivar_cotizacion: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/admin/configuracion-precios/guardar', methods=['POST'])
def guardar_configuracion_precios():
    data = request.get_json()
    campo = data.get('campo')
    valor = data.get('valor')
    
    if not campo:
        return jsonify({'ok': False, 'error': 'Campo requerido'})
    
    campos_validos = ['usar_precio', 'recargo_escaso', 'recargo_1_cotizacion', 'redondeo_superior',
                      'ganancia_min_escaso', 'ganancia_max_escaso', 'pedido_min_domicilio_gratis',
                      'umbral_brecha_3cot', 'umbral_brecha_4cot',
                      'permitir_publicar_sin_cotizaciones']
    
    if campo not in campos_validos:
        return jsonify({'ok': False, 'error': 'Campo no vlido'})
    
    db = get_db_connection()
    
    # Verificar si existe configuracin
    config_existe = db.execute("SELECT COUNT(*) as count FROM CONFIGURACION_PRECIOS").fetchone()
    
    if config_existe['count'] == 0:
        # Insertar nueva configuracin
        db.execute(f"""
            INSERT INTO CONFIGURACION_PRECIOS 
            (usar_precio, recargo_escaso, redondeo_superior, ganancia_min_escaso, ganancia_max_escaso)
            VALUES (?, ?, ?, ?, ?)
        """, ('minimo', 30, 100, 2000, 10000))
        db.commit()
    
    # Actualizar campo especfico
    db.execute(f"UPDATE CONFIGURACION_PRECIOS SET {campo} = ?", (valor,))
    db.commit()
    db.close()

    return jsonify({'ok': True})


@app.route('/admin/precios/calcular-precio-producto')
def calcular_precio_producto():
    """
    Calcula el precio de un solo producto segn polticas.
    Parmetros: medicamento_id, fabricante_id
    """
    medicamento_id = request.args.get('medicamento_id', type=int)
    fabricante_id = request.args.get('fabricante_id', type=int)

    if not medicamento_id or not fabricante_id:
        return jsonify({'precio_calculado': None, 'error': 'Parmetros faltantes'}), 400

    db = get_db_connection()
    precio_calculado = calcular_precio_segun_politica(medicamento_id, fabricante_id, db)
    db.close()

    return jsonify({'precio_calculado': precio_calculado})


@app.route('/admin/precios/calcular-politicas-masivo')
def calcular_politicas_masivo():
    """
    Retorna lista de productos con sus precios actuales y precios calculados segn polticas.
    No modifica BD, solo retorna datos para review.
    """
    db = get_db_connection()

    # Obtener todos los productos con cotizaciones
    productos = db.execute("""
        SELECT
            p.medicamento_id,
            p.fabricante_id,
            p.precio as precio_actual,
            m.nombre as medicamento_nombre,
            f.nombre as fabricante_nombre,
            (SELECT COUNT(*) FROM precios_competencia pc
             WHERE pc.medicamento_id = p.medicamento_id
             AND pc.fabricante_id = p.fabricante_id) as num_cotizaciones
        FROM precios p
        INNER JOIN medicamentos m ON p.medicamento_id = m.id
        INNER JOIN fabricantes f ON p.fabricante_id = f.id
        WHERE EXISTS (SELECT 1 FROM precios_competencia pc
                      WHERE pc.medicamento_id = p.medicamento_id
                      AND pc.fabricante_id = p.fabricante_id)
        ORDER BY num_cotizaciones DESC, p.precio DESC
    """).fetchall()

    resultados = []

    for prod in productos:
        # Calcular precio segn polticas
        precio_nuevo = calcular_precio_segun_politica(prod['medicamento_id'], prod['fabricante_id'], db)

        if precio_nuevo is not None:
            # Solo incluir si el precio es diferente (redondeo para comparacin)
            if round(prod['precio_actual']) != round(precio_nuevo):
                # Obtener cotizaciones para mostrar
                cotizaciones = db.execute("""
                    SELECT precio FROM precios_competencia
                    WHERE medicamento_id = ? AND fabricante_id = ?
                    ORDER BY precio ASC
                """, (prod['medicamento_id'], prod['fabricante_id'])).fetchall()

                precios_cot = [c['precio'] for c in cotizaciones]

                resultados.append({
                    'medicamento_id': prod['medicamento_id'],
                    'fabricante_id': prod['fabricante_id'],
                    'nombre': f"{prod['medicamento_nombre']} - {prod['fabricante_nombre']}",
                    'precio_actual': prod['precio_actual'],
                    'precio_nuevo': precio_nuevo,
                    'cambio': precio_nuevo - prod['precio_actual'],
                    'num_cotizaciones': prod['num_cotizaciones'],
                    'cotizaciones': precios_cot
                })

    db.close()

    return jsonify({'productos': resultados})


@app.route('/admin/precios/aplicar-politica-producto', methods=['POST'])
def aplicar_politica_producto():
    """
    Aplica el nuevo precio a un producto especfico.
    """
    data = request.get_json()
    med_id = data.get('medicamento_id')
    fab_id = data.get('fabricante_id')
    precio_nuevo = data.get('precio_nuevo')

    if not med_id or not fab_id or not precio_nuevo:
        return jsonify({'ok': False})

    db = get_db_connection()
    db.execute("""
        UPDATE precios
        SET precio = ?
        WHERE medicamento_id = ? AND fabricante_id = ?
    """, (precio_nuevo, med_id, fab_id))
    db.commit()
    db.close()

    return jsonify({'ok': True})


@app.route('/admin/precios/exportar-comparativa-excel')
def exportar_comparativa_excel():
    """
    Exporta tabla comparativa de precios: Productos vs Competidores
    Columnas: ID | Nombre Completo | Laboratorio | Mi Precio | [Competidor 1] | [Competidor 2] | ...
    """
    db = get_db_connection()

    competidores = db.execute("SELECT DISTINCT t.id, UPPER(TRIM(t.nombre)) as nombre FROM terceros t JOIN precios_competencia pc ON t.id = pc.competidor_id ORDER BY nombre").fetchall()
    competidores_list = sorted([c['nombre'] for c in competidores])

    productos = db.execute("SELECT m.id as med_id, f.id as fab_id, UPPER(TRIM(m.nombre)) as nombre_completo, UPPER(TRIM(f.nombre)) as laboratorio, p.precio as mi_precio FROM medicamentos m JOIN precios p ON m.id = p.medicamento_id JOIN fabricantes f ON p.fabricante_id = f.id ORDER BY nombre_completo").fetchall()

    productos_data = []
    for prod in productos:
        cotizaciones = db.execute("SELECT UPPER(TRIM(t.nombre)) as competidor, pc.precio FROM precios_competencia pc JOIN terceros t ON pc.competidor_id = t.id WHERE pc.medicamento_id = ? AND pc.fabricante_id = ?", (prod['med_id'], prod['fab_id'])).fetchall()





























        # Crear dict de cotizaciones por competidor
        cot_dict = {cot['competidor']: cot['precio'] for cot in cotizaciones}

        productos_data.append({
            'nombre_completo': prod['nombre_completo'],
            'laboratorio': prod['laboratorio'],
            'mi_precio': prod['mi_precio'],
            'cotizaciones': cot_dict
        })

    db.close()

    # 4. Crear archivo Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Comparativa Precios"

    # 5. Encabezados
    headers = ['Nombre Completo', 'Laboratorio', 'Mi Precio'] + competidores_list
    ws.append(headers)

    # Estilo para encabezados
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # 6. Datos
    for prod in productos_data:
        row = [
            prod['nombre_completo'],
            prod['laboratorio'],
            prod['mi_precio']
        ]

        # Agregar precios de cada competidor (en orden)
        for comp_nombre in competidores_list:
            precio = prod['cotizaciones'].get(comp_nombre, None)
            row.append(precio if precio else '')

        ws.append(row)

    # 7. Ajustar anchos de columna
    ws.column_dimensions['A'].width = 50  # Nombre Completo
    ws.column_dimensions['B'].width = 20  # Laboratorio
    ws.column_dimensions['C'].width = 12  # Mi Precio

    for i, _ in enumerate(competidores_list, start=4):
        col_letter = ws.cell(row=1, column=i).column_letter
        ws.column_dimensions[col_letter].width = 15

    # 8. Formato de nmeros
    for row in ws.iter_rows(min_row=2, min_col=3, max_col=len(headers)):
        for cell in row:
            if cell.value and isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0'

    # 9. Guardar en memoria y enviar
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"comparativa_precios_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


@app.route('/admin/fabricantes/listar-con-stats')
def listar_fabricantes_con_stats():
    """
    Retorna todos los fabricantes con estadsticas de uso en BD
    """
    db = get_db_connection()

    fabricantes = db.execute("""
        SELECT
            f.id,
            f.nombre,
            (SELECT COUNT(*) FROM precios WHERE fabricante_id = f.id) as total_precios,
            (SELECT COUNT(*) FROM existencias WHERE fabricante_id = f.id) as total_existencias,
            (SELECT COUNT(*) FROM precios_competencia WHERE fabricante_id = f.id) as total_competencia
        FROM fabricantes f
        ORDER BY f.nombre
    """).fetchall()

    db.close()

    return jsonify({
        'ok': True,
        'fabricantes': [dict(f) for f in fabricantes]
    })


@app.route('/admin/fabricantes/estadisticas-fusion', methods=['POST'])
def estadisticas_fusion_fabricantes():
    """
    Retorna estadsticas de impacto de fusionar fabricantes
    """
    data = request.get_json()
    ids = data.get('ids', [])

    if not ids:
        return jsonify({'ok': False, 'error': 'IDs requeridos'})

    db = get_db_connection()
    placeholders = ','.join('?' * len(ids))

    total_precios = db.execute(f"""
        SELECT COUNT(*) as total FROM precios
        WHERE fabricante_id IN ({placeholders})
    """, ids).fetchone()['total']

    total_existencias = db.execute(f"""
        SELECT COUNT(*) as total FROM existencias
        WHERE fabricante_id IN ({placeholders})
    """, ids).fetchone()['total']

    total_competencia = db.execute(f"""
        SELECT COUNT(*) as total FROM precios_competencia
        WHERE fabricante_id IN ({placeholders})
    """, ids).fetchone()['total']

    db.close()

    return jsonify({
        'ok': True,
        'total_precios': total_precios,
        'total_existencias': total_existencias,
        'total_competencia': total_competencia
    })


@app.route('/admin/fabricantes/fusionar', methods=['POST'])
def fusionar_fabricantes():
    """
    Fusiona fabricantes duplicados en uno solo (maestro).
    Actualiza todas las referencias en BD y elimina duplicados.
    """
    data = request.get_json()
    maestro_id = data.get('maestro_id')
    duplicados_ids = data.get('duplicados_ids', [])

    if not maestro_id or not duplicados_ids:
        return jsonify({'ok': False, 'error': 'Datos incompletos'})

    # Verificar que el maestro no est en duplicados
    if maestro_id in duplicados_ids:
        return jsonify({'ok': False, 'error': 'El maestro no puede estar en la lista de duplicados'})

    db = get_db_connection()

    try:
        # Estadsticas para retornar
        precios_actualizados = 0
        existencias_actualizadas = 0
        competencia_actualizada = 0

        placeholders = ','.join('?' * len(duplicados_ids))

        # 1. Actualizar tabla PRECIOS
        # Primero verificar si hay conflictos (mismo medicamento_id + maestro_id ya existe)
        for dup_id in duplicados_ids:
            # Obtener precios del duplicado
            precios_dup = db.execute("""
                SELECT medicamento_id, precio, fecha_actualizacion, stock_fabricante, imagen
                FROM precios WHERE fabricante_id = ?
            """, (dup_id,)).fetchall()

            for precio in precios_dup:
                # Verificar si ya existe este medicamento con el fabricante maestro
                existe = db.execute("""
                    SELECT id FROM precios
                    WHERE medicamento_id = ? AND fabricante_id = ?
                """, (precio['medicamento_id'], maestro_id)).fetchone()

                if existe:
                    # Ya existe, actualizar si el duplicado tiene info ms reciente
                    db.execute("""
                        UPDATE precios
                        SET precio = CASE WHEN ? > fecha_actualizacion THEN ? ELSE precio END,
                            fecha_actualizacion = CASE WHEN ? > fecha_actualizacion THEN ? ELSE fecha_actualizacion END,
                            stock_fabricante = COALESCE(?, stock_fabricante),
                            imagen = COALESCE(?, imagen)
                        WHERE medicamento_id = ? AND fabricante_id = ?
                    """, (
                        precio['fecha_actualizacion'], precio['precio'],
                        precio['fecha_actualizacion'], precio['fecha_actualizacion'],
                        precio['stock_fabricante'], precio['imagen'],
                        precio['medicamento_id'], maestro_id
                    ))
                    precios_actualizados += 1
                else:
                    # No existe, cambiar el fabricante_id
                    db.execute("""
                        UPDATE precios
                        SET fabricante_id = ?
                        WHERE medicamento_id = ? AND fabricante_id = ?
                    """, (maestro_id, precio['medicamento_id'], dup_id))
                    precios_actualizados += 1

        # Eliminar precios duplicados que quedaron
        db.execute(f"""
            DELETE FROM precios
            WHERE fabricante_id IN ({placeholders})
        """, duplicados_ids)

        # 2. Actualizar tabla EXISTENCIAS
        db.execute(f"""
            UPDATE existencias
            SET fabricante_id = ?
            WHERE fabricante_id IN ({placeholders})
        """, [maestro_id] + duplicados_ids)
        existencias_actualizadas = db.total_changes

        # 3. Actualizar tabla PRECIOS_COMPETENCIA
        db.execute(f"""
            UPDATE precios_competencia
            SET fabricante_id = ?
            WHERE fabricante_id IN ({placeholders})
        """, [maestro_id] + duplicados_ids)
        competencia_actualizada = db.total_changes

        # 4. Eliminar fabricantes duplicados
        db.execute(f"""
            DELETE FROM fabricantes
            WHERE id IN ({placeholders})
        """, duplicados_ids)
        fabricantes_eliminados = len(duplicados_ids)

        db.commit()
        db.close()

        return jsonify({
            'ok': True,
            'precios_actualizados': precios_actualizados,
            'existencias_actualizadas': existencias_actualizadas,
            'competencia_actualizada': competencia_actualizada,
            'fabricantes_eliminados': fabricantes_eliminados
        })

    except Exception as e:
        db.rollback()
        db.close()
        return jsonify({'ok': False, 'error': str(e)})


@app.route('/admin/precio-sugerido/guardar', methods=['POST'])
def guardar_precio_sugerido():
    data = request.get_json()
    medicamento_id = data.get('medicamento_id')
    fabricante_id = data.get('fabricante_id')
    precio = data.get('precio')
    
    if not medicamento_id or not fabricante_id or not precio:
        return jsonify({'ok': False, 'error': 'Datos incompletos'})
    
    db = get_db_connection()
    
    # Actualizar precio
    db.execute(
        "UPDATE precios SET precio = ? WHERE medicamento_id = ? AND fabricante_id = ?",
        (float(precio), medicamento_id, fabricante_id)
    )
    
    db.commit()
    db.close()

    return jsonify({'ok': True})


@app.route('/admin/medicamento/editar', methods=['POST'])
def editar_medicamento():
    """Edita el nombre y/o fabricante de un medicamento existente"""
    data = request.get_json()
    medicamento_id = data.get('medicamento_id')
    fabricante_id_original = data.get('fabricante_id_original')
    nuevo_nombre = data.get('nuevo_nombre')
    nuevo_fabricante_id = data.get('nuevo_fabricante_id')

    if not medicamento_id or not fabricante_id_original or not nuevo_nombre or not nuevo_fabricante_id:
        return jsonify({'ok': False, 'error': 'Datos incompletos'})

    db = get_db_connection()

    try:
        # Verificar el nombre actual del medicamento
        medicamento_actual = db.execute("""
            SELECT nombre FROM medicamentos WHERE id = ?
        """, (medicamento_id,)).fetchone()

        # Solo actualizar nombre si realmente cambi
        if medicamento_actual and medicamento_actual['nombre'] != nuevo_nombre:
            # Verificar si ya existe OTRO medicamento con ese nombre
            nombre_duplicado = db.execute("""
                SELECT id FROM medicamentos
                WHERE nombre = ? AND id != ?
            """, (nuevo_nombre, medicamento_id)).fetchone()

            if nombre_duplicado:
                db.close()
                return jsonify({
                    'ok': False,
                    'error': f' Ya existe otro medicamento con el nombre:\n"{nuevo_nombre}"\n\nDebes usar un nombre diferente.'
                })

            # Si no hay duplicado, actualizar
            db.execute("""
                UPDATE medicamentos
                SET nombre = ?
                WHERE id = ?
            """, (nuevo_nombre, medicamento_id))

        # Si cambi el fabricante, necesitamos mover el registro de precios
        if str(fabricante_id_original) != str(nuevo_fabricante_id):
            # Verificar si ya existe precio para la nueva combinacin
            precio_existente = db.execute("""
                SELECT precio FROM precios
                WHERE medicamento_id = ? AND fabricante_id = ?
            """, (medicamento_id, nuevo_fabricante_id)).fetchone()

            if precio_existente:
                # Si ya existe, eliminar el precio antiguo
                db.execute("""
                    DELETE FROM precios
                    WHERE medicamento_id = ? AND fabricante_id = ?
                """, (medicamento_id, fabricante_id_original))
            else:
                # Si no existe, actualizar el fabricante_id
                db.execute("""
                    UPDATE precios
                    SET fabricante_id = ?
                    WHERE medicamento_id = ? AND fabricante_id = ?
                """, (nuevo_fabricante_id, medicamento_id, fabricante_id_original))

            # Mover cotizaciones de competencia
            db.execute("""
                UPDATE precios_competencia
                SET fabricante_id = ?
                WHERE medicamento_id = ? AND fabricante_id = ?
            """, (nuevo_fabricante_id, medicamento_id, fabricante_id_original))

        db.commit()
        db.close()
        return jsonify({'ok': True})

    except Exception as e:
        db.close()
        return jsonify({'ok': False, 'error': str(e)})


# ltima revisin: 2025-12-12 15:30
# Usado en: templates/precios_dinamicos.html (evento input en tercero-nombre)
# Busca terceros por nombre con normalizacin case-insensitive cuando usuario escribe
@app.route('/admin/terceros/buscar', methods=['GET'])
@admin_required
def buscar_terceros():
    query = request.args.get('q', '').strip()

    if not query or len(query) < 2:
        return jsonify({'terceros': []})

    db = get_db_connection()
    terceros = db.execute("""
        SELECT id, nombre, telefono, direccion, url_busqueda_base
        FROM terceros
        WHERE LOWER(nombre) LIKE LOWER(?)
        ORDER BY LOWER(nombre)
        LIMIT 10
    """, (f'%{query}%',)).fetchall()
    db.close()

    return jsonify({'terceros': [dict(t) for t in terceros]})


# ltima revisin: 2025-12-16
# Usado en: templates/precios_dinamicos.html (funcin cargarUltimosTerceros)
# Muestra ltimos 4 terceros distintos usados en cotizaciones
@app.route('/admin/terceros/ultimos', methods=['GET'])
@admin_required
def ultimos_terceros():
    """Obtiene los ltimos N terceros distintos usados en cotizaciones (precios_competencia)"""
    try:
        limit = request.args.get('limit', 4, type=int)

        db = get_db_connection()
        terceros = db.execute("""
            SELECT t.id, t.nombre, t.telefono, t.direccion, t.url_busqueda_base
            FROM terceros t
            INNER JOIN precios_competencia pc ON t.id = pc.competidor_id
            GROUP BY t.id, t.nombre, t.telefono, t.direccion, t.url_busqueda_base
            ORDER BY MAX(pc.fecha_actualizacion) DESC
            LIMIT ?
        """, (limit,)).fetchall()
        db.close()

        return jsonify({'terceros': [dict(t) for t in terceros]})
    except Exception as e:
        print(f" Error en /admin/terceros/ultimos: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e), 'terceros': []}), 500


@app.route('/admin/terceros/todos', methods=['GET'])
@admin_required
def todos_terceros():
    """Obtiene todos los terceros con sus URLs de bsqueda"""
    try:
        db = get_db_connection()
        # PostgreSQL usa LOWER() en lugar de COLLATE NOCASE
        terceros = db.execute("""
            SELECT id, nombre, telefono, direccion, url_busqueda_base
            FROM terceros
            ORDER BY LOWER(nombre) ASC
        """).fetchall()
        db.close()

        return jsonify({'terceros': [dict(t) for t in terceros]})
    except Exception as e:
        print(f" Error en /admin/terceros/todos: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e), 'terceros': []}), 500


@app.route('/admin/precio-competencia/editar', methods=['POST'])
def editar_precio_competencia():
    data = request.get_json()
    id_cotizacion = data.get('id')
    precio = data.get('precio')
    url = data.get('url')  # URL opcional

    if not id_cotizacion or not precio or float(precio) <= 0:
        return jsonify({'ok': False})

    db = get_db_connection()
    fecha_actual = datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')

    if url:
        db.execute("""
            UPDATE precios_competencia
            SET precio = ?, url = ?, fecha_actualizacion = ?
            WHERE id = ?
        """, (float(precio), url, fecha_actual, id_cotizacion))
    else:
        db.execute("""
            UPDATE precios_competencia
            SET precio = ?, fecha_actualizacion = ?
            WHERE id = ?
        """, (float(precio), fecha_actual, id_cotizacion))

    db.commit()
    db.close()

    return jsonify({'ok': True})

@app.route('/admin/precio-competencia/eliminar', methods=['POST'])
def eliminar_precio_competencia():
    data = request.get_json()
    id_cotizacion = data.get('id')

    if not id_cotizacion:
        return jsonify({'ok': False})

    db = get_db_connection()

    # Obtener medicamento_id y fabricante_id antes de eliminar
    cotizacion = db.execute(
        "SELECT medicamento_id, fabricante_id FROM precios_competencia WHERE id = ?",
        (id_cotizacion,)
    ).fetchone()

    if not cotizacion:
        db.close()
        return jsonify({'ok': False})

    med_id = cotizacion['medicamento_id']
    fab_id = cotizacion['fabricante_id']

    # Eliminar la cotizacin
    db.execute("DELETE FROM precios_competencia WHERE id = ?", (id_cotizacion,))
    db.commit()

    # Recalcular precio segn polticas
    precio_nuevo = calcular_precio_segun_politica(med_id, fab_id, db)

    if precio_nuevo is not None:
        # Actualizar precio en tabla precios
        db.execute("""
            UPDATE precios
            SET precio = ?
            WHERE medicamento_id = ? AND fabricante_id = ?
        """, (precio_nuevo, med_id, fab_id))
        db.commit()

    # Obtener contadores actualizados
    cotizaciones_con_url = db.execute(
        "SELECT COUNT(*) as count FROM precios_competencia WHERE medicamento_id = ? AND fabricante_id = ? AND url IS NOT NULL AND url != ''",
        (med_id, fab_id)
    ).fetchone()['count']

    cotizaciones_total = db.execute(
        "SELECT COUNT(*) as count FROM precios_competencia WHERE medicamento_id = ? AND fabricante_id = ?",
        (med_id, fab_id)
    ).fetchone()['count']

    db.close()

    return jsonify({
        'ok': True,
        'precio_nuevo': precio_nuevo,
        'cotizaciones_con_url': cotizaciones_con_url,
        'cotizaciones_total': cotizaciones_total
    })


@app.route('/admin/imagen-producto/obtener', methods=['GET'])
def obtener_imagen_producto():
    """Obtiene la imagen de un producto (medicamento + fabricante) desde la tabla precios"""
    medicamento_id = request.args.get('medicamento_id')
    fabricante_id = request.args.get('fabricante_id')

    if not medicamento_id or not fabricante_id:
        return jsonify({'imagen_url': None})

    db = get_db_connection()
    precio = db.execute("""
        SELECT imagen
        FROM precios
        WHERE medicamento_id = ? AND fabricante_id = ?
    """, (medicamento_id, fabricante_id)).fetchone()
    db.close()

    if precio and precio['imagen']:
        imagen = precio['imagen']
        # Si es base64, devolver directamente
        if imagen.startswith('data:image'):
            return jsonify({'imagen_url': imagen})
        # Si es nombre de archivo, construir la ruta
        else:
            return jsonify({'imagen_url': f'/static/uploads/{imagen}'})

    return jsonify({'imagen_url': None})


@app.route('/admin/imagen-producto/guardar', methods=['POST'])
def guardar_imagen_producto():
    """Guarda la imagen de un producto (medicamento + fabricante) en base64 en la tabla precios"""
    data = request.get_json()
    medicamento_id = data.get('medicamento_id')
    fabricante_id = data.get('fabricante_id')
    imagen_base64 = data.get('imagen_base64')

    print(f" Guardando imagen - Med ID: {medicamento_id}, Fab ID: {fabricante_id}, Imagen length: {len(imagen_base64) if imagen_base64 else 0}")

    if not medicamento_id or not fabricante_id or not imagen_base64:
        print(f" Faltan datos - Med: {bool(medicamento_id)}, Fab: {bool(fabricante_id)}, Img: {bool(imagen_base64)}")
        return jsonify({'ok': False, 'error': 'Faltan datos requeridos'})

    try:
        db = get_db_connection()
        # Verificar si existe el registro en precios
        exists = db.execute("""
            SELECT id FROM precios
            WHERE medicamento_id = ? AND fabricante_id = ?
        """, (medicamento_id, fabricante_id)).fetchone()

        if exists:
            print(f" Registro existe, actualizando imagen...")
            # Actualizar imagen
            db.execute("""
                UPDATE precios
                SET imagen = ?
                WHERE medicamento_id = ? AND fabricante_id = ?
            """, (imagen_base64, medicamento_id, fabricante_id))
        else:
            print(f" No existe registro de precio para Med {medicamento_id} + Fab {fabricante_id}")
            return jsonify({'ok': False, 'error': 'No existe registro de precio para este producto'})

        db.commit()
        db.close()

        print(f" Imagen guardada exitosamente")
        return jsonify({'ok': True})
    except Exception as e:
        print(f" Error al guardar imagen: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'ok': False, 'error': str(e)})


# --- FIN RUTAS PARA PRecios actualizador con listado ---



# ============================================
# RUTAS PARA GESTIN DE NAVEGACIN
# ============================================

@app.route('/admin/navegacion')
def admin_navegacion():
    """Pgina del administrador de navegacin"""
    return render_template('navegacion_admin.html')

@app.route('/api/navegacion', methods=['GET'])
def get_navegacion():
    """Obtener todas las URLs del men"""
    conn = get_db_connection()
    navegacion = conn.execute('''
        SELECT id, nombre_corto, url, descripcion, activo, orden,
               fecha_creacion, fecha_actualizacion, fecha_ultimo_uso
        FROM NAVEGACION_MENU
        ORDER BY orden ASC, nombre_corto ASC
    ''').fetchall()
    conn.close()

    return jsonify([dict(row) for row in navegacion])

@app.route('/api/navegacion/activas', methods=['GET'])
def get_navegacion_activas():
    """Obtener solo las URLs activas"""
    conn = get_db_connection()
    navegacion = conn.execute('''
        SELECT id, nombre_corto, url, descripcion, orden
        FROM NAVEGACION_MENU
        WHERE activo = '1'
        ORDER BY orden ASC, nombre_corto ASC
    ''').fetchall()
    conn.close()
    
    return jsonify([dict(row) for row in navegacion])

@app.route('/api/navegacion', methods=['POST'])
def crear_navegacion():
    """Crear una nueva entrada de navegacin"""
    try:
        data = request.get_json()

        nombre_corto = data.get('nombre_corto', '').strip()
        url = data.get('url', '').strip()
        descripcion = data.get('descripcion', '').strip()
        orden = data.get('orden', 0)

        # Validaciones
        if not nombre_corto:
            return jsonify({'error': 'El nombre corto es obligatorio'}), 400

        if len(nombre_corto.split()) > 3:
            return jsonify({'error': 'El nombre debe tener mximo 3 palabras'}), 400

        if not url:
            return jsonify({'error': 'La URL es obligatoria'}), 400

        # Asegurar que la URL comience con /
        if not url.startswith('http://') and not url.startswith('https://') and not url.startswith('/'):
            url = '/' + url

        conn = get_db_connection()

        # Obtener el prximo ID
        cursor = conn.execute('SELECT COALESCE(MAX(id), 0) + 1 FROM NAVEGACION_MENU')
        nueva_id = cursor.fetchone()[0]

        conn.execute('''
            INSERT INTO NAVEGACION_MENU (id, nombre_corto, url, descripcion, orden)
            VALUES (?, ?, ?, ?, ?)
        ''', (nueva_id, nombre_corto, url, descripcion, orden))

        conn.commit()
        conn.close()

        print(f" URL guardada: {nombre_corto} -> {url}")

        return jsonify({
            'success': True,
            'id': nueva_id,
            'mensaje': 'URL agregada correctamente'
        }), 201
    except Exception as e:
        print(f" Error al guardar URL: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Error al guardar: {str(e)}'}), 500

@app.route('/api/navegacion/<int:id>', methods=['PUT'])
def actualizar_navegacion(id):
    """Actualizar una entrada de navegacin"""
    data = request.get_json()
    
    nombre_corto = data.get('nombre_corto', '').strip()
    url = data.get('url', '').strip()
    descripcion = data.get('descripcion', '').strip()
    activo = data.get('activo', 1)
    orden = data.get('orden', 0)
    
    # Validaciones
    if nombre_corto and len(nombre_corto.split()) > 3:
        return jsonify({'error': 'El nombre debe tener mximo 3 palabras'}), 400
    
    if url and not url.startswith('http://') and not url.startswith('https://') and not url.startswith('/'):
        url = '/' + url
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    cursor.execute('''
        UPDATE NAVEGACION_MENU
        SET nombre_corto = ?,
            url = ?,
            descripcion = ?,
            activo = ?,
            orden = ?,
            fecha_actualizacion = datetime('now', 'localtime')
        WHERE id = ?
    ''', (nombre_corto, url, descripcion, activo, orden, id))
    
    conn.commit()
    conn.close()
    
    return jsonify({
        'success': True,
        'mensaje': 'URL actualizada correctamente'
    })

@app.route('/api/navegacion/<int:id>', methods=['DELETE'])
def eliminar_navegacion(id):
    """Eliminar una entrada de navegacin"""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    cursor.execute('DELETE FROM NAVEGACION_MENU WHERE id = ?', (id,))
    
    conn.commit()
    conn.close()
    
    return jsonify({
        'success': True,
        'mensaje': 'URL eliminada correctamente'
    })

@app.route('/api/navegacion/<int:id>/toggle', methods=['PUT'])
def toggle_navegacion(id):
    """Activar/desactivar una entrada de navegacin"""
    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute('''
        UPDATE NAVEGACION_MENU
        SET activo = NOT activo,
            fecha_actualizacion = datetime('now', 'localtime')
        WHERE id = ?
    ''', (id,))

    conn.commit()
    conn.close()

    return jsonify({
        'success': True,
        'mensaje': 'Estado actualizado correctamente'
    })

@app.route('/api/navegacion/<int:id>/uso', methods=['PUT'])
def registrar_uso_navegacion(id):
    """Registrar uso de una entrada de navegacin"""
    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute('''
        UPDATE NAVEGACION_MENU
        SET fecha_ultimo_uso = datetime('now', 'localtime')
        WHERE id = ?
    ''', (id,))

    conn.commit()
    conn.close()

    return jsonify({
        'success': True,
        'mensaje': 'Uso registrado correctamente'
    })

# -------------------------------------------------------------------
#  INICIA PASTILLERO ---
# -------------------------------------------------------------------

# ============================================
#  RUTAS API DEL PASTILLERO
# ============================================

def obtener_pastillero_activo(usuario_id):
    """
    Obtener el pastillero activo del usuario.
    Prioridad:
    1. Pastillero guardado en sesión (pastillero_activo_id)
    2. Primer pastillero donde es propietario o miembro
    """
    # Si hay un pastillero guardado en sesión, usarlo
    if 'pastillero_activo_id' in session:
        pastillero_id = session['pastillero_activo_id']

        # Verificar que el usuario todavía tiene acceso
        db = get_db_connection()
        acceso = db.execute('''
            SELECT id FROM relaciones_pastillero
            WHERE pastillero_id = %s AND usuario_id = %s
        ''', (pastillero_id, usuario_id)).fetchone()
        db.close()

        if acceso:
            return pastillero_id
        else:
            # Ya no tiene acceso, eliminar de sesión
            del session['pastillero_activo_id']

    # Obtener el primer pastillero donde es propietario o miembro
    db = get_db_connection()
    pastillero = db.execute('''
        SELECT p.id, p.nombre, rp.tipo
        FROM pastilleros p
        INNER JOIN relaciones_pastillero rp ON p.id = rp.pastillero_id
        WHERE rp.usuario_id = %s
        AND rp.tipo IN ('propietario', 'miembro')
        ORDER BY rp.tipo DESC, p.id ASC
        LIMIT 1
    ''', (usuario_id,)).fetchone()

    db.close()

    if pastillero:
        return pastillero['id']

    return None


@app.route('/api/crear-usuario-pastillero', methods=['POST'])
def api_crear_usuario_pastillero():
    """Crear usuario rápido para acceder al pastillero sin checkout"""
    data = request.get_json()

    nombre = data.get('nombre', '').strip()
    telefono = data.get('telefono', '').strip()

    if not nombre or not telefono:
        return jsonify({'ok': False, 'error': 'Nombre y teléfono son requeridos'}), 400

    try:
        conn = get_db_connection()

        # Verificar si ya existe el teléfono
        tercero_existente = conn.execute("""
            SELECT id, nombre FROM terceros WHERE telefono = ? LIMIT 1
        """, (telefono,)).fetchone()

        if tercero_existente:
            tercero_id = tercero_existente['id']
            print(f"✅ Tercero existente encontrado: ID {tercero_id} ({tercero_existente['nombre']})")

            # Actualizar nombre si cambió
            if tercero_existente['nombre'] != nombre:
                conn.execute("""
                    UPDATE terceros
                    SET nombre = ?, fecha_actualizacion = CURRENT_TIMESTAMP
                    WHERE id = ?
                """, (nombre, tercero_id))
                conn.commit()
                print(f"  ✏️ Nombre actualizado a: {nombre}")
        else:
            # Crear nuevo tercero
            cursor_seq = conn.execute("SELECT COALESCE(MAX(id), 0) + 1 FROM terceros")
            next_tercero_id = cursor_seq.fetchone()[0]

            print(f"➕ Creando nuevo tercero con ID {next_tercero_id}...")
            conn.execute("""
                INSERT INTO terceros (id, nombre, telefono, fecha_creacion)
                VALUES (?, ?, ?, CURRENT_TIMESTAMP)
            """, (next_tercero_id, nombre, telefono))
            tercero_id = next_tercero_id
            conn.commit()
            print(f"✅ Tercero creado con ID: {tercero_id}")

        conn.close()

        # Establecer sesión
        session['usuario_id'] = tercero_id
        print(f"🔐 Sesión establecida para usuario_id: {tercero_id}")

        return jsonify({
            'ok': True,
            'usuario_id': tercero_id,
            'nombre': nombre,
            'telefono': telefono
        })

    except Exception as e:
        print(f"❌ Error al crear usuario para pastillero: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'ok': False, 'error': str(e)}), 500


@app.route('/api/restaurar-sesion', methods=['POST'])
def api_restaurar_sesion():
    """Restaurar sesión del usuario desde localStorage"""
    data = request.get_json()
    usuario_id = data.get('usuario_id')

    if not usuario_id:
        return jsonify({'ok': False, 'error': 'usuario_id requerido'}), 400

    try:
        conn = get_db_connection()

        # Verificar que el usuario existe
        tercero = conn.execute("""
            SELECT id, nombre, telefono FROM terceros WHERE id = ? LIMIT 1
        """, (usuario_id,)).fetchone()

        conn.close()

        if tercero:
            # Restaurar sesión
            session['usuario_id'] = tercero['id']
            print(f"🔄 Sesión restaurada para usuario_id: {tercero['id']} ({tercero['nombre']})")

            return jsonify({
                'ok': True,
                'usuario_id': tercero['id'],
                'nombre': tercero['nombre'],
                'telefono': tercero['telefono']
            })
        else:
            return jsonify({'ok': False, 'error': 'Usuario no encontrado'}), 404

    except Exception as e:
        print(f"❌ Error al restaurar sesión: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'ok': False, 'error': str(e)}), 500


@app.route('/api/pastillero', methods=['GET', 'POST'])
def api_pastillero():
    """Obtener todos los medicamentos del pastillero del usuario"""
    if 'usuario_id' not in session:
        return jsonify({'ok': False, 'error': 'No autenticado'}), 401

    usuario_id = session['usuario_id']

    # Obtener pastillero activo del usuario
    pastillero_id = obtener_pastillero_activo(usuario_id)
    if not pastillero_id:
        return jsonify({'ok': True, 'medicamentos': []})  # Usuario sin pastillero

    #  Aceptar parmetros por GET o POST
    if request.method == 'POST':
        data = request.get_json() or {}
        orden = data.get('orden', 'recientes')
        sintomas_filtro = data.get('sintomas')
    else:  # GET
        orden = request.args.get('orden', 'recientes')
        sintomas_filtro = request.args.get('sintomas')

    try:
        conn = get_db_connection()

        # Determinar ORDER BY segn el parmetro
        if orden == 'alfabetico':
            # PostgreSQL: usar LOWER() en lugar de COLLATE NOCASE
            order_by = 'ORDER BY LOWER(p.nombre) ASC'
        else:  # recientes (por defecto)
            order_by = 'ORDER BY p.fecha_actualizado DESC'

        #  Construir filtro WHERE para sntomas
        where_sintomas = ""
        params = [pastillero_id]
        
        if sintomas_filtro:
            # Convertir string de sntomas separados por coma en lista
            sintomas_list = [normalizar_texto(s.strip()) for s in sintomas_filtro.split(',')]
            
            # Crear condiciones OR para cada sntoma
            sintomas_conditions = ' OR '.join([
                "LOWER(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(s.nombre, '', 'a'), '', 'e'), '', 'i'), '', 'o'), '', 'u'), '', 'n')) LIKE ?" 
                for _ in sintomas_list
            ])
            where_sintomas = f'AND ({sintomas_conditions})'
            
            # Agregar parmetros para cada sntoma
            for sintoma in sintomas_list:
                params.append(f'%{sintoma}%')
        
        query = f'''
            SELECT
                p.id,
                p.nombre,
                p.cantidad,
                p.unidad,
                p.medicamento_id,
                p.horas_entre_tomas,
                p.proxima_toma,
                p.recordatorio_activo,
                STRING_AGG(DISTINCT s.nombre, ',') as "sintomas_str"
            FROM pastillero_usuarios p
            LEFT JOIN medicamentos m ON p.medicamento_id = m.id
            LEFT JOIN medicamento_sintoma ms ON m.id = ms.medicamento_id
            LEFT JOIN sintomas s ON ms.sintoma_id = s.id
            WHERE p.pastillero_id = %s
            {where_sintomas}
            GROUP BY p.id, p.nombre, p.cantidad, p.unidad, p.medicamento_id, p.horas_entre_tomas, p.proxima_toma, p.recordatorio_activo
            {order_by}
        '''

        medicamentos = conn.execute(query, params).fetchall()

        medicamentos_list = []
        for med in medicamentos:
            medicamentos_list.append({
                'id': med['id'],
                'nombre': med['nombre'],
                'cantidad': med['cantidad'],
                'unidad': med['unidad'],
                'medicamento_id': med['medicamento_id'],
                'sintomas': med['sintomas_str'].split(',') if med['sintomas_str'] else [],
                'horas_entre_tomas': med['horas_entre_tomas'],
                'proxima_toma': med['proxima_toma'].isoformat() if med['proxima_toma'] else None,
                'recordatorio_activo': med['recordatorio_activo'] or False
            })
        
        conn.close()
        return jsonify({'ok': True, 'medicamentos': medicamentos_list})
    except Exception as e:
        print(f"Error al obtener pastillero: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'ok': False, 'error': str(e)}), 500


@app.route('/api/obtener-datos-usuario-sesion', methods=['GET'])
def api_obtener_datos_usuario_sesion():
    """Obtener datos del usuario desde la sesión activa"""
    if 'usuario_id' not in session:
        return jsonify({'ok': False, 'error': 'No hay sesión activa'})

    usuario_id = session['usuario_id']
    nombre = session.get('nombre', '')
    telefono = session.get('telefono', '')

    return jsonify({
        'ok': True,
        'usuario_id': usuario_id,
        'nombre': nombre,
        'telefono': telefono
    })


@app.route('/api/generar-token-vinculacion', methods=['POST'])
def api_generar_token_vinculacion():
    """Generar token único para vincular otro dispositivo"""
    if 'usuario_id' not in session:
        return jsonify({'ok': False, 'error': 'No hay sesión activa'}), 401

    usuario_id = session['usuario_id']
    nombre = session.get('nombre', '')
    telefono = session.get('telefono', '')

    try:
        # Generar token único
        token = str(uuid.uuid4())

        # Guardar en BD
        db = get_db_connection()

        # Limpiar tokens antiguos (más de 10 minutos)
        db.execute('''
            DELETE FROM tokens_vinculacion
            WHERE fecha_creacion < NOW() - INTERVAL '10 minutes'
        ''')

        # Insertar nuevo token
        db.execute('''
            INSERT INTO tokens_vinculacion (token, usuario_id, nombre, telefono)
            VALUES (%s, %s, %s, %s)
        ''', (token, usuario_id, nombre, telefono))

        db.commit()
        db.close()

        # Generar URL completa
        url_vinculacion = f"https://tuc-tuc.onrender.com/vincular/{token}"

        return jsonify({
            'ok': True,
            'token': token,
            'url': url_vinculacion
        })

    except Exception as e:
        print(f"Error al generar token: {e}")
        return jsonify({'ok': False, 'error': str(e)}), 500


@app.route('/vincular/<token>')
def vincular_dispositivo(token):
    """Página para vincular dispositivo mediante token QR"""
    return render_template('tienda_home.html', token_vinculacion=token)


@app.route('/api/validar-token-vinculacion/<token>', methods=['POST'])
def api_validar_token_vinculacion(token):
    """Validar token y crear sesión en nuevo dispositivo"""
    try:
        db = get_db_connection()

        # Buscar token
        token_data = db.execute('''
            SELECT usuario_id, nombre, telefono, usado, fecha_creacion
            FROM tokens_vinculacion
            WHERE token = %s
        ''', (token,)).fetchone()

        if not token_data:
            db.close()
            return jsonify({'ok': False, 'error': 'Token inválido'}), 404

        # Verificar que no esté usado
        if token_data['usado']:
            db.close()
            return jsonify({'ok': False, 'error': 'Este token ya fue utilizado'}), 400

        # Verificar que no haya expirado (5 minutos)
        fecha_creacion = token_data['fecha_creacion']
        if isinstance(fecha_creacion, str):
            from dateutil import parser
            fecha_creacion = parser.parse(fecha_creacion)

        tiempo_transcurrido = datetime.utcnow() - fecha_creacion.replace(tzinfo=None)
        if tiempo_transcurrido.total_seconds() > 300:  # 5 minutos
            db.close()
            return jsonify({'ok': False, 'error': 'El token ha expirado. Genera uno nuevo.'}), 400

        # Marcar token como usado
        db.execute('''
            UPDATE tokens_vinculacion
            SET usado = TRUE, fecha_uso = NOW()
            WHERE token = %s
        ''', (token,))

        db.commit()
        db.close()

        # Crear sesión
        session['usuario_id'] = token_data['usuario_id']
        session['nombre'] = token_data['nombre']
        session['telefono'] = token_data['telefono']
        session.permanent = True

        return jsonify({
            'ok': True,
            'usuario_id': token_data['usuario_id'],
            'nombre': token_data['nombre'],
            'telefono': token_data['telefono']
        })

    except Exception as e:
        print(f"Error al validar token: {e}")
        traceback.print_exc()
        return jsonify({'ok': False, 'error': str(e)}), 500


@app.route('/api/pastillero/count', methods=['GET'])
def api_pastillero_count():
    """Obtener cantidad de medicamentos en el pastillero"""
    if 'usuario_id' not in session:
        return jsonify({'ok': True, 'count': 0, 'nombre': 'Mi pastillero'})

    usuario_id = session['usuario_id']

    # Obtener pastillero activo del usuario
    pastillero_id = obtener_pastillero_activo(usuario_id)
    if not pastillero_id:
        return jsonify({'ok': True, 'count': 0, 'nombre': 'Mi pastillero'})

    try:
        conn = get_db_connection()

        # Obtener count y nombre del pastillero
        result = conn.execute('''
            SELECT
                COUNT(pu.id) as count,
                p.nombre as nombre_pastillero
            FROM pastilleros p
            LEFT JOIN pastillero_usuarios pu ON p.id = pu.pastillero_id
            WHERE p.id = %s
            GROUP BY p.id, p.nombre
        ''', (pastillero_id,)).fetchone()

        # Si no hay resultado, obtener solo el nombre del pastillero
        if not result:
            nombre = conn.execute('''
                SELECT nombre FROM pastilleros WHERE id = %s
            ''', (pastillero_id,)).fetchone()

            conn.close()
            return jsonify({
                'ok': True,
                'count': 0,
                'nombre': nombre['nombre'] if nombre else 'Mi pastillero'
            })

        conn.close()
        return jsonify({
            'ok': True,
            'count': result['count'],
            'nombre': result['nombre_pastillero']
        })
    except Exception as e:
        print(f"Error al contar pastillero: {e}")
        return jsonify({'ok': True, 'count': 0, 'nombre': 'Mi pastillero'})


@app.route('/api/pastillero/verificar/<medicamento_id>', methods=['GET'])
def api_pastillero_verificar(medicamento_id):
    """Verificar si un medicamento ya existe en el pastillero"""
    if 'usuario_id' not in session:
        return jsonify({'ok': False, 'existe': False})

    usuario_id = session['usuario_id']

    pastillero_id = obtener_pastillero_activo(usuario_id)
    if not pastillero_id:
        return jsonify({'ok': False, 'error': 'No tienes un pastillero activo'}), 400

    try:
        conn = get_db_connection()

        # Si medicamento_id es 'null', buscar por nombre (parmetro adicional necesario)
        # Por ahora solo buscar por ID si no es null
        if medicamento_id == 'null':
            # No podemos verificar sin el nombre, retornar que no existe
            conn.close()
            return jsonify({'ok': True, 'existe': False})

        medicamento_id = int(medicamento_id)

        resultado = conn.execute('''
            SELECT cantidad, unidad, nombre
            FROM pastillero_usuarios
            WHERE pastillero_id = ? AND medicamento_id = ?
        ''', (pastillero_id, medicamento_id)).fetchone()

        conn.close()

        if resultado:
            return jsonify({
                'ok': True,
                'existe': True,
                'cantidad': resultado['cantidad'],
                'unidad': resultado['unidad'],
                'nombre': resultado['nombre']
            })
        else:
            return jsonify({'ok': True, 'existe': False})
    except Exception as e:
        print(f"Error al verificar medicamento: {e}")
        return jsonify({'ok': False, 'existe': False})


@app.route('/api/pastillero/agregar', methods=['POST'])
def api_pastillero_agregar():
    """Agregar un medicamento existente al pastillero"""
    if 'usuario_id' not in session:
        return jsonify({'ok': False, 'error': 'No autenticado'}), 401

    usuario_id = session['usuario_id']

    # Obtener pastillero activo del usuario
    pastillero_id = obtener_pastillero_activo(usuario_id)
    if not pastillero_id:
        return jsonify({'ok': False, 'error': 'No tienes un pastillero activo'}), 400

    data = request.get_json()

    medicamento_id = data.get('medicamento_id')
    cantidad = data.get('cantidad', 1)
    unidad = data.get('unidad', 'pastillas')
    nombre = data.get('nombre')  #  Para medicamentos personales

    # medicamento_id puede ser None/null para medicamentos personales

    conn = get_db_connection()
    try:
        #  CASO 1: Medicamento personal (NULL) - buscar por nombre
        if medicamento_id is None or str(medicamento_id).lower() == 'null':
            if not nombre:
                conn.close()
                return jsonify({'ok': False, 'error': 'Falta nombre del medicamento'}), 400

            # Buscar si ya existe este medicamento personal por nombre
            existe = conn.execute('''
                SELECT id, cantidad FROM pastillero_usuarios
                WHERE pastillero_id = %s AND medicamento_id IS NULL AND nombre = %s
            ''', (pastillero_id, nombre)).fetchone()

            if existe:
                # Actualizar cantidad del existente
                conn.execute('''
                    UPDATE pastillero_usuarios
                    SET cantidad = cantidad + %s
                    WHERE id = %s
                ''', (cantidad, existe['id']))
            else:
                # Crear nuevo medicamento personal
                conn.execute('''
                    INSERT INTO pastillero_usuarios (pastillero_id, medicamento_id, nombre, cantidad, unidad)
                    VALUES (%s, NULL, %s, %s, %s)
                ''', (pastillero_id, nombre, cantidad, unidad))

        #  CASO 2: Medicamento oficial (con ID)
        else:
            medicamento_id = int(medicamento_id)

            # Verificar si ya existe en el pastillero (por ID o por nombre)
            existe = conn.execute('''
                SELECT id, cantidad FROM pastillero_usuarios
                WHERE pastillero_id = %s AND medicamento_id = %s
            ''', (pastillero_id, medicamento_id)).fetchone()

            if existe:
                # Actualizar cantidad
                conn.execute('''
                    UPDATE pastillero_usuarios
                    SET cantidad = cantidad + %s
                    WHERE id = %s
                ''', (cantidad, existe['id']))
            else:
                # Obtener nombre del medicamento de la BD
                medicamento = conn.execute('SELECT nombre FROM medicamentos WHERE id = %s', (medicamento_id,)).fetchone()

                if not medicamento:
                    conn.close()
                    return jsonify({'ok': False, 'error': 'Medicamento no encontrado'}), 404

                #  Extraer nombre base normalizado (sin presentacin)
                nombre_normalizado = extraer_nombre_base(medicamento['nombre'])

                #  Verificar si existe un medicamento personal con el mismo nombre (vincular)
                personal_existente = conn.execute('''
                    SELECT id, cantidad FROM pastillero_usuarios
                    WHERE pastillero_id = %s AND medicamento_id IS NULL AND nombre = %s
                ''', (pastillero_id, nombre_normalizado)).fetchone()

                if personal_existente:
                    # Vincular el medicamento personal con el ID oficial y sumar cantidad
                    conn.execute('''
                        UPDATE pastillero_usuarios
                        SET medicamento_id = %s, cantidad = cantidad + %s
                        WHERE id = %s
                    ''', (medicamento_id, cantidad, personal_existente['id']))
                else:
                    # Insertar nuevo con nombre normalizado
                    conn.execute('''
                        INSERT INTO pastillero_usuarios (pastillero_id, medicamento_id, nombre, cantidad, unidad)
                        VALUES (%s, %s, %s, %s, %s)
                    ''', (pastillero_id, medicamento_id, nombre_normalizado, cantidad, unidad))
        
        conn.commit()
        conn.close()
        return jsonify({'ok': True})
    except Exception as e:
        conn.rollback()
        conn.close()
        print(f"Error al agregar al pastillero: {e}")
        return jsonify({'ok': False, 'error': str(e)}), 500


@app.route('/api/pastillero/crear', methods=['POST'])
def api_pastillero_crear():
    """Crear un medicamento nuevo en el pastillero (no est en la BD)"""
    if 'usuario_id' not in session:
        return jsonify({'ok': False, 'error': 'No autenticado'}), 401

    usuario_id = session['usuario_id']

    pastillero_id = obtener_pastillero_activo(usuario_id)
    if not pastillero_id:
        return jsonify({'ok': False, 'error': 'No tienes un pastillero activo'}), 400

    data = request.get_json()

    nombre = data.get('nombre', '').strip()
    cantidad = data.get('cantidad', 1)
    unidad = data.get('unidad', 'pastillas')

    if not nombre:
        return jsonify({'ok': False, 'error': 'Falta nombre del medicamento'}), 400

    conn = get_db_connection()
    try:
        # Insertar medicamento sin medicamento_id (ser NULL)
        conn.execute('''
            INSERT INTO pastillero_usuarios (pastillero_id, medicamento_id, nombre, cantidad, unidad)
            VALUES (%s, NULL, %s, %s, %s)
        ''', (pastillero_id, nombre, cantidad, unidad))

        #  Crear alerta para el admin
        conn.execute('''
            INSERT INTO alertas_admin (tipo, mensaje, usuario_id, fecha_creacion)
            VALUES ('medicamento_faltante', %s, %s, CURRENT_TIMESTAMP)
        ''', (f'Usuario agreg medicamento no registrado: {nombre}', usuario_id))

        conn.commit()
        conn.close()
        return jsonify({'ok': True})
    except Exception as e:
        conn.rollback()
        conn.close()
        print(f"Error al crear medicamento en pastillero: {e}")
        return jsonify({'ok': False, 'error': str(e)}), 500


@app.route('/api/pastillero/<int:medicamento_id>/tomar', methods=['POST'])
def api_pastillero_tomar(medicamento_id):
    """Restar 1 unidad (usuario tom el medicamento)"""
    if 'usuario_id' not in session:
        return jsonify({'ok': False, 'error': 'No autenticado'}), 401

    usuario_id = session['usuario_id']

    pastillero_id = obtener_pastillero_activo(usuario_id)
    if not pastillero_id:
        return jsonify({'ok': False, 'error': 'No tienes un pastillero activo'}), 400

    conn = get_db_connection()
    try:
        # Verificar que el medicamento pertenece al pastillero
        medicamento = conn.execute('''
            SELECT cantidad, unidad FROM pastillero_usuarios
            WHERE id = ? AND pastillero_id = ?
        ''', (medicamento_id, pastillero_id)).fetchone()

        if not medicamento:
            conn.close()
            return jsonify({'ok': False, 'error': 'Medicamento no encontrado'}), 404

        nueva_cantidad = medicamento['cantidad'] - 1

        if nueva_cantidad <= 0:
            # Eliminar si llega a 0
            conn.execute('DELETE FROM pastillero_usuarios WHERE id = ?', (medicamento_id,))
        else:
            # Actualizar cantidad
            conn.execute('''
                UPDATE pastillero_usuarios
                SET cantidad = ?
                WHERE id = ?
            ''', (nueva_cantidad, medicamento_id))

        conn.commit()
        conn.close()
        return jsonify({
            'ok': True,
            'nueva_cantidad': max(0, nueva_cantidad),
            'unidad': medicamento['unidad']
        })
    except Exception as e:
        conn.rollback()
        conn.close()
        print(f"Error al tomar medicamento: {e}")
        return jsonify({'ok': False, 'error': str(e)}), 500


@app.route('/api/pastillero/<int:medicamento_id>/agregar', methods=['POST'])
def api_pastillero_agregar_cantidad(medicamento_id):
    """Agregar cantidad a un medicamento existente en el pastillero"""
    if 'usuario_id' not in session:
        return jsonify({'ok': False, 'error': 'No autenticado'}), 401

    usuario_id = session['usuario_id']

    pastillero_id = obtener_pastillero_activo(usuario_id)
    if not pastillero_id:
        return jsonify({'ok': False, 'error': 'No tienes un pastillero activo'}), 400

    data = request.get_json()
    cantidad = data.get('cantidad', 1)

    conn = get_db_connection()
    try:
        # Verificar que el medicamento pertenece al pastillero
        medicamento = conn.execute('''
            SELECT cantidad, unidad FROM pastillero_usuarios
            WHERE id = ? AND pastillero_id = ?
        ''', (medicamento_id, pastillero_id)).fetchone()

        if not medicamento:
            conn.close()
            return jsonify({'ok': False, 'error': 'Medicamento no encontrado'}), 404

        nueva_cantidad = medicamento['cantidad'] + cantidad

        # Actualizar cantidad
        conn.execute('''
            UPDATE pastillero_usuarios
            SET cantidad = cantidad + ?
            WHERE id = ?
        ''', (cantidad, medicamento_id))

        conn.commit()
        conn.close()
        return jsonify({
            'ok': True,
            'nueva_cantidad': nueva_cantidad,
            'unidad': medicamento['unidad']
        })
    except Exception as e:
        conn.rollback()
        conn.close()
        print(f"Error al agregar cantidad: {e}")
        return jsonify({'ok': False, 'error': str(e)}), 500


# ============================================
#  ENDPOINTS DE RECORDATORIOS
# ============================================

@app.route('/api/pastillero/<int:medicamento_id>/activar-recordatorio', methods=['POST'])
def api_activar_recordatorio(medicamento_id):
    """Activar recordatorio para un medicamento del pastillero"""
    if 'usuario_id' not in session:
        return jsonify({'ok': False, 'error': 'No autenticado'}), 401

    usuario_id = session['usuario_id']

    pastillero_id = obtener_pastillero_activo(usuario_id)
    if not pastillero_id:
        return jsonify({'ok': False, 'error': 'No tienes un pastillero activo'}), 400

    data = request.get_json()

    horas_entre_tomas = data.get('horas_entre_tomas')
    proxima_toma = data.get('proxima_toma')  # ISO format: "2025-12-23T20:00:00"

    if not horas_entre_tomas or not proxima_toma:
        return jsonify({'ok': False, 'error': 'Faltan datos requeridos'}), 400

    try:
        conn = get_db_connection()

        # Verificar que el medicamento pertenece al pastillero
        medicamento = conn.execute('''
            SELECT id, nombre FROM pastillero_usuarios
            WHERE id = %s AND pastillero_id = %s
        ''', (medicamento_id, pastillero_id)).fetchone()

        if not medicamento:
            conn.close()
            return jsonify({'ok': False, 'error': 'Medicamento no encontrado'}), 404

        # Actualizar recordatorio
        conn.execute('''
            UPDATE pastillero_usuarios
            SET horas_entre_tomas = %s,
                proxima_toma = %s,
                recordatorio_activo = TRUE
            WHERE id = %s AND pastillero_id = %s
        ''', (horas_entre_tomas, proxima_toma, medicamento_id, pastillero_id))

        conn.commit()
        conn.close()

        print(f"[OK] Recordatorio activado para medicamento {medicamento_id} ({medicamento['nombre']})")

        return jsonify({
            'ok': True,
            'mensaje': 'Recordatorio activado correctamente'
        })

    except Exception as e:
        if 'conn' in locals():
            conn.rollback()
            conn.close()
        print(f"Error al activar recordatorio: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'ok': False, 'error': str(e)}), 500


@app.route('/api/pastillero/<int:medicamento_id>/desactivar-recordatorio', methods=['POST'])
def api_desactivar_recordatorio(medicamento_id):
    """Desactivar recordatorio para un medicamento"""
    if 'usuario_id' not in session:
        return jsonify({'ok': False, 'error': 'No autenticado'}), 401

    usuario_id = session['usuario_id']

    pastillero_id = obtener_pastillero_activo(usuario_id)
    if not pastillero_id:
        return jsonify({'ok': False, 'error': 'No tienes un pastillero activo'}), 400

    try:
        conn = get_db_connection()

        conn.execute('''
            UPDATE pastillero_usuarios
            SET recordatorio_activo = FALSE
            WHERE id = %s AND pastillero_id = %s
        ''', (medicamento_id, pastillero_id))

        conn.commit()
        conn.close()

        return jsonify({'ok': True, 'mensaje': 'Recordatorio desactivado'})

    except Exception as e:
        if 'conn' in locals():
            conn.rollback()
            conn.close()
        print(f"Error al desactivar recordatorio: {e}")
        return jsonify({'ok': False, 'error': str(e)}), 500


@app.route('/api/pastillero/contactos', methods=['GET'])
def api_obtener_contactos_adicionales():
    """Obtener lista de contactos adicionales del usuario"""
    if 'usuario_id' not in session:
        return jsonify({'ok': False, 'error': 'No autenticado'}), 401

    usuario_id = session['usuario_id']

    try:
        conn = get_db_connection()

        contactos = conn.execute('''
            SELECT
                pca.id,
                pca.contacto_id,
                t.nombre,
                t.telefono,
                t.telegram_chat_id,
                pca.fecha_agregado
            FROM pastillero_contactos_adicionales pca
            INNER JOIN terceros t ON pca.contacto_id = t.id
            WHERE pca.usuario_id = %s
            ORDER BY pca.fecha_agregado DESC
        ''', (usuario_id,)).fetchall()

        conn.close()

        contactos_list = []
        for c in contactos:
            contactos_list.append({
                'id': c['id'],
                'contacto_id': c['contacto_id'],
                'nombre': c['nombre'],
                'telefono': c['telefono'],
                'telegram_vinculado': bool(c['telegram_chat_id']),
                'fecha_agregado': c['fecha_agregado'].isoformat() if c['fecha_agregado'] else None
            })

        return jsonify({'ok': True, 'contactos': contactos_list})

    except Exception as e:
        if 'conn' in locals():
            conn.close()
        print(f"Error al obtener contactos: {e}")
        return jsonify({'ok': False, 'error': str(e)}), 500


@app.route('/api/pastillero/contactos/agregar', methods=['POST'])
def api_agregar_contacto_adicional():
    """Agregar un contacto adicional para recordatorios"""
    if 'usuario_id' not in session:
        return jsonify({'ok': False, 'error': 'No autenticado'}), 401

    usuario_id = session['usuario_id']
    data = request.get_json()

    nombre = data.get('nombre', '').strip()
    telefono = data.get('telefono', '').strip()

    if not nombre or not telefono:
        return jsonify({'ok': False, 'error': 'Nombre y teléfono son requeridos'}), 400

    # Normalizar teléfono (remover espacios, guiones, etc)
    telefono = ''.join(filter(str.isdigit, telefono))

    if len(telefono) < 10:
        return jsonify({'ok': False, 'error': 'Teléfono inválido'}), 400

    try:
        conn = get_db_connection()

        # Buscar si el contacto ya existe en terceros
        tercero_existente = conn.execute('''
            SELECT id FROM terceros WHERE telefono = %s LIMIT 1
        ''', (telefono,)).fetchone()

        if tercero_existente:
            contacto_id = tercero_existente['id']

            # Actualizar nombre si cambió
            conn.execute('''
                UPDATE terceros
                SET nombre = %s, fecha_actualizacion = CURRENT_TIMESTAMP
                WHERE id = %s
            ''', (nombre, contacto_id))
        else:
            # Crear nuevo tercero
            conn.execute('''
                INSERT INTO terceros (nombre, telefono, fecha_creacion)
                VALUES (%s, %s, CURRENT_TIMESTAMP)
            ''', (nombre, telefono))

            # Obtener el ID recién creado
            contacto_id = conn.execute('''
                SELECT id FROM terceros WHERE telefono = %s LIMIT 1
            ''', (telefono,)).fetchone()['id']

        # Verificar que no se esté agregando a sí mismo
        if contacto_id == usuario_id:
            conn.close()
            return jsonify({'ok': False, 'error': 'No puedes agregarte a ti mismo como contacto'}), 400

        # Verificar que no esté duplicado
        existe_relacion = conn.execute('''
            SELECT id FROM pastillero_contactos_adicionales
            WHERE usuario_id = %s AND contacto_id = %s
        ''', (usuario_id, contacto_id)).fetchone()

        if existe_relacion:
            conn.close()
            return jsonify({'ok': False, 'error': 'Este contacto ya está agregado'}), 400

        # Crear la relación
        conn.execute('''
            INSERT INTO pastillero_contactos_adicionales (usuario_id, contacto_id)
            VALUES (%s, %s)
        ''', (usuario_id, contacto_id))

        conn.commit()
        conn.close()

        return jsonify({
            'ok': True,
            'mensaje': 'Contacto agregado correctamente',
            'contacto_id': contacto_id
        })

    except Exception as e:
        if 'conn' in locals():
            conn.rollback()
            conn.close()
        print(f"Error al agregar contacto: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'ok': False, 'error': str(e)}), 500


@app.route('/api/pastillero/contactos/<int:contacto_relacion_id>/eliminar', methods=['POST'])
def api_eliminar_contacto_adicional(contacto_relacion_id):
    """Eliminar un contacto adicional"""
    if 'usuario_id' not in session:
        return jsonify({'ok': False, 'error': 'No autenticado'}), 401

    usuario_id = session['usuario_id']

    try:
        conn = get_db_connection()

        # Verificar que la relación pertenece al usuario
        relacion = conn.execute('''
            SELECT id FROM pastillero_contactos_adicionales
            WHERE id = %s AND usuario_id = %s
        ''', (contacto_relacion_id, usuario_id)).fetchone()

        if not relacion:
            conn.close()
            return jsonify({'ok': False, 'error': 'Contacto no encontrado'}), 404

        # Eliminar la relación
        conn.execute('''
            DELETE FROM pastillero_contactos_adicionales
            WHERE id = %s
        ''', (contacto_relacion_id,))

        conn.commit()
        conn.close()

        return jsonify({'ok': True, 'mensaje': 'Contacto eliminado correctamente'})

    except Exception as e:
        if 'conn' in locals():
            conn.rollback()
            conn.close()
        print(f"Error al eliminar contacto: {e}")
        return jsonify({'ok': False, 'error': str(e)}), 500


@app.route('/api/vincular-telegram', methods=['POST'])
def api_vincular_telegram():
    """Vincular chat_id de Telegram con usuario"""
    data = request.get_json()

    telefono = data.get('telefono')
    chat_id = data.get('chat_id')

    if not telefono or not chat_id:
        return jsonify({'ok': False, 'error': 'Faltan datos requeridos'}), 400

    try:
        conn = get_db_connection()

        # Buscar tercero por teléfono
        tercero = conn.execute('''
            SELECT id, nombre FROM terceros WHERE telefono = %s LIMIT 1
        ''', (telefono,)).fetchone()

        if not tercero:
            conn.close()
            return jsonify({'ok': False, 'error': 'Teléfono no encontrado'}), 404

        # Actualizar telegram_chat_id
        conn.execute('''
            UPDATE terceros
            SET telegram_chat_id = %s
            WHERE id = %s
        ''', (str(chat_id), tercero['id']))

        conn.commit()
        conn.close()

        print(f"[OK] Telegram vinculado: chat_id={chat_id} -> usuario={tercero['nombre']} (tel: {telefono})")

        return jsonify({
            'ok': True,
            'nombre': tercero['nombre'],
            'mensaje': f'Vinculado correctamente con {tercero["nombre"]}'
        })

    except Exception as e:
        if 'conn' in locals():
            conn.rollback()
            conn.close()
        print(f"Error al vincular Telegram: {e}")
        return jsonify({'ok': False, 'error': str(e)}), 500


@app.route('/telegram/webhook', methods=['POST'])
def telegram_webhook():
    """
    Webhook para recibir mensajes del bot de Telegram.
    Maneja comandos como /vincular y callbacks de botones interactivos.
    """
    try:
        update = request.get_json()

        # Log para debugging
        print(f"[TELEGRAM] Webhook recibido: {update}")

        # Verificar si es un mensaje de texto (comando)
        if 'message' in update:
            message = update['message']
            chat_id = message['chat']['id']
            text = message.get('text', '')

            # Comando /start
            if text.startswith('/start'):
                enviar_mensaje_telegram(
                    chat_id,
                    "¡Hola! Soy el bot de TucTuc Medicamentos.\n\n"
                    "Para vincular tu cuenta y recibir recordatorios de tus medicamentos:\n"
                    "Envía: /vincular TU_TELEFONO\n\n"
                    "Ejemplo: /vincular 3166686397"
                )

            # Comando /michatid - Para obtener el chat_id del usuario
            elif text.startswith('/michatid'):
                enviar_mensaje_telegram(
                    chat_id,
                    f"📱 Tu Chat ID de Telegram es:\n\n<code>{chat_id}</code>\n\n"
                    f"Puedes usar este ID para vinculación manual si tienes problemas."
                )
                print(f"[INFO] Usuario solicitó chat_id: {chat_id}")

            # Comando /vincular TELEFONO
            elif text.startswith('/vincular'):
                partes = text.split()

                if len(partes) != 2:
                    enviar_mensaje_telegram(
                        chat_id,
                        "❌ Formato incorrecto.\n\n"
                        "Uso correcto: /vincular TU_TELEFONO\n"
                        "Ejemplo: /vincular 3166686397"
                    )
                    return jsonify({'ok': True})

                telefono = partes[1].strip()

                # Normalizar teléfono (remover espacios, guiones, paréntesis)
                telefono = ''.join(filter(str.isdigit, telefono))

                # Validar que tenga al menos 10 dígitos
                if len(telefono) < 10:
                    enviar_mensaje_telegram(
                        chat_id,
                        "❌ Teléfono inválido. Debe tener al menos 10 dígitos.\n\n"
                        "Ejemplo: /vincular 3166686397"
                    )
                    return jsonify({'ok': True})

                # Vincular en la base de datos
                conn = get_db_connection()

                # Buscar usuario por teléfono
                tercero = conn.execute('''
                    SELECT id, nombre FROM terceros WHERE telefono = %s LIMIT 1
                ''', (telefono,)).fetchone()

                if not tercero:
                    conn.close()
                    enviar_mensaje_telegram(
                        chat_id,
                        f"❌ No encontramos una cuenta con el teléfono {telefono}.\n\n"
                        "Asegúrate de:\n"
                        "1. Haber creado tu pastillero en la app\n"
                        "2. Usar el mismo teléfono registrado\n\n"
                        "¿Necesitas ayuda? Visita: https://tuc-tuc.onrender.com"
                    )
                    return jsonify({'ok': True})

                # Actualizar telegram_chat_id
                conn.execute('''
                    UPDATE terceros
                    SET telegram_chat_id = %s
                    WHERE id = %s
                ''', (str(chat_id), tercero['id']))

                conn.commit()
                conn.close()

                print(f"[OK] Telegram vinculado: chat_id={chat_id} -> {tercero['nombre']} (tel: {telefono})")

                # Enviar confirmación
                enviar_mensaje_telegram(
                    chat_id,
                    f"✅ ¡Vinculado correctamente!\n\n"
                    f"Hola {tercero['nombre']}, ahora recibirás recordatorios de tus medicamentos.\n\n"
                    f"Para activar recordatorios:\n"
                    f"1. Abre tu pastillero en https://tuc-tuc.onrender.com\n"
                    f"2. Presiona el botón ⏰ en el medicamento\n"
                    f"3. Configura cada cuántas horas tomas el medicamento\n\n"
                    f"¡Listo! Te enviaré recordatorios aquí en Telegram."
                )

        # Verificar si es un callback de botón (para los botones "Ya tomé" / "Cancelar hoy")
        elif 'callback_query' in update:
            callback = update['callback_query']
            chat_id = callback['message']['chat']['id']
            callback_data = callback['data']
            message_id = callback['message']['message_id']

            # El formato del callback será: "tomar_MED_ID" o "cancelar_MED_ID"
            partes = callback_data.split('_')
            accion = partes[0]
            medicamento_id = int(partes[1])

            if accion == 'tomar':
                # Marcar como tomado y restar 1 pastilla
                conn = get_db_connection()

                # Obtener chat_id del usuario
                tercero = conn.execute('''
                    SELECT id FROM terceros WHERE telegram_chat_id = %s LIMIT 1
                ''', (str(chat_id),)).fetchone()

                if not tercero:
                    conn.close()
                    # Responder al callback
                    responder_callback(callback['id'], "❌ Usuario no encontrado")
                    return jsonify({'ok': True})

                # Obtener medicamento del pastillero
                med = conn.execute('''
                    SELECT p.nombre, p.cantidad, p.horas_entre_tomas
                    FROM pastillero_usuarios p
                    WHERE p.id = %s AND p.usuario_id = %s
                    LIMIT 1
                ''', (medicamento_id, tercero['id'])).fetchone()

                if not med:
                    conn.close()
                    responder_callback(callback['id'], "❌ Medicamento no encontrado")
                    return jsonify({'ok': True})

                nueva_cantidad = max(0, med['cantidad'] - 1)

                # Restar 1 pastilla
                conn.execute('''
                    UPDATE pastillero_usuarios
                    SET cantidad = %s
                    WHERE id = %s
                ''', (nueva_cantidad, medicamento_id))

                # Calcular próxima toma
                from datetime import datetime, timedelta
                proxima_toma = datetime.now() + timedelta(hours=med['horas_entre_tomas'])

                conn.execute('''
                    UPDATE pastillero_usuarios
                    SET proxima_toma = %s
                    WHERE id = %s
                ''', (proxima_toma, medicamento_id))

                conn.commit()
                conn.close()

                # Responder al callback
                responder_callback(callback['id'], "✅ ¡Registrado!")

                # Editar el mensaje para confirmación
                mensaje_confirmacion = f"✅ <b>Confirmado</b>\n\n{med['nombre']}\n\nQuedan: {nueva_cantidad} pastillas"

                # Alerta si quedan pocas pastillas
                if nueva_cantidad <= 3 and nueva_cantidad > 0:
                    mensaje_confirmacion += f"\n\n⚠️ <b>¡Te quedan solo {nueva_cantidad} pastillas!</b>\n"
                    mensaje_confirmacion += "¿Quieres hacer un pedido?\n"
                    mensaje_confirmacion += "👉 https://tuc-tuc.onrender.com"
                elif nueva_cantidad == 0:
                    mensaje_confirmacion += "\n\n⚠️ <b>¡Se te acabaron las pastillas!</b>\n"
                    mensaje_confirmacion += "Haz tu pedido aquí:\n"
                    mensaje_confirmacion += "👉 https://tuc-tuc.onrender.com"

                editar_mensaje_telegram(chat_id, message_id, mensaje_confirmacion)

            elif accion == 'cancelar':
                # Solo marcar el recordatorio como visto
                responder_callback(callback['id'], "❌ Cancelado por hoy")

                # Editar el mensaje
                editar_mensaje_telegram(
                    chat_id,
                    message_id,
                    "❌ <b>Recordatorio cancelado por hoy</b>\n\nTe recordaré en la próxima toma programada."
                )

        return jsonify({'ok': True})

    except Exception as e:
        print(f"[ERROR] Error en webhook de Telegram: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'ok': False, 'error': str(e)}), 500


def enviar_mensaje_telegram(chat_id, texto, reply_markup=None):
    """Envía un mensaje de texto a un chat de Telegram"""
    try:
        conn = get_db_connection()
        config = conn.execute('SELECT telegram_token FROM CONFIGURACION_SISTEMA WHERE id = 1').fetchone()
        conn.close()

        if not config:
            # Usar token hardcoded como fallback
            token = "8486881295:AAFjs-SU74er_shs4KnQYImMtyU5OTXycng"
        else:
            token = config[0]

        url = f"https://api.telegram.org/bot{token}/sendMessage"
        data = {
            'chat_id': chat_id,
            'text': texto,
            'parse_mode': 'HTML'
        }

        if reply_markup:
            data['reply_markup'] = reply_markup

        response = requests.post(url, json=data, timeout=10)

        if response.status_code == 200:
            print(f"[OK] Mensaje enviado a chat_id={chat_id}")
            return True
        else:
            print(f"[ERROR] No se pudo enviar mensaje: {response.status_code}")
            print(f"Response: {response.text}")
            return False

    except Exception as e:
        print(f"[ERROR] Error enviando mensaje Telegram: {e}")
        return False


def editar_mensaje_telegram(chat_id, message_id, texto):
    """Edita un mensaje existente en Telegram"""
    try:
        conn = get_db_connection()
        config = conn.execute('SELECT telegram_token FROM CONFIGURACION_SISTEMA WHERE id = 1').fetchone()
        conn.close()

        if not config:
            token = "8486881295:AAFjs-SU74er_shs4KnQYImMtyU5OTXycng"
        else:
            token = config[0]

        url = f"https://api.telegram.org/bot{token}/editMessageText"
        data = {
            'chat_id': chat_id,
            'message_id': message_id,
            'text': texto,
            'parse_mode': 'HTML'
        }

        response = requests.post(url, json=data, timeout=10)

        if response.status_code == 200:
            print(f"[OK] Mensaje editado: chat_id={chat_id}, msg_id={message_id}")
            return True
        else:
            print(f"[ERROR] No se pudo editar mensaje: {response.status_code}")
            return False

    except Exception as e:
        print(f"[ERROR] Error editando mensaje: {e}")
        return False


def responder_callback(callback_id, texto):
    """Responde a un callback query (notificación emergente)"""
    try:
        conn = get_db_connection()
        config = conn.execute('SELECT telegram_token FROM CONFIGURACION_SISTEMA WHERE id = 1').fetchone()
        conn.close()

        if not config:
            token = "8486881295:AAFjs-SU74er_shs4KnQYImMtyU5OTXycng"
        else:
            token = config[0]

        url = f"https://api.telegram.org/bot{token}/answerCallbackQuery"
        data = {
            'callback_query_id': callback_id,
            'text': texto,
            'show_alert': False
        }

        response = requests.post(url, json=data, timeout=10)
        return response.status_code == 200

    except Exception as e:
        print(f"[ERROR] Error respondiendo callback: {e}")
        return False


# ============================================
#  RUTA DE BSQUEDA PARA AUTOCOMPLETADO
# ============================================





def extraer_nombre_base(nombre_completo):
    """
    Extrae el nombre base del medicamento sin la presentacin.
    Ejemplos:
    - "Ibuprofeno 400mg x 100 tabletas" -> "Ibuprofeno 400mg"
    - "Acetaminofn 500mg x 10 tabletas - blister" -> "Acetaminofn 500mg"
    - "Loratadina 10mg jarabe 120ml" -> "Loratadina 10mg"
    
    Estrategia:
    1. Buscar el patrn de concentracin (400mg, 500mg, etc)
    2. Tomar todo hasta la concentracin + la concentracin misma
    3. Eliminar cualquier cosa despus (x100, tabletas, blister, etc)
    """
    # Patrn para detectar concentracin: nmero + unidad (mg, g, ml, %, etc)
    patron_concentracion = r'\d+\.?\d*\s*(mg|g|ml|mcg|ui|%|mEq)'
    
    match = re.search(patron_concentracion, nombre_completo, re.IGNORECASE)
    
    if match:
        # Encontramos concentracin, tomar hasta ah
        fin_concentracion = match.end()
        nombre_base = nombre_completo[:fin_concentracion].strip()
        return nombre_base
    
    # Si no hay concentracin, tomar solo la primera palabra (nombre del medicamento)
    primera_palabra = nombre_completo.split()[0] if nombre_completo.split() else nombre_completo
    return primera_palabra

@app.route('/api/productos/buscar', methods=['GET'])
def buscar_productos_simple():
    """
    Bsqueda agrupada de productos para autocompletado.
    - Normaliza bsqueda (tildes, maysculas)
    - Agrupa por nombre base (sin presentacin)
    - Vincula con el producto de menor precio
    -  Incluye medicamentos personales del pastillero (medicamento_id NULL)
    """
    query = request.args.get('q', '').strip()
    
    if len(query) < 2:
        return jsonify({'ok': True, 'productos': []})
    
    # Verificar si el usuario est autenticado para buscar medicamentos personales
    usuario_id = session.get('usuario_id')
    
    try:
        # Normalizar la bsqueda del usuario
        query_normalizado = normalizar_texto(query)
        
        conn = get_db_connection()
        
        # 1 Buscar medicamentos OFICIALES de la BD (con precios)
        productos_raw = conn.execute('''
            SELECT DISTINCT
                m.id as medicamento_id,
                m.nombre,
                f.nombre as fabricante,
                f.id as fabricante_id,
                p.precio,
                p.id as precio_id,
                'oficial' as tipo
            FROM medicamentos m
            INNER JOIN precios p ON m.id = p.medicamento_id
            LEFT JOIN fabricantes f ON p.fabricante_id = f.id
            WHERE LOWER(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(
                m.nombre, '', 'a'), '', 'e'), '', 'i'), '', 'o'), '', 'u'), '', 'n')
            ) LIKE ?
            AND m.activo = '1'
            AND p.precio > 0
            ORDER BY m.nombre, p.precio ASC
        ''', (f'%{query_normalizado}%',)).fetchall()
        
        # 2 Buscar medicamentos PERSONALES del pastillero (NULL)
        medicamentos_personales = []
        if usuario_id:
            medicamentos_personales = conn.execute('''
                SELECT DISTINCT
                    NULL as medicamento_id,
                    nombre,
                    NULL as fabricante,
                    NULL as fabricante_id,
                    0 as precio,
                    NULL as precio_id,
                    'personal' as tipo
                FROM pastillero_usuarios
                WHERE usuario_id = ?
                AND medicamento_id IS NULL
                AND LOWER(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(
                    nombre, '', 'a'), '', 'e'), '', 'i'), '', 'o'), '', 'u'), '', 'n')
                ) LIKE ?
            ''', (usuario_id, f'%{query_normalizado}%')).fetchall()
        
        conn.close()
        
        # 3 Agrupar por nombre base y seleccionar el de menor precio
        productos_agrupados = {}
        
        # Procesar medicamentos oficiales
        for p in productos_raw:
            nombre_base = extraer_nombre_base(p['nombre'])
            
            # Si ya existe este nombre base, comparar precios
            if nombre_base in productos_agrupados:
                # Quedarnos con el de menor precio
                if p['precio'] < productos_agrupados[nombre_base]['precio']:
                    productos_agrupados[nombre_base] = {
                        'medicamento_id': p['medicamento_id'],
                        'fabricante_id': p['fabricante_id'],
                        'nombre': nombre_base,
                        'fabricante': p['fabricante'] if p['fabricante'] else 'Genrico',
                        'precio': p['precio'],
                        'precio_id': p['precio_id'],
                        'tipo': 'oficial'
                    }
            else:
                # Primera vez que vemos este nombre base
                productos_agrupados[nombre_base] = {
                    'medicamento_id': p['medicamento_id'],
                    'fabricante_id': p['fabricante_id'],
                    'nombre': nombre_base,
                    'fabricante': p['fabricante'] if p['fabricante'] else 'Genrico',
                    'precio': p['precio'],
                    'precio_id': p['precio_id'],
                    'tipo': 'oficial'
                }
        
        # 4 Agregar medicamentos personales (solo si no existen en oficiales)
        for p in medicamentos_personales:
            nombre_base = p['nombre']  # Ya est normalizado en el pastillero
            
            # Solo agregar si NO existe ya en productos oficiales
            if nombre_base not in productos_agrupados:
                productos_agrupados[nombre_base] = {
                    'medicamento_id': None,  # NULL para medicamentos personales
                    'fabricante_id': None,
                    'nombre': nombre_base,
                    'tipo': 'personal',
                    'es_personal': True  #  Flag para identificar en frontend
                }
        
        # Convertir a lista y limitar a 10 resultados
        productos_list = list(productos_agrupados.values())[:10]
        
        # Remover campos innecesarios antes de enviar
        for prod in productos_list:
            prod.pop('precio', None)
            prod.pop('precio_id', None)
            prod.pop('fabricante', None)
            prod.pop('fabricante_id', None)
            prod.pop('tipo', None)
        
        return jsonify({'ok': True, 'productos': productos_list})
        
    except Exception as e:
        print(f"Error buscando productos: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'ok': True, 'productos': []})


# -------------------------------------------------------------------
# --- TERMINA PASTILLERO ---
# -------------------------------------------------------------------


# -------------------------------------------------------------------
# --- SISTEMA DE CATEGORÍAS PARAMETRIZABLES ---
# -------------------------------------------------------------------

@app.route('/api/admin/categorias', methods=['GET'])
@admin_required
def obtener_categorias():
    """Obtiene todas las categorías para el admin"""
    try:
        conn = get_db_connection()
        categorias = conn.execute("""
            SELECT c.*,
                   COUNT(DISTINCT mc.medicamento_id) as num_productos
            FROM categorias c
            LEFT JOIN medicamento_categoria mc ON c.id = mc.categoria_id
            GROUP BY c.id
            ORDER BY c.orden ASC, c.nombre ASC
        """).fetchall()
        conn.close()

        return jsonify({
            'ok': True,
            'categorias': [dict(cat) for cat in categorias]
        })
    except Exception as e:
        print(f"Error obteniendo categorías: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'ok': False, 'error': str(e)}), 500


@app.route('/api/admin/categorias', methods=['POST'])
@admin_required
def crear_categoria():
    """Crea una nueva categoría"""
    try:
        data = request.get_json()
        nombre = data.get('nombre', '').strip()
        descripcion = data.get('descripcion', '').strip()
        imagen = data.get('imagen', '').strip()
        orden = data.get('orden', 0)
        activo = data.get('activo', True)
        es_destacada = data.get('es_destacada', False)

        if not nombre:
            return jsonify({'ok': False, 'error': 'El nombre es requerido'}), 400

        conn = get_db_connection()

        # Si es destacada, desactivar las demás
        if es_destacada:
            conn.execute("UPDATE categorias SET es_destacada = FALSE")

        cursor = conn.execute("""
            INSERT INTO categorias (nombre, descripcion, imagen, orden, activo, es_destacada)
            VALUES (%s, %s, %s, %s, %s, %s)
            RETURNING id
        """, (nombre, descripcion, imagen or None, orden, activo, es_destacada))

        categoria_id = cursor.fetchone()[0]
        conn.commit()
        conn.close()

        return jsonify({
            'ok': True,
            'categoria_id': categoria_id,
            'message': 'Categoría creada exitosamente'
        })
    except Exception as e:
        print(f"Error creando categoría: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'ok': False, 'error': str(e)}), 500


@app.route('/api/admin/categorias/<int:categoria_id>', methods=['PUT'])
@admin_required
def actualizar_categoria(categoria_id):
    """Actualiza una categoría existente"""
    try:
        data = request.get_json()
        nombre = data.get('nombre', '').strip()
        descripcion = data.get('descripcion', '').strip()
        imagen = data.get('imagen', '').strip()
        orden = data.get('orden', 0)
        activo = data.get('activo', True)
        es_destacada = data.get('es_destacada', False)

        if not nombre:
            return jsonify({'ok': False, 'error': 'El nombre es requerido'}), 400

        conn = get_db_connection()

        # Si es destacada, desactivar las demás
        if es_destacada:
            conn.execute("UPDATE categorias SET es_destacada = FALSE WHERE id != %s", (categoria_id,))

        conn.execute("""
            UPDATE categorias
            SET nombre = %s,
                descripcion = %s,
                imagen = %s,
                orden = %s,
                activo = %s,
                es_destacada = %s,
                fecha_actualizacion = CURRENT_TIMESTAMP
            WHERE id = %s
        """, (nombre, descripcion, imagen or None, orden, activo, es_destacada, categoria_id))

        conn.commit()
        conn.close()

        return jsonify({
            'ok': True,
            'message': 'Categoría actualizada exitosamente'
        })
    except Exception as e:
        print(f"Error actualizando categoría: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'ok': False, 'error': str(e)}), 500


@app.route('/api/admin/categorias/<int:categoria_id>', methods=['DELETE'])
@admin_required
def eliminar_categoria(categoria_id):
    """Elimina una categoría"""
    try:
        conn = get_db_connection()

        # Verificar que no sea la última categoría
        count = conn.execute("SELECT COUNT(*) FROM categorias").fetchone()[0]
        if count <= 1:
            conn.close()
            return jsonify({'ok': False, 'error': 'No puedes eliminar la última categoría'}), 400

        conn.execute("DELETE FROM categorias WHERE id = %s", (categoria_id,))
        conn.commit()
        conn.close()

        return jsonify({
            'ok': True,
            'message': 'Categoría eliminada exitosamente'
        })
    except Exception as e:
        print(f"Error eliminando categoría: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'ok': False, 'error': str(e)}), 500


@app.route('/api/admin/categorias/<int:categoria_id>/productos', methods=['GET'])
@admin_required
def obtener_productos_categoria(categoria_id):
    """Obtiene los productos de una categoría"""
    try:
        conn = get_db_connection()
        productos = conn.execute("""
            SELECT m.id, m.nombre
            FROM medicamentos m
            INNER JOIN medicamento_categoria mc ON m.id = mc.medicamento_id
            WHERE mc.categoria_id = %s
            ORDER BY m.nombre
        """, (categoria_id,)).fetchall()
        conn.close()

        return jsonify({
            'ok': True,
            'productos': [dict(p) for p in productos]
        })
    except Exception as e:
        print(f"Error obteniendo productos de categoría: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'ok': False, 'error': str(e)}), 500


@app.route('/api/admin/categorias/<int:categoria_id>/productos', methods=['POST'])
@admin_required
def agregar_producto_categoria(categoria_id):
    """Agrega un producto a una categoría"""
    try:
        data = request.get_json()
        medicamento_id = data.get('medicamento_id')

        if not medicamento_id:
            return jsonify({'ok': False, 'error': 'medicamento_id es requerido'}), 400

        conn = get_db_connection()

        # Verificar si ya existe
        existe = conn.execute("""
            SELECT 1 FROM medicamento_categoria
            WHERE medicamento_id = %s AND categoria_id = %s
        """, (medicamento_id, categoria_id)).fetchone()

        if existe:
            conn.close()
            return jsonify({'ok': False, 'error': 'El producto ya está en esta categoría'}), 400

        conn.execute("""
            INSERT INTO medicamento_categoria (medicamento_id, categoria_id)
            VALUES (%s, %s)
        """, (medicamento_id, categoria_id))

        conn.commit()
        conn.close()

        return jsonify({
            'ok': True,
            'message': 'Producto agregado a la categoría'
        })
    except Exception as e:
        print(f"Error agregando producto a categoría: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'ok': False, 'error': str(e)}), 500


@app.route('/api/admin/categorias/<int:categoria_id>/productos/<int:medicamento_id>', methods=['DELETE'])
@admin_required
def eliminar_producto_categoria(categoria_id, medicamento_id):
    """Elimina un producto de una categoría"""
    try:
        conn = get_db_connection()
        conn.execute("""
            DELETE FROM medicamento_categoria
            WHERE categoria_id = %s AND medicamento_id = %s
        """, (categoria_id, medicamento_id))
        conn.commit()
        conn.close()

        return jsonify({
            'ok': True,
            'message': 'Producto eliminado de la categoría'
        })
    except Exception as e:
        print(f"Error eliminando producto de categoría: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'ok': False, 'error': str(e)}), 500


# Endpoint para buscar medicamentos (usado en admin de categorías)
@app.route('/api/medicamentos/buscar', methods=['GET'])
@admin_required
def buscar_medicamentos_admin():
    """Busca medicamentos por nombre para el admin - Solo productos con precio"""
    try:
        query = request.args.get('q', '').strip()

        conn = get_db_connection()

        # Si no hay query o es muy corto, devolver TODOS los medicamentos
        if not query or len(query) < 2:
            medicamentos = conn.execute("""
                SELECT DISTINCT
                    m.id,
                    m.nombre,
                    f.nombre as fabricante_nombre,
                    p.precio
                FROM medicamentos m
                INNER JOIN precios p ON m.id = p.medicamento_id
                INNER JOIN fabricantes f ON p.fabricante_id = f.id
                WHERE p.precio > 0
                ORDER BY m.nombre, f.nombre
            """).fetchall()
        else:
            # Normalizar búsqueda: remover acentos y convertir a minúsculas
            query_normalizado = query.lower()

            medicamentos = conn.execute("""
                SELECT DISTINCT
                    m.id,
                    m.nombre,
                    f.nombre as fabricante_nombre,
                    p.precio
                FROM medicamentos m
                INNER JOIN precios p ON m.id = p.medicamento_id
                INNER JOIN fabricantes f ON p.fabricante_id = f.id
                WHERE unaccent(LOWER(m.nombre)) LIKE unaccent(%s)
                AND p.precio > 0
                ORDER BY m.nombre, f.nombre
                LIMIT 20
            """, (f'%{query_normalizado}%',)).fetchall()

        conn.close()

        return jsonify({
            'ok': True,
            'medicamentos': [{
                'id': m['id'],
                'nombre': m['nombre'],
                'fabricante': m['fabricante_nombre'],
                'precio': float(m['precio'])
            } for m in medicamentos]
        })
    except Exception as e:
        print(f"Error buscando medicamentos: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'ok': False, 'error': str(e)}), 500


# Endpoint público para obtener categorías activas (para la tienda)
@app.route('/api/categorias', methods=['GET'])
def obtener_categorias_publicas():
    """Obtiene categorías activas para mostrar en la tienda"""
    print("=" * 70)
    print(" API /api/categorias - VERSION: 2025-12-23 18:00 FIXED")
    print("=" * 70)
    try:
        conn = get_db_connection()
        categorias = conn.execute("""
            SELECT id, nombre, descripcion, imagen, orden
            FROM categorias
            WHERE activo = TRUE
            ORDER BY orden ASC, nombre ASC
        """).fetchall()

        # Obtener categoría destacada
        destacada = conn.execute("""
            SELECT id FROM categorias
            WHERE es_destacada = TRUE AND activo = TRUE
            LIMIT 1
        """).fetchone()

        conn.close()

        return jsonify({
            'ok': True,
            'categorias': [dict(cat) for cat in categorias],
            'categoria_destacada_id': destacada['id'] if destacada else None
        })
    except Exception as e:
        print(f"Error obteniendo categorías públicas: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'ok': False, 'error': str(e)}), 500


# -------------------------------------------------------------------
# --- ENDPOINTS DE CHAT Y MENSAJERÍA ---
# -------------------------------------------------------------------

@app.route('/api/chat/buscar-usuarios', methods=['GET'])
def api_chat_buscar_usuarios():
    """Buscar usuarios por nombre o teléfono para iniciar conversación"""
    if 'usuario_id' not in session:
        return jsonify({'ok': False, 'error': 'No autenticado'}), 401

    query = request.args.get('q', '').strip()
    if len(query) < 2:
        return jsonify({'ok': True, 'usuarios': []})

    usuario_id = session['usuario_id']

    try:
        conn = get_db_connection()

        # Buscar usuarios por nombre o teléfono (excluyendo al usuario actual)
        usuarios = conn.execute('''
            SELECT id, nombre, telefono
            FROM terceros
            WHERE id != %s
            AND (
                LOWER(nombre) LIKE %s
                OR telefono LIKE %s
            )
            ORDER BY nombre
            LIMIT 10
        ''', (usuario_id, f'%{query.lower()}%', f'%{query}%')).fetchall()

        conn.close()

        return jsonify({
            'ok': True,
            'usuarios': [dict(u) for u in usuarios]
        })
    except Exception as e:
        print(f"Error al buscar usuarios: {e}")
        return jsonify({'ok': False, 'error': str(e)}), 500


@app.route('/api/chat/conversaciones', methods=['GET'])
def api_chat_conversaciones():
    """Obtener lista de conversaciones del usuario"""
    if 'usuario_id' not in session:
        return jsonify({'ok': False, 'error': 'No autenticado'}), 401

    usuario_id = session['usuario_id']

    try:
        conn = get_db_connection()

        # Obtener conversaciones únicas con último mensaje
        conversaciones = conn.execute('''
            WITH ultima_interaccion AS (
                SELECT
                    CASE
                        WHEN remitente_id = %s THEN destinatario_id
                        ELSE remitente_id
                    END as otro_usuario_id,
                    MAX(fecha) as ultima_fecha
                FROM mensajes
                WHERE remitente_id = %s OR destinatario_id = %s
                GROUP BY otro_usuario_id
            )
            SELECT
                t.id,
                t.nombre,
                t.telefono,
                ui.ultima_fecha,
                (
                    SELECT COUNT(*)
                    FROM mensajes m
                    WHERE m.remitente_id = t.id
                    AND m.destinatario_id = %s
                    AND m.estado = 'pendiente'
                ) as no_leidos
            FROM ultima_interaccion ui
            INNER JOIN terceros t ON t.id = ui.otro_usuario_id
            ORDER BY ui.ultima_fecha DESC
        ''', (usuario_id, usuario_id, usuario_id, usuario_id)).fetchall()

        conn.close()

        return jsonify({
            'ok': True,
            'conversaciones': [dict(c) for c in conversaciones]
        })
    except Exception as e:
        print(f"Error al obtener conversaciones: {e}")
        return jsonify({'ok': False, 'error': str(e)}), 500


@app.route('/api/chat/mensajes/<int:otro_usuario_id>', methods=['GET'])
def api_chat_mensajes(otro_usuario_id):
    """Obtener mensajes de una conversación específica"""
    if 'usuario_id' not in session:
        return jsonify({'ok': False, 'error': 'No autenticado'}), 401

    usuario_id = session['usuario_id']

    try:
        conn = get_db_connection()

        # Marcar mensajes como leídos
        conn.execute('''
            UPDATE mensajes
            SET estado = 'leido'
            WHERE remitente_id = %s
            AND destinatario_id = %s
            AND estado = 'pendiente'
        ''', (otro_usuario_id, usuario_id))

        # Obtener mensajes de la conversación
        mensajes = conn.execute('''
            SELECT
                m.id,
                m.remitente_id,
                m.destinatario_id,
                m.mensaje,
                m.tipo,
                m.pastillero_id,
                m.estado,
                m.fecha,
                t_rem.nombre as remitente_nombre,
                t_dest.nombre as destinatario_nombre,
                p.nombre as pastillero_nombre
            FROM mensajes m
            INNER JOIN terceros t_rem ON m.remitente_id = t_rem.id
            INNER JOIN terceros t_dest ON m.destinatario_id = t_dest.id
            LEFT JOIN pastilleros p ON m.pastillero_id = p.id
            WHERE (m.remitente_id = %s AND m.destinatario_id = %s)
               OR (m.remitente_id = %s AND m.destinatario_id = %s)
            ORDER BY m.fecha ASC
        ''', (usuario_id, otro_usuario_id, otro_usuario_id, usuario_id)).fetchall()

        conn.commit()
        conn.close()

        return jsonify({
            'ok': True,
            'mensajes': [dict(m) for m in mensajes]
        })
    except Exception as e:
        print(f"Error al obtener mensajes: {e}")
        return jsonify({'ok': False, 'error': str(e)}), 500


@app.route('/api/chat/enviar', methods=['POST'])
def api_chat_enviar():
    """Enviar un mensaje"""
    if 'usuario_id' not in session:
        return jsonify({'ok': False, 'error': 'No autenticado'}), 401

    usuario_id = session['usuario_id']
    data = request.get_json()

    destinatario_id = data.get('destinatario_id')
    mensaje = data.get('mensaje', '').strip()
    tipo = data.get('tipo', 'texto')
    pastillero_id = data.get('pastillero_id')

    if not destinatario_id or not mensaje:
        return jsonify({'ok': False, 'error': 'Faltan datos'}), 400

    try:
        conn = get_db_connection()

        conn.execute('''
            INSERT INTO mensajes (remitente_id, destinatario_id, mensaje, tipo, pastillero_id, estado, fecha)
            VALUES (%s, %s, %s, %s, %s, 'pendiente', CURRENT_TIMESTAMP)
        ''', (usuario_id, destinatario_id, mensaje, tipo, pastillero_id))

        conn.commit()
        conn.close()

        return jsonify({'ok': True})
    except Exception as e:
        print(f"Error al enviar mensaje: {e}")
        return jsonify({'ok': False, 'error': str(e)}), 500


# -------------------------------------------------------------------
# --- ENDPOINTS DE GESTIÓN DE PASTILLEROS ---
# -------------------------------------------------------------------

@app.route('/api/pastilleros/mis-pastilleros', methods=['GET'])
def api_mis_pastilleros():
    """Obtener todos los pastilleros del usuario (propietario o miembro)"""
    if 'usuario_id' not in session:
        return jsonify({'ok': False, 'error': 'No autenticado'}), 401

    usuario_id = session['usuario_id']

    try:
        conn = get_db_connection()

        pastilleros = conn.execute('''
            SELECT
                p.id,
                p.nombre,
                rp.tipo,
                p.creado_por_usuario_id,
                t.nombre as creador_nombre,
                (SELECT COUNT(*) FROM pastillero_usuarios pu WHERE pu.pastillero_id = p.id) as total_medicamentos
            FROM pastilleros p
            INNER JOIN relaciones_pastillero rp ON p.id = rp.pastillero_id
            LEFT JOIN terceros t ON p.creado_por_usuario_id = t.id
            WHERE rp.usuario_id = %s
            ORDER BY rp.tipo DESC, p.id ASC
        ''', (usuario_id,)).fetchall()

        conn.close()

        # Obtener el pastillero activo actual
        pastillero_activo = obtener_pastillero_activo(usuario_id)

        return jsonify({
            'ok': True,
            'pastilleros': [dict(p) for p in pastilleros],
            'activo_id': pastillero_activo
        })
    except Exception as e:
        print(f"Error al obtener pastilleros: {e}")
        return jsonify({'ok': False, 'error': str(e)}), 500


@app.route('/api/pastilleros/cambiar-activo', methods=['POST'])
def api_cambiar_pastillero_activo():
    """Cambiar el pastillero activo del usuario"""
    if 'usuario_id' not in session:
        return jsonify({'ok': False, 'error': 'No autenticado'}), 401

    usuario_id = session['usuario_id']
    data = request.get_json()
    pastillero_id = data.get('pastillero_id')

    if not pastillero_id:
        return jsonify({'ok': False, 'error': 'Falta pastillero_id'}), 400

    try:
        conn = get_db_connection()

        # Verificar que el usuario tiene acceso a este pastillero
        acceso = conn.execute('''
            SELECT id FROM relaciones_pastillero
            WHERE pastillero_id = %s AND usuario_id = %s
        ''', (pastillero_id, usuario_id)).fetchone()

        if not acceso:
            conn.close()
            return jsonify({'ok': False, 'error': 'No tienes acceso a este pastillero'}), 403

        # Guardar en sesión (por ahora, en el futuro podría ser en BD)
        session['pastillero_activo_id'] = pastillero_id

        conn.close()
        return jsonify({'ok': True})
    except Exception as e:
        print(f"Error al cambiar pastillero activo: {e}")
        return jsonify({'ok': False, 'error': str(e)}), 500


@app.route('/api/pastilleros/crear', methods=['POST'])
def api_crear_pastillero():
    """Crear un nuevo pastillero"""
    if 'usuario_id' not in session:
        return jsonify({'ok': False, 'error': 'No autenticado'}), 401

    usuario_id = session['usuario_id']
    data = request.get_json()
    nombre = data.get('nombre', '').strip()

    if not nombre:
        return jsonify({'ok': False, 'error': 'Falta nombre del pastillero'}), 400

    try:
        conn = get_db_connection()

        # Crear pastillero
        conn.execute('''
            INSERT INTO pastilleros (nombre, creado_por_usuario_id)
            VALUES (%s, %s)
        ''', (nombre, usuario_id))

        # Obtener el ID del pastillero creado
        pastillero_id = conn.execute('''
            SELECT id FROM pastilleros
            WHERE creado_por_usuario_id = %s
            ORDER BY id DESC LIMIT 1
        ''', (usuario_id,)).fetchone()['id']

        # Crear relación de propietario
        conn.execute('''
            INSERT INTO relaciones_pastillero (pastillero_id, usuario_id, tipo)
            VALUES (%s, %s, 'propietario')
        ''', (pastillero_id, usuario_id))

        conn.commit()
        conn.close()

        return jsonify({'ok': True, 'pastillero_id': pastillero_id})
    except Exception as e:
        print(f"Error al crear pastillero: {e}")
        return jsonify({'ok': False, 'error': str(e)}), 500


@app.route('/api/pastilleros/compartir', methods=['POST'])
def api_compartir_pastillero():
    """Enviar invitación para compartir pastillero"""
    if 'usuario_id' not in session:
        return jsonify({'ok': False, 'error': 'No autenticado'}), 401

    usuario_id = session['usuario_id']
    data = request.get_json()

    pastillero_id = data.get('pastillero_id')
    destinatario_id = data.get('destinatario_id')
    tipo_acceso = data.get('tipo_acceso', 'miembro')  # 'miembro' o 'autorizado'

    if not pastillero_id or not destinatario_id:
        return jsonify({'ok': False, 'error': 'Faltan datos'}), 400

    try:
        conn = get_db_connection()

        # Verificar que el usuario es propietario del pastillero
        es_propietario = conn.execute('''
            SELECT id FROM relaciones_pastillero
            WHERE pastillero_id = %s AND usuario_id = %s AND tipo = 'propietario'
        ''', (pastillero_id, usuario_id)).fetchone()

        if not es_propietario:
            conn.close()
            return jsonify({'ok': False, 'error': 'Solo el propietario puede compartir'}), 403

        # Obtener nombre del pastillero
        pastillero = conn.execute('''
            SELECT nombre FROM pastilleros WHERE id = %s
        ''', (pastillero_id,)).fetchone()

        # Crear mensaje de invitación
        tipo_msg = 'invitacion_compartir' if tipo_acceso == 'miembro' else 'solicitud_acceso'
        mensaje_texto = f"Te invito a {'compartir' if tipo_acceso == 'miembro' else 'ver'} mi pastillero '{pastillero['nombre']}'"

        conn.execute('''
            INSERT INTO mensajes (remitente_id, destinatario_id, mensaje, tipo, pastillero_id, estado, fecha)
            VALUES (%s, %s, %s, %s, %s, 'pendiente', CURRENT_TIMESTAMP)
        ''', (usuario_id, destinatario_id, mensaje_texto, tipo_msg, pastillero_id))

        conn.commit()
        conn.close()

        return jsonify({'ok': True})
    except Exception as e:
        print(f"Error al compartir pastillero: {e}")
        return jsonify({'ok': False, 'error': str(e)}), 500


@app.route('/api/pastilleros/responder-invitacion', methods=['POST'])
def api_responder_invitacion():
    """Aceptar o rechazar invitación de pastillero"""
    if 'usuario_id' not in session:
        return jsonify({'ok': False, 'error': 'No autenticado'}), 401

    usuario_id = session['usuario_id']
    data = request.get_json()

    mensaje_id = data.get('mensaje_id')
    aceptar = data.get('aceptar', False)

    if not mensaje_id:
        return jsonify({'ok': False, 'error': 'Falta mensaje_id'}), 400

    try:
        conn = get_db_connection()

        # Obtener el mensaje
        mensaje = conn.execute('''
            SELECT * FROM mensajes
            WHERE id = %s AND destinatario_id = %s
        ''', (mensaje_id, usuario_id)).fetchone()

        if not mensaje:
            conn.close()
            return jsonify({'ok': False, 'error': 'Mensaje no encontrado'}), 404

        if aceptar:
            # Determinar el tipo de acceso
            tipo = 'miembro' if mensaje['tipo'] == 'invitacion_compartir' else 'autorizado'

            # Agregar al pastillero
            conn.execute('''
                INSERT INTO relaciones_pastillero (pastillero_id, usuario_id, tipo)
                VALUES (%s, %s, %s)
                ON CONFLICT DO NOTHING
            ''', (mensaje['pastillero_id'], usuario_id, tipo))

            # Marcar mensaje como aceptado
            conn.execute('''
                UPDATE mensajes SET estado = 'aceptado' WHERE id = %s
            ''', (mensaje_id,))
        else:
            # Marcar mensaje como rechazado
            conn.execute('''
                UPDATE mensajes SET estado = 'rechazado' WHERE id = %s
            ''', (mensaje_id,))

        conn.commit()
        conn.close()

        return jsonify({'ok': True})
    except Exception as e:
        print(f"Error al responder invitación: {e}")
        return jsonify({'ok': False, 'error': str(e)}), 500


# -------------------------------------------------------------------
# --- ZONA 7: INICIALIZACIN Y EJECUCIN DEL SERVIDOR ---
# -------------------------------------------------------------------

# ESTE CÓDIGO SE DEBE INSERTAR EN 1_medicamentos.py ANTES DEL if __name__ == '__main__'

# -------------------------------------------------------------------
# --- ZONA: SUGERIR SÍNTOMAS (ADMIN) ---
# -------------------------------------------------------------------

# Importar funciones auxiliares
from sugerir_sintomas_helpers import (
    normalizar, detectar_diagnosticos_en_texto,
    extraer_sugeridos_de_texto_avanzado, validar_texto_medicamento,
    normalizar_sintomas_lista, REGLAS_DIAGNOSTICOS
)


@app.route('/admin/sugerir-sintomas', methods=['GET'])
@app.route('/admin/sugerir-sintomas/<int:med_id>', methods=['GET'])
@admin_required
def admin_sugerir_sintomas(med_id=None):
    """Vista principal para sugerir síntomas a medicamentos"""
    try:
        conn = get_db_connection()

        # Si no se especifica medicamento, buscar el primero pendiente
        if med_id is None:
            primer_pendiente = conn.execute("""
                SELECT m.id
                FROM medicamentos m
                LEFT JOIN medicamento_sintoma ms ON m.id = ms.medicamento_id
                WHERE ms.sintoma_id IS NULL AND m.activo = 'TRUE'
                ORDER BY
                    CASE WHEN m.componente_activo_id IS NULL THEN 0 ELSE 1 END,
                    (SELECT CASE WHEN p.precio > 0 THEN 0 ELSE 1 END FROM precios p WHERE p.medicamento_id = m.id LIMIT 1),
                    m.nombre
                LIMIT 1
            """).fetchone()

            if primer_pendiente:
                conn.close()
                return redirect(url_for('admin_sugerir_sintomas', med_id=primer_pendiente['id']))
            else:
                conn.close()
                return render_template('admin_sugerir_sintomas.html',
                                     medicamento=None,
                                     medicamentos_pendientes=0,
                                     mensaje="✅ No hay medicamentos pendientes")

        # Obtener medicamento actual
        medicamento = conn.execute("""
            SELECT m.id, m.nombre, m.componente_activo_id,
                   ca.nombre as componente_activo_nombre
            FROM medicamentos m
            LEFT JOIN medicamentos ca ON m.componente_activo_id = ca.id
            WHERE m.id = %s
        """, (med_id,)).fetchone()

        if not medicamento:
            conn.close()
            return "Medicamento no encontrado", 404

        # Obtener lista de medicamentos pendientes (para el selector)
        medicamentos_list = conn.execute("""
            SELECT m.id, m.nombre, m.componente_activo_id,
                   CASE WHEN p.precio > 0 THEN 1 ELSE 0 END as tiene_precio
            FROM medicamentos m
            LEFT JOIN medicamento_sintoma ms ON m.id = ms.medicamento_id
            LEFT JOIN precios p ON p.medicamento_id = m.id
            WHERE ms.sintoma_id IS NULL AND m.activo = 'TRUE'
            ORDER BY
                CASE WHEN m.componente_activo_id IS NULL THEN 0 ELSE 1 END,
                CASE WHEN p.precio > 0 THEN 0 ELSE 1 END,
                m.nombre
            LIMIT 200
        """).fetchall()

        # Contar total pendientes
        total_pendientes = conn.execute("""
            SELECT COUNT(DISTINCT m.id) as total
            FROM medicamentos m
            LEFT JOIN medicamento_sintoma ms ON m.id = ms.medicamento_id
            WHERE ms.sintoma_id IS NULL AND m.activo = 'TRUE'
        """).fetchone()['total']

        # Determinar término de búsqueda (componente activo si existe, sino nombre)
        termino_busqueda = medicamento['nombre']
        if medicamento['componente_activo_nombre']:
            termino_busqueda = medicamento['componente_activo_nombre']

        conn.close()

        return render_template('admin_sugerir_sintomas.html',
                             medicamento=medicamento,
                             medicamentos_list=medicamentos_list,
                             total_pendientes=total_pendientes,
                             termino_busqueda=termino_busqueda)

    except Exception as e:
        print(f"Error en admin_sugerir_sintomas: {e}")
        traceback.print_exc()
        return jsonify({'ok': False, 'error': str(e)}), 500


@app.route('/admin/sugerir-sintomas/filtrar', methods=['GET'])
@admin_required
def filtrar_medicamentos_sugerir():
    """Filtra medicamentos según tipo y precio"""
    try:
        filtro_tipo = request.args.get('tipo', 'todos')  # todos, genericos, comerciales
        filtro_precio = request.args.get('precio', 'todos')  # todos, con, sin

        conn = get_db_connection()

        where_clauses = ["ms.sintoma_id IS NULL", "m.activo = 'TRUE'"]

        if filtro_tipo == 'genericos':
            where_clauses.append("m.componente_activo_id IS NULL")
        elif filtro_tipo == 'comerciales':
            where_clauses.append("m.componente_activo_id IS NOT NULL")

        if filtro_precio == 'con':
            where_clauses.append("p.precio > 0")
        elif filtro_precio == 'sin':
            where_clauses.append("(p.precio IS NULL OR p.precio <= 0)")

        where_sql = " AND ".join(where_clauses)

        medicamentos = conn.execute(f"""
            SELECT DISTINCT m.id, m.nombre, m.componente_activo_id,
                   CASE WHEN p.precio > 0 THEN 1 ELSE 0 END as tiene_precio
            FROM medicamentos m
            LEFT JOIN medicamento_sintoma ms ON m.id = ms.medicamento_id
            LEFT JOIN precios p ON p.medicamento_id = m.id
            WHERE {where_sql}
            ORDER BY
                CASE WHEN m.componente_activo_id IS NULL THEN 0 ELSE 1 END,
                CASE WHEN p.precio > 0 THEN 0 ELSE 1 END,
                m.nombre
            LIMIT 200
        """).fetchall()

        conn.close()

        return jsonify({
            'ok': True,
            'medicamentos': [{
                'id': m['id'],
                'nombre': m['nombre'],
                'es_generico': m['componente_activo_id'] is None,
                'tiene_precio': m['tiene_precio'] == 1
            } for m in medicamentos]
        })

    except Exception as e:
        print(f"Error filtrando medicamentos: {e}")
        return jsonify({'ok': False, 'error': str(e)}), 500


@app.route('/admin/sugerir-sintomas/procesar-texto/<int:med_id>', methods=['POST'])
@admin_required
def procesar_texto_pegado(med_id):
    """Procesa el texto pegado y detecta síntomas/diagnósticos"""
    try:
        data = request.get_json()
        texto = data.get('texto', '').strip()

        if not texto or len(texto) < 20:
            return jsonify({'ok': False, 'error': 'Texto muy corto (mínimo 20 caracteres)'}), 400

        conn = get_db_connection()

        # Obtener medicamento
        medicamento = conn.execute("""
            SELECT m.id, m.nombre, ca.nombre as componente_activo_nombre
            FROM medicamentos m
            LEFT JOIN medicamentos ca ON m.componente_activo_id = ca.id
            WHERE m.id = %s
        """, (med_id,)).fetchone()

        if not medicamento:
            conn.close()
            return jsonify({'ok': False, 'error': 'Medicamento no encontrado'}), 404

        # VALIDAR si el texto corresponde al medicamento
        nombre_validar = medicamento['componente_activo_nombre'] or medicamento['nombre']
        coincide, confianza = validar_texto_medicamento(texto, nombre_validar)

        # Detectar diagnósticos
        diagnosticos_detectados_raw = detectar_diagnosticos_en_texto(texto)

        # Procesar diagnósticos
        diagnosticos_resultado = []
        sintomas_derivados = set()

        for d_raw in diagnosticos_detectados_raw:
            d_nombre = d_raw['nombre']
            d_sintomas = d_raw['sintomas']

            # Verificar si el diagnóstico existe en la BD
            d_bd = conn.execute(
                'SELECT id FROM diagnosticos WHERE LOWER(descripcion) = %s',
                (d_nombre.lower(),)
            ).fetchone()

            d_id = d_bd['id'] if d_bd else None

            diagnosticos_resultado.append({
                'nombre': d_nombre.title(),
                'id': d_id,
                'sintomas': d_sintomas,
                'nuevo': not d_id
            })

            # Agregar síntomas derivados
            for sintoma in d_sintomas:
                sintomas_derivados.add(normalizar(sintoma))

        # Extraer síntomas directos del texto
        sintomas_heuristica = extraer_sugeridos_de_texto_avanzado(texto)
        for s in sintomas_heuristica:
            sintomas_derivados.add(normalizar(s))

        # Obtener todos los síntomas de la BD
        sintomas_db = conn.execute('SELECT id, nombre FROM sintomas').fetchall()

        # Matchear síntomas detectados con los de la BD
        sintomas_resultado = []
        for s_norm in sorted(sintomas_derivados):
            encontrado = next(
                (x for x in sintomas_db if normalizar(x['nombre']) == s_norm),
                None
            )

            if encontrado:
                sintomas_resultado.append({
                    'label': encontrado['nombre'],
                    'id': encontrado['id'],
                    'nuevo': False
                })
            else:
                sintomas_resultado.append({
                    'label': s_norm.title(),
                    'id': None,
                    'nuevo': True
                })

        conn.close()

        return jsonify({
            'ok': True,
            'validacion': {
                'coincide': coincide,
                'confianza': confianza,
                'mensaje': f"Confianza: {confianza}%" if not coincide else "Texto validado"
            },
            'diagnosticos': diagnosticos_resultado,
            'sintomas': sintomas_resultado
        })

    except Exception as e:
        print(f"Error procesando texto: {e}")
        traceback.print_exc()
        return jsonify({'ok': False, 'error': str(e)}), 500


@app.route('/admin/sugerir-sintomas/guardar/<int:med_id>', methods=['POST'])
@admin_required
def guardar_sugerencias_sintomas(med_id):
    """Guarda los síntomas y diagnósticos seleccionados"""
    try:
        data = request.get_json()
        sintomas_ids = data.get('sintomas', [])  # Lista de IDs
        diagnosticos_ids = data.get('diagnosticos', [])  # Lista de IDs
        nuevos_sintomas = data.get('nuevos_sintomas', [])  # Lista de nombres
        nuevos_diagnosticos = data.get('nuevos_diagnosticos', [])  # Lista de {nombre, sintomas}

        conn = get_db_connection()

        # 1. Crear nuevos síntomas si es necesario
        sintomas_creados_ids = []
        for nombre_sintoma in nuevos_sintomas:
            # Verificar que no exista
            existe = conn.execute(
                'SELECT id FROM sintomas WHERE LOWER(nombre) = %s',
                (nombre_sintoma.lower(),)
            ).fetchone()

            if not existe:
                conn.execute(
                    'INSERT INTO sintomas (nombre) VALUES (%s)',
                    (nombre_sintoma.title(),)
                )
                conn.commit()

                nuevo_id = conn.execute(
                    'SELECT id FROM sintomas WHERE LOWER(nombre) = %s',
                    (nombre_sintoma.lower(),)
                ).fetchone()['id']

                sintomas_creados_ids.append(nuevo_id)

        # 2. Crear nuevos diagnósticos si es necesario
        diagnosticos_creados_ids = []
        for diag_data in nuevos_diagnosticos:
            nombre_diag = diag_data['nombre']

            # Verificar que no exista
            existe = conn.execute(
                'SELECT id FROM diagnosticos WHERE LOWER(descripcion) = %s',
                (nombre_diag.lower(),)
            ).fetchone()

            if not existe:
                conn.execute(
                    'INSERT INTO diagnosticos (descripcion) VALUES (%s)',
                    (nombre_diag.title(),)
                )
                conn.commit()

                nuevo_id = conn.execute(
                    'SELECT id FROM diagnosticos WHERE LOWER(descripcion) = %s',
                    (nombre_diag.lower(),)
                ).fetchone()['id']

                diagnosticos_creados_ids.append(nuevo_id)

        # 3. Asignar síntomas al medicamento
        todos_sintomas_ids = sintomas_ids + sintomas_creados_ids

        for sintoma_id in todos_sintomas_ids:
            # Verificar que no exista la relación
            existe = conn.execute(
                'SELECT 1 FROM medicamento_sintoma WHERE medicamento_id = %s AND sintoma_id = %s',
                (med_id, sintoma_id)
            ).fetchone()

            if not existe:
                conn.execute(
                    'INSERT INTO medicamento_sintoma (medicamento_id, sintoma_id) VALUES (%s, %s)',
                    (med_id, sintoma_id)
                )

        # 4. Asignar diagnósticos al medicamento
        todos_diagnosticos_ids = diagnosticos_ids + diagnosticos_creados_ids

        for diag_id in todos_diagnosticos_ids:
            # Verificar que no exista la relación
            existe = conn.execute(
                'SELECT 1 FROM medicamento_diagnostico WHERE medicamento_id = %s AND diagnostico_id = %s',
                (med_id, diag_id)
            ).fetchone()

            if not existe:
                conn.execute(
                    'INSERT INTO medicamento_diagnostico (medicamento_id, diagnostico_id) VALUES (%s, %s)',
                    (med_id, diag_id)
                )

        conn.commit()

        # 5. Buscar siguiente medicamento pendiente
        siguiente = conn.execute("""
            SELECT m.id
            FROM medicamentos m
            LEFT JOIN medicamento_sintoma ms ON m.id = ms.medicamento_id
            WHERE ms.sintoma_id IS NULL AND m.activo = 'TRUE' AND m.id != %s
            ORDER BY
                CASE WHEN m.componente_activo_id IS NULL THEN 0 ELSE 1 END,
                (SELECT CASE WHEN p.precio > 0 THEN 0 ELSE 1 END FROM precios p WHERE p.medicamento_id = m.id LIMIT 1),
                m.nombre
            LIMIT 1
        """, (med_id,)).fetchone()

        conn.close()

        return jsonify({
            'ok': True,
            'siguiente_id': siguiente['id'] if siguiente else None,
            'sintomas_guardados': len(todos_sintomas_ids),
            'diagnosticos_guardados': len(todos_diagnosticos_ids)
        })

    except Exception as e:
        print(f"Error guardando síntomas: {e}")
        traceback.print_exc()
        return jsonify({'ok': False, 'error': str(e)}), 500


@app.route('/admin/diagnostico-sintomas', methods=['GET'])
@admin_required
def diagnostico_sintomas():
    """Endpoint temporal para diagnosticar medicamentos pendientes"""
    try:
        conn = get_db_connection()

        # 1. Total medicamentos
        total = conn.execute("SELECT COUNT(*) as total FROM medicamentos").fetchone()

        # 2. Medicamentos activos
        activos = conn.execute("SELECT COUNT(*) as total FROM medicamentos WHERE activo = 'TRUE'").fetchone()

        # 3. Con síntomas
        con_sintomas = conn.execute("""
            SELECT COUNT(DISTINCT m.id) as total
            FROM medicamentos m
            INNER JOIN medicamento_sintoma ms ON m.id = ms.medicamento_id
        """).fetchone()

        # 4. Sin síntomas (activos)
        sin_sintomas = conn.execute("""
            SELECT COUNT(DISTINCT m.id) as total
            FROM medicamentos m
            LEFT JOIN medicamento_sintoma ms ON m.id = ms.medicamento_id
            WHERE ms.sintoma_id IS NULL AND m.activo = 'TRUE'
        """).fetchone()

        # 5. Valores del campo activo
        valores_activo = conn.execute("""
            SELECT DISTINCT activo, COUNT(*) as cantidad
            FROM medicamentos
            GROUP BY activo
        """).fetchall()

        # 6. Primeros 5 sin síntomas
        primeros = conn.execute("""
            SELECT m.id, m.nombre, m.activo, m.componente_activo_id
            FROM medicamentos m
            LEFT JOIN medicamento_sintoma ms ON m.id = ms.medicamento_id
            WHERE ms.sintoma_id IS NULL
            ORDER BY m.nombre
            LIMIT 5
        """).fetchall()

        conn.close()

        return jsonify({
            'ok': True,
            'total_medicamentos': total['total'],
            'medicamentos_activos': activos['total'],
            'con_sintomas': con_sintomas['total'],
            'sin_sintomas_activos': sin_sintomas['total'],
            'valores_activo': [{'valor': v['activo'], 'cantidad': v['cantidad']} for v in valores_activo],
            'primeros_5_sin_sintomas': [{'id': m['id'], 'nombre': m['nombre'], 'activo': m['activo']} for m in primeros]
        })

    except Exception as e:
        print(f"Error en diagnóstico: {e}")
        traceback.print_exc()
        return jsonify({'ok': False, 'error': str(e)}), 500


if __name__ == '__main__':
    #  LLAMADA AL INICIALIZADOR DE DATOS EXTERNO
    #initialize_full_db()#
    
    # Despus de inicializar, puedes ejecutar Flask
    app.run(debug=True, host='0.0.0.0')
