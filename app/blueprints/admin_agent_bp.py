import io
import json
import secrets
from datetime import datetime, timezone

from flask import Blueprint, jsonify, redirect, render_template, request, send_file, session, url_for

from ..db import get_db_connection
from .auth import admin_required, solo_admin

bp = Blueprint('admin_agent', __name__)

_tablas_listas = False


def _crear_tablas(conn):
    global _tablas_listas
    if _tablas_listas:
        return
    sqls = [
        """CREATE TABLE IF NOT EXISTS admin_agent_sesiones (
            id SERIAL PRIMARY KEY,
            cliente_id VARCHAR(100) NOT NULL,
            token VARCHAR(100) NOT NULL UNIQUE,
            activo BOOLEAN DEFAULT TRUE,
            nombre VARCHAR(200),
            ip_local VARCHAR(50),
            ruta_bd VARCHAR(500),
            ultimo_ping TIMESTAMPTZ DEFAULT NOW(),
            created_at TIMESTAMPTZ DEFAULT NOW()
        )""",
        """CREATE TABLE IF NOT EXISTS admin_agent_consultas (
            id SERIAL PRIMARY KEY,
            sesion_id INTEGER REFERENCES admin_agent_sesiones(id),
            tipo VARCHAR(50) NOT NULL,
            parametros JSONB DEFAULT '{}',
            respuesta JSONB,
            estado VARCHAR(20) DEFAULT 'pendiente',
            created_at TIMESTAMPTZ DEFAULT NOW(),
            respondida_at TIMESTAMPTZ
        )""",
        """CREATE TABLE IF NOT EXISTS admin_agent_permisos (
            id SERIAL PRIMARY KEY,
            usuario_id INTEGER NOT NULL,
            cliente_id VARCHAR(100) NOT NULL,
            UNIQUE(usuario_id, cliente_id)
        )""",
        """CREATE TABLE IF NOT EXISTS admin_agent_reportes (
            id SERIAL PRIMARY KEY,
            cliente_id VARCHAR(100),
            tipo VARCHAR(50),
            estado VARCHAR(20),
            detalle TEXT,
            ip VARCHAR(50),
            created_at TIMESTAMPTZ DEFAULT NOW()
        )""",
        """CREATE TABLE IF NOT EXISTS reporte_permisos (
            id SERIAL PRIMARY KEY,
            reporte_id VARCHAR(100) NOT NULL,
            cliente_id VARCHAR(100) NOT NULL,
            fuente VARCHAR(10) NOT NULL DEFAULT 'local',
            UNIQUE(reporte_id, cliente_id, fuente)
        )""",
    ]
    for sql in sqls:
        conn.execute(sql)
    conn.commit()
    for alter in [
        'ALTER TABLE admin_agent_sesiones ADD COLUMN IF NOT EXISTS nombre VARCHAR(200)',
        'ALTER TABLE admin_agent_sesiones ADD COLUMN IF NOT EXISTS ip_local VARCHAR(50)',
        'ALTER TABLE admin_agent_sesiones ADD COLUMN IF NOT EXISTS ruta_bd VARCHAR(500)',
        'ALTER TABLE reporte_permisos ADD COLUMN IF NOT EXISTS usuario_id INTEGER',
        'ALTER TABLE reporte_permisos ALTER COLUMN cliente_id DROP NOT NULL',
    ]:
        try:
            conn.execute(alter)
            conn.commit()
        except Exception:
            pass
    _tablas_listas = True


def _puede_ver(usuario_id, rol, cliente_id, conn):
    if rol == 'Administrador':
        return True
    row = conn.execute(
        "SELECT 1 FROM admin_agent_permisos WHERE usuario_id=%s AND cliente_id=%s",
        (usuario_id, cliente_id)
    ).fetchone()
    return bool(row)


# ── API agente (llamada desde admin_agent.py en PC cliente) ───────────────────

@bp.route('/api/admin-agent/checkin', methods=['POST'])
def checkin():
    data = request.get_json() or {}
    cliente_id = data.get('cliente_id', '').strip()
    if not cliente_id:
        return jsonify({'ok': False, 'error': 'cliente_id requerido'}), 400
    nombre   = (data.get('nombre') or cliente_id).strip()
    ip_local = (data.get('ip_local') or '').strip()
    ruta_bd  = (data.get('ruta_bd') or '').strip()
    conn = get_db_connection()
    try:
        _crear_tablas(conn)
        # Si ya hay sesión activa con ping reciente (< 25s), reutilizarla
        existing = conn.execute(
            "SELECT token FROM admin_agent_sesiones "
            "WHERE cliente_id=%s AND activo=TRUE "
            "AND ultimo_ping > NOW() - INTERVAL '25 seconds' "
            "ORDER BY ultimo_ping DESC LIMIT 1",
            (cliente_id,)
        ).fetchone()
        if existing:
            conn.execute(
                "UPDATE admin_agent_sesiones SET nombre=%s, ip_local=%s WHERE token=%s",
                (nombre, ip_local, existing['token'])
            )
            conn.commit()
            return jsonify({'ok': True, 'token': existing['token'], 'reused': True})
        token = secrets.token_hex(24)
        conn.execute("UPDATE admin_agent_sesiones SET activo=FALSE WHERE cliente_id=%s", (cliente_id,))
        conn.execute(
            "INSERT INTO admin_agent_sesiones (cliente_id, token, nombre, ip_local, ruta_bd) VALUES (%s,%s,%s,%s,%s)",
            (cliente_id, token, nombre, ip_local, ruta_bd)
        )
        conn.commit()
        return jsonify({'ok': True, 'token': token})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500
    finally:
        conn.close()


@bp.route('/api/admin-agent/reporte', methods=['POST'])
def reporte():
    """Recibe reportes de instalación y arranque desde PCs clientes."""
    data = request.get_json() or {}
    cliente_id = data.get('cliente_id', '')
    tipo       = data.get('tipo', '')
    estado     = data.get('estado', '')
    detalle    = data.get('detalle', '')
    ip         = data.get('ip', '')
    conn = get_db_connection()
    try:
        _crear_tablas(conn)
        conn.execute(
            "INSERT INTO admin_agent_reportes (cliente_id, tipo, estado, detalle, ip) VALUES (%s,%s,%s,%s,%s)",
            (cliente_id, tipo, estado, detalle, ip)
        )
        # Notificar a Merlin via chat_mensajes canal='captura':
        # - instalacion: siempre (primera vez es crítico saberlo)
        # - arranque_*: solo si hay error (arranques ok son rutinarios)
        _notificar = tipo == 'instalacion' or (tipo.startswith('arranque') and estado != 'ok')
        if _notificar:
            resumen_detalle = detalle[-1500:] if len(detalle) > 1500 else detalle
            texto = (
                f"[REPORTE {cliente_id.upper()}] tipo={tipo} estado={estado} ip={ip}\n"
                f"{resumen_detalle}"
            )
            try:
                conn.execute(
                    "INSERT INTO chat_mensajes (rol, contenido, canal) VALUES ('user',%s,'captura')",
                    (texto,)
                )
            except Exception:
                pass  # Si chat_mensajes no existe en este entorno, no bloquear
        conn.commit()
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500
    finally:
        conn.close()


@bp.route('/api/admin-agent/reportes', methods=['GET'])
@admin_required
def ver_reportes():
    """Lista los últimos reportes para que Merlin pueda consultarlos."""
    conn = get_db_connection()
    try:
        _crear_tablas(conn)
        rows = conn.execute(
            "SELECT cliente_id, tipo, estado, detalle, ip, created_at FROM admin_agent_reportes ORDER BY created_at DESC LIMIT 50"
        ).fetchall()
        return jsonify({'ok': True, 'reportes': [dict(r) for r in rows]})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500
    finally:
        conn.close()


@bp.route('/api/admin-agent/checkout', methods=['POST'])
def checkout():
    token = (request.get_json() or {}).get('token', '')
    conn = get_db_connection()
    try:
        _crear_tablas(conn)
        conn.execute("UPDATE admin_agent_sesiones SET activo=FALSE WHERE token=%s", (token,))
        conn.commit()
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500
    finally:
        conn.close()


@bp.route('/api/admin-agent/ping', methods=['POST'])
def ping():
    data    = request.get_json() or {}
    token   = data.get('token', '')
    ruta_bd = (data.get('ruta_bd') or '').strip() or None
    nombre  = (data.get('nombre') or '').strip() or None
    conn = get_db_connection()
    try:
        _crear_tablas(conn)
        sesion = conn.execute(
            "UPDATE admin_agent_sesiones SET ultimo_ping=NOW(), ruta_bd=COALESCE(%s,ruta_bd), nombre=COALESCE(%s,nombre) "
            "WHERE token=%s AND activo=TRUE RETURNING id",
            (ruta_bd, nombre, token)
        ).fetchone()
        if not sesion:
            return jsonify({'ok': False, 'error': 'sesión inválida'}), 401
        consulta = conn.execute(
            "SELECT id, tipo, parametros FROM admin_agent_consultas "
            "WHERE sesion_id=%s AND estado='pendiente' ORDER BY id ASC LIMIT 1",
            (sesion['id'],)
        ).fetchone()
        if consulta:
            conn.execute("UPDATE admin_agent_consultas SET estado='procesando' WHERE id=%s", (consulta['id'],))
            conn.commit()
            return jsonify({'ok': True, 'consulta': {
                'id': consulta['id'],
                'tipo': consulta['tipo'],
                'parametros': consulta['parametros'],
            }})
        conn.commit()
        return jsonify({'ok': True, 'consulta': None})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500
    finally:
        conn.close()


@bp.route('/api/admin-agent/respuesta', methods=['POST'])
def respuesta():
    data = request.get_json() or {}
    token       = data.get('token', '')
    consulta_id = data.get('consulta_id')
    resp_data   = data.get('respuesta')
    error       = data.get('error')
    conn = get_db_connection()
    try:
        _crear_tablas(conn)
        sesion = conn.execute(
            "SELECT id FROM admin_agent_sesiones WHERE token=%s AND activo=TRUE", (token,)
        ).fetchone()
        if not sesion:
            return jsonify({'ok': False, 'error': 'sesión inválida'}), 401
        from psycopg2.extras import Json as PgJson
        payload = PgJson(resp_data if resp_data is not None else {'error': error})
        estado  = 'error' if error else 'lista'
        conn.execute(
            "UPDATE admin_agent_consultas SET respuesta=%s, estado=%s, respondida_at=NOW() "
            "WHERE id=%s AND sesion_id=%s",
            (payload, estado, consulta_id, sesion['id'])
        )
        conn.commit()
        return jsonify({'ok': True})
    except Exception as e:
        import traceback, logging
        logging.error(f"respuesta 500: {e}\n{traceback.format_exc()}")
        return jsonify({'ok': False, 'error': str(e)}), 500
    finally:
        conn.close()


# ── API browser ───────────────────────────────────────────────────────────────

@bp.route('/api/admin-agent/consultar', methods=['POST'])
def consultar():
    if 'usuario_id' not in session:
        return jsonify({'ok': False, 'error': 'No autenticado'}), 401
    data       = request.get_json() or {}
    cliente_id = data.get('cliente_id', '').strip()
    tipo       = data.get('tipo', '').strip()
    parametros = data.get('parametros', {})
    if not cliente_id or not tipo:
        return jsonify({'ok': False, 'error': 'cliente_id y tipo requeridos'}), 400
    conn = get_db_connection()
    try:
        _crear_tablas(conn)
        if not _puede_ver(session['usuario_id'], session.get('rol', ''), cliente_id, conn):
            return jsonify({'ok': False, 'error': 'No autorizado'}), 403
        sesion = conn.execute(
            "SELECT id FROM admin_agent_sesiones WHERE cliente_id=%s AND activo=TRUE ORDER BY id DESC LIMIT 1",
            (cliente_id,)
        ).fetchone()
        if not sesion:
            return jsonify({'ok': False, 'error': 'Agente no conectado'}), 404
        row = conn.execute(
            "INSERT INTO admin_agent_consultas (sesion_id, tipo, parametros) VALUES (%s,%s,%s) RETURNING id",
            (sesion['id'], tipo, json.dumps(parametros))
        ).fetchone()
        conn.commit()
        return jsonify({'ok': True, 'consulta_id': row['id']})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500
    finally:
        conn.close()


@bp.route('/api/admin-agent/resultado/<int:consulta_id>', methods=['GET'])
def resultado(consulta_id):
    if 'usuario_id' not in session:
        return jsonify({'ok': False, 'error': 'No autenticado'}), 401
    conn = get_db_connection()
    try:
        _crear_tablas(conn)
        row = conn.execute(
            "SELECT estado, respuesta FROM admin_agent_consultas WHERE id=%s", (consulta_id,)
        ).fetchone()
        if not row:
            return jsonify({'ok': False, 'error': 'No encontrada'}), 404
        resp = {'ok': True, 'estado': row['estado'], 'respuesta': row['respuesta']}
        if row['estado'] == 'error':
            r = row['respuesta']
            if isinstance(r, dict):
                resp['error'] = r.get('error', str(r))
            else:
                try:
                    resp['error'] = json.loads(r).get('error', r)
                except Exception:
                    resp['error'] = str(r)
        if row['estado'] in ('lista', 'error'):
            try:
                conn.execute("DELETE FROM admin_agent_consultas WHERE id=%s", (consulta_id,))
                conn.commit()
            except Exception:
                pass
        return jsonify(resp)
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500
    finally:
        conn.close()


@bp.route('/api/admin-agent/excel', methods=['POST'])
def excel():
    if 'usuario_id' not in session:
        return jsonify({'ok': False, 'error': 'No autenticado'}), 401
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        return jsonify({'ok': False, 'error': 'openpyxl no instalado'}), 500

    data    = request.get_json() or {}
    rows    = data.get('rows', [])
    titulo  = data.get('titulo', 'REG_CTAS')
    desde   = data.get('desde', '')
    hasta   = data.get('hasta', '')
    empresa = data.get('empresa', '')
    agente  = data.get('agente', '')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'REG_CTAS'

    hdr_font = Font(bold=True, color='003F7F')
    hdr_fill = PatternFill('solid', fgColor='DDEEFF')
    tot_font = Font(bold=True)

    ws.append([titulo])
    ws['A1'].font = Font(bold=True, size=13)
    filtro = f'Desde: {desde}    Hasta: {hasta}'
    if empresa:
        filtro += f'    Empresa: {empresa}'
    if agente:
        filtro += f'    Agente: {agente}'
    ws.append([filtro])
    ws.append([])

    cabecera = ['CONSEC', 'FECHA', 'LAPSO', 'TIPO', 'DOCUMENTO',
                'CUENTA', 'NIT', 'TERCERO', 'DÉBITO', 'CRÉDITO', 'DETALLE', 'ANULADO']
    ws.append(cabecera)
    hdr_row = ws.max_row
    for cell in ws[hdr_row]:
        cell.font = hdr_font
        cell.fill = hdr_fill

    for r in rows:
        tipo_txt = str(r.get('tipo') or '')
        if r.get('tipo_nom'):
            tipo_txt += ' ' + r['tipo_nom']
        ws.append([
            r.get('consecutivo', ''),
            (r.get('fecha') or '')[:10],
            r.get('lapso', ''),
            tipo_txt.strip(),
            r.get('documento', ''),
            r.get('cuenta', ''),
            r.get('nit', ''),
            r.get('tercero_nom') or r.get('tercero', ''),
            r.get('debito') or 0,
            r.get('credito') or 0,
            r.get('detalle', ''),
            'Sí' if r.get('anulado') else '',
        ])

    tot_deb = sum((r.get('debito') or 0) for r in rows)
    tot_cre = sum((r.get('credito') or 0) for r in rows)
    ws.append([f'TOTAL ({len(rows)} registros)', '', '', '', '', '', '', '', tot_deb, tot_cre, '', ''])
    for cell in ws[ws.max_row]:
        cell.font = tot_font

    for col, width in zip('ABCDEFGHIJKL', [10, 12, 10, 16, 14, 16, 14, 32, 14, 14, 32, 7]):
        ws.column_dimensions[col].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    import datetime as _dt
    fname = f'REG_CTAS_{_dt.date.today().isoformat()}.xlsx'
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@bp.route('/api/admin-agent/estado/<cliente_id>', methods=['GET'])
def estado(cliente_id):
    if 'usuario_id' not in session:
        return jsonify({'ok': False, 'error': 'No autenticado'}), 401
    conn = get_db_connection()
    try:
        _crear_tablas(conn)
        if not _puede_ver(session['usuario_id'], session.get('rol', ''), cliente_id, conn):
            return jsonify({'ok': False, 'error': 'No autorizado'}), 403
        sesion = conn.execute(
            "SELECT ultimo_ping FROM admin_agent_sesiones "
            "WHERE cliente_id=%s AND activo=TRUE ORDER BY id DESC LIMIT 1",
            (cliente_id,)
        ).fetchone()
        if not sesion:
            return jsonify({'ok': True, 'conectado': False})
        ping = sesion['ultimo_ping']
        if ping.tzinfo is None:
            ping = ping.replace(tzinfo=timezone.utc)
        lag = (datetime.now(timezone.utc) - ping).total_seconds()
        return jsonify({'ok': True, 'conectado': lag < 30})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500
    finally:
        conn.close()


@bp.route('/api/admin-agent/agentes', methods=['GET'])
def agentes():
    if 'usuario_id' not in session:
        return jsonify({'ok': False, 'error': 'No autenticado'}), 401
    conn = get_db_connection()
    try:
        _crear_tablas(conn)
        if session.get('rol') == 'Administrador':
            rows = conn.execute("""
                SELECT DISTINCT ON (cliente_id)
                    cliente_id, nombre, ip_local, ruta_bd, activo, ultimo_ping
                FROM admin_agent_sesiones
                WHERE ultimo_ping > NOW() - INTERVAL '24 hours'
                ORDER BY cliente_id, id DESC
            """).fetchall()
        else:
            rows = conn.execute("""
                SELECT DISTINCT ON (s.cliente_id)
                    s.cliente_id, s.nombre, s.ip_local, s.ruta_bd, s.activo, s.ultimo_ping
                FROM admin_agent_sesiones s
                JOIN admin_agent_permisos p ON p.cliente_id = s.cliente_id
                WHERE p.usuario_id = %s
                  AND s.ultimo_ping > NOW() - INTERVAL '24 hours'
                ORDER BY s.cliente_id, s.id DESC
            """, (session['usuario_id'],)).fetchall()
        now = datetime.now(timezone.utc)
        result = []
        for r in rows:
            ping = r['ultimo_ping']
            if ping and ping.tzinfo is None:
                ping = ping.replace(tzinfo=timezone.utc)
            lag = int((now - ping).total_seconds()) if ping else 9999
            result.append({
                'cliente_id': r['cliente_id'],
                'nombre':     r['nombre'] or r['cliente_id'],
                'ip_local':   r['ip_local'] or '',
                'ruta_bd':    r['ruta_bd'] or '',
                'conectado':  bool(r['activo']) and lag < 30,
                'lag':        lag,
            })
        return jsonify({'ok': True, 'agentes': result})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500
    finally:
        conn.close()


@bp.route('/api/admin-agent/permisos/usuario/<int:uid>', methods=['GET'])
@solo_admin
def permisos_get(uid):
    conn = get_db_connection()
    try:
        _crear_tablas(conn)
        rows = conn.execute(
            "SELECT cliente_id FROM admin_agent_permisos WHERE usuario_id=%s ORDER BY cliente_id", (uid,)
        ).fetchall()
        return jsonify({'ok': True, 'clientes': [r['cliente_id'] for r in rows]})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500
    finally:
        conn.close()


@bp.route('/api/admin-agent/permisos/usuario/<int:uid>', methods=['POST'])
@solo_admin
def permisos_set(uid):
    clientes = (request.get_json() or {}).get('clientes', [])
    conn = get_db_connection()
    try:
        _crear_tablas(conn)
        conn.execute("DELETE FROM admin_agent_permisos WHERE usuario_id=%s", (uid,))
        for cid in clientes:
            conn.execute(
                "INSERT INTO admin_agent_permisos (usuario_id, cliente_id) VALUES (%s,%s) ON CONFLICT DO NOTHING",
                (uid, cid.strip())
            )
        conn.commit()
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500
    finally:
        conn.close()


# ── Páginas HTML ──────────────────────────────────────────────────────────────

@bp.route('/admin/agentes')
@solo_admin
def agentes_page():
    conn = get_db_connection()
    try:
        _crear_tablas(conn)
        rows = conn.execute(
            "SELECT id, nombre, usuario FROM usuarios WHERE rol='ClienteVFP' ORDER BY nombre"
        ).fetchall()
        usuarios = [{'id': r['id'], 'nombre': r['nombre'], 'usuario': r['usuario']} for r in rows]
        return render_template('admin_agentes.html', usuarios=usuarios)
    except Exception as e:
        return str(e), 500
    finally:
        conn.close()


# ── Permisos de reportes SAR por cliente ──────────────────────────────────────

REPORTES_CATALOGO = [
    {'id': 'inventario',       'nombre': 'Rotación de Inventario', 'categoria': 'Inventario'},
    {'id': 'ventas_clientes',  'nombre': 'Ventas por Clientes',    'categoria': 'Ventas'},
    {'id': 'consulta_cuentas', 'nombre': 'Consulta de Cuentas',    'categoria': 'Contabilidad'},
    {'id': 'reg_ctas',         'nombre': 'Movimientos REG_CTAS',   'categoria': 'Contabilidad'},
]


# ── Permisos de reportes — por usuario ───────────────────────────────────────

@bp.route('/api/admin-agent/mis-reportes-permisos', methods=['GET'])
def mis_reportes_permisos():
    """SarReportes llama este endpoint — usa la sesión autenticada."""
    uid = session.get('usuario_id')
    if not uid:
        return jsonify({'ok': False, 'error': 'No autenticado'}), 401
    conn = get_db_connection()
    try:
        _crear_tablas(conn)
        rows = conn.execute(
            "SELECT reporte_id, fuente FROM reporte_permisos WHERE usuario_id=%s",
            (uid,)
        ).fetchall()
        return jsonify({'ok': True, 'permisos': [{'reporte_id': r['reporte_id'], 'fuente': r['fuente']} for r in rows]})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500
    finally:
        conn.close()


@bp.route('/api/admin-agent/reportes-permisos-u/<int:uid>', methods=['GET'])
@solo_admin
def reportes_permisos_get_u(uid):
    conn = get_db_connection()
    try:
        _crear_tablas(conn)
        rows = conn.execute(
            "SELECT reporte_id, fuente FROM reporte_permisos WHERE usuario_id=%s",
            (uid,)
        ).fetchall()
        return jsonify({'ok': True, 'permisos': [{'reporte_id': r['reporte_id'], 'fuente': r['fuente']} for r in rows]})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500
    finally:
        conn.close()


@bp.route('/api/admin-agent/reportes-permisos-u/<int:uid>', methods=['POST'])
@solo_admin
def reportes_permisos_set_u(uid):
    permisos = (request.get_json() or {}).get('permisos', [])
    conn = get_db_connection()
    try:
        _crear_tablas(conn)
        conn.execute("DELETE FROM reporte_permisos WHERE usuario_id=%s", (uid,))
        for p in permisos:
            rid    = str(p.get('reporte_id', '')).strip()
            fuente = str(p.get('fuente', 'local')).strip()
            if rid and fuente in ('local', 'remoto'):
                conn.execute(
                    "INSERT INTO reporte_permisos (reporte_id, usuario_id, fuente) VALUES (%s,%s,%s) "
                    "ON CONFLICT DO NOTHING",
                    (rid, uid, fuente)
                )
        conn.commit()
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500
    finally:
        conn.close()


# ── Agentes por usuario ───────────────────────────────────────────────────────

@bp.route('/api/admin-agent/usuario-agentes/<int:uid>', methods=['GET'])
@solo_admin
def usuario_agentes_get(uid):
    conn = get_db_connection()
    try:
        _crear_tablas(conn)
        rows = conn.execute(
            "SELECT cliente_id FROM admin_agent_permisos WHERE usuario_id=%s",
            (uid,)
        ).fetchall()
        return jsonify({'ok': True, 'agentes': [r['cliente_id'] for r in rows]})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500
    finally:
        conn.close()


@bp.route('/api/admin-agent/usuario-agentes/<int:uid>', methods=['POST'])
@solo_admin
def usuario_agentes_set(uid):
    agentes = (request.get_json() or {}).get('agentes', [])
    conn = get_db_connection()
    try:
        _crear_tablas(conn)
        conn.execute("DELETE FROM admin_agent_permisos WHERE usuario_id=%s", (uid,))
        for aid in agentes:
            aid = str(aid).strip()
            if aid:
                conn.execute(
                    "INSERT INTO admin_agent_permisos (usuario_id, cliente_id) VALUES (%s,%s) ON CONFLICT DO NOTHING",
                    (uid, aid)
                )
        conn.commit()
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500
    finally:
        conn.close()


@bp.route('/api/admin-agent/agentes-activos', methods=['GET'])
@solo_admin
def agentes_activos():
    conn = get_db_connection()
    try:
        _crear_tablas(conn)
        rows = conn.execute("""
            SELECT DISTINCT ON (cliente_id) cliente_id, nombre, ultimo_ping
            FROM admin_agent_sesiones
            WHERE ultimo_ping > NOW() - INTERVAL '30 days'
            ORDER BY cliente_id, id DESC
        """).fetchall()
        now = datetime.now(timezone.utc)
        agentes = []
        for r in rows:
            up = r['ultimo_ping']
            if up:
                if up.tzinfo is None:
                    up = up.replace(tzinfo=timezone.utc)
                diff = (now - up).total_seconds()
                online = diff < 300
                mins = int(diff / 60)
                if mins < 60:
                    ultimo = f'hace {mins}min'
                elif mins < 1440:
                    ultimo = f'hace {mins // 60}h'
                else:
                    ultimo = f'hace {mins // 1440}d'
            else:
                online = False
                ultimo = 'desconocido'
            agentes.append({
                'id':     r['cliente_id'],
                'nombre': r['nombre'] or r['cliente_id'],
                'online': online,
                'ultimo': ultimo,
            })
        return jsonify({'ok': True, 'agentes': agentes})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500
    finally:
        conn.close()


# ── CRUD usuarios SAR ─────────────────────────────────────────────────────────

@bp.route('/api/admin-agent/sar-usuarios', methods=['GET'])
@solo_admin
def sar_usuarios_list():
    conn = get_db_connection()
    try:
        rows = conn.execute(
            "SELECT id, usuario, nombre, rol FROM usuarios "
            "WHERE rol IN ('Administrador','ClienteVFP') ORDER BY nombre"
        ).fetchall()
        return jsonify({'ok': True, 'usuarios': [dict(r) for r in rows]})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500
    finally:
        conn.close()


@bp.route('/api/admin-agent/sar-usuarios', methods=['POST'])
@solo_admin
def sar_usuarios_crear():
    data    = request.get_json() or {}
    usuario = str(data.get('usuario',  '')).strip()
    password = str(data.get('password', '')).strip()
    nombre  = str(data.get('nombre',   '')).strip()
    rol     = str(data.get('rol', 'ClienteVFP')).strip()
    if not usuario or not password or not nombre:
        return jsonify({'ok': False, 'error': 'usuario, password y nombre son requeridos'}), 400
    if rol not in ('Administrador', 'ClienteVFP'):
        return jsonify({'ok': False, 'error': 'rol inválido'}), 400
    conn = get_db_connection()
    try:
        if conn.execute("SELECT 1 FROM usuarios WHERE usuario=%s", (usuario,)).fetchone():
            return jsonify({'ok': False, 'error': 'Usuario ya existe'}), 409
        row = conn.execute(
            "INSERT INTO usuarios (usuario, password, nombre, rol) VALUES (%s,%s,%s,%s) RETURNING id",
            (usuario, password, nombre, rol)
        ).fetchone()
        conn.commit()
        return jsonify({'ok': True, 'id': row['id']})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500
    finally:
        conn.close()


@bp.route('/api/admin-agent/sar-usuarios/<int:uid>', methods=['PATCH'])
@solo_admin
def sar_usuarios_update(uid):
    data = request.get_json() or {}
    conn = get_db_connection()
    try:
        if data.get('password'):
            conn.execute("UPDATE usuarios SET password=%s WHERE id=%s", (data['password'], uid))
        if data.get('nombre'):
            conn.execute("UPDATE usuarios SET nombre=%s WHERE id=%s", (data['nombre'], uid))
        conn.commit()
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500
    finally:
        conn.close()


@bp.route('/api/admin-agent/sar-usuarios/<int:uid>', methods=['DELETE'])
@solo_admin
def sar_usuarios_delete(uid):
    if uid == session.get('usuario_id'):
        return jsonify({'ok': False, 'error': 'No puedes eliminarte a ti mismo'}), 400
    conn = get_db_connection()
    try:
        conn.execute("DELETE FROM reporte_permisos WHERE usuario_id=%s", (uid,))
        conn.execute("DELETE FROM admin_agent_permisos WHERE usuario_id=%s", (uid,))
        conn.execute("DELETE FROM usuarios WHERE id=%s AND rol != 'Administrador'", (uid,))
        conn.commit()
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500
    finally:
        conn.close()


# ── Página de permisos ────────────────────────────────────────────────────────

@bp.route('/admin/permisos-reportes')
@solo_admin
def permisos_reportes_page():
    return render_template('admin_permisos_reportes.html', reportes=REPORTES_CATALOGO)


# ── Versiones de agentes — auto-update ────────────────────────────────────────
# Para publicar nueva versión: actualizar aquí + subir EXE al release de GitHub.
# La URL base apunta al release tag correspondiente.

VERSIONES_AGENTES = {
    'sar_reportes': {
        'version': '1.2.2',
        'url': 'https://github.com/Rafaelolivares07/TUC-TUC/releases/download/agentes/SarReportes.exe',
    },
    'admin_agent': {
        'version': '1.2.2',
        'url': 'https://github.com/Rafaelolivares07/TUC-TUC/releases/download/agentes/AdminAgent.exe',
    },
}


@bp.route('/api/version/agentes', methods=['GET'])
def api_version_agentes():
    """Público — sin auth. Los EXEs locales lo consultan al arrancar."""
    return jsonify({'ok': True, **VERSIONES_AGENTES})
