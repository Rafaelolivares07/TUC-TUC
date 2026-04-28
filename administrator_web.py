"""
administrator_web.py — Módulo web local para Administrator VFP
Sirve formularios HTML desde localhost para consultar DBF.
Abre desde VFP con: RUN /N chrome.exe http://localhost:5002/ventas_clientes

Estrategia de migración:
  - Hoy: lee directamente de DBF local
  - Futuro: cambia leer_ruta_bd() y las funciones de datos por consultas PostgreSQL
  - El frontend (templates) no cambia
"""

import os, sys, struct, datetime, io
import dbf
from flask import Flask, render_template, request, jsonify, send_file

try:
    import numpy as np
    HAS_NUMPY = True
except ImportError:
    HAS_NUMPY = False

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

app = Flask(__name__)

@app.before_request
def _lowercase_url():
    from flask import redirect, request
    if request.path != request.path.lower():
        return redirect(request.path.lower(), 301)

RUTA_DBF_FILE = r"C:\S.A.R\RutaBaseDatos\ruta.dbf"
ENC = 'cp1252'

# ── Offsets PROD_FACT1 (record_size=179) ──────────────────────────────────────
PF_REC_SIZE  = 179
PF_CANTIDAD  = 49   # N(10,4)
PF_PRECIO    = 59   # N(10,2)
PF_DESCUENTO = 69   # N(10,0)
PF_EMPRESA   = 91   # C(4)
PF_FECHAHORA = 95   # T(8)  — Julian Day 4 bytes LE + ms 4 bytes LE
PF_POR_IVA   = 103  # N(10,2)
PF_CLIENTE   = 126  # N(10)

# ── Offsets TERCEROS (record_size=739) ────────────────────────────────────────
TER_REC_SIZE = 739
TER_COD_TER  = 1    # N(10)
TER_NOMBRE   = 11   # C(50)
TER_IDENT    = 61   # C(15)

# ── Offsets REG_CTAS (record_size=345) ────────────────────────────────────────
RC_REC_SIZE = 345
RC_CUENTA   = 11   # C(15)
RC_LAPSO    = 26   # D(8)  YYYYMMDD ASCII
RC_TERCERO  = 42   # N(10)
RC_EMPRESA  = 52   # C(10)
RC_TOTDEB   = 108  # N(10)
RC_TOTCRE   = 118  # N(10)
RC_ANULADO  = 344  # N(1)


# ════════════════════════════════════════════════════════════════════════════════
#  CAPA DE DATOS — reemplazar estas funciones al migrar a PostgreSQL
# ════════════════════════════════════════════════════════════════════════════════

def leer_ruta_bd():
    t = dbf.Table(RUTA_DBF_FILE, ignore_memos=True)
    t.open(dbf.READ_ONLY)
    ruta = None
    for r in t:
        ruta = r['RUTA'].strip()
        break
    t.close()
    return os.path.dirname(ruta)


def _date_to_jd(d):
    """Convierte datetime.date a Julian Day Number (mismo algoritmo que VFP)."""
    a = (14 - d.month) // 12
    y = d.year + 4800 - a
    m = d.month + 12*a - 3
    return d.day + (153*m + 2)//5 + 365*y + y//4 - y//100 + y//400 - 32045


def _parse_n(b, offset, length):
    v = b[offset:offset+length].strip(b' ')
    try:
        return float(v) if v else 0.0
    except Exception:
        return 0.0


def leer_terceros(ruta_bd):
    """Retorna dict {cod_ter: {nombre, identificacion}} leyendo TERCEROS en binario."""
    path = os.path.join(ruta_bd, 'TERCEROS.DBF')
    with open(path, 'rb') as f:
        hdr = f.read(32)
        num = struct.unpack_from('<I', hdr, 4)[0]
        hsz = struct.unpack_from('<H', hdr, 8)[0]
        f.seek(hsz)
        raw = f.read(num * TER_REC_SIZE)
    result = {}
    n = len(raw) // TER_REC_SIZE
    for i in range(n):
        b = raw[i*TER_REC_SIZE:(i+1)*TER_REC_SIZE]
        if b[0] == 0x2A:
            continue
        cod    = b[TER_COD_TER:TER_COD_TER+10].strip(b' ').decode(ENC, 'replace').strip()
        nombre = b[TER_NOMBRE:TER_NOMBRE+50].strip(b' ').decode(ENC, 'replace').strip()
        ident  = b[TER_IDENT:TER_IDENT+15].strip(b' ').decode(ENC, 'replace').strip()
        if cod:
            result[cod] = {'nombre': nombre, 'identificacion': ident}
    return result


def leer_ventas_clientes(ruta_bd, desde, hasta, empresa=''):
    """
    Lee PROD_FACT1 y agrupa por CLIENTE.
    Retorna dict {cod_ter: {cantidad, subtotal, iva}}.
    """
    path = os.path.join(ruta_bd, 'PROD_FACT1.DBF')
    jd_desde = _date_to_jd(desde)
    jd_hasta = _date_to_jd(hasta)

    with open(path, 'rb') as f:
        hdr = f.read(32)
        num = struct.unpack_from('<I', hdr, 4)[0]
        hsz = struct.unpack_from('<H', hdr, 8)[0]
        f.seek(hsz)
        raw = f.read(num * PF_REC_SIZE)

    if HAS_NUMPY:
        data = np.frombuffer(raw[:len(raw) // PF_REC_SIZE * PF_REC_SIZE], dtype=np.uint8)
        n    = len(data) // PF_REC_SIZE
        data = data.reshape(n, PF_REC_SIZE)

        mask = data[:, 0] != 0x2A

        # Filtro fecha — Julian Day en FECHAHORA (4 bytes LE)
        jd = (data[:, PF_FECHAHORA].astype(np.int64) |
              (data[:, PF_FECHAHORA+1].astype(np.int64) << 8) |
              (data[:, PF_FECHAHORA+2].astype(np.int64) << 16) |
              (data[:, PF_FECHAHORA+3].astype(np.int64) << 24))
        mask &= (jd >= jd_desde) & (jd <= jd_hasta)

        # Filtro empresa
        emp = empresa.strip()
        if emp:
            emp_b = emp.encode(ENC).ljust(4)[:4]
            emp_match = np.all(data[:, PF_EMPRESA:PF_EMPRESA+4] ==
                               np.frombuffer(emp_b, dtype=np.uint8), axis=1)
            mask &= emp_match

        indices = np.where(mask)[0]
    else:
        indices = range(len(raw) // PF_REC_SIZE)

    clientes = {}
    emp = empresa.strip()
    for idx in indices:
        b = raw[int(idx)*PF_REC_SIZE:(int(idx)+1)*PF_REC_SIZE]
        if not HAS_NUMPY:
            if b[0] == 0x2A:
                continue
            if emp and b[PF_EMPRESA:PF_EMPRESA+4].rstrip(b' ').decode(ENC,'replace').strip() != emp:
                continue
        cliente = b[PF_CLIENTE:PF_CLIENTE+10].strip(b' ').decode(ENC, 'replace').strip()
        if not cliente or cliente == '0':
            continue
        cantidad = _parse_n(b, PF_CANTIDAD, 10)
        precio   = _parse_n(b, PF_PRECIO, 10)
        descto   = _parse_n(b, PF_DESCUENTO, 10)
        por_iva  = _parse_n(b, PF_POR_IVA, 10)
        subtotal = cantidad * precio - descto
        iva      = cantidad * precio * (por_iva / 100)
        if cliente not in clientes:
            clientes[cliente] = {'cantidad': 0.0, 'subtotal': 0.0, 'iva': 0.0}
        clientes[cliente]['cantidad'] += cantidad
        clientes[cliente]['subtotal'] += subtotal
        clientes[cliente]['iva']      += iva

    return clientes


def calcular_puestos(clientes_dict, terceros):
    """Construye lista de filas con puestos y % participación, ordenada por nombre."""
    rows = []
    for cod, d in clientes_dict.items():
        ter = terceros.get(cod, {})
        rows.append({
            'codigo':         cod,
            'identificacion': ter.get('identificacion', ''),
            'nombre':         ter.get('nombre', cod),
            'cantidad':       round(d['cantidad'], 2),
            'subtotal':       round(d['subtotal'], 2),
            'iva':            round(d['iva'], 2),
        })

    rows = [r for r in rows if r['nombre']]
    total_cant = sum(r['cantidad'] for r in rows) or 1
    total_val  = sum(r['subtotal'] for r in rows) or 1

    # Puesto por cantidades
    rows.sort(key=lambda r: r['cantidad'], reverse=True)
    for i, r in enumerate(rows):
        r['puesto_cant'] = i + 1
        r['por_cant']    = round(r['cantidad'] / total_cant * 100, 2)

    # Puesto por valores
    puesto_val = {r['codigo']: i+1
                  for i, r in enumerate(sorted(rows, key=lambda r: r['subtotal'], reverse=True))}
    for r in rows:
        r['puesto_val'] = puesto_val[r['codigo']]
        r['por_val']    = round(r['subtotal'] / total_val * 100, 2)

    rows.sort(key=lambda r: r['nombre'])
    return rows


def leer_cuentas(ruta_bd, q=''):
    """Retorna lista [{codigo, nombre}] para autocomplete de cuentas."""
    path = os.path.join(ruta_bd, 'CUENTAS.DBF')
    t = dbf.Table(path, ignore_memos=True)
    t.open(dbf.READ_ONLY)
    fields = [f.upper() for f in t.field_names]
    f_cod = 'CUENTA' if 'CUENTA' in fields else fields[0]
    f_nom = 'NOMBRE' if 'NOMBRE' in fields else (fields[1] if len(fields) > 1 else fields[0])
    results = []
    q_low = q.lower().strip()
    for r in t:
        cod = str(r[f_cod]).strip()
        nom = str(r[f_nom]).strip()
        if not cod:
            continue
        if q_low and q_low not in cod.lower() and q_low not in nom.lower():
            continue
        results.append({'codigo': cod, 'nombre': nom})
        if len(results) >= 20:
            break
    t.close()
    return results


def leer_reg_ctas(ruta_bd, cuenta, empresa, desde, hasta):
    """Lee REG_CTAS y agrupa por TERCERO. Retorna dict {cod_ter: {tot_deb, tot_cre}}."""
    path = os.path.join(ruta_bd, 'REG_CTAS.DBF')
    desde_i = int(desde.strftime('%Y%m%d'))
    hasta_i = int(hasta.strftime('%Y%m%d'))
    cuenta_upper = cuenta.upper().strip()

    with open(path, 'rb') as f:
        hdr = f.read(32)
        num = struct.unpack_from('<I', hdr, 4)[0]
        hsz = struct.unpack_from('<H', hdr, 8)[0]
        f.seek(hsz)
        raw = f.read(num * RC_REC_SIZE)

    if HAS_NUMPY:
        data = np.frombuffer(raw[:len(raw)//RC_REC_SIZE*RC_REC_SIZE], dtype=np.uint8)
        n    = len(data) // RC_REC_SIZE
        data = data.reshape(n, RC_REC_SIZE)

        mask = data[:, 0] != 0x2A

        cta_b = cuenta_upper.encode(ENC)
        mask &= np.all(data[:, RC_CUENTA:RC_CUENTA+len(cta_b)] ==
                       np.frombuffer(cta_b, dtype=np.uint8), axis=1)

        lapso = (
            (data[:, RC_LAPSO  ].astype(np.int64) - 48) * 10000000 +
            (data[:, RC_LAPSO+1].astype(np.int64) - 48) * 1000000  +
            (data[:, RC_LAPSO+2].astype(np.int64) - 48) * 100000   +
            (data[:, RC_LAPSO+3].astype(np.int64) - 48) * 10000    +
            (data[:, RC_LAPSO+4].astype(np.int64) - 48) * 1000     +
            (data[:, RC_LAPSO+5].astype(np.int64) - 48) * 100      +
            (data[:, RC_LAPSO+6].astype(np.int64) - 48) * 10       +
            (data[:, RC_LAPSO+7].astype(np.int64) - 48)
        )
        mask &= (lapso >= desde_i) & (lapso <= hasta_i)

        emp = empresa.strip()
        if emp:
            emp_b = emp.encode(ENC)
            mask &= np.all(data[:, RC_EMPRESA:RC_EMPRESA+len(emp_b)] ==
                           np.frombuffer(emp_b, dtype=np.uint8), axis=1)

        mask &= (data[:, RC_ANULADO] != ord('1'))
        indices = np.where(mask)[0]
    else:
        indices = range(len(raw) // RC_REC_SIZE)

    resultado = {}
    emp = empresa.strip()
    for idx in indices:
        b = raw[int(idx)*RC_REC_SIZE:(int(idx)+1)*RC_REC_SIZE]
        if not HAS_NUMPY:
            if b[0] == 0x2A: continue
            cta = b[RC_CUENTA:RC_CUENTA+len(cuenta_upper)].decode(ENC,'replace').upper()
            if not cta.startswith(cuenta_upper): continue
            try:
                lapso_i = int(b[RC_LAPSO:RC_LAPSO+8].decode('ascii','replace'))
            except: continue
            if not (desde_i <= lapso_i <= hasta_i): continue
            if emp and b[RC_EMPRESA:RC_EMPRESA+10].rstrip(b' ').decode(ENC,'replace').strip() != emp: continue
            if b[RC_ANULADO:RC_ANULADO+1] == b'1': continue

        ter = b[RC_TERCERO:RC_TERCERO+10].strip(b' ').decode(ENC,'replace').strip()
        if not ter or ter == '0': continue
        deb = _parse_n(b, RC_TOTDEB, 10)
        cre = _parse_n(b, RC_TOTCRE, 10)
        if ter not in resultado:
            resultado[ter] = {'tot_deb': 0.0, 'tot_cre': 0.0}
        resultado[ter]['tot_deb'] += deb
        resultado[ter]['tot_cre'] += cre

    return resultado


def construir_filas_cuentas(reg_dict, terceros):
    """Construye filas {identificacion, nombre, tot_deb, tot_cre, neto} ordenadas por nombre."""
    rows = []
    for cod, d in reg_dict.items():
        ter = terceros.get(cod, {})
        rows.append({
            'identificacion': ter.get('identificacion', ''),
            'nombre':         ter.get('nombre', cod),
            'tot_deb':        round(d['tot_deb'], 2),
            'tot_cre':        round(d['tot_cre'], 2),
            'neto':           round(d['tot_deb'] - d['tot_cre'], 2),
        })
    rows = [r for r in rows if r['nombre']]
    rows.sort(key=lambda r: r['nombre'])
    return rows


def leer_empresas(ruta_bd):
    """Retorna lista de empresas únicas de PROD_FACT1."""
    path = os.path.join(ruta_bd, 'PROD_FACT1.DBF')
    with open(path, 'rb') as f:
        hdr = f.read(32)
        num = struct.unpack_from('<I', hdr, 4)[0]
        hsz = struct.unpack_from('<H', hdr, 8)[0]
        f.seek(hsz)
        raw = f.read(num * PF_REC_SIZE)
    empresas = set()
    n = len(raw) // PF_REC_SIZE
    for i in range(0, n, 50):  # muestra cada 50 registros para velocidad
        b = raw[i*PF_REC_SIZE:(i+1)*PF_REC_SIZE]
        if b[0] == 0x2A:
            continue
        emp = b[PF_EMPRESA:PF_EMPRESA+4].rstrip(b' ').decode(ENC, 'replace').strip()
        if emp:
            empresas.add(emp)
    return sorted(empresas)


# ════════════════════════════════════════════════════════════════════════════════
#  RUTAS FLASK
# ════════════════════════════════════════════════════════════════════════════════

@app.route('/')
def index():
    return '<a href="/ventas_clientes">Ventas por Clientes</a><br><a href="/consulta_cuentas">Consulta de Cuentas</a>'


@app.route('/consulta_cuentas')
def consulta_cuentas():
    hoy   = datetime.date.today()
    desde = hoy.replace(day=1).isoformat()
    hasta = hoy.isoformat()
    try:
        ruta_bd  = leer_ruta_bd()
        empresas = leer_empresas(ruta_bd)
    except Exception:
        empresas = []
    return render_template('adm_consulta_cuentas.html', desde=desde, hasta=hasta, empresas=empresas)


@app.route('/api/cuentas_autocomplete')
def api_cuentas_autocomplete():
    q = request.args.get('q', '').strip()
    if len(q) < 2:
        return jsonify({'ok': True, 'cuentas': []})
    try:
        ruta_bd = leer_ruta_bd()
        cuentas = leer_cuentas(ruta_bd, q)
        return jsonify({'ok': True, 'cuentas': cuentas})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e), 'cuentas': []})


@app.route('/api/consulta_cuentas')
def api_consulta_cuentas():
    cuenta  = request.args.get('cuenta', '').strip()
    empresa = request.args.get('empresa', '')
    desde_s = request.args.get('desde', '')
    hasta_s = request.args.get('hasta', '')
    if not cuenta:
        return jsonify({'ok': False, 'error': 'Seleccione una cuenta'})
    try:
        desde = datetime.date.fromisoformat(desde_s)
        hasta = datetime.date.fromisoformat(hasta_s)
    except Exception as e:
        return jsonify({'ok': False, 'error': f'Fecha inválida: {e}'})
    try:
        ruta_bd  = leer_ruta_bd()
        terceros = leer_terceros(ruta_bd)
        reg      = leer_reg_ctas(ruta_bd, cuenta, empresa, desde, hasta)
        rows     = construir_filas_cuentas(reg, terceros)
        total_deb = sum(r['tot_deb'] for r in rows)
        total_cre = sum(r['tot_cre'] for r in rows)
        return jsonify({
            'ok': True,
            'rows': rows,
            'totales': {
                'terceros': len(rows),
                'tot_deb':  round(total_deb, 2),
                'tot_cre':  round(total_cre, 2),
                'neto':     round(total_deb - total_cre, 2),
            }
        })
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})


@app.route('/api/consulta_cuentas/excel')
def api_consulta_cuentas_excel():
    if not HAS_OPENPYXL:
        return 'openpyxl no instalado', 500
    cuenta  = request.args.get('cuenta', '').strip()
    empresa = request.args.get('empresa', '')
    desde_s = request.args.get('desde', '')
    hasta_s = request.args.get('hasta', '')
    try:
        desde    = datetime.date.fromisoformat(desde_s)
        hasta    = datetime.date.fromisoformat(hasta_s)
        ruta_bd  = leer_ruta_bd()
        terceros = leer_terceros(ruta_bd)
        reg      = leer_reg_ctas(ruta_bd, cuenta, empresa, desde, hasta)
        rows     = construir_filas_cuentas(reg, terceros)
        total_deb = sum(r['tot_deb'] for r in rows)
        total_cre = sum(r['tot_cre'] for r in rows)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Consulta de Cuentas'
        hdr_font = Font(bold=True, color='003F7F')
        hdr_fill = PatternFill('solid', fgColor='DDEEFF')
        tot_font = Font(bold=True)

        ws.append([f'CONSULTA CUENTA: {cuenta}'])
        ws['A1'].font = Font(bold=True, size=13)
        ws.append([f'Desde: {desde}    Hasta: {hasta}    Empresa: {empresa or "Todas"}'])
        ws.append([])
        cabecera = ['IDENTIFICACION', 'NOMBRE', 'DÉBITO', 'CRÉDITO', 'NETO']
        ws.append(cabecera)
        for cell in ws[ws.max_row]:
            cell.font = hdr_font
            cell.fill = hdr_fill

        for r in rows:
            ws.append([r['identificacion'], r['nombre'], r['tot_deb'], r['tot_cre'], r['neto']])
            fila = ws[ws.max_row]
            for ci in [2, 3, 4]:
                fila[ci].number_format = '#,##0.00'

        ws.append([])
        ws.append(['', 'TOTALES', round(total_deb,2), round(total_cre,2), round(total_deb-total_cre,2)])
        for c in ws[ws.max_row]: c.font = tot_font
        ws.append(['', f'Generado: {datetime.datetime.now().strftime("%Y-%m-%d %H:%M")}'])

        for i, w in enumerate([18, 42, 14, 14, 14], 1):
            ws.column_dimensions[ws.cell(1, i).column_letter].width = w

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        fname = f'consulta_cuenta_{cuenta}_{desde}_{hasta}.xlsx'
        return send_file(buf, as_attachment=True, download_name=fname,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        return str(e), 500


@app.route('/ventas_clientes')
def ventas_clientes():
    hoy    = datetime.date.today()
    desde  = hoy.replace(day=1).isoformat()
    hasta  = hoy.isoformat()
    try:
        ruta_bd  = leer_ruta_bd()
        empresas = leer_empresas(ruta_bd)
    except Exception:
        empresas = []
    return render_template('adm_ventas_clientes.html',
                           desde=desde, hasta=hasta, empresas=empresas)


@app.route('/api/ventas_clientes')
def api_ventas_clientes():
    desde_s = request.args.get('desde', '')
    hasta_s = request.args.get('hasta', '')
    empresa = request.args.get('empresa', '')
    try:
        desde = datetime.date.fromisoformat(desde_s)
        hasta = datetime.date.fromisoformat(hasta_s)
    except Exception as e:
        return jsonify({'ok': False, 'error': f'Fecha inválida: {e}'})
    try:
        ruta_bd  = leer_ruta_bd()
        terceros = leer_terceros(ruta_bd)
        clientes = leer_ventas_clientes(ruta_bd, desde, hasta, empresa)
        rows     = calcular_puestos(clientes, terceros)
        total_clientes = len(rows)
        total_cant     = sum(r['cantidad'] for r in rows)
        total_val      = sum(r['subtotal'] for r in rows)
        total_iva      = sum(r['iva'] for r in rows)
        return jsonify({
            'ok': True,
            'rows': rows,
            'totales': {
                'clientes':      total_clientes,
                'cantidad':      round(total_cant, 2),
                'subtotal':      round(total_val, 2),
                'iva':           round(total_iva, 2),
                'prom_cliente':  round(total_val / total_clientes, 2) if total_clientes else 0,
                'prom_producto': round(total_val / total_cant, 2) if total_cant else 0,
            }
        })
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})


@app.route('/api/ventas_clientes/excel')
def api_ventas_clientes_excel():
    if not HAS_OPENPYXL:
        return 'openpyxl no instalado — pip install openpyxl', 500
    desde_s = request.args.get('desde', '')
    hasta_s = request.args.get('hasta', '')
    empresa = request.args.get('empresa', '')
    try:
        desde    = datetime.date.fromisoformat(desde_s)
        hasta    = datetime.date.fromisoformat(hasta_s)
        ruta_bd  = leer_ruta_bd()
        terceros = leer_terceros(ruta_bd)
        clientes = leer_ventas_clientes(ruta_bd, desde, hasta, empresa)
        rows     = calcular_puestos(clientes, terceros)
        total_clientes = len(rows)
        total_cant     = sum(r['cantidad'] for r in rows)
        total_val      = sum(r['subtotal'] for r in rows)
        total_iva      = sum(r['iva'] for r in rows)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Ventas por Clientes'

        hdr_font = Font(bold=True, color='003F7F')
        hdr_fill = PatternFill('solid', fgColor='DDEEFF')
        tot_font = Font(bold=True)

        ws.append(['VENTAS POR CLIENTES'])
        ws['A1'].font = Font(bold=True, size=13)
        ws.append([f'Desde: {desde}    Hasta: {hasta}    Empresa: {empresa or "Todas"}'])
        ws.append([])

        cabecera = ['IDENTIFICACION', 'NOMBRE', 'CANTIDAD', '% CANT', 'PUESTO CANT',
                    'SUBTOTAL', '% VAL', 'PUESTO VAL']
        ws.append(cabecera)
        for cell in ws[ws.max_row]:
            cell.font = hdr_font
            cell.fill = hdr_fill

        for r in rows:
            ws.append([
                r['identificacion'], r['nombre'],
                r['cantidad'], r['por_cant'] / 100, r['puesto_cant'],
                r['subtotal'],  r['por_val'] / 100,  r['puesto_val'],
            ])
            fila = ws[ws.max_row]
            fila[2].number_format = '#,##0.00'   # cantidad
            fila[3].number_format = '0.00%'       # % cant
            fila[5].number_format = '#,##0.00'   # subtotal
            fila[6].number_format = '0.00%'       # % val

        ws.append([])
        ws.append(['', 'TOTAL VENTAS', round(total_cant, 2), '', '', round(total_val, 2)])
        for c in ws[ws.max_row]: c.font = tot_font
        ws.append(['', 'CANTIDAD DE CLIENTES', total_clientes])
        ws.append(['', 'PROMEDIO POR CLIENTE', '', '', '', round(total_val/total_clientes, 2) if total_clientes else 0])
        ws.append(['', 'PROMEDIO POR PRODUCTO', round(total_val/total_cant, 2) if total_cant else 0])
        ws.append(['', 'TOTAL IVA', '', '', '', round(total_iva, 2)])
        ws.append(['', f'Generado: {datetime.datetime.now().strftime("%Y-%m-%d %H:%M")}'])

        for i, w in enumerate([18, 42, 12, 10, 12, 16, 10, 12], 1):
            ws.column_dimensions[ws.cell(1, i).column_letter].width = w

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        fname = f'ventas_clientes_{desde}_{hasta}.xlsx'
        return send_file(buf, as_attachment=True, download_name=fname,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        return str(e), 500


def _reportar_arranque():
    """POST al servidor (ngrok) cuando administrator_web arranca. Lee URL de admin_agent.ini."""
    try:
        import configparser, requests, socket as _s
        from pathlib import Path as _P
        cfg = configparser.ConfigParser()
        ini = _P(__file__).parent / 'admin_agent.ini'
        cfg.read(str(ini), encoding='utf-8')
        servidor = cfg.get('agent', 'servidor', fallback='').strip().rstrip('/')
        if not servidor:
            return
        try:
            ip = _s.gethostbyname(_s.gethostname())
        except Exception:
            ip = ''
        requests.post(
            f'{servidor}/api/admin-agent/reporte',
            json={
                'cliente_id': 'pilar',
                'tipo':       'arranque_web',
                'estado':     'ok',
                'detalle':    'administrator_web.py arrancó en localhost:5002',
                'ip':         ip,
            },
            timeout=10,
        )
    except Exception:
        pass


if __name__ == '__main__':
    import socket as _socket
    # El reloader de Flask arranca dos procesos; el lock solo aplica al padre
    if os.environ.get('WERKZEUG_RUN_MAIN') != 'true':
        _lock = _socket.socket(_socket.AF_INET, _socket.SOCK_STREAM)
        try:
            _lock.bind(('127.0.0.1', 47836))
            _lock.listen(1)
        except OSError:
            sys.exit(0)
        print('=' * 55)
        print('  Administrator Web — localhost:5002')
        print('=' * 55)
        _reportar_arranque()
    app.run(host='127.0.0.1', port=5002, debug=True, use_reloader=True)
